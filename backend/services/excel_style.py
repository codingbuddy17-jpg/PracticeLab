"""
One look for every workbook this application produces.

Seven writers grew independently and three header colours came with them:
#4F46E5 on the assessment and auditor exports, #1E3A5F on two more, and
#1F3864 on the answer-key templates — the last two being near-identical navies
that differ by a digit, which is drift rather than choice. One export had no
formatting at all.

The palette here is the one the Word handover documents use, so a workbook and
a document from this application look like they came from the same place.

Use `header_row` on the row that holds the column names and `fit_columns`
before saving. Both are safe to call twice.
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# The house palette. NAVY and ACCENT are lifted from docs/_spec_docx_build.js
# so the two families of deliverable match.
NAVY = "1F3864"
ACCENT = "2E75B6"
GREY = "595959"
INPUT_YELLOW = "FFFDE7"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=ACCENT)
INPUT_FILL = PatternFill("solid", fgColor=INPUT_YELLOW)

WHITE_FONT = Font(color="FFFFFF", bold=True, size=10)
DARK_FONT = Font(bold=True, size=10)

THIN = Side(style="thin", color="D9D9D9")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Semantic cell fills. These are NOT drift — a passing row and a failing row
# should not look alike — so they stay distinct from the header palette.
PASS_FILL = PatternFill("solid", fgColor="ECFDF5")
FAIL_FILL = PatternFill("solid", fgColor="FFF1F2")
PASS_FONT = Font(color="166534", bold=True, size=10)
FAIL_FONT = Font(color="991B1B", bold=True, size=10)


def header_row(ws, row: int = 1, *, fill=None, freeze: bool = True,
               autofilter: bool = True) -> None:
    """
    Style one row as the column headings, and make it behave like one.

    Freezing matters more than the colour: a heading that scrolls away leaves a
    reader counting columns to work out what they are looking at, and every
    export here is long enough for that to happen. The filter is what makes a
    workbook usable rather than merely readable.
    """
    fill = fill or HEADER_FILL
    last_col = ws.max_column
    if not last_col:
        return
    for col in range(1, last_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = WHITE_FONT
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    ws.row_dimensions[row].height = 26
    if freeze:
        ws.freeze_panes = ws.cell(row=row + 1, column=1)
    if autofilter and ws.max_row > row:
        ws.auto_filter.ref = "A%d:%s%d" % (
            row, get_column_letter(last_col), ws.max_row)


def fit_columns(ws, *, minimum: int = 10, maximum: int = 52) -> None:
    """
    Width from the content, bounded at both ends.

    Unbounded, one long rationale or a pasted note makes a column wider than
    the screen and pushes everything else out of view — which is how a
    perfectly correct export becomes unreadable.
    """
    for col in range(1, ws.max_column + 1):
        longest = 0
        for cell in ws[get_column_letter(col)]:
            v = cell.value
            if v is None:
                continue
            longest = max(longest, max(len(s) for s in str(v).split("\n")))
        ws.column_dimensions[get_column_letter(col)].width = max(
            minimum, min(maximum, longest + 2))


def finish(ws, *, header: int = 1, **kw) -> None:
    """Both, in the order they need to happen. The usual call."""
    fit_columns(ws)
    header_row(ws, header, **kw)
