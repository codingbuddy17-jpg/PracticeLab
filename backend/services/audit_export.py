"""
Excel exports for the auditor module.

Every sheet carries its denominator. A rate on two opportunities and a rate on
twenty look identical in a spreadsheet cell, and once a figure has been pasted
into a deck nobody can tell which they are reading — so the counts travel with
the percentages rather than being left behind on the screen they came from.

NA is written as the literal text "NA", never as 0 or a blank. A cohort that
has never met a spurious code has no Delete accuracy, and both alternatives say
something false about them.
"""

import io
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=12)
NOTE_FONT = Font(italic=True, size=9, color="6B7280")


def _na(value: Optional[float]) -> Any:
    """None means the opportunity never existed — say so rather than imply zero."""
    return "NA" if value is None else value


def _sheet(wb: Workbook, title: str, headers: list[str], rows: list[list],
           note: Optional[str] = None):
    ws = wb.create_sheet(title[:31])
    r = 1
    if note:
        ws.cell(row=1, column=1, value=note).font = NOTE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
        r = 3
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, row in enumerate(rows, start=r + 1):
        for c, v in enumerate(row, start=1):
            ws.cell(row=i, column=c, value=v)
    for c, h in enumerate(headers, start=1):
        width = max(len(str(h)) + 2,
                    *(len(str(row[c - 1])) + 2 for row in rows)) if rows else len(str(h)) + 2
        ws.column_dimensions[get_column_letter(c)].width = min(max(width, 10), 42)
    ws.freeze_panes = ws.cell(row=r + 1, column=1)
    return ws


def _finish(wb: Workbook) -> bytes:
    if wb.active and wb.active.max_row == 1 and wb.active.max_column == 1 \
            and wb.active["A1"].value is None:
        wb.remove(wb.active)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── batch results ────────────────────────────────────────────────────────────

def export_batch_results(batch_name: str, charts: list[dict],
                         summary: Optional[dict] = None) -> bytes:
    """
    One row per audited chart, plus the findings behind them.

    Chart type is included because this is a trainer's export — the auditor
    never sees it, but whoever reads this needs to know which charts had
    nothing to find.
    """
    wb = Workbook()

    if summary:
        _sheet(wb, "Summary",
               ["Measure", "Value", "Basis"],
               [
                   ["Audit accuracy", _na(summary.get("audit_accuracy")),
                    "average of chart scores"],
                   ["Clean charts", summary.get("clean_charts", 0), "no errors introduced"],
                   ["Clean accuracy", _na(summary.get("clean_accuracy")), "restraint"],
                   ["Opportunity charts", summary.get("opportunity_charts", 0), ""],
                   ["Opportunity accuracy", _na(summary.get("opportunity_accuracy")),
                    "detection"],
                   ["Add accuracy", _na(summary.get("add_accuracy")),
                    f"{summary.get('add_found', 0)}/{summary.get('add_planted', 0)} pooled"],
                   ["Revise accuracy", _na(summary.get("revise_accuracy")),
                    f"{summary.get('revise_found', 0)}/{summary.get('revise_planted', 0)} pooled"],
                   ["Delete accuracy", _na(summary.get("delete_accuracy")),
                    f"{summary.get('delete_found', 0)}/{summary.get('delete_planted', 0)} pooled"],
                   ["DRG-impacting accuracy", _na(summary.get("drg_accuracy")),
                    f"{summary.get('drg_found', 0)}/{summary.get('drg_planted', 0)} pooled"],
                   ["Query accuracy", _na(summary.get("query_accuracy")),
                    f"{summary.get('query_correct', 0)}/{summary.get('query_charts', 0)} charts"],
                   ["Findings on correct items", summary.get("over_calls", 0), ""],
                   ["Found but corrected wrongly", summary.get("detected_not_corrected", 0),
                    "reported, not scored"],
                   ["Opportunities", summary.get("opportunities", 0), ""],
                   ["Verdict", summary.get("pass_fail") or "withheld",
                    summary.get("verdict_withheld_reason") or ""],
               ],
               note=f"Audit batch: {batch_name}")

    _sheet(wb, "Chart_Results",
           ["Auditor", "Emp ID", "Chart", "Chart type", "Audit accuracy %",
            "Add found", "Add total", "Add %",
            "Revise found", "Revise total", "Revise %",
            "Delete found", "Delete total", "Delete %",
            "DRG found", "DRG total",
            "Over-calls", "Over-call tier", "Deduction %",
            "Query expected", "Query raised", "Query correct",
            "Detected not corrected"],
           [[
               c.get("auditor_name"), c.get("emp_id"), c.get("chart_number"),
               "Clean" if c.get("is_clean") else "Opportunity",
               c.get("audit_accuracy"),
               c.get("add_found"), c.get("add_planted"), _na(c.get("add_accuracy")),
               c.get("revise_found"), c.get("revise_planted"), _na(c.get("revise_accuracy")),
               c.get("delete_found"), c.get("delete_planted"), _na(c.get("delete_accuracy")),
               c.get("drg_impacting_found"), c.get("drg_impacting_planted"),
               c.get("over_calls"), c.get("over_call_tier") or "",
               c.get("over_call_deduction"),
               _bool(c.get("query_expected")), _bool(c.get("query_flagged")),
               _bool(c.get("query_correct")),
               c.get("detected_not_corrected"),
           ] for c in charts],
           note="Accuracy columns show found and total beside the percentage. "
                "NA means nothing of that kind was in the chart.")

    detail = []
    for c in charts:
        for entry in (c.get("outcomes") or []):
            p = entry.get("planting") or {}
            detail.append([
                c.get("auditor_name"), c.get("chart_number"),
                p.get("section"), p.get("action"), p.get("field") or "code",
                (p.get("line") + 1) if isinstance(p.get("line"), int) else "",
                p.get("claim_value") or "", p.get("correct_value") or "",
                entry.get("outcome"),
                "Yes" if p.get("drg_impacting") else "No",
                "Real coder mistake" if p.get("origin") == "observed" else "Generated",
                p.get("observed_coders") or "",
            ])
    _sheet(wb, "Findings_Detail",
           ["Auditor", "Chart", "Section", "Action", "Field", "Line",
            "On the claim", "Correct value", "Outcome", "DRG impact",
            "Source", "Coders who made it"],
           detail,
           note="One row per error introduced, and what the auditor did about it.")

    return _finish(wb)


def _bool(v) -> str:
    return "" if v is None else ("Yes" if v else "No")


# ── analytics ────────────────────────────────────────────────────────────────

def export_analytics(overview: dict, batches: list[dict], auditors: list[dict],
                     detection: dict, specialties: Optional[list[dict]] = None,
                     chart_signals: Optional[list[dict]] = None) -> bytes:
    """The analytics tabs, one sheet each."""
    wb = Workbook()

    _sheet(wb, "Overview", ["Measure", "Value", "Basis"],
           [
               ["Charts scored", overview.get("charts", 0), ""],
               ["Auditors", overview.get("auditors", 0), ""],
               ["Batches", overview.get("batches", 0), ""],
               ["Audit Score", _na(overview.get("audit_score")),
                "weighted score used for verdict"],
               ["Review Score", _na(overview.get("review_score")),
                overview.get("review_basis", "")],
               ["Error Detection Rate", _na(overview.get("audit_accuracy")),
                overview.get("audit_accuracy_basis", "")],
               ["Clean Chart Score", _na(overview.get("clean_accuracy")),
                f"{overview.get('clean_charts', 0)} charts — restraint"],
               ["Opportunity Chart Score", _na(overview.get("opportunity_accuracy")),
                f"{overview.get('opportunity_charts', 0)} charts — detection"],
               ["Add accuracy", _na((overview.get("add") or {}).get("accuracy")),
                _pooled(overview.get("add"))],
               ["Revise accuracy", _na((overview.get("revise") or {}).get("accuracy")),
                _pooled(overview.get("revise"))],
               ["Delete accuracy", _na((overview.get("delete") or {}).get("accuracy")),
                _pooled(overview.get("delete"))],
               ["DRG-impacting accuracy", _na(overview.get("drg_accuracy")),
                f"{overview.get('drg_found', 0)}/{overview.get('drg_planted', 0)}"],
               ["Query accuracy", _na(overview.get("query_accuracy")),
                f"{overview.get('query_correct', 0)}/{overview.get('query_charts', 0)} charts"],
               ["Findings on correct items", overview.get("over_calls", 0), ""],
               ["Found but corrected wrongly", overview.get("detected_not_corrected", 0),
                "reported, not scored"],
               ["Opportunities", overview.get("opportunities", 0), ""],
           ],
           note="Audit Score is weighted. Error Detection Rate and Review Score "
                "are shown separately because they answer different questions.")

    cols = ["Charts", "Audit Score", "Review Score", "Error Detection Rate",
            "Clean Chart Score", "Opportunity Chart Score",
            "Add found", "Add total", "Add %",
            "Revise found", "Revise total", "Revise %",
            "Delete found", "Delete total", "Delete %",
            "DRG %", "Over-calls", "Opportunities", "Verdict"]

    def _row(r: dict) -> list:
        return [
            r.get("charts"), _na(r.get("audit_score")), _na(r.get("review_score")),
            _na(r.get("audit_accuracy")),
            _na(r.get("clean_accuracy")), _na(r.get("opportunity_accuracy")),
            (r.get("add") or {}).get("found"), (r.get("add") or {}).get("planted"),
            _na((r.get("add") or {}).get("accuracy")),
            (r.get("revise") or {}).get("found"), (r.get("revise") or {}).get("planted"),
            _na((r.get("revise") or {}).get("accuracy")),
            (r.get("delete") or {}).get("found"), (r.get("delete") or {}).get("planted"),
            _na((r.get("delete") or {}).get("accuracy")),
            _na(r.get("drg_accuracy")), r.get("over_calls"), r.get("opportunities"),
            r.get("pass_fail") or "withheld",
        ]

    if specialties is not None:
        _sheet(wb, "By_Specialty", ["Specialty", "Auditors", "Batches"] + cols,
               [[s.get("specialty"), s.get("auditors"), s.get("batches")] + _row(s)
                for s in specialties],
               note="Weakest first. Specialties with no scored audit results do not appear.")

    _sheet(wb, "By_Batch", ["Batch", "Specialty", "Status", "Auditors"] + cols,
           [[b.get("name"), b.get("specialty"), b.get("status"), b.get("auditors")]
            + _row(b) for b in batches])

    _sheet(wb, "By_Auditor", ["Auditor", "Emp ID", "Batches"] + cols,
           [[a.get("auditor_name"), a.get("emp_id"), a.get("batches")] + _row(a)
            for a in auditors],
           note="Weakest first. Component accuracies pool across every batch the "
                "auditor appears in.")

    _sheet(wb, "Detection_Patterns",
           ["Kind of error", "Introduced", "Caught", "Missed",
            "Corrected wrongly", "Caught %"],
           [[k.get("label"), k.get("planted"), k.get("found"), k.get("missed"),
             k.get("detected_not_corrected"), _na(k.get("accuracy"))]
            for k in (detection.get("by_kind") or [])],
           note="Worst first — which errors a cohort cannot see. Only kinds "
                f"introduced at least {detection.get('min_for_pattern', 5)} times "
                "are worth acting on.")

    _sheet(wb, "Real_vs_Generated",
           ["Source", "Introduced", "Caught", "Missed", "Caught %"],
           [[o.get("label"), o.get("planted"), o.get("found"), o.get("missed"),
             _na(o.get("accuracy"))]
            for o in (detection.get("by_origin") or [])],
           note="Errors taken from real coder submissions against ones the system "
                "generated. Only the first describes the job.")

    if chart_signals is not None:
        _sheet(wb, "Chart_Signals",
               ["Chart", "Specialty", "Category", "Attempts", "Audit %",
                "Clean", "Opportunity", "Opportunities", "Missed", "Over-calls",
                "Found but corrected wrongly", "Signal"],
               [[c.get("chart_number"), c.get("specialty"), c.get("category"),
                 c.get("attempts"), _na(c.get("audit_accuracy")),
                 c.get("clean_charts"), c.get("opportunity_charts"),
                 c.get("opportunities"), c.get("missed"), c.get("over_calls"),
                 c.get("detected_not_corrected"), c.get("signal")]
                for c in chart_signals],
               note="Chart/key QA view, weakest first.")

    return _finish(wb)


def _pooled(cell: Optional[dict]) -> str:
    cell = cell or {}
    return f"{cell.get('found', 0)}/{cell.get('planted', 0)} pooled"


# ── the error library ────────────────────────────────────────────────────────

def export_key_sets(sets: list[dict]) -> bytes:
    """Every authored error set, one row per error."""
    wb = Workbook()
    rows = []
    for k in sets:
        for m in (k.get("mutations") or []):
            rows.append([
                k.get("chart_number"), k.get("name"),
                m.get("section"), m.get("action"), m.get("field") or "code",
                (m.get("line") + 1) if isinstance(m.get("line"), int) else "",
                m.get("claim_value") or "", m.get("correct_value") or "",
                _bool(k.get("query_expected")), k.get("query_rationale") or "",
                "Yes" if k.get("always_plant") else "No",
                k.get("authored_by"),
            ])
        if not (k.get("mutations") or []):
            rows.append([
                k.get("chart_number"), k.get("name"), "", "", "", "", "", "",
                _bool(k.get("query_expected")), k.get("query_rationale") or "",
                "Yes" if k.get("always_plant") else "No", k.get("authored_by"),
            ])
    _sheet(wb, "Audit_Keys",
           ["Chart", "Set", "Section", "Action", "Field", "Line",
            "Shown on the claim", "Correct value", "Query expected",
            "Query rationale", "Always used", "Authored by"],
           rows,
           note="What the auditor must find. A set with no error rows is a correct "
                "claim that still needs a query.")
    return _finish(wb)
