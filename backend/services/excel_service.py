"""
Excel generation and parsing for PracticeLab.

Trainers fill answer keys in Excel; coders do not. Coding is done in the
PracticeLab interface, which is the only route a submission takes. The offline
route — a workbook per coder, filled and emailed back for grading — was removed
along with the workbook builder, its ZIP packer and its parser: it had no
endpoint behind it, and it predated diagnosis pointers, ED Single Path levels
and Dx-only specialties, so anyone who switched it back on would have graded
Surgery and ED Profee coders wrong from the first batch.
"""
import io
from typing import Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from services.grading_engine import canonical_pointers, norm_units


# ── Style helpers ─────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
INPUT_FILL = PatternFill("solid", fgColor="FFFDE7")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=10)
DARK_FONT = Font(bold=True, size=10)
THIN = Side(style="thin", color="BDBDBD")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header(ws, col, row, value, fill=None, font=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill or HEADER_FILL
    cell.font = font or WHITE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    return cell


def _input_cell(ws, col, row, value=None):
    # Use value=None (not "") so openpyxl writes a plain styled empty cell.
    # value="" creates t="inlineStr" cells which Excel for Mac rewrites in a
    # format openpyxl cannot read back, causing all filled values to read as None.
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell


# ── IP Answer Key template ────────────────────────────────────────────────────

IP_SDX_COUNT = 20
IP_PCS_COUNT = 8
OP_SDX_COUNT = 20
OP_CPT_COUNT = 10


def _build_ip_ak_headers(ws):
    """Write IP answer key header row. Returns column map."""
    col = 1
    _header(ws, col, 1, "Chart_Number"); col += 1
    _header(ws, col, 1, "PDx_Code"); col += 1
    _header(ws, col, 1, "PDx_POA\n(Y/N/U/W/1)"); col += 1
    for i in range(1, IP_SDX_COUNT + 1):
        _header(ws, col, 1, f"SDx_{i}"); col += 1
        _header(ws, col, 1, f"SDx_{i}_POA"); col += 1
        _header(ws, col, 1, f"SDx_{i}_CCMCC\n(MCC/CC/-)"); col += 1
    for i in range(1, IP_PCS_COUNT + 1):
        _header(ws, col, 1, f"PCS_{i}"); col += 1
    ws.row_dimensions[1].height = 36


def _build_op_ak_headers(ws, with_pointers: bool = False, single_path: bool = False,
                         dx_only: bool = False, with_units: bool = False):
    col = 1
    _header(ws, col, 1, "Chart_Number"); col += 1
    if single_path:
        # ED Single Path: both levels are coded from one chart and often differ
        _header(ws, col, 1, "Facility_ED_Level"); col += 1
        _header(ws, col, 1, "Profee_ED_Level"); col += 1
    _header(ws, col, 1, "PDx_Code"); col += 1
    for i in range(1, OP_SDX_COUNT + 1):
        _header(ws, col, 1, f"SDx_{i}"); col += 1
    # Ancillary/radiology is diagnosis-only in practice — CPTs are auto-coded
    # upstream — so offering CPT columns invites trainers to fill cells that
    # cannot score.
    for i in range(1, 0 if dx_only else OP_CPT_COUNT + 1):
        _header(ws, col, 1, f"CPT_{i}"); col += 1
        _header(ws, col, 1, f"CPT_{i}_Modifier"); col += 1
        # Units. Left blank the line is one unit, which is what a coder writing
        # a single procedure does — so a key that says nothing about units does
        # not grade them, and existing keys are unaffected. Rubric-graded work
        # (Edits, Denials) gets no column: a cell that cannot score is worse
        # than no cell.
        if with_units:
            _header(ws, col, 1, f"CPT_{i}_Units")
            ws.cell(1, col).comment = Comment(
                "How many units of this procedure. Leave blank for a single unit.\n"
                "Fill it only where the count matters — bilateral procedures, "
                "add-on codes billed more than once.\n\n"
                "A key that leaves this blank does not grade units at all.",
                "PracticeLab")
            col += 1
        if with_pointers:
            # Professional claims (CMS-1500 Box 24E): which Dx justify this line.
            # NUMBERS index the Dx list, as coders refer to them — 1 = PDx,
            # 2 = SDx_1, 3 = SDx_2 ... Letters are still accepted on upload so
            # keys written before the switch keep grading.
            _header(ws, col, 1, f"CPT_{i}_DxPointers")
            ws.cell(1, col).comment = Comment(
                "Which diagnoses justify this line, by NUMBER: 1 = PDx, "
                "2 = SDx_1, 3 = SDx_2 ...\nUp to 4 per line, first is primary. "
                "Example: 1,2\n\nLetters (A,B) from older keys are still accepted.",
                "PracticeLab")
            col += 1
    ws.row_dimensions[1].height = 36


def generate_coder_list_template() -> bytes:
    """Blank coder list template: Coder_Name | Emp_ID columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Coder_List"

    _header(ws, 1, 1, "Coder_Name")
    _header(ws, 2, 1, "Emp_ID")
    ws.row_dimensions[1].height = 28

    info = [
        "Enter one coder per row.",
        "Coder_Name: Enter exactly as you want it to appear on reports (e.g. Smith, John A  or  Sarah Johnson).",
        "Emp_ID: Required. Used as the unique identifier for trend analytics across batches.",
        "Duplicate Emp_IDs in this file will be skipped (first row kept).",
    ]
    for r in range(2, 7):
        _input_cell(ws, 1, r)
        _input_cell(ws, 2, r)

    note = wb.create_sheet("Instructions")
    for i, line in enumerate(info, 1):
        c = note.cell(row=i, column=1, value=line)
        c.font = Font(size=11)
    note.column_dimensions["A"].width = 90

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_coder_list(file_bytes: bytes) -> list[dict]:
    """
    Parse uploaded coder list Excel.
    Returns [{name, emp_id}] — deduped by emp_id (first occurrence wins).
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    seen_ids: set[str] = set()
    coders = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or (not row[0] and not row[1]):
            continue
        name = str(row[0] or "").strip()
        emp_id = str(row[1] or "").strip().upper()
        if not name or not emp_id:
            continue
        if emp_id in seen_ids:
            continue
        seen_ids.add(emp_id)
        coders.append({"name": name, "emp_id": emp_id})

    return coders


def generate_answer_key_template(specialty: str, with_pointers: bool = False,
                                 single_path: bool = False,
                                 dx_only: bool = False,
                                 with_units: bool = False) -> bytes:
    """
    Returns bytes of blank answer key Excel file.
    specialty: 'IP' for inpatient, 'OP' for outpatient.
    with_pointers: add a Dx-pointer column per CPT line (professional claims).
    """
    wb = Workbook()
    ws = wb.active
    is_ip = specialty.upper() == "IP"
    ws.title = f"{'IP' if is_ip else 'OP'}_Answer_Key"

    if is_ip:
        _build_ip_ak_headers(ws)
        total_cols = 3 + IP_SDX_COUNT * 3 + IP_PCS_COUNT
    else:
        _build_op_ak_headers(ws, with_pointers=with_pointers, single_path=single_path,
                             dx_only=dx_only, with_units=with_units)
        # Per CPT line: code + modifier, plus one column each for units and
        # pointers where that specialty has them. Getting this wrong leaves the
        # tail columns unstyled and the widths short.
        per_cpt = 2 + (1 if with_units else 0) + (1 if with_pointers else 0)
        cpt_cols = 0 if dx_only else OP_CPT_COUNT * per_cpt
        total_cols = 2 + (2 if single_path else 0) + OP_SDX_COUNT + cpt_cols

    # 10 blank input rows
    for r in range(2, 12):
        for c in range(1, total_cols + 1):
            _input_cell(ws, c, r)

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    for c in range(4, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # Instructions sheet
    info = wb.create_sheet("Instructions")
    instructions = [
        ("PracticeLab Answer Key Template", True),
        ("", False),
        ("HOW TO USE:", True),
        ("1. Fill one row per chart. Chart_Number must match exactly what is in PracticeLab.", False),
        ("2. PDx_Code: Principal diagnosis ICD-10-CM code (e.g. J18.9)", False),
    ]
    if is_ip:
        instructions += [
            ("3. PDx_POA: Present on Admission indicator — Y, N, U, W, or 1", False),
            ("4. SDx: Secondary diagnoses. Fill from left, leave unused columns blank.", False),
            ("5. SDx_POA: POA indicator per secondary diagnosis.", False),
            ("6. SDx_CCMCC: Enter MCC, CC, or - (dash for neither).", False),
            ("7. PCS: ICD-10-PCS procedure codes (7 characters). Fill from left.", False),
            ("8. Do NOT enter column headers or extra rows.", False),
            ("", False),
            ("SCORING REFERENCE (IP-DRG):", True),
            ("  PDx (code + POA): 20 pts", False),
            ("  SDx: 20 pts  |  PCS: 20 pts  |  DRG (trainer review): 40 pts", False),
            ("  Pass threshold: 80%", False),
        ]
    else:
        instructions += [
            ("3. SDx: Secondary diagnoses (code only, no POA for OP).", False),
            ("4. CPT: Procedure codes. CPT_Modifier is optional (e.g. 59, LT, RT).", False),
            ("5. Do NOT enter column headers or extra rows.", False),
            ("", False),
            ("SCORING REFERENCE (OP):", True),
            ("  PDx: 25 pts  |  SDx: 25 pts  |  CPT (code + modifier): 50 pts", False),
            ("  Pass threshold: 90%", False),
        ]
    instructions += [
        ("", False),
        ("Upload this completed file via PracticeLab → Answer Keys → Upload.", False),
    ]
    for r, (text, bold) in enumerate(instructions, 1):
        cell = info.cell(row=r, column=1, value=text)
        cell.font = Font(bold=bold, size=11)
    info.column_dimensions["A"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Consolidated answer key export ───────────────────────────────────────────

def export_all_answer_keys(answer_keys: list) -> bytes:
    """
    Export all stored answer keys as a filled Excel workbook.
    answer_keys: list of (AnswerKey, Chart) tuples from DB query.

    ONE SHEET PER SPECIALTY, each using that specialty's own upload layout.
    Grouping everything into a generic "OP" sheet dropped the columns that only
    some specialties have — Surgery lost its Dx pointers, ED Single Path lost
    both level columns — so an export → edit → re-upload round trip silently
    discarded them.
    """
    from routers.practicelab_pkg.shared import (
        _is_ip, _uses_pointers, _is_single_path, _is_dx_only, _uses_units,
    )

    wb = Workbook()
    wb.remove(wb.active)

    groups: dict = {}
    for ak, ch in answer_keys:
        groups.setdefault(ch.specialty, []).append((ak, ch))

    def _fill_sheet(ws, rows, spec):
        is_ip = _is_ip(spec)
        with_pointers = _uses_pointers(spec)
        single_path = _is_single_path(spec)
        dx_only = _is_dx_only(spec)
        with_units = _uses_units(spec)

        if is_ip:
            _build_ip_ak_headers(ws)
        else:
            _build_op_ak_headers(ws, with_pointers=with_pointers,
                                 single_path=single_path, dx_only=dx_only,
                                 with_units=with_units)
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14

        for row_num, (ak, ch) in enumerate(
                sorted(rows, key=lambda x: x[1].chart_number), start=2):
            col = 1
            _input_cell(ws, col, row_num, ch.chart_number); col += 1

            if is_ip:
                _input_cell(ws, col, row_num, ak.pdx_code or ""); col += 1
                _input_cell(ws, col, row_num, ak.pdx_poa or ""); col += 1
                sdx_list = ak.sdx or []
                for i in range(IP_SDX_COUNT):
                    entry = sdx_list[i] if i < len(sdx_list) else {}
                    _input_cell(ws, col, row_num, entry.get("code", "")); col += 1
                    _input_cell(ws, col, row_num, entry.get("poa", "")); col += 1
                    _input_cell(ws, col, row_num, entry.get("ccmcc", "")); col += 1
                pcs_list = ak.pcs or []
                for i in range(IP_PCS_COUNT):
                    entry = pcs_list[i] if i < len(pcs_list) else {}
                    _input_cell(ws, col, row_num, entry.get("code", "")); col += 1
                continue

            if single_path:
                _input_cell(ws, col, row_num, ak.facility_level or ""); col += 1
                _input_cell(ws, col, row_num, ak.profee_level or ""); col += 1

            _input_cell(ws, col, row_num, ak.pdx_code or ""); col += 1
            sdx_list = ak.sdx or []
            for i in range(OP_SDX_COUNT):
                entry = sdx_list[i] if i < len(sdx_list) else {}
                _input_cell(ws, col, row_num, entry.get("code", "")); col += 1

            if dx_only:
                continue
            cpt_list = ak.cpt or []
            for i in range(OP_CPT_COUNT):
                entry = cpt_list[i] if i < len(cpt_list) else {}
                _input_cell(ws, col, row_num, entry.get("code", "")); col += 1
                _input_cell(ws, col, row_num, entry.get("modifier", "")); col += 1
                if with_units:
                    # Blank where the key never stated units, so a round-trip
                    # export → re-upload does not invent a claim it did not make.
                    _input_cell(ws, col, row_num, entry.get("units", "")); col += 1
                if with_pointers:
                    _input_cell(ws, col, row_num,
                                ",".join(entry.get("pointers", []) or [])); col += 1

    for spec in sorted(groups, key=lambda s: s.value):
        # Sheet names cannot contain / and are capped at 31 chars
        title = spec.value.replace("/", "-")[:31]
        _fill_sheet(wb.create_sheet(title), groups[spec], spec)

    if not wb.sheetnames:
        ws = wb.create_sheet("No Data")
        ws.cell(1, 1, "No answer keys found.")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Coder answer sheet ────────────────────────────────────────────────────────

# ── Answer key upload parser ──────────────────────────────────────────────────

def _rows_preferring_cached_values(file_bytes: bytes, sheet_index: int = 0) -> list[tuple]:
    """
    Read a sheet's rows, taking each cell's cached value when the file has one
    and its literal content otherwise.

    Neither mode alone is safe. data_only=True returns None for every cell of a
    file written by Numbers or Google Sheets, because those tools save no cached
    formula results — the whole key parses as empty. data_only=False returns the
    FORMULA for any computed cell, so a trainer who assembles a code with
    =CONCATENATE(...) has the literal "=CONCATENATE(\"E11\",\".9\")" stored as
    their principal diagnosis, and every coder is then marked wrong against it.

    Reading both and preferring the cached value wherever one exists gives the
    right answer for both kinds of file.
    """
    literal = load_workbook(io.BytesIO(file_bytes), data_only=False).worksheets[sheet_index]
    cached = load_workbook(io.BytesIO(file_bytes), data_only=True).worksheets[sheet_index]

    lit_rows = list(literal.iter_rows(values_only=True))
    cache_rows = list(cached.iter_rows(values_only=True))

    out = []
    for i, lit in enumerate(lit_rows):
        cac = cache_rows[i] if i < len(cache_rows) else ()
        merged = []
        for j, val in enumerate(lit):
            cached_val = cac[j] if j < len(cac) else None
            is_formula = isinstance(val, str) and val.startswith("=")
            merged.append(cached_val if (is_formula and cached_val is not None) else val)
        out.append(tuple(merged))
    return out

def parse_answer_key_upload(file_bytes: bytes, specialty: str, with_pointers: bool = False,
                            single_path: bool = False, dx_only: bool = False) -> list[dict]:
    """
    Parse a trainer-uploaded answer key Excel file.
    Returns list of dicts, one per non-empty row:
      IP: {chart_number, pdx_code, pdx_poa, sdx:[{code,poa,ccmcc}], pcs:[{code}]}
      OP: {chart_number, pdx_code, sdx:[{code}], cpt:[{code,modifier}]}
    """
    # worksheets[0] is always the data sheet; Instructions is created last so it
    # becomes wb.active, but its index is always 1.
    all_rows = _rows_preferring_cached_values(file_bytes, 0)
    is_ip = specialty.upper() == "IP"
    # Scan ALL rows; skip the header row by name so this works regardless of
    # whether data starts at row 2 or somewhere else (e.g. Numbers adds a blank row).
    HEADER_NAMES = {"chart_number", "chart number", "chart#", "chartnumber"}
    results = []

    # Where each CPT field sits, read from the header row rather than counted
    # off a fixed stride. Units added a column mid-block, so a key filled from
    # last month's template has a different layout to one filled from today's —
    # a stride would read its modifiers as units.
    header_at: dict[str, int] = {}
    for idx, name in enumerate(all_rows[0] if all_rows else ()):
        if name is None:
            continue
        key = str(name).split("\n")[0].strip().lower()
        header_at.setdefault(key, idx)
    if not any(name in header_at for name in HEADER_NAMES):
        raise ValueError("Missing required column: Chart_Number")

    def _cell(row, idx) -> str:
        """Safely read a cell value, returning "" for None/empty/'None' sentinel."""
        val = row[idx] if idx < len(row) else None
        if val is None:
            return ""
        s = str(val).strip()
        return "" if s.lower() == "none" else s

    for row in all_rows:
        if not row or row[0] is None:
            continue
        chart_number = _cell(row, 0)
        if not chart_number or chart_number.lower().replace(" ", "_") in HEADER_NAMES:
            continue

        if is_ip:
            pdx_code = _cell(row, 1)
            pdx_poa = _cell(row, 2).upper()
            sdx = []
            col = 3
            for _ in range(IP_SDX_COUNT):
                code = _cell(row, col)
                poa = _cell(row, col + 1).upper()
                ccmcc = _cell(row, col + 2).upper()
                if code:
                    sdx.append({"code": code, "poa": poa, "ccmcc": ccmcc or "-"})
                col += 3
            pcs = []
            for _ in range(IP_PCS_COUNT):
                code = _cell(row, col)
                if code:
                    pcs.append({"code": code})
                col += 1
            results.append({
                "chart_number": chart_number,
                "pdx_code": pdx_code,
                "pdx_poa": pdx_poa,
                "sdx": sdx,
                "pcs": pcs,
            })
        else:
            base = 3 if single_path else 1
            facility_level = str(row[1] or "").strip() if single_path else ""
            profee_level = str(row[2] or "").strip() if single_path else ""
            pdx_code = str(row[base] or "").strip()
            sdx = []
            col = base + 1
            for _ in range(OP_SDX_COUNT):
                code = _cell(row, col)
                if code:
                    sdx.append({"code": code})
                col += 1
            cpt = []
            step = 3 if with_pointers else 2
            for i in range(0 if dx_only else OP_CPT_COUNT):
                named = header_at.get(f"cpt_{i + 1}")
                base_col = named if named is not None else col
                units_col = header_at.get(f"cpt_{i + 1}_units")
                mod_col = header_at.get(f"cpt_{i + 1}_modifier", base_col + 1)
                ptr_col = header_at.get(f"cpt_{i + 1}_dxpointers",
                                        base_col + 2 if units_col is None else None)

                code = _cell(row, base_col)
                if code:
                    entry = {"code": code, "modifier": _cell(row, mod_col)}
                    # Absent means "do not grade units", which is not the same
                    # as an explicit 1 — so an empty cell adds nothing.
                    if units_col is not None:
                        raw_units = _cell(row, units_col)
                        if raw_units:
                            entry["units"] = norm_units(raw_units)
                    if with_pointers and ptr_col is not None:
                        # isalpha() here silently DROPPED every numeric
                        # pointer, which is now the documented form — an
                        # uploaded key would have graded every line unlinked.
                        raw = _cell(row, ptr_col).upper()
                        ptrs = [x.strip() for x in raw.replace(" ", ",").split(",") if x.strip()]
                        entry["pointers"] = canonical_pointers(ptrs)
                    cpt.append(entry)
                col += step
            row_out = {
                "chart_number": chart_number,
                "pdx_code": pdx_code,
                "sdx": sdx,
                "cpt": cpt,
            }
            if single_path:
                row_out["facility_level"] = facility_level
                row_out["profee_level"] = profee_level
            results.append(row_out)

    return results


# ── E/M answer key Excel parser ──────────────────────────────────────────────

def parse_em_answer_key_upload(file_bytes: bytes) -> list[dict]:
    """
    Parse a trainer-filled E/M answer key Excel (from the bulk template).

    Column layout (A=0 … AY=50):
      A  chart_number
      B  copa_self_limited      C  copa_stable_acute       D  copa_stable_chronic
      E  copa_acute_uncomplicated F copa_chronic_exacerbation G copa_undiagnosed_new
      H  copa_acute_systemic    I  copa_acute_complicated_injury
      J  copa_chronic_severe    K  copa_threat_to_life
      L  copa_level_override (blank = auto-derive)
      M  dr_prior_external_notes N dr_review_test_results  O  dr_order_tests
      P  dr_independent_historian (Y/N)
      Q  dr_independent_interpretation (Y/N)
      R  dr_external_discussion (Y/N)
      S  dr_level_override
      T  risk_low               U  risk_prescription_drug_mgmt
      V  risk_minor_surgery_with_factors W risk_elective_major_no_factors
      X  risk_hospitalization   Y  risk_sdoh
      Z  risk_drug_intensive_monitoring AA risk_elective_major_with_factors
      AB risk_emergency_major_surgery   AC risk_hospitalization_escalation
      AD risk_dnr_deescalate           AE risk_parenteral_controlled
      AF risk_level_override
      AG em_code               AH em_modifier
      AI patient_type (New / Established / NA)
      AJ-AQ dx_codes (Primary + 7 additional = 8 total)
      AR-AY procedure_cpts (4 CPTs × code+modifier)
      AZ entered_by
      BA level_method (MDM / Time)   BB total_time (minutes)
      BC-BF procedure CPT Dx pointers (one cell per CPT, e.g. "1,2")
      BG-BJ procedure CPT units (blank = not graded on units)
      BK    em_category (blank = derive from the E/M code)
      BL    critical_care_minutes
    """
    all_rows = _rows_preferring_cached_values(file_bytes, 0)

    HEADER_NAMES = {"chart_number", "chart number", "chart#", "chartnumber"}

    # Fields are located by header text, falling back to the position they have
    # always had. A column inserted by hand no longer shifts everything after it
    # into the wrong field.
    at = _em_column_map(all_rows[0] if all_rows else ())

    def _raw(row, idx, default=""):
        val = row[idx] if (idx is not None and idx < len(row)) else None
        if val is None:
            return default
        s = str(val).strip()
        return default if s.lower() in ("none", "") else s

    def _f(row, field, default=""):
        return _raw(row, at.get(field), default)

    def _int(row, field) -> int:
        try:
            return max(0, int(float(_f(row, field, "0") or "0")))
        except (ValueError, TypeError):
            return 0

    def _bool(row, field) -> bool:
        return _f(row, field).upper() in ("Y", "YES", "TRUE", "1")

    results = []
    for row in all_rows:
        if not row or row[0] is None:
            continue
        chart_number = _f(row, "chart_number")
        if not chart_number or chart_number.lower().replace(" ", "_") in HEADER_NAMES:
            continue

        copa_override = _f(row, "copa_level_override")
        dr_override = _f(row, "dr_level_override")

        # Risk booleans (T–AE = indices 19–30)
        risk_fields = [
            "risk_low", "risk_prescription_drug_mgmt", "risk_minor_surgery_with_factors",
            "risk_elective_major_no_factors", "risk_hospitalization", "risk_sdoh",
            "risk_drug_intensive_monitoring", "risk_elective_major_with_factors",
            "risk_emergency_major_surgery", "risk_hospitalization_escalation",
            "risk_dnr_deescalate", "risk_parenteral_controlled",
        ]
        risk_override = _f(row, "risk_level_override")

        patient_type = _f(row, "patient_type").upper().strip() or "NA"
        if patient_type not in ("NEW", "ESTABLISHED", "NA"):
            patient_type = "NA"

        dx_codes = [c for c in (_f(row, f"dx_{i}") for i in range(1, 9)) if c]

        # Emitted as dicts so pointers and units survive; the grader also
        # accepts the legacy "code:modifier" string form.
        procedure_cpts = []
        for slot in range(1, 5):
            code = _f(row, f"cpt_{slot}")
            if not code:
                continue
            raw_ptr = _f(row, f"cpt_{slot}_pointers").upper()
            pointers = [x.strip()[:1] for x in raw_ptr.replace(" ", ",").split(",")
                        if x.strip() and x.strip()[0].isalpha()][:4]
            line = {
                "code": code,
                "modifier": _f(row, f"cpt_{slot}_modifier"),
                "pointers": pointers,
            }
            raw_units = _f(row, f"cpt_{slot}_units")
            if raw_units:
                line["units"] = norm_units(raw_units)
            procedure_cpts.append(line)

        em_category = _f(row, "em_category").strip().lower().replace(" ", "_").replace("&", "and")

        def _minutes(field):
            raw = _f(row, field)
            try:
                return int(float(raw)) if raw else None
            except (TypeError, ValueError):
                return None

        critical_care_minutes = _minutes("critical_care_minutes")
        entered_by = _f(row, "entered_by")

        level_method = (_f(row, "level_method") or "MDM").upper().strip()
        if level_method not in ("MDM", "TIME"):
            level_method = "MDM"
        total_time = _minutes("total_time")

        results.append({
            "chart_number": chart_number,
            "em_category":                 em_category,
            "critical_care_minutes":       critical_care_minutes,
            **{f: _int(row, f) for f in (
                "copa_self_limited", "copa_stable_acute", "copa_stable_chronic",
                "copa_acute_uncomplicated", "copa_chronic_exacerbation",
                "copa_undiagnosed_new", "copa_acute_systemic",
                "copa_acute_complicated_injury", "copa_chronic_severe",
                "copa_threat_to_life")},
            "copa_level_override":         copa_override,
            "copa_level_overridden":       bool(copa_override),
            "dr_prior_external_notes":     _int(row, "dr_prior_external_notes"),
            "dr_review_test_results":      _int(row, "dr_review_test_results"),
            "dr_order_tests":              _int(row, "dr_order_tests"),
            "dr_independent_historian":    _bool(row, "dr_independent_historian"),
            "dr_independent_interpretation": _bool(row, "dr_independent_interpretation"),
            "dr_external_discussion":      _bool(row, "dr_external_discussion"),
            "dr_level_override":           dr_override,
            "dr_level_overridden":         bool(dr_override),
            **{field: _bool(row, field) for field in risk_fields},
            "risk_level_override":         risk_override,
            "risk_level_overridden":       bool(risk_override),
            "em_code":                     _f(row, "em_code"),
            "em_modifier":                 _f(row, "em_modifier"),
            "patient_type":                patient_type,
            "level_method":                level_method,
            "total_time":                  total_time,
            "dx_codes":                    dx_codes,
            "procedure_cpts":              procedure_cpts,
            "entered_by":                  entered_by,
        })

    return results


# ── Results Excel export ──────────────────────────────────────────────────────

def export_batch_results(batch_name: str, results: list[dict]) -> bytes:
    """
    Export grading results to Excel with three sheets:
    Batch_Summary, Results_Summary (per coder), Feedback_Detail
    """
    wb = Workbook()
    wb.remove(wb.active)

    # ── Batch Summary sheet
    bs = wb.create_sheet("Batch_Summary")
    bs["A1"] = "Batch"
    bs["B1"] = batch_name
    bs["A1"].font = bs["B1"].font = Font(bold=True)

    # Aggregate from results
    coders = {}
    for r in results:
        name = r["coder_name"]
        if name not in coders:
            coders[name] = {"scores": [], "pass": 0, "fail": 0}
        if r.get("total_score") is not None:
            coders[name]["scores"].append(r["total_score"])
            if r.get("pass_fail") == "PASS":
                coders[name]["pass"] += 1
            else:
                coders[name]["fail"] += 1

    total_coders = len(coders)
    # Only count coders with at least one finalized score; ties go to FAIL
    passed = sum(1 for v in coders.values() if (v["pass"] + v["fail"]) > 0 and v["pass"] > v["fail"])
    all_scores = [s for v in coders.values() for s in v["scores"]]
    avg = round(sum(all_scores) / len(all_scores), 1) if all_scores else 0

    bs["A3"] = "Total Coders"; bs["B3"] = total_coders
    bs["A4"] = "Passed"; bs["B4"] = passed
    bs["A5"] = "Failed"; bs["B5"] = total_coders - passed
    bs["A6"] = "Pass Rate"; bs["B6"] = f"{round(passed/total_coders*100, 1)}%" if total_coders else "N/A"
    bs["A7"] = "Avg Grading Score"; bs["B7"] = f"{avg}%"
    for r in range(3, 8):
        bs.cell(r, 1).font = Font(bold=True)
    bs.column_dimensions["A"].width = 20
    bs.column_dimensions["B"].width = 20

    # ── Results Summary sheet
    rs = wb.create_sheet("Results_Summary")
    is_ip = any(r.get("specialty") == "IP-DRG" for r in results)
    headers = ["Coder", "Charts", "Avg PDx", "Avg SDx"]
    if is_ip:
        headers += ["Avg PCS", "Avg DRG"]
    else:
        headers.append("Avg CPT")
    headers += ["Avg Total", "Pass/Fail"]

    for c, h in enumerate(headers, 1):
        _header(rs, c, 1, h)

    # Per-coder aggregation
    coder_detail: dict[str, dict] = {}
    for r in results:
        name = r["coder_name"]
        if name not in coder_detail:
            coder_detail[name] = {
                "count": 0, "scored": 0, "pdx": 0, "sdx": 0, "pcs": 0, "cpt": 0,
                "drg": 0, "total": 0, "passed": 0,
            }
        d = coder_detail[name]
        d["count"] += 1
        d["pdx"] += r.get("pdx_score", 0)
        d["sdx"] += r.get("sdx_score", 0)
        d["pcs"] += r.get("pcs_score") or 0
        d["cpt"] += r.get("cpt_score") or 0
        d["drg"] += r.get("drg_score") or 0
        if r.get("total_score") is not None:
            d["total"] += r["total_score"]
            d["scored"] += 1
        if r.get("pass_fail") == "PASS":
            d["passed"] += 1

    for row_num, (name, d) in enumerate(coder_detail.items(), 2):
        cnt = d["count"] or 1
        scored = d["scored"]
        pf = "PENDING" if scored == 0 else ("PASS" if d["passed"] > scored / 2 else "FAIL")
        row_data = [name, d["count"],
                    round(d["pdx"] / cnt, 1), round(d["sdx"] / cnt, 1)]
        if is_ip:
            row_data += [round(d["pcs"] / cnt, 1), round(d["drg"] / cnt, 1)]
        else:
            row_data.append(round(d["cpt"] / cnt, 1))
        row_data += [round(d["total"] / cnt, 1), pf]
        for c, val in enumerate(row_data, 1):
            cell = rs.cell(row=row_num, column=c, value=val)
            cell.border = THIN_BORDER
            if c == len(row_data):
                pf_color = "00A000" if val == "PASS" else "6B7280" if val == "PENDING" else "CC0000"
                cell.font = Font(bold=True, color=pf_color)
    for c in range(1, len(headers) + 1):
        rs.column_dimensions[get_column_letter(c)].width = 14

    # ── Feedback Detail sheet
    fd = wb.create_sheet("Feedback_Detail")
    fb_headers = ["Coder", "Chart_Number", "Section", "Issue_Type", "AK_Code", "Coder_Code", "Detail"]
    for c, h in enumerate(fb_headers, 1):
        _header(fd, c, 1, h)

    fb_row = 2
    for r in results:
        for fb in r.get("feedback", []):
            fd.cell(fb_row, 1, r["coder_name"]).border = THIN_BORDER
            fd.cell(fb_row, 2, r.get("chart_number", "")).border = THIN_BORDER
            fd.cell(fb_row, 3, fb.get("section", "")).border = THIN_BORDER
            fd.cell(fb_row, 4, fb.get("issue_type", "")).border = THIN_BORDER
            fd.cell(fb_row, 5, fb.get("ak_code", "")).border = THIN_BORDER
            fd.cell(fb_row, 6, fb.get("coder_code", "")).border = THIN_BORDER
            fd.cell(fb_row, 7, fb.get("detail", "")).border = THIN_BORDER
            fb_row += 1
    for c, w in enumerate([18, 16, 10, 18, 14, 14, 30], 1):
        fd.column_dimensions[get_column_letter(c)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Coder performance export ─────────────────────────────────────────────────

def export_coder_performance(rows: list, feedback_rows: list) -> bytes:
    """
    Coder performance, one row per graded chart.

    Deliberately LONG format — one row per graded result, every dimension its
    own column — so the file can be sliced any way, which is the reason to
    want Excel rather than the PDF.
    A pre-formatted report would defeat the purpose.

    rows: dicts, one per GradingResult, already flattened by the caller.
    feedback_rows: dicts, one per feedback item, for error-pattern analysis.
    """
    wb = Workbook()

    # ── Sheet 1: the fact table ──
    ws = wb.active
    ws.title = "Results"
    cols = [
        ("Coder", 22), ("Emp ID", 12), ("Batch", 26), ("Assignment Type", 16),
        ("Batch Date", 13), ("Chart", 13), ("Specialty", 16), ("Category", 20),
        ("Difficulty", 13), ("Graded On", 13), ("Score %", 10), ("Result", 10),
        ("PDx", 8), ("SDx", 8), ("PCS", 8), ("CPT", 8), ("DRG", 8),
        ("DPO Dx %", 11), ("DPO POA %", 11), ("DPO Proc %", 11), ("DPO Overall %", 14),
    ]
    for i, (label, width) in enumerate(cols, start=1):
        _header(ws, i, 1, label)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"          # headers stay put while scrolling
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    for r, row in enumerate(rows, start=2):
        for i, key in enumerate([
            "coder_name", "emp_id", "batch_name", "assignment_type", "batch_date",
            "chart_number", "specialty", "category", "difficulty", "graded_on",
            "total_score", "pass_fail", "pdx_score", "sdx_score", "pcs_score",
            "cpt_score", "drg_score", "dpo_dx", "dpo_poa", "dpo_proc", "dpo_overall",
        ], start=1):
            ws.cell(row=r, column=i, value=row.get(key))

    # ── Sheet 2: per-coder summary, for the reader who just wants the answer ──
    ws2 = wb.create_sheet("By Coder")
    sum_cols = [("Coder", 22), ("Emp ID", 12), ("Charts", 10), ("Passed", 10),
                ("Pass Rate %", 13), ("Avg Score %", 13), ("Batches", 10),
                ("First Graded", 13), ("Last Graded", 13)]
    for i, (label, width) in enumerate(sum_cols, start=1):
        _header(ws2, i, 1, label)
        ws2.column_dimensions[get_column_letter(i)].width = width
    ws2.freeze_panes = "A2"

    agg: dict = {}
    for row in rows:
        key = row.get("emp_id") or row.get("coder_name")
        a = agg.setdefault(key, {"coder": row.get("coder_name"), "emp": row.get("emp_id"),
                                 "scores": [], "passed": 0, "batches": set(), "dates": []})
        if row.get("total_score") is not None:
            a["scores"].append(row["total_score"])
        if str(row.get("pass_fail", "")).upper() == "PASS":
            a["passed"] += 1
        a["batches"].add(row.get("batch_name"))
        if row.get("graded_on"):
            a["dates"].append(row["graded_on"])

    for r, a in enumerate(sorted(agg.values(), key=lambda x: (x["coder"] or "").lower()), start=2):
        n = len(a["scores"])
        ws2.cell(r, 1, a["coder"]); ws2.cell(r, 2, a["emp"])
        ws2.cell(r, 3, n); ws2.cell(r, 4, a["passed"])
        ws2.cell(r, 5, round(a["passed"] / n * 100, 1) if n else None)
        ws2.cell(r, 6, round(sum(a["scores"]) / n, 1) if n else None)
        ws2.cell(r, 7, len(a["batches"]))
        ws2.cell(r, 8, min(a["dates"]) if a["dates"] else None)
        ws2.cell(r, 9, max(a["dates"]) if a["dates"] else None)

    # ── Sheet 3: errors, so "what does this coder repeat" can be sliced ──
    ws3 = wb.create_sheet("Errors")
    err_cols = [("Coder", 22), ("Emp ID", 12), ("Batch", 26), ("Chart", 13),
                ("Specialty", 16), ("Category", 20), ("Section", 12),
                ("Issue", 16), ("Expected", 14), ("Coded", 14), ("Detail", 40)]
    for i, (label, width) in enumerate(err_cols, start=1):
        _header(ws3, i, 1, label)
        ws3.column_dimensions[get_column_letter(i)].width = width
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(err_cols))}1"

    for r, f in enumerate(feedback_rows, start=2):
        for i, key in enumerate([
            "coder_name", "emp_id", "batch_name", "chart_number", "specialty",
            "category", "section", "issue_type", "ak_code", "coder_code", "detail",
        ], start=1):
            ws3.cell(row=r, column=i, value=f.get(key))

    if not rows:
        ws.cell(2, 1, "No results match the selected filters.")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_batch_list(rows: list) -> bytes:
    """
    The batch list, exactly as the panel shows it.

    Deliberately NOT the coder-performance export. That one is long-format
    one row per graded result, every dimension its own column —
    which is the right shape for slicing performance and the wrong shape for
    "give me the list I am looking at". This is one row per batch, the columns
    on screen, in the order they appear.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Batches"

    cols = [
        ("Name", 32), ("Type", 12), ("Specialty", 16), ("Status", 10),
        ("Coders", 9), ("Charts / Coder", 14), ("Cycles", 9),
        ("Graded Results", 15), ("Days Open", 11),
        ("Created By", 18), ("Created", 13), ("Closed", 13),
    ]
    for i, (label, width) in enumerate(cols, start=1):
        _header(ws, i, 1, label)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    if not rows:
        ws.cell(2, 1, "No batches match the selected filters.")
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    keys = ["name", "type", "specialty", "status", "coder_count", "charts_per_coder",
            "cycles", "graded_count", "days_open", "created_by", "created_at", "closed_at"]
    for r, row in enumerate(rows, start=2):
        for i, key in enumerate(keys, start=1):
            ws.cell(r, i, row.get(key))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_batch_analytics(rows: list) -> bytes:
    """
    The Analytics -> By Batch table, as shown.

    Third batch-shaped export, and the three answer different questions:
      - coder-performance : one row per graded RESULT, sliceable
      - batch list        : one row per batch, the roster view (coders,
                            cycles, status) — about SET-UP
      - this one          : one row per batch, the performance view (scores,
                            pass rates, chart counts) — about OUTCOMES
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Performance"

    cols = [
        ("Batch", 30), ("Type", 10), ("Specialty", 16), ("Created", 12),
        ("Coders", 9), ("Charts", 9), ("Graded Results", 15),
        ("Pass Mark %", 12), ("Avg Grading Score %", 19), ("Chart Pass Rate %", 17),
        ("Below Target", 13), ("Last Graded", 13),
    ]
    for i, (label, width) in enumerate(cols, start=1):
        _header(ws, i, 1, label)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    if not rows:
        ws.cell(2, 1, "No batches match the selected filters.")
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    keys = ["batch_name", "type", "specialty", "created_at", "coder_count",
            "chart_count", "graded_count", "pass_threshold", "avg_score",
            "pass_rate", "below_target", "last_graded_at"]
    for r, row in enumerate(rows, start=2):
        for i, key in enumerate(keys, start=1):
            ws.cell(r, i, row.get(key))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_grid(sheet_title: str, row_header: str, columns: list, rows: list) -> bytes:
    """
    A wide grid: one row per entity, one column per column, values in between.

    Wide is the right shape here precisely because it is the wrong shape for
    pivoting. These sheets exist so a trainer can read the whole grid at once —
    every coder against every batch — which the screen cannot do past about
    fourteen columns. The long-format export already covers slicing.

    columns: [{"key": ..., "label": ..., "sub": ...}]
    rows:    [{"label": ..., "sub": ..., "values": {key: value}, "overall": ...}]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    _header(ws, 1, 1, row_header)
    ws.column_dimensions["A"].width = 28
    for i, col in enumerate(columns, start=2):
        label = col["label"] + (f"\n{col['sub']}" if col.get("sub") else "")
        _header(ws, i, 1, label)
        ws.column_dimensions[get_column_letter(i)].width = 16
    overall_col = len(columns) + 2
    _header(ws, overall_col, 1, "Overall")
    ws.column_dimensions[get_column_letter(overall_col)].width = 12

    ws.row_dimensions[1].height = 30
    # Both panes: the row labels and the header must survive scrolling in a
    # grid that is wide and long at the same time.
    ws.freeze_panes = "B2"

    if not rows:
        ws.cell(2, 1, "No data matches the selected filters.")
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    for r, row in enumerate(rows, start=2):
        ws.cell(r, 1, row["label"] + (f" ({row['sub']})" if row.get("sub") else ""))
        for i, col in enumerate(columns, start=2):
            v = row["values"].get(col["key"])
            # Blank, not zero: a coder with no result in a batch did not score
            # nothing, and a 0 would drag any average taken over the column.
            if v is not None:
                ws.cell(r, i, v)
        if row.get("overall") is not None:
            ws.cell(r, overall_col, row["overall"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_chart_signals(rows: list) -> bytes:
    """One row per chart, with the signal and the bar it was judged against."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Chart Signals"

    cols = [("Chart", 14), ("Signal", 16), ("Specialty", 16), ("Topic", 24),
            ("Attempts", 10), ("Avg Score %", 12), ("Pass Rate %", 12),
            ("Pass Mark %", 12), ("Distinct Codes Missed", 21)]
    for i, (label, width) in enumerate(cols, start=1):
        _header(ws, i, 1, label)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    if not rows:
        ws.cell(2, 1, "No charts match the selected filters.")
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    keys = ["chart_number", "teaching_label", "specialty", "category",
            "attempt_count", "avg_score", "pass_rate", "pass_threshold", "error_variety"]
    for r, row in enumerate(rows, start=2):
        for i, key in enumerate(keys, start=1):
            ws.cell(r, i, row.get(key))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_error_analysis(data: dict) -> bytes:
    """
    Error analysis across five sheets.

    Deliberately not one flat table. The tab answers several questions that
    have different shapes — what the findings are, where errors concentrate,
    which codes matter and how each is spread — and flattening them into one
    sheet would mean a column set that fits none of them.

    The Codes sheet carries the pattern verdict, because a code's count without
    its spread cannot be acted on, and that is the whole point of the tab.
    """
    wb = Workbook()

    # ── 1. Insights ──
    ws = wb.active
    ws.title = "Insights"
    _header(ws, 1, 1, "Finding")
    _header(ws, 2, 1, "Type")
    ws.column_dimensions["A"].width = 120
    ws.column_dimensions["B"].width = 14
    ws.freeze_panes = "A2"
    r = 2
    for n in (data.get("commentary") or []):
        cell = ws.cell(r, 1, n.get("text"))
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(r, 2, n.get("kind"))
        ws.row_dimensions[r].height = 30
        r += 1
    if r == 2:
        ws.cell(2, 1, "No findings — not enough graded work yet.")

    r += 1
    for label, key in [("Total errors", "total_errors"), ("Graded charts", "graded_charts"),
                       ("Errors per chart", "errors_per_chart"), ("Distinct codes", "total_codes")]:
        ws.cell(r, 1, label)
        ws.cell(r, 2, data.get(key))
        r += 1

    def _sheet(title, cols, rows, keys):
        w = wb.create_sheet(title[:31])
        for i, (label, width) in enumerate(cols, start=1):
            _header(w, i, 1, label)
            w.column_dimensions[get_column_letter(i)].width = width
        w.freeze_panes = "A2"
        w.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
        for ri, row in enumerate(rows, start=2):
            for ci, k in enumerate(keys, start=1):
                v = row.get(k)
                w.cell(ri, ci, ", ".join(v) if isinstance(v, list) else v)
        if not rows:
            w.cell(2, 1, "No data for the selected filters.")
        return w

    _sheet("By Issue Type", [("Issue Type", 22), ("Errors", 12), ("Share %", 12)],
           data.get("by_issue_type") or [], ["type", "count", "pct"])
    _sheet("By Section", [("Section", 18), ("Errors", 12), ("Share %", 12)],
           data.get("by_section") or [], ["section", "count", "pct"])
    _sheet("By Specialty",
           [("Specialty", 20), ("Errors", 12), ("Graded Charts", 15),
            ("Errors per Chart", 18), ("Enough to Rank", 15)],
           data.get("by_specialty") or [],
           ["specialty", "errors", "charts", "errors_per_chart", "rankable"])

    codes = data.get("codes") or []
    # The reference columns are what make this sheet pivotable: a trainer can
    # slice by chapter or root operation in Excel without any of it being built
    # into the app. They are blank for CPT, which this application does not
    # hold, and for codes absent from the loaded edition.
    w = _sheet("Codes",
               [("Code", 16), ("Description", 54), ("Code System", 13),
                ("Chapter", 40), ("CC/MCC", 10),
                ("Root Operation", 18), ("Approach", 22), ("Body System", 22),
                ("Device", 22), ("Qualifier", 18),
                ("Times", 10), ("Coders", 10), ("Charts", 10),
                ("Pattern", 14), ("What it means", 62), ("Section", 12),
                ("Issue Mix", 34), ("Specialties", 22), ("Per Coder", 12),
                ("Last Seen", 13)],
               [{**c, "issue_mix": " · ".join(f"{b['type'].replace('_', ' ')} {b['count']}"
                                              for b in c.get("issue_breakdown") or [])}
                for c in codes],
               ["code", "description", "code_system", "chapter", "cc_mcc",
                "root_operation", "approach", "body_system", "device", "qualifier",
                "count", "coders_affected", "charts_affected", "pattern",
                "pattern_reason", "top_section", "issue_mix", "specialties",
                "per_coder", "last_seen"])
    for ri in range(2, len(codes) + 2):
        w.cell(ri, 15).alignment = Alignment(wrap_text=True, vertical="top")
        w.cell(ri, 2).alignment = Alignment(wrap_text=True, vertical="top")

    # ── the clinical axes, one sheet each ────────────────────────────────────
    _axis_cols = [("Group", 46), ("Errors", 10), ("Coders", 10), ("Charts", 10),
                  ("Share %", 10)]
    _axis_keys = ["label", "count", "coders_affected", "charts_affected", "share"]
    if data.get("by_chapter"):
        _sheet("By Chapter", _axis_cols, data["by_chapter"], _axis_keys)
    if data.get("by_ccmcc"):
        _sheet("By CC-MCC", _axis_cols, data["by_ccmcc"], _axis_keys)
    for axis, rows in (data.get("by_pcs_axis") or {}).items():
        if rows:
            _sheet("PCS " + axis.replace("_", " ").title(), _axis_cols, rows, _axis_keys)

    # The evidence behind each verdict. On screen this opens one code at a
    # time; in a sheet there is no reason to make someone click two hundred
    # times, and "Team-wide" without the names is a claim you cannot check.
    _sheet("Code by Coder",
           [("Code", 16), ("Coder", 24), ("Emp ID", 14), ("Times", 10), ("Charts", 10)],
           data.get("code_coders") or [],
           ["code", "coder_name", "emp_id", "count", "charts"])
    _sheet("Code by Chart",
           [("Code", 16), ("Chart", 16), ("Specialty", 18), ("Topic", 22),
            ("Times", 10), ("Coders", 10)],
           data.get("code_charts") or [],
           ["code", "chart_number", "specialty", "category", "count", "coders"])

    if data.get("trend"):
        _sheet("Trend", [("Month", 12), ("Errors", 12)],
               data["trend"], ["month", "total"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── E/M answer key column table ──────────────────────────────────────────────
#
# One ordered list, written by the template generator and read by the parser.
#
# The parser used to index this sheet purely by position — AR was 43, BA was 52
# — with no check that the column at 43 was the one it wanted. A single inserted
# column shifts every field after it, and the parse still "succeeds": the key
# stores a modifier where a code belongs and grades every coder against it, with
# nothing anywhere reporting a problem. Naming the columns here lets the parser
# find each field by its header and fall back to position only for a file that
# has no header row at all.
#
# APPEND ONLY. A file filled from an older template must keep parsing, and the
# positional fallback depends on these indices never moving.
EM_KEY_COLUMNS: list[tuple[str, str]] = [
    ("chart_number", "Chart Number"),
    ("copa_self_limited", "COPA: Self-limited/Minor Problems (count)"),
    ("copa_stable_acute", "COPA: Stable Acute Illness (count)"),
    ("copa_stable_chronic", "COPA: Stable Chronic Illness (count)"),
    ("copa_acute_uncomplicated", "COPA: Acute Uncomplicated (count)"),
    ("copa_chronic_exacerbation", "COPA: Chronic Exacerbation (count)"),
    ("copa_undiagnosed_new", "COPA: Undiagnosed New Problem (count)"),
    ("copa_acute_systemic", "COPA: Acute w/ Systemic Symptoms (count)"),
    ("copa_acute_complicated_injury", "COPA: Acute Complicated Injury (count)"),
    ("copa_chronic_severe", "COPA: Chronic Severe Exacerbation (count)"),
    ("copa_threat_to_life", "COPA: Threat to Life/Function (0 or 1)"),
    ("copa_level_override", "COPA Level Override (leave blank to auto-derive)"),
    ("dr_prior_external_notes", "DR: Prior External Notes (count)"),
    ("dr_review_test_results", "DR: Review Test Results (count)"),
    ("dr_order_tests", "DR: Order Tests (count)"),
    ("dr_independent_historian", "DR: Independent Historian (Y/N)"),
    ("dr_independent_interpretation", "DR: Independent Interpretation (Y/N)"),
    ("dr_external_discussion", "DR: External Discussion (Y/N)"),
    ("dr_level_override", "DR Level Override (leave blank to auto-derive)"),
    ("risk_low", "Risk: Low (Y/N)"),
    ("risk_prescription_drug_mgmt", "Risk: Prescription Drug Mgmt (Y/N)"),
    ("risk_minor_surgery_with_factors", "Risk: Minor Surgery w/ Risk Factors (Y/N)"),
    ("risk_elective_major_no_factors", "Risk: Elective Major - No Risk Factors (Y/N)"),
    ("risk_hospitalization", "Risk: Hospitalization (Y/N)"),
    ("risk_sdoh", "Risk: SDOH Limitation (Y/N)"),
    ("risk_drug_intensive_monitoring", "Risk: Drug Intensive Toxicity Monitoring (Y/N)"),
    ("risk_elective_major_with_factors", "Risk: Elective Major w/ Risk Factors (Y/N)"),
    ("risk_emergency_major_surgery", "Risk: Emergency Major Surgery (Y/N)"),
    ("risk_hospitalization_escalation", "Risk: Hospitalization/Escalation (Y/N)"),
    ("risk_dnr_deescalate", "Risk: DNR / De-escalate (Y/N)"),
    ("risk_parenteral_controlled", "Risk: Parenteral Controlled Substance (Y/N)"),
    ("risk_level_override", "Risk Level Override (leave blank to auto-derive)"),
    ("em_code", "E/M Code"),
    ("em_modifier", "E/M Modifier (e.g. 25)"),
    ("patient_type", "Patient Type (New / Established / NA)"),
    ("dx_1", "Primary Dx Code"),
    ("dx_2", "Additional Dx 2"),
    ("dx_3", "Additional Dx 3"),
    ("dx_4", "Additional Dx 4"),
    ("dx_5", "Additional Dx 5"),
    ("dx_6", "Additional Dx 6"),
    ("dx_7", "Additional Dx 7"),
    ("dx_8", "Additional Dx 8"),
    ("cpt_1", "Procedure CPT 1"),
    ("cpt_1_modifier", "Procedure CPT 1 Modifier"),
    ("cpt_2", "Procedure CPT 2"),
    ("cpt_2_modifier", "Procedure CPT 2 Modifier"),
    ("cpt_3", "Procedure CPT 3"),
    ("cpt_3_modifier", "Procedure CPT 3 Modifier"),
    ("cpt_4", "Procedure CPT 4"),
    ("cpt_4_modifier", "Procedure CPT 4 Modifier"),
    ("entered_by", "Entered By"),
    ("level_method", "Level By (MDM / Time)"),
    ("total_time", "Total Time (minutes)"),
    ("cpt_1_pointers", "CPT 1 Dx Pointers (e.g. 1,2)"),
    ("cpt_2_pointers", "CPT 2 Dx Pointers (e.g. 1,2)"),
    ("cpt_3_pointers", "CPT 3 Dx Pointers (e.g. 1,2)"),
    ("cpt_4_pointers", "CPT 4 Dx Pointers (e.g. 1,2)"),
    ("cpt_1_units", "CPT 1 Units (blank = 1)"),
    ("cpt_2_units", "CPT 2 Units (blank = 1)"),
    ("cpt_3_units", "CPT 3 Units (blank = 1)"),
    ("cpt_4_units", "CPT 4 Units (blank = 1)"),
    ("em_category", "Encounter Category (blank = from E/M code)"),
    ("critical_care_minutes", "Critical Care Total Minutes (critical care only)"),
]

EM_KEY_FIELD_INDEX = {field: i for i, (field, _) in enumerate(EM_KEY_COLUMNS)}


def _em_column_map(header_row) -> dict:
    """
    field -> column index for THIS file.

    Matched on the header text the template wrote. Anything the header row does
    not name keeps its position from EM_KEY_COLUMNS, so a file with no headers
    at all still parses exactly as it did.
    """
    seen = {}
    for idx, cell in enumerate(header_row or ()):
        if cell is None:
            continue
        seen.setdefault(" ".join(str(cell).split()).lower(), idx)
    return {
        field: seen.get(" ".join(header.split()).lower(), default_idx)
        for default_idx, (field, header) in enumerate(EM_KEY_COLUMNS)
    }
