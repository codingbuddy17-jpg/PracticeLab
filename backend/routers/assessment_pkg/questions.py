"""
Assessment Question Bank router.
Handles upload, listing, editing, status changes, template download, pool preview.
"""
import io
import re
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from database import get_db
from models import AssessmentQuestion, AssessmentAuditLog
from config import settings
from routers.assessment_pkg.security import require_passphrase


def _audit(db: Session, trainer: str, action: str, specialty: Optional[str] = None, details: Optional[str] = None):
    db.add(AssessmentAuditLog(trainer_name=trainer, action=action, specialty=specialty, details=details))
    db.commit()

router = APIRouter()

SPECIALTY_PREFIX = {
    "ICD10CM": "ICDX",
    "Surgery": "SURG",
    "ED Facility": "EDFC",
    "ED Profee": "EDPF",
    "Ancillary": "ANCL",
    "IP-DRG": "IPDR",
    "E&M": "EMC",
    "E&M - Multispecialty": "EMMS",
    "IVR": "IVR",
    "Anesthesia": "ANES",
}

VALID_SPECIALTIES = set(SPECIALTY_PREFIX.keys())
VALID_DIFFICULTIES = {"Easy", "Medium", "Hard"}
VALID_ANSWERS = {"A", "B", "C", "D"}
VALID_TYPES = {"Conceptual", "Scenario", "Rule-based"}

TEMPLATE_HEADERS = [
    "Question_ID", "Question_Text", "Option_A", "Option_B", "Option_C", "Option_D",
    "Correct_Answer", "Difficulty", "Topic", "Question_Type", "Active_Status", "Shuffle_Options",
]

SAMPLE_ROWS: dict = {
    "ICD10CM": [
        ["", "A patient is diagnosed with Type 2 diabetes mellitus with diabetic chronic kidney disease, stage 3. What is the correct ICD-10-CM code?",
         "E11.22", "E11.65", "E13.22", "E11.29", "A", "Easy", "Diabetes", "Conceptual", "Active", "Yes"],
        ["", "Acute appendicitis without abscess, without peritonitis. Select the correct code.",
         "K35.80", "K35.2", "K37", "K35.89", "A", "Easy", "Abdominal", "Conceptual", "Active", "Yes"],
        ["", "A patient has essential hypertension and chronic kidney disease stage 3. How should this be coded?",
         "I12.9 and N18.3", "I10 and N18.3", "I12.9 alone", "I10 alone", "A", "Medium", "Cardiovascular", "Rule-based", "Active", "Yes"],
    ],
    "Surgery": [
        ["", "Which CPT code represents a laparoscopic appendectomy?",
         "44970", "44950", "44960", "44979", "A", "Easy", "Laparoscopic", "Conceptual", "Active", "Yes"],
        ["", "Open repair of initial inguinal hernia, reducible, patient age 5 years or older. Correct CPT?",
         "49505", "49500", "49507", "49520", "A", "Medium", "Hernia", "Conceptual", "Active", "Yes"],
        ["", "Arthroscopic partial medial meniscectomy, right knee. CPT code?",
         "29881", "29880", "29882", "29876", "A", "Medium", "Orthopedic", "Conceptual", "Active", "Yes"],
    ],
}

DEFAULT_SAMPLE = [
    ["", "Sample question text — replace with your question.",
     "Option A text", "Option B text", "Option C text", "Option D text",
     "A", "Medium", "General", "Conceptual", "Active", "Yes"],
    ["", "Another sample question demonstrating the format.",
     "First choice", "Second choice", "Third choice", "Fourth choice",
     "B", "Easy", "General", "Scenario", "Active", "Yes"],
    ["", "Hard difficulty rule-based sample question.",
     "Answer option A", "Answer option B", "Answer option C", "Answer option D",
     "C", "Hard", "Guidelines", "Rule-based", "Active", "No"],
]


def option_problem(a: str, b: str, c: str, d: str) -> Optional[str]:
    """
    Why these four options cannot be used, or None if they are fine.

    Blank and duplicate choices are not untidy data — they are a grading
    defect. Duplicate text gives a question two right answers while only one
    letter is keyed, and because options are shuffled per coder, WHICH one is
    keyed varies between them: two coders picking the identical text get
    different marks. A blank choice reaches the coder as an empty option.

    Shared with the standalone parser so both entry points refuse the same
    thing — this is exactly the rule that otherwise gets fixed where it was
    found and forgotten in the other path.
    """
    labels = ("Option_A", "Option_B", "Option_C", "Option_D")
    values = (a, b, c, d)

    blanks = [lbl for lbl, v in zip(labels, values) if not (v or "").strip()]
    if blanks:
        return f"blank {', '.join(blanks)} — every option needs text"

    seen: Dict[str, str] = {}
    for lbl, v in zip(labels, values):
        key = " ".join(v.split()).casefold()
        if key in seen:
            return (f"duplicate option text in {seen[key]} and {lbl} — "
                    "two identical choices cannot both be graded")
        seen[key] = lbl
    return None


def _next_qid(db: Session, specialty: str) -> str:
    prefix = SPECIALTY_PREFIX.get(specialty, "QST")
    pattern = f"{prefix}-%"
    rows = db.query(AssessmentQuestion.question_id).filter(
        AssessmentQuestion.question_id.like(pattern)
    ).all()
    max_num = 0
    for (qid,) in rows:
        m = re.search(r"-(\d+)$", qid)
        if m:
            max_num = max(max_num, int(m.group(1)))
    return f"{prefix}-{max_num + 1:03d}"


@router.get("/questions/stats")
def question_stats(db: Session = Depends(get_db)):
    """Per-specialty active/inactive/total counts."""
    rows = db.query(
        AssessmentQuestion.specialty,
        AssessmentQuestion.status,
        func.count(AssessmentQuestion.id),
    ).group_by(AssessmentQuestion.specialty, AssessmentQuestion.status).all()

    agg: dict = {}
    for specialty, status, cnt in rows:
        if specialty not in agg:
            agg[specialty] = {"specialty": specialty, "total": 0, "active": 0, "inactive": 0}
        agg[specialty]["total"] += cnt
        if status == "Active":
            agg[specialty]["active"] += cnt
        else:
            agg[specialty]["inactive"] += cnt

    return list(agg.values())


@router.get("/questions/template")
def download_template(specialty: str = Query(default="ICD10CM")):
    """Download blank xlsx template with 3 sample rows for the given specialty."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    samples = SAMPLE_ROWS.get(specialty, DEFAULT_SAMPLE)
    for row_idx, row_data in enumerate(samples, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Column widths
    widths = [15, 60, 30, 30, 30, 30, 14, 12, 20, 15, 14, 16]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"assessment_template_{specialty.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _readiness(total_active: int, per_paper: int, coders: int) -> Dict[str, Any]:
    """
    What this pool will actually produce for a paper of `per_paper` questions
    sat by `coders` people, with per-coder randomisation on.

    Readiness is not an absolute count. A 30-question pool is thin for a
    25-question paper and generous for a 5-question one, so a fixed "50+ is
    healthy" rule answers the wrong question.

    The number that matters is PAIRWISE OVERLAP — how much of their paper any
    two coders hold in common — because that is what decides whether comparing
    notes defeats the assessment. Drawing n of p independently, two coders
    share about n x (n/p) questions, so the overlap share is simply n/p.

    Uniqueness (items no OTHER coder holds) is reported too, since that is what
    compute_randomisation_stats measures after generation and the two should be
    comparable. It is deliberately NOT the headline: with 40 coders even a
    healthy 250-question pool scores ~2% uniqueness, which reads as alarming
    when the paper is in fact fine.
    """
    if per_paper <= 0 or coders <= 0:
        return {}

    if total_active < per_paper:
        return {
            "per_paper": per_paper,
            "coders": coders,
            "verdict": "insufficient",
            "headline": f"Only {total_active} active question(s) — a {per_paper}-question paper cannot be built.",
            "advice": "Add questions, widen the topic filter, or lower the question count.",
            "overlap_pct": None,
            "shared_questions": None,
            "uniqueness_pct": None,
            "pool_for_light_overlap": per_paper * 4,
            "pool_for_strong": per_paper * 10,
        }

    overlap = per_paper / total_active
    shared = round(overlap * per_paper, 1)
    uniqueness = round(((1 - overlap) ** (coders - 1)) * 100, 1) if coders > 1 else 100.0
    overlap_pct = round(overlap * 100, 1)

    if overlap >= 0.5:
        verdict, advice = "near_identical", (
            f"Aim for {per_paper * 4}+ active questions to get overlap under 25%.")
    elif overlap >= 0.25:
        verdict, advice = "heavy_overlap", (
            f"Workable, but {per_paper * 4}+ questions would drop overlap below 25%.")
    elif overlap >= 0.10:
        verdict, advice = "workable", (
            f"Fine for most purposes; {per_paper * 10}+ would make papers largely distinct.")
    else:
        verdict, advice = "strong", "Papers will be largely distinct."

    return {
        "per_paper": per_paper,
        "coders": coders,
        "verdict": verdict,
        "headline": (f"Any two coders will share about {shared:g} of their {per_paper} "
                     f"questions ({overlap_pct}%)."),
        "advice": advice,
        "overlap_pct": overlap_pct,
        "shared_questions": shared,
        "uniqueness_pct": uniqueness,
        "pool_for_light_overlap": per_paper * 4,
        "pool_for_strong": per_paper * 10,
    }


def _difficulty_outlook(by_diff: Dict[str, int], total_active: int, per_paper: int) -> Dict[str, Any]:
    """
    What generation will DO with this mix — not whether the mix is tidy.

    Neither imbalance errors, which is the trap. In difficulty_mode="auto" the
    ratios are derived FROM the pool, so a 95/5 Easy/Hard bank silently yields
    an easy paper that looks deliberate. In "manual" mode _stratified_pick
    cascades any shortfall into the next bucket, so a requested 30% Hard
    quietly becomes Easy and Medium. Say which, rather than calling it
    "unbalanced".
    """
    if not total_active:
        return {}

    counts = {d: by_diff.get(d, 0) for d in ("Easy", "Medium", "Hard")}
    auto_mix = {d: round(c / total_active * 100) for d, c in counts.items()}

    # An even three-way manual split is the common case; a bucket that cannot
    # cover its third is the one that will cascade.
    third = per_paper / 3 if per_paper else 0
    short = [d for d, c in counts.items() if third and c < third]

    notes = []
    dominant = max(auto_mix, key=lambda d: auto_mix[d]) if total_active else None
    if dominant and auto_mix[dominant] >= 70:
        notes.append(
            f"On automatic difficulty, papers will be about {auto_mix[dominant]}% {dominant} "
            f"— the mix is taken from the pool, so this is what coders will sit.")
    if short and per_paper:
        notes.append(
            "If you set the difficulty mix manually, "
            + ", ".join(f"{d} ({counts[d]} available)" for d in short)
            + f" cannot cover an even share of a {per_paper}-question paper; "
              "the shortfall is filled from the other levels rather than refused.")

    return {"auto_mix_pct": auto_mix, "short_buckets": short, "notes": notes}


@router.get("/questions/pool-summary")
def pool_summary(
    specialty: str = Query(...),
    questions_per_paper: int = Query(default=0, ge=0, le=500),
    coders: int = Query(default=0, ge=0, le=5000),
    db: Session = Depends(get_db),
):
    """
    Counts for a specialty, plus what they mean for an intended paper.
    No question content is exposed — counts only.
    """
    if specialty not in VALID_SPECIALTIES:
        # Previously an unknown specialty returned zeroes, which reads exactly
        # like a real but empty pool.
        raise HTTPException(
            status_code=400,
            detail=f"Unknown specialty '{specialty}'. Expected one of: "
                   + ", ".join(sorted(VALID_SPECIALTIES)),
        )

    base = [AssessmentQuestion.specialty == specialty]
    filters = base + [AssessmentQuestion.status == "Active"]

    total_active = db.query(func.count(AssessmentQuestion.id)).filter(*filters).scalar() or 0
    # Inactive is actionable in its own right: reactivating 80 retired questions
    # is a great deal cheaper than authoring them.
    total_inactive = db.query(func.count(AssessmentQuestion.id)).filter(
        *base, AssessmentQuestion.status != "Active"
    ).scalar() or 0

    by_topic = db.query(
        AssessmentQuestion.topic,
        func.count(AssessmentQuestion.id),
    ).filter(*filters).group_by(AssessmentQuestion.topic).all()

    by_diff = db.query(
        AssessmentQuestion.difficulty,
        func.count(AssessmentQuestion.id),
    ).filter(*filters).group_by(AssessmentQuestion.difficulty).all()

    diff_map = {d: c for d, c in by_diff}

    return {
        "specialty": specialty,
        "total_active": total_active,
        "total_inactive": total_inactive,
        "by_topic": sorted(
            [{"topic": t or "Uncategorized", "count": c} for t, c in by_topic],
            key=lambda x: -x["count"],
        ),
        "by_difficulty": diff_map,
        "readiness": _readiness(total_active, questions_per_paper, coders),
        "difficulty_outlook": _difficulty_outlook(diff_map, total_active, questions_per_paper),
    }


@router.get("/questions/export")
def export_questions(
    specialty: str = Query(...),
    passphrase: str = Query(...),
    trainer_name: str = Query(default="Trainer"),
    db: Session = Depends(get_db),
):
    """Passphrase-protected XLSX export of full question bank for a specialty."""
    if passphrase != settings.MASTER_ADMIN_PASSPHRASE:
        raise HTTPException(status_code=403, detail="Invalid passphrase")

    qs = db.query(AssessmentQuestion).filter(
        AssessmentQuestion.specialty == specialty,
    ).order_by(AssessmentQuestion.question_id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = specialty[:31]
    _write_specialty_sheet(ws, qs)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"questions_{specialty.replace(' ', '_').replace('&', 'and')}.xlsx"
    _audit(db, trainer_name, "download", specialty, f"Downloaded {len(qs)} questions")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _write_specialty_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, qs: list) -> None:
    """Write header + question rows into a worksheet (reused for single and multi-tab exports)."""
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, q in enumerate(qs, start=2):
        row_vals = [
            q.question_id, q.question_text,
            q.option_a, q.option_b, q.option_c, q.option_d,
            q.correct_answer, q.difficulty, q.topic or "",
            q.question_type, q.status,
            "Yes" if q.shuffle_options else "No",
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    widths = [15, 60, 30, 30, 30, 30, 14, 12, 20, 15, 14, 16]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w


@router.get("/questions/export-all")
def export_all_questions(
    passphrase: str = Query(...),
    trainer_name: str = Query(default="Trainer"),
    specialty: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    """Export full question bank as a multi-tab XLSX (one tab per specialty).
    If specialty is provided, exports only that specialty as a single tab.
    Passphrase-protected. Audit-logged.
    """
    if passphrase != settings.MASTER_ADMIN_PASSPHRASE:
        raise HTTPException(status_code=403, detail="Invalid passphrase")

    specialties_to_export: List[str] = [specialty] if specialty else sorted(SPECIALTY_PREFIX.keys())

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    total_questions = 0
    for sp in specialties_to_export:
        qs = (
            db.query(AssessmentQuestion)
            .filter(AssessmentQuestion.specialty == sp)
            .order_by(AssessmentQuestion.question_id)
            .all()
        )
        # Tab name: max 31 chars (Excel limit), strip special chars
        tab_name = sp[:31]
        ws = wb.create_sheet(title=tab_name)
        _write_specialty_sheet(ws, qs)
        total_questions += len(qs)

    if not wb.worksheets:
        raise HTTPException(status_code=404, detail="No questions found to export.")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    if specialty:
        filename = f"questions_{specialty.replace(' ', '_').replace('&', 'and')}.xlsx"
        audit_detail = f"Downloaded {total_questions} questions for {specialty}"
        _audit(db, trainer_name, "download", specialty, audit_detail)
    else:
        filename = "question_bank_all_specialties.xlsx"
        audit_detail = f"Downloaded full inventory — {total_questions} questions across {len(specialties_to_export)} specialties"
        _audit(db, trainer_name, "download", None, audit_detail)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/questions/pool-preview")
def pool_preview(
    specialty: str = Query(...),
    topic_filter: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    """Return available counts broken down by difficulty for one specialty."""
    q = db.query(
        AssessmentQuestion.difficulty,
        func.count(AssessmentQuestion.id),
    ).filter(
        AssessmentQuestion.specialty == specialty,
        AssessmentQuestion.status == "Active",
    )
    if topic_filter:
        topics = [t.strip() for t in topic_filter.split(",") if t.strip()]
        if len(topics) == 1:
            q = q.filter(AssessmentQuestion.topic.ilike(f"%{topics[0]}%"))
        elif topics:
            from sqlalchemy import or_
            q = q.filter(or_(*[AssessmentQuestion.topic.ilike(f"%{t}%") for t in topics]))
    rows = q.group_by(AssessmentQuestion.difficulty).all()

    counts = {"Easy": 0, "Medium": 0, "Hard": 0}
    for diff, cnt in rows:
        if diff in counts:
            counts[diff] = cnt

    return {
        "specialty": specialty,
        "active_count": sum(counts.values()),
        "easy": counts["Easy"],
        "medium": counts["Medium"],
        "hard": counts["Hard"],
    }


@router.get("/questions")
def list_questions(
    specialty: Optional[str] = Query(default=None),
    difficulty: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    topic: Optional[str] = Query(default=None),
    search: Optional[str] = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
    db: Session = Depends(get_db),
    _: None = Depends(require_passphrase),
):
    q = db.query(AssessmentQuestion)
    if specialty:
        q = q.filter(AssessmentQuestion.specialty == specialty)
    if difficulty:
        q = q.filter(AssessmentQuestion.difficulty == difficulty)
    if status:
        q = q.filter(AssessmentQuestion.status == status)
    if topic:
        q = q.filter(AssessmentQuestion.topic.ilike(f"%{topic}%"))
    if search:
        q = q.filter(AssessmentQuestion.question_text.ilike(f"%{search}%"))

    total = q.count()
    items = q.order_by(AssessmentQuestion.question_id).offset((page - 1) * page_size).limit(page_size).all()

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "results": [_q_out(item) for item in items],
    }


@router.post("/questions/upload")
def upload_questions(
    specialty: str = Form(...),
    uploaded_by: str = Form(...),
    file: UploadFile = File(...),
    dry_run: bool = Form(default=False),
    db: Session = Depends(get_db),
):
    """
    Upload .xlsx of questions for one specialty. Upserts by question_id.

    dry_run reports exactly what WOULD happen and writes nothing. The work was
    already all done before the commit, so the preview costs one rollback — and
    it is the honest way to warn about overwrites, since it counts the rows
    that will actually be replaced rather than estimating them client-side.
    """
    if specialty not in VALID_SPECIALTIES:
        raise HTTPException(status_code=400, detail=f"Unknown specialty: {specialty}")

    try:
        contents = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
    except Exception as exc:
        # "File is not a zip file" is what openpyxl says when handed a .xls or a
        # .csv, and it means nothing to a trainer who saved from Excel's default
        # dropdown. Name the actual mistake.
        hint = ("This does not look like a .xlsx file. If you saved from Excel, "
                "choose 'Excel Workbook (.xlsx)' — not .xls, .csv or Numbers."
                if "zip" in str(exc).lower() else str(exc))
        raise HTTPException(status_code=400, detail=f"Could not read the file. {hint}")

    headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]

    def col(name: str) -> Optional[int]:
        try:
            return headers.index(name) + 1
        except ValueError:
            return None

    col_map = {h: col(h) for h in TEMPLATE_HEADERS}
    missing = [h for h in ["Question_Text", "Option_A", "Option_B", "Option_C", "Option_D", "Correct_Answer", "Difficulty"] if not col_map.get(h)]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing}")

    created: List[str] = []
    updated: List[str] = []
    duplicates: List[str] = []  # blank-ID rows whose question_text already exists
    skipped: List[str] = []
    errors: List[str] = []

    # Read the bank ONCE. Both of these were inside the row loop: _next_qid
    # loaded every question_id for the specialty on every blank-ID row, and the
    # duplicate check queried per row. On a bank of a few thousand that is
    # millions of rows fetched for a single upload, and the request can outlive
    # the gateway — which surfaces to the trainer as a bare "upload failed"
    # with no reason, because a proxy timeout carries no JSON body.
    prefix = SPECIALTY_PREFIX.get(specialty, "QST")
    next_num = 0
    for (qid_existing,) in db.query(AssessmentQuestion.question_id).filter(
        AssessmentQuestion.question_id.like(f"{prefix}-%")
    ).all():
        m = re.search(r"-(\d+)$", qid_existing)
        if m:
            next_num = max(next_num, int(m.group(1)))

    existing_text = {
        t: q for q, t in db.query(
            AssessmentQuestion.question_id, AssessmentQuestion.question_text
        ).filter(AssessmentQuestion.specialty == specialty).all()
    }
    by_qid = {
        q.question_id: q for q in db.query(AssessmentQuestion).filter(
            AssessmentQuestion.question_id.like(f"{prefix}-%")
        ).all()
    }
    # Texts added earlier in THIS file, so one upload cannot smuggle in its own
    # duplicates — the per-row query used to catch that via autoflush.
    seen_text: Dict[str, str] = {}

    for row_idx in range(2, ws.max_row + 1):
        def cell_val(col_name: str) -> str:
            c = col_map.get(col_name)
            if c is None:
                return ""
            v = ws.cell(row=row_idx, column=c).value
            return str(v).strip() if v is not None else ""

        q_text = cell_val("Question_Text")
        if not q_text:
            continue

        correct = cell_val("Correct_Answer").upper()
        if correct not in VALID_ANSWERS:
            errors.append(f"Row {row_idx}: invalid Correct_Answer '{correct}'")
            continue

        difficulty = cell_val("Difficulty").capitalize()
        if difficulty == "":
            difficulty = "Medium"
        if difficulty not in VALID_DIFFICULTIES:
            errors.append(f"Row {row_idx}: invalid Difficulty '{difficulty}'")
            continue

        opt_a, opt_b = cell_val("Option_A"), cell_val("Option_B")
        opt_c, opt_d = cell_val("Option_C"), cell_val("Option_D")
        problem = option_problem(opt_a, opt_b, opt_c, opt_d)
        if problem:
            errors.append(f"Row {row_idx}: {problem}")
            continue

        # A blank cell is a default; a misspelled one is a mistake. Silently
        # recategorising "Conceptial" hides the typo in a bank nobody re-reads.
        q_type = cell_val("Question_Type")
        if not q_type:
            q_type = "Conceptual"
        elif q_type not in VALID_TYPES:
            errors.append(f"Row {row_idx}: invalid Question_Type '{q_type}' — expected one of: "
                          + ", ".join(sorted(VALID_TYPES)))
            continue

        # "Actve" used to fall through to Inactive, so a typo retired a
        # question with nothing on screen saying so.
        active_status = cell_val("Active_Status")
        if active_status.lower() in ("", "active", "1", "yes", "true"):
            status = "Active"
        elif active_status.lower() in ("inactive", "0", "no", "false", "retired"):
            status = "Inactive"
        else:
            errors.append(f"Row {row_idx}: invalid Active_Status '{active_status}' — "
                          "expected Active or Inactive")
            continue

        shuffle_val = cell_val("Shuffle_Options").lower()
        shuffle_options = shuffle_val not in ("no", "0", "false", "n")

        qid = cell_val("Question_ID")
        topic = cell_val("Topic") or None

        # A supplied id must belong to the specialty being uploaded. The check
        # below this only fires for ids that ALREADY exist, so a new
        # "SURG-999" uploaded under ICD10CM was created as an ICD10CM question
        # wearing a Surgery prefix.
        if qid and "-" in qid and not qid.upper().startswith(f"{prefix.upper()}-"):
            errors.append(f"Row {row_idx}: Question_ID '{qid}' does not match the "
                          f"'{prefix}-' prefix for {specialty} — leave it blank to "
                          "have one assigned")
            continue

        if not qid:
            # Duplicate text guard: if this exact question text already exists for this specialty, skip it
            match_id = existing_text.get(q_text) or seen_text.get(q_text)
            if match_id:
                duplicates.append(f"Row {row_idx}: identical question text already exists as {match_id} — skipped to prevent duplicate")
                continue
            next_num += 1
            qid = f"{prefix}-{next_num:03d}"
            seen_text[q_text] = qid
            # We generated this above the highest id in the bank, so there is
            # nothing to look up — skip the round trip.
            existing = None
        else:
            existing = by_qid.get(qid) or db.query(AssessmentQuestion).filter(
                AssessmentQuestion.question_id == qid).first()

        if existing:
            if existing.specialty != specialty:
                errors.append(f"Row {row_idx}: Question_ID '{qid}' belongs to specialty '{existing.specialty}', not '{specialty}' — skipped.")
                continue
            existing.question_text = q_text
            existing.option_a = opt_a
            existing.option_b = opt_b
            existing.option_c = opt_c
            existing.option_d = opt_d
            existing.correct_answer = correct
            existing.difficulty = difficulty
            existing.topic = topic
            existing.question_type = q_type
            existing.status = status
            existing.shuffle_options = shuffle_options
            existing.uploaded_by = uploaded_by
            updated.append(qid)
        else:
            aq = AssessmentQuestion(
                question_id=qid,
                specialty=specialty,
                question_text=q_text,
                option_a=opt_a,
                option_b=opt_b,
                option_c=opt_c,
                option_d=opt_d,
                correct_answer=correct,
                difficulty=difficulty,
                topic=topic,
                question_type=q_type,
                status=status,
                shuffle_options=shuffle_options,
                uploaded_by=uploaded_by,
            )
            db.add(aq)
            created.append(qid)

    if dry_run:
        # Nothing is written, including the audit row — a preview is not an
        # event worth recording.
        db.rollback()
    else:
        db.commit()
        _audit(db, uploaded_by, "upload", specialty,
               f"Created {len(created)}, updated {len(updated)}, "
               f"{len(duplicates)} duplicates skipped, {len(errors)} errors")

    total_stored = len(created) + len(updated)
    return {
        "dry_run": dry_run,
        "stored": total_stored,
        "stored_ids": created + updated,
        "created": len(created),
        "created_ids": created,
        "updated": len(updated),
        "updated_ids": updated,
        "duplicates": len(duplicates),
        "duplicate_warnings": duplicates,
        "skipped": len(skipped),
        "errors": errors,
    }


@router.put("/questions/{question_id}/status")
def update_question_status(
    question_id: str,
    status: str = Query(...),
    updated_by: str = Query(...),
    db: Session = Depends(get_db),
    _: None = Depends(require_passphrase),
):
    q = db.query(AssessmentQuestion).filter(AssessmentQuestion.question_id == question_id).first()
    if not q:
        raise HTTPException(status_code=404, detail="Question not found")
    if status not in {"Active", "Inactive"}:
        raise HTTPException(status_code=400, detail="status must be Active or Inactive")
    q.status = status
    db.commit()
    action = "retire" if status == "Inactive" else "reactivate"
    _audit(db, updated_by, action, q.specialty, f"{question_id} → {status}")
    return {"question_id": question_id, "status": status}


@router.put("/questions/{question_id}")
def update_question(
    question_id: str,
    payload: dict,
    db: Session = Depends(get_db),
    _: None = Depends(require_passphrase),
):
    q = db.query(AssessmentQuestion).filter(AssessmentQuestion.question_id == question_id).first()
    if not q:
        raise HTTPException(status_code=404, detail="Question not found")

    allowed = {"question_text", "option_a", "option_b", "option_c", "option_d",
               "correct_answer", "difficulty", "topic", "question_type"}
    updated_by = payload.get("updated_by", "Trainer")

    # Validate against the SAME rules the upload applies. This endpoint used to
    # setattr whatever it was handed, so the Question Bank screen was a second
    # door into the bank with no lock on it: the duplicate and blank options
    # the upload refuses could still be typed in here, along with a
    # correct_answer of "Z" or a difficulty of "banana".
    proposed = {k: (str(v).strip() if isinstance(v, str) else v)
                for k, v in payload.items() if k in allowed}

    def field(name: str):
        return proposed[name] if name in proposed else getattr(q, name)

    if not str(field("question_text") or "").strip():
        raise HTTPException(status_code=400, detail="Question_Text cannot be blank")

    problem = option_problem(field("option_a"), field("option_b"),
                             field("option_c"), field("option_d"))
    if problem:
        raise HTTPException(status_code=400, detail=f"Cannot save: {problem}")

    if str(field("correct_answer")).upper() not in VALID_ANSWERS:
        raise HTTPException(status_code=400,
                            detail=f"correct_answer must be one of: {', '.join(sorted(VALID_ANSWERS))}")
    if str(field("difficulty")) not in VALID_DIFFICULTIES:
        raise HTTPException(status_code=400,
                            detail=f"difficulty must be one of: {', '.join(sorted(VALID_DIFFICULTIES))}")
    if str(field("question_type")) not in VALID_TYPES:
        raise HTTPException(status_code=400,
                            detail=f"question_type must be one of: {', '.join(sorted(VALID_TYPES))}")

    for key, val in proposed.items():
        setattr(q, key, val)
    q.correct_answer = str(q.correct_answer).upper()

    db.commit()
    _audit(db, str(updated_by), "edit", q.specialty, f"Edited {question_id}")
    return _q_out(q)


@router.delete("/questions/{question_id}")
def delete_question(
    question_id: str,
    passphrase: str = Query(...),
    db: Session = Depends(get_db),
):
    if passphrase != settings.MASTER_ADMIN_PASSPHRASE:
        raise HTTPException(status_code=403, detail="Invalid passphrase")
    q = db.query(AssessmentQuestion).filter(AssessmentQuestion.question_id == question_id).first()
    if not q:
        raise HTTPException(status_code=404, detail="Question not found")
    db.delete(q)
    db.commit()
    return {"deleted": question_id}


@router.post("/questions/parse-standalone")
async def parse_standalone_questions(file: UploadFile = File(...)):
    """Parse an Excel file of questions and return them as JSON without saving to the bank."""
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
    except Exception as exc:
        # "File is not a zip file" is what openpyxl says when handed a .xls or a
        # .csv, and it means nothing to a trainer who saved from Excel's default
        # dropdown. Name the actual mistake.
        hint = ("This does not look like a .xlsx file. If you saved from Excel, "
                "choose 'Excel Workbook (.xlsx)' — not .xls, .csv or Numbers."
                if "zip" in str(exc).lower() else str(exc))
        raise HTTPException(status_code=400, detail=f"Could not read the file. {hint}")

    headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]

    def col(name: str):
        try: return headers.index(name) + 1
        except ValueError: return None

    col_map = {h: col(h) for h in TEMPLATE_HEADERS}
    missing = [h for h in ["Question_Text", "Option_A", "Option_B", "Option_C", "Option_D", "Correct_Answer"] if not col_map.get(h)]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing}")

    questions = []
    errors = []
    for row_idx in range(2, ws.max_row + 1):
        def cell_val(col_name: str) -> str:
            c = col_map.get(col_name)
            if not c: return ""
            v = ws.cell(row=row_idx, column=c).value
            return str(v).strip() if v is not None else ""

        q_text = cell_val("Question_Text")
        if not q_text:
            continue
        correct = cell_val("Correct_Answer").upper()
        if correct not in ("A", "B", "C", "D"):
            errors.append(f"Row {row_idx}: invalid Correct_Answer '{correct}' — skipped")
            continue
        difficulty = cell_val("Difficulty").capitalize() or "Medium"
        if difficulty not in ("Easy", "Medium", "Hard"):
            difficulty = "Medium"
        opt_a, opt_b = cell_val("Option_A"), cell_val("Option_B")
        opt_c, opt_d = cell_val("Option_C"), cell_val("Option_D")
        # Standalone questions never enter the bank, but they are shuffled and
        # graded by exactly the same code, so they carry the same defect.
        problem = option_problem(opt_a, opt_b, opt_c, opt_d)
        if problem:
            errors.append(f"Row {row_idx}: {problem} — skipped")
            continue
        shuffle_val = cell_val("Shuffle_Options").lower()
        shuffle_options = shuffle_val not in ("no", "0", "false", "n")
        questions.append({
            "question_text": q_text,
            "option_a": opt_a,
            "option_b": opt_b,
            "option_c": opt_c,
            "option_d": opt_d,
            "correct_answer": correct,
            "difficulty": difficulty,
            "topic": cell_val("Topic") or "",
            "shuffle_options": shuffle_options,
        })

    if not questions:
        raise HTTPException(status_code=400, detail="No valid question rows found. Check the template format.")

    return {"questions": questions, "count": len(questions), "errors": errors}


def _q_out(q: AssessmentQuestion) -> dict:
    return {
        "id": q.id,
        "question_id": q.question_id,
        "specialty": q.specialty,
        "question_text": q.question_text,
        "option_a": q.option_a,
        "option_b": q.option_b,
        "option_c": q.option_c,
        "option_d": q.option_d,
        "correct_answer": q.correct_answer,
        "difficulty": q.difficulty,
        "topic": q.topic,
        "question_type": q.question_type,
        "status": q.status,
        "shuffle_options": q.shuffle_options,
        "last_used_at": q.last_used_at.isoformat() if q.last_used_at else None,
        "uploaded_by": q.uploaded_by,
        "created_at": q.created_at.isoformat() if q.created_at else None,
    }
