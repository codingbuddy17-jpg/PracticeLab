"""
Assessment export router — PDF test papers, answer keys, summary.
"""
import io
import zipfile
from typing import List, Dict, Any

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)
from reportlab.lib import colors

import json as _json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_db
from routers.assessment_pkg.security import require_passphrase
from models import (
    GeneratedAssessment, GeneratedAssessmentStudent,
    AssessmentSession, AssessmentResponse, AssessmentResult,
)
from services.download_headers import content_disposition

router = APIRouter()


def _get_assessment_or_404(assessment_id: int, db: Session) -> GeneratedAssessment:
    a = db.query(GeneratedAssessment).filter(GeneratedAssessment.id == assessment_id).first()
    if not a:
        raise HTTPException(status_code=404, detail="Assessment not found")
    return a


def _build_student_pdf(assessment_name: str, student_label: str, questions: List[Dict[str, Any]]) -> bytes:
    """Build a clean test paper PDF (no answers shown)."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        rightMargin=0.75 * inch,
        leftMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=4,
        textColor=colors.HexColor("#1e1b4b"),
    )
    sub_style = ParagraphStyle(
        "sub",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.HexColor("#6b7280"),
        spaceAfter=16,
    )
    q_style = ParagraphStyle(
        "question",
        parent=styles["Normal"],
        fontSize=11,
        leading=14,
        spaceAfter=4,
        textColor=colors.HexColor("#111827"),
    )
    opt_style = ParagraphStyle(
        "option",
        parent=styles["Normal"],
        fontSize=10,
        leading=13,
        leftIndent=20,
        spaceAfter=2,
        textColor=colors.HexColor("#374151"),
    )
    ans_style = ParagraphStyle(
        "answer",
        parent=styles["Normal"],
        fontSize=10,
        leading=13,
        leftIndent=20,
        spaceAfter=2,
        textColor=colors.HexColor("#16a34a"),
    )

    story = [
        Paragraph(assessment_name, title_style),
        Paragraph(f"{student_label} &nbsp;|&nbsp; {len(questions)} Questions", sub_style),
    ]

    for i, q in enumerate(questions, start=1):
        story.append(Paragraph(f"<b>{i}.</b> {q['question_text']}", q_style))
        story.append(Paragraph(f"A. {q['option_a']}", opt_style))
        story.append(Paragraph(f"B. {q['option_b']}", opt_style))
        story.append(Paragraph(f"C. {q['option_c']}", opt_style))
        story.append(Paragraph(f"D. {q['option_d']}", opt_style))
        story.append(Spacer(1, 8))

    doc.build(story)
    buf.seek(0)
    return buf.read()


def _build_answer_key_pdf(assessment_name: str, students: List[GeneratedAssessmentStudent]) -> bytes:
    """Build consolidated answer key PDF for all students."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        rightMargin=0.75 * inch,
        leftMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=8,
        textColor=colors.HexColor("#1e1b4b"),
    )
    student_style = ParagraphStyle(
        "student",
        parent=styles["Heading2"],
        fontSize=12,
        spaceBefore=12,
        spaceAfter=6,
        textColor=colors.HexColor("#4f46e5"),
    )

    story = [Paragraph(f"Answer Key — {assessment_name}", title_style)]

    for student in students:
        story.append(Paragraph(student.student_label, student_style))
        questions = student.questions_json
        if isinstance(questions, str):
            import json
            questions = json.loads(questions)

        table_data = [["#", "QID", "Answer", "Specialty", "Topic"]]
        for i, q in enumerate(questions, start=1):
            table_data.append([
                str(i),
                q.get("question_id", ""),
                q.get("correct_answer", ""),
                q.get("specialty", ""),
                q.get("topic", "") or "",
            ])

        tbl = Table(
            table_data,
            colWidths=[0.4 * inch, 1.0 * inch, 0.7 * inch, 1.4 * inch, 2.2 * inch],
            repeatRows=1,
        )
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e5e7eb")),
            ("ALIGN", (0, 0), (2, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(tbl)
        story.append(PageBreak())

    doc.build(story)
    buf.seek(0)
    return buf.read()


@router.get("/{assessment_id}/export-pdf")
def export_student_pdfs(assessment_id: int, db: Session = Depends(get_db),
                        _: None = Depends(require_passphrase)):
    """Download ZIP of per-student test PDFs (no answers)."""
    assessment = _get_assessment_or_404(assessment_id, db)
    students = db.query(GeneratedAssessmentStudent).filter(
        GeneratedAssessmentStudent.assessment_id == assessment_id
    ).all()

    if not students:
        raise HTTPException(status_code=404, detail="No student data found for this assessment")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for student in students:
            questions = student.questions_json
            if isinstance(questions, str):
                import json
                questions = json.loads(questions)
            pdf_bytes = _build_student_pdf(assessment.assessment_name, student.student_label, questions)
            filename = f"{student.student_label}_{assessment.assessment_name.replace(' ', '_')}.pdf"
            zf.writestr(filename, pdf_bytes)

    zip_buf.seek(0)
    safe_name = assessment.assessment_name.replace(" ", "_")
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers=content_disposition(f"{safe_name}_tests.zip", "tests.zip"),
    )


@router.get("/{assessment_id}/export-answer-key")
def export_answer_key(assessment_id: int, db: Session = Depends(get_db),
                      _: None = Depends(require_passphrase)):
    """Download combined answer key PDF for all students."""
    assessment = _get_assessment_or_404(assessment_id, db)
    students = db.query(GeneratedAssessmentStudent).filter(
        GeneratedAssessmentStudent.assessment_id == assessment_id
    ).all()

    if not students:
        raise HTTPException(status_code=404, detail="No student data found")

    pdf_bytes = _build_answer_key_pdf(assessment.assessment_name, students)
    safe_name = assessment.assessment_name.replace(" ", "_")
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers=content_disposition(f"{safe_name}_answer_key.pdf", "answer_key.pdf"),
    )


@router.get("/{assessment_id}/export-responses.xlsx")
def export_responses_excel(assessment_id: int, db: Session = Depends(get_db)):
    """
    Download full per-question response breakdown for all coders — like MS Forms quiz export.
    One row per coder. Columns: Coder Name, Employee ID, Submitted At, Score %, Time (min),
    then for each question: Q# Text | Selected | Correct Answer | Result (✓/✗)
    """
    assessment = _get_assessment_or_404(assessment_id, db)

    # Load student question sets (question text + correct answer per slot)
    students = db.query(GeneratedAssessmentStudent).filter(
        GeneratedAssessmentStudent.assessment_id == assessment_id
    ).all()
    slot_questions: dict[str, list] = {}
    for s in students:
        qs = s.questions_json
        if isinstance(qs, str):
            try:
                qs = _json.loads(qs)
            except Exception:
                qs = []
        slot_questions[s.student_label] = qs if isinstance(qs, list) else []

    # Load all sessions for this assessment
    sessions = db.query(AssessmentSession).filter(
        AssessmentSession.assessment_id == assessment_id
    ).order_by(AssessmentSession.coder_name).all()

    if not sessions:
        raise HTTPException(status_code=404, detail="No sessions found for this assessment")

    # Build response lookup: session_id -> {question_index: selected_answer}
    all_session_ids = [s.id for s in sessions]
    responses = db.query(AssessmentResponse).filter(
        AssessmentResponse.session_id.in_(all_session_ids)
    ).all()
    resp_map: dict[int, dict[int, str | None]] = {}
    for r in responses:
        resp_map.setdefault(r.session_id, {})[r.question_index] = r.selected_answer

    # Results lookup: session_id -> AssessmentResult
    results = db.query(AssessmentResult).filter(
        AssessmentResult.session_id.in_(all_session_ids)
    ).all()
    result_map = {r.session_id: r for r in results}

    # Find max question count across all coders
    max_q = max((len(slot_questions.get(s.coder_name, [])) for s in sessions), default=0)

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Responses"

    # Styles
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor="1E3A5F")
    q_hdr_fill = PatternFill("solid", fgColor="374151")
    pass_fill = PatternFill("solid", fgColor="D1FAE5")
    fail_fill = PatternFill("solid", fgColor="FEE2E2")
    correct_fill = PatternFill("solid", fgColor="ECFDF5")
    wrong_fill   = PatternFill("solid", fgColor="FFF1F2")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    def hdr(cell, val, fill=None):
        cell.value = val
        cell.font = hdr_font
        cell.fill = fill or hdr_fill
        cell.alignment = center
        cell.border = border

    # ── Header row 1: fixed columns + question group labels ──────────────────
    fixed_cols = ["Coder Name", "Employee ID", "Status", "Submitted At",
                  "Score %", "Correct", "Total Q", "Time Taken", "Time (min)"]
    for ci, label in enumerate(fixed_cols, 1):
        hdr(ws.cell(row=1, column=ci), label)

    q_start_col = len(fixed_cols) + 1
    for qi in range(max_q):
        base = q_start_col + qi * 4
        hdr(ws.cell(row=1, column=base),     f"Q{qi+1} — Question",    q_hdr_fill)
        hdr(ws.cell(row=1, column=base + 1), f"Q{qi+1} Selected",      q_hdr_fill)
        hdr(ws.cell(row=1, column=base + 2), f"Q{qi+1} Correct Ans",   q_hdr_fill)
        hdr(ws.cell(row=1, column=base + 3), f"Q{qi+1} Result",        q_hdr_fill)
        ws.merge_cells(start_row=1, start_column=base, end_row=1, end_column=base)

    ws.row_dimensions[1].height = 36

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_i, sess in enumerate(sessions, 2):
        result = result_map.get(sess.id)
        resp   = resp_map.get(sess.id, {})
        questions = slot_questions.get(sess.coder_name, [])

        submitted_str = sess.submitted_at.strftime("%Y-%m-%d %H:%M") if sess.submitted_at else "—"
        secs = result.time_taken_seconds if result and result.time_taken_seconds else None
        time_hms  = (f"{secs // 3600:02}:{(secs % 3600) // 60:02}:{secs % 60:02}" if secs else "—")
        time_min  = round(secs / 60, 1) if secs else "—"
        score_pct = round(result.score_pct, 1) if result else None
        correct   = result.correct_count if result else "—"
        total_q   = result.total_questions if result else len(questions)
        status    = sess.status.capitalize() if sess.status else "—"

        row_fill = pass_fill if (score_pct is not None and score_pct >= 80) else (fail_fill if score_pct is not None else None)

        def cell(col, val, fill=None, align=None):
            c = ws.cell(row=row_i, column=col, value=val)
            c.border = border
            c.alignment = align or center
            if fill:
                c.fill = fill
            if row_fill and not fill:
                c.fill = row_fill
            return c

        cell(1, sess.coder_name, align=left)
        cell(2, sess.employee_id or "—")
        cell(3, status)
        cell(4, submitted_str)
        cell(5, f"{score_pct}%" if score_pct is not None else "—")
        cell(6, correct)
        cell(7, total_q)
        cell(8, time_hms)
        cell(9, time_min)

        for qi, q in enumerate(questions):
            base = q_start_col + qi * 4
            selected = resp.get(qi)
            correct_ans = q.get("correct_answer", "")

            # Map letter to full option text
            opt_map = {"A": q.get("option_a",""), "B": q.get("option_b",""),
                       "C": q.get("option_c",""), "D": q.get("option_d","")}
            selected_text  = f"{selected} — {opt_map.get(selected,'')}" if selected else "—"
            correct_text   = f"{correct_ans} — {opt_map.get(correct_ans,'')}" if correct_ans else "—"
            is_correct     = selected == correct_ans if selected else None
            result_symbol  = "✓ Correct" if is_correct else ("✗ Wrong" if is_correct is False else "—")
            q_fill = correct_fill if is_correct else (wrong_fill if is_correct is False else None)

            c1 = ws.cell(row=row_i, column=base, value=q.get("question_text", ""))
            c1.border = border; c1.alignment = left
            c2 = ws.cell(row=row_i, column=base+1, value=selected_text)
            c2.border = border; c2.alignment = left; c2.fill = q_fill or PatternFill()
            c3 = ws.cell(row=row_i, column=base+2, value=correct_text)
            c3.border = border; c3.alignment = left
            c4 = ws.cell(row=row_i, column=base+3, value=result_symbol)
            c4.border = border; c4.alignment = center; c4.fill = q_fill or PatternFill()
            if is_correct is True:
                c4.font = Font(bold=True, color="166534")
            elif is_correct is False:
                c4.font = Font(bold=True, color="991B1B")

        ws.row_dimensions[row_i].height = 40

    # ── Column widths ─────────────────────────────────────────────────────────
    fixed_widths = [22, 14, 12, 18, 10, 10, 10, 12, 10]
    for ci, w in enumerate(fixed_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for qi in range(max_q):
        base = q_start_col + qi * 4
        ws.column_dimensions[get_column_letter(base)].width = 48      # question text
        ws.column_dimensions[get_column_letter(base+1)].width = 28    # selected
        ws.column_dimensions[get_column_letter(base+2)].width = 28    # correct
        ws.column_dimensions[get_column_letter(base+3)].width = 12    # ✓/✗

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = assessment.assessment_name.replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=content_disposition(f"{safe}_responses.xlsx", "responses.xlsx"),
    )


@router.get("/{assessment_id}/session/{session_id}/review")
def get_session_review(assessment_id: int, session_id: int, db: Session = Depends(get_db)):
    """
    Return per-question review for one coder session — question, all options,
    what they selected, correct answer, whether they got it right.
    Used for the in-app post-submission review screen.
    """
    sess = db.query(AssessmentSession).filter(
        AssessmentSession.id == session_id,
        AssessmentSession.assessment_id == assessment_id,
    ).first()
    if not sess:
        raise HTTPException(status_code=404, detail="Session not found")
    if sess.status not in ("submitted", "auto_submitted"):
        raise HTTPException(status_code=400, detail="Session not yet submitted")

    # Get questions for this coder
    slot = db.query(GeneratedAssessmentStudent).filter(
        GeneratedAssessmentStudent.id == sess.student_slot_id
    ).first()
    questions = []
    if slot:
        qs = slot.questions_json
        if isinstance(qs, str):
            try:
                qs = _json.loads(qs)
            except Exception:
                qs = []
        questions = qs if isinstance(qs, list) else []

    # Get responses
    responses = db.query(AssessmentResponse).filter(
        AssessmentResponse.session_id == session_id
    ).order_by(AssessmentResponse.question_index).all()
    resp_by_idx = {r.question_index: r for r in responses}

    # Get result
    result = db.query(AssessmentResult).filter(
        AssessmentResult.session_id == session_id
    ).first()

    review_questions = []
    for qi, q in enumerate(questions):
        resp = resp_by_idx.get(qi)
        selected = resp.selected_answer if resp else None
        correct  = q.get("correct_answer", "")
        opt_map  = {"A": q.get("option_a",""), "B": q.get("option_b",""),
                    "C": q.get("option_c",""), "D": q.get("option_d","")}
        review_questions.append({
            "index": qi,
            "question_id": q.get("question_id",""),
            "question_text": q.get("question_text",""),
            "difficulty": q.get("difficulty",""),
            "topic": q.get("topic",""),
            "options": [
                {"letter": k, "text": v,
                 "selected": k == selected,
                 "is_correct": k == correct}
                for k, v in opt_map.items()
            ],
            "selected_answer": selected,
            "correct_answer": correct,
            "is_correct": selected == correct if selected else False,
            "answered": selected is not None,
        })

    return {
        "assessment_id": assessment_id,
        "session_id": session_id,
        "coder_name": sess.coder_name,
        "employee_id": sess.employee_id,
        "submitted_at": sess.submitted_at.isoformat() if sess.submitted_at else None,
        "score_pct": result.score_pct if result else None,
        "correct_count": result.correct_count if result else None,
        "total_questions": result.total_questions if result else len(questions),
        "time_taken_seconds": result.time_taken_seconds if result else None,
        "questions": review_questions,
    }


@router.get("/{assessment_id}/summary")
def get_assessment_summary(assessment_id: int, db: Session = Depends(get_db)):
    """Return stored assessment data for history/detail view."""
    assessment = _get_assessment_or_404(assessment_id, db)
    students = db.query(GeneratedAssessmentStudent).filter(
        GeneratedAssessmentStudent.assessment_id == assessment_id
    ).all()

    q_count = 0
    if students:
        first_qs = students[0].questions_json
        if isinstance(first_qs, list):
            q_count = len(first_qs)
        elif isinstance(first_qs, str):
            import json
            try:
                q_count = len(json.loads(first_qs))
            except Exception:
                q_count = 0

    return {
        "id": assessment.id,
        "assessment_name": assessment.assessment_name,
        "student_count": assessment.student_count,
        "generated_by": assessment.generated_by,
        "generated_at": assessment.generated_at.isoformat() if assessment.generated_at else None,
        "config_id": assessment.config_id,
        "questions_per_student": q_count,
        "students": [s.student_label for s in students],
    }
