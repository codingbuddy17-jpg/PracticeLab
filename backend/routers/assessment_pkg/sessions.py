"""
Assessment Sessions router — trainer-side session management.
Handles creating sessions per coder, listing status, coder list upload.
"""
import io
import random
import string
from datetime import datetime, timezone, timedelta
from services.timeutil import as_utc
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from database import get_db
from models import (
    GeneratedAssessment, GeneratedAssessmentStudent, AssessmentSession,
    AssessmentResult, AssessmentResponse,
)

router = APIRouter()

SESSION_EXPIRY_HOURS = 8


def _make_token() -> str:
    """Generate a readable unique session token like ASM-X7K2M9QP."""
    chars = string.ascii_uppercase + string.digits
    suffix = "".join(random.choices(chars, k=8))
    return f"ASM-{suffix}"


class CoderItem(BaseModel):
    coder_name: str
    employee_id: Optional[str] = None


class CreateSessionsRequest(BaseModel):
    assessment_id: int
    duration_minutes: int
    coders: List[CoderItem]


@router.post("/sessions/create")
def create_sessions(req: CreateSessionsRequest, db: Session = Depends(get_db)):
    """Create one session token per coder for a generated assessment."""
    assessment = db.query(GeneratedAssessment).filter(
        GeneratedAssessment.id == req.assessment_id
    ).first()
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")

    if req.duration_minutes < 5 or req.duration_minutes > 480:
        raise HTTPException(status_code=400, detail="Duration must be between 5 and 480 minutes")

    if not req.coders:
        raise HTTPException(status_code=400, detail="At least one coder required")

    # Check if sessions already exist for this assessment
    existing = db.query(AssessmentSession).filter(
        AssessmentSession.assessment_id == req.assessment_id
    ).count()
    if existing > 0:
        raise HTTPException(
            status_code=409,
            detail=f"Sessions already exist for this assessment ({existing} sessions). Delete them first to regenerate."
        )

    expires_at = datetime.now(timezone.utc) + timedelta(hours=SESSION_EXPIRY_HOURS)

    created = []
    for coder in req.coders:
        name = coder.coder_name.strip()
        if not name:
            continue
        # Ensure token uniqueness
        token = None
        for _ in range(10):
            candidate = _make_token()
            if not db.query(AssessmentSession).filter(AssessmentSession.session_token == candidate).first():
                token = candidate
                break
        if not token:
            raise HTTPException(status_code=500, detail="Failed to generate unique session token. Please retry.")

        # Try to find this coder's dedicated student slot by name match
        slot = db.query(GeneratedAssessmentStudent).filter(
            GeneratedAssessmentStudent.assessment_id == req.assessment_id,
            GeneratedAssessmentStudent.student_label == name,
        ).first()

        session = AssessmentSession(
            session_token=token,
            assessment_id=req.assessment_id,
            coder_name=name,
            employee_id=coder.employee_id.strip() if coder.employee_id else None,
            duration_minutes=req.duration_minutes,
            expires_at=expires_at,
            status="pending",
            student_slot_id=slot.id if slot else None,
        )
        db.add(session)
        created.append({
            "coder_name": name,
            "employee_id": coder.employee_id,
            "session_token": token,
        })

    db.commit()
    return {
        "assessment_id": req.assessment_id,
        "duration_minutes": req.duration_minutes,
        "expires_at": expires_at.isoformat(),
        "sessions": created,
    }


@router.get("/sessions/parse-coders")
def download_coder_template():
    """Download a blank XLSX template for bulk coder upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coders"
    headers = ["Coder_Name", "Employee_ID"]
    fill = PatternFill("solid", fgColor="4F46E5")
    font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = font
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
    # Sample rows
    ws.cell(row=2, column=1, value="Jane Smith")
    ws.cell(row=2, column=2, value="EMP001")
    ws.cell(row=3, column=1, value="John Doe")
    ws.cell(row=3, column=2, value="EMP002")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="coder_list_template.xlsx"'},
    )


@router.post("/sessions/parse-coders")
async def parse_coder_file(file: UploadFile = File(...)):
    """Parse an uploaded XLSX coder list and return [{coder_name, employee_id}]."""
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {exc}")

    headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]

    def col(name: str) -> Optional[int]:
        try:
            return headers.index(name) + 1
        except ValueError:
            return None

    name_col = col("Coder_Name")
    id_col = col("Employee_ID")
    if not name_col:
        raise HTTPException(status_code=400, detail="Missing required column: Coder_Name")

    coders = []
    for row in range(2, ws.max_row + 1):
        name = str(ws.cell(row=row, column=name_col).value or "").strip()
        if not name:
            continue
        emp_id = str(ws.cell(row=row, column=id_col).value or "").strip() if id_col else ""
        coders.append({"coder_name": name, "employee_id": emp_id or None})

    return {"coders": coders, "count": len(coders)}


@router.get("/{assessment_id}/sessions")
def list_sessions(assessment_id: int, db: Session = Depends(get_db)):
    """List all sessions for an assessment with current status and scores."""
    sessions = db.query(AssessmentSession).filter(
        AssessmentSession.assessment_id == assessment_id
    ).order_by(AssessmentSession.coder_name).all()

    # Which sessions carry a trainer correction, in one query rather than per
    # row — the badge is on every row, so a per-row lookup would be an N+1.
    corrected_ids = {
        r[0] for r in db.query(AssessmentResponse.session_id)
        .filter(AssessmentResponse.session_id.in_([s.id for s in sessions] or [-1]),
                AssessmentResponse.override_is_correct.isnot(None))
        .distinct().all()
    }

    now = datetime.now(timezone.utc)
    rows = []
    for s in sessions:
        # A pending session that never started simply lapses — nothing to score.
        effective_status = "expired" if (
            s.status == "pending" and now > as_utc(s.expires_at)
        ) else s.status

        result = db.query(AssessmentResult).filter(AssessmentResult.session_id == s.id).first()
        rows.append({
            "session_id": s.id,
            "session_token": s.session_token,
            "coder_name": s.coder_name,
            "employee_id": s.employee_id,
            "status": effective_status,
            "duration_minutes": s.duration_minutes,
            "expires_at": s.expires_at.isoformat(),
            "started_at": s.started_at.isoformat() if s.started_at else None,
            "submitted_at": s.submitted_at.isoformat() if s.submitted_at else None,
            "auto_submitted": s.auto_submitted,
            "corrected": s.id in corrected_ids,
            "score_pct": result.score_pct if result else None,
            "correct_count": result.correct_count if result else None,
            "total_questions": result.total_questions if result else None,
            "time_taken_seconds": result.time_taken_seconds if result else None,
        })

    # The bar this paper was judged against, so the screen colours scores
    # against the real mark instead of a hardcoded 80/90 that may belong to a
    # different assessment entirely.
    assessment = db.query(GeneratedAssessment).filter(
        GeneratedAssessment.id == assessment_id).first()
    pass_threshold = float(assessment.pass_threshold) if (
        assessment and assessment.pass_threshold) else 90.0

    return {
        "assessment_id": assessment_id,
        "pass_threshold": pass_threshold,
        "sessions": rows,
    }


@router.delete("/{assessment_id}/sessions")
def delete_sessions(assessment_id: int, db: Session = Depends(get_db)):
    """Delete all sessions for an assessment (only if none submitted yet)."""
    sessions = db.query(AssessmentSession).filter(
        AssessmentSession.assessment_id == assessment_id
    ).all()
    submitted = sum(1 for s in sessions if s.status == "submitted")

    # A scored session is a result regardless of what its status column says —
    # status is one flag and can be stale, while an AssessmentResult row is the
    # work itself. Checking only the status would delete graded work whose
    # status had not been written for any reason.
    graded = db.query(AssessmentResult).filter(
        AssessmentResult.session_id.in_([s.id for s in sessions] or [-1])
    ).count()

    if submitted or graded:
        raise HTTPException(
            status_code=409,
            detail=(f"Cannot delete — {max(submitted, graded)} session(s) have results. "
                    "Deleting would lose them."),
        )
    db.query(AssessmentSession).filter(
        AssessmentSession.assessment_id == assessment_id
    ).delete(synchronize_session=False)
    db.commit()
    return {"deleted": True}
