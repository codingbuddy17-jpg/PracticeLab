"""E/M MDM scoring module — answer key management, grading engine, scoring config."""
from __future__ import annotations

import json
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import text

from database import get_db
from models import Chart, Specialty
from services.excel_service import parse_em_answer_key_upload
from .shared import MASTER_PASSPHRASE, _find_chart

router = APIRouter()

# ── Specialties that use MDM-based E/M scoring ────────────────────────────────
EM_SPECIALTIES = {Specialty.EM, Specialty.ED_PROFEE}


def _is_em(specialty: Specialty) -> bool:
    return specialty in EM_SPECIALTIES


# ── COPA level derivation ─────────────────────────────────────────────────────

def derive_copa_level(d: dict) -> str:
    """Derive COPA level from element counts per 2023 AMA MDM table."""
    if d.get("copa_threat_to_life", 0) >= 1 or d.get("copa_chronic_severe", 0) >= 1:
        return "High"
    if (d.get("copa_chronic_exacerbation", 0) >= 1
            or d.get("copa_stable_chronic", 0) >= 2
            or d.get("copa_undiagnosed_new", 0) >= 1
            or d.get("copa_acute_systemic", 0) >= 1
            or d.get("copa_acute_complicated_injury", 0) >= 1):
        return "Moderate"
    if (d.get("copa_self_limited", 0) >= 2
            or d.get("copa_stable_chronic", 0) >= 1
            or d.get("copa_acute_uncomplicated", 0) >= 1
            or d.get("copa_stable_acute", 0) >= 1):
        return "Low"
    if d.get("copa_self_limited", 0) >= 1:
        return "Minimal"
    return "Minimal"


# ── Data Review level derivation ──────────────────────────────────────────────

def _cat1_count(d: dict) -> int:
    return (
        (1 if d.get("dr_prior_external_notes", 0) >= 1 else 0)
        + (1 if d.get("dr_review_test_results", 0) >= 1 else 0)
        + (1 if d.get("dr_order_tests", 0) >= 1 else 0)
        + (1 if d.get("dr_independent_historian") else 0)
    )


def derive_dr_level(d: dict) -> str:
    """Derive Data Review level from elements per 2023 AMA MDM table."""
    cat1 = _cat1_count(d)
    cat2 = bool(d.get("dr_independent_interpretation"))
    cat3 = bool(d.get("dr_external_discussion"))

    cat1_mod = cat1 >= 3
    cats_met = sum([cat1_mod, cat2, cat3])

    if cats_met >= 2:
        return "Extensive"
    if cats_met >= 1:
        return "Moderate"
    if cat1 >= 2 or cat2 or cat3:
        return "Limited"
    return "Minimal"


# ── Risk level derivation ─────────────────────────────────────────────────────

def derive_risk_level(d: dict) -> str:
    """Derive Risk level — highest single element present wins."""
    high_fields = [
        "risk_drug_intensive_monitoring", "risk_elective_major_with_factors",
        "risk_emergency_major_surgery", "risk_hospitalization_escalation",
        "risk_dnr_deescalate", "risk_parenteral_controlled",
    ]
    moderate_fields = [
        "risk_prescription_drug_mgmt", "risk_minor_surgery_with_factors",
        "risk_elective_major_no_factors", "risk_hospitalization", "risk_sdoh",
    ]
    if any(d.get(f) for f in high_fields):
        return "High"
    if any(d.get(f) for f in moderate_fields):
        return "Moderate"
    if d.get("risk_low"):
        return "Low"
    return "Minimal"


# ── Overall MDM level (2-of-3 rule) ──────────────────────────────────────────

_LEVEL_ORDER = {"Minimal": 0, "Low": 1, "Moderate": 2, "Extensive": 3, "High": 3}
_LEVEL_NAMES = ["Minimal", "Low", "Moderate", "High"]


def derive_mdm_level(copa: str, dr: str, risk: str) -> str:
    """2-of-3 rule: overall level = second-lowest of the three component levels."""
    scores = sorted([_LEVEL_ORDER.get(copa, 0), _LEVEL_ORDER.get(dr, 0), _LEVEL_ORDER.get(risk, 0)])
    return _LEVEL_NAMES[min(scores[1], 3)]


# ── EM code → MDM level mapping ───────────────────────────────────────────────

_EM_CODE_LEVEL = {
    "99202": "Straightforward", "99212": "Straightforward",
    "99203": "Low", "99213": "Low",
    "99204": "Moderate", "99214": "Moderate",
    "99205": "High", "99215": "High",
    "99281": "Minimal", "99282": "Straightforward",
    "99283": "Low", "99284": "Moderate", "99285": "High",
}


def em_code_to_level(code: str) -> Optional[str]:
    return _EM_CODE_LEVEL.get(code.strip())


# ── Encounter categories ──────────────────────────────────────────────────────
#
# Not every E/M encounter is levelled by MDM. A preventive visit is levelled by
# the patient's age and whether they are new or established; critical care is
# levelled by total time; a nurse visit has no level at all. The MDM tables
# were being demanded for all of them, because the three level columns are NOT
# NULL — so a trainer keying a preventive visit had to invent a COPA level, and
# the coder then had to guess the same invention to score the 30 reasoning
# points. That tests neither of them on anything real.
#
# Category comes from the E/M code, which is the one thing both the key and the
# coder always state.

OFFICE = "office"
INPATIENT_OBSERVATION = "inpatient_observation"
EMERGENCY = "emergency"
PREVENTIVE = "preventive"
CRITICAL_CARE = "critical_care"
OTHER = "other"

EM_CATEGORIES = (OFFICE, INPATIENT_OBSERVATION, EMERGENCY,
                 PREVENTIVE, CRITICAL_CARE, OTHER)

EM_CATEGORY_LABELS = {
    OFFICE: "Office / outpatient",
    INPATIENT_OBSERVATION: "Inpatient & observation",
    EMERGENCY: "Emergency department",
    PREVENTIVE: "Preventive medicine",
    CRITICAL_CARE: "Critical care",
    OTHER: "Other",
}

# The categories whose code selection is driven by the 2023 MDM tables, and so
# the only ones where the COPA / Data Review / Risk checkoff is the work being
# assessed.
MDM_CATEGORIES = {OFFICE, INPATIENT_OBSERVATION, EMERGENCY}

# Explicit code sets, per CPT 2026. Ranges are spelled out rather than
# range-matched on the numeric value: E/M numbering is not contiguous by
# category (99238 is a discharge, 99241 was deleted), and a range test would
# quietly swallow codes added between them in a future edition.
_CATEGORY_CODES = {
    OFFICE: {
        "99202", "99203", "99204", "99205",
        "99211", "99212", "99213", "99214", "99215",
    },
    INPATIENT_OBSERVATION: {
        "99221", "99222", "99223",          # initial inpatient / observation
        "99231", "99232", "99233",          # subsequent
        "99234", "99235", "99236",          # admit & discharge same day
        "99238", "99239",                   # discharge day management
    },
    EMERGENCY: {"99281", "99282", "99283", "99284", "99285"},
    PREVENTIVE: {
        "99381", "99382", "99383", "99384", "99385", "99386", "99387",  # new
        "99391", "99392", "99393", "99394", "99395", "99396", "99397",  # established
        "99401", "99402", "99403", "99404",          # counselling, individual
        "99411", "99412", "99429",
    },
    CRITICAL_CARE: {
        "99291", "99292",                            # adult, time-based
        "99468", "99469", "99471", "99472", "99475", "99476",   # neonatal/paediatric
        "99477", "99478", "99479", "99480",           # intensive, low birth weight
    },
}

_CODE_TO_CATEGORY = {code: cat for cat, codes in _CATEGORY_CODES.items() for code in codes}

# Critical care is billed in one initial unit plus add-on units. 99292 is the
# add-on, so a long encounter is 99291 once and 99292 as many times as the time
# supports — the units column on the CPT line is how that count is stated.
CRITICAL_CARE_INITIAL = "99291"
CRITICAL_CARE_ADDON = "99292"


def em_category(code: str) -> str:
    """
    Which kind of encounter this E/M code represents.

    Anything unrecognised is OTHER, never a guess. Home visits, nursing
    facility, transitional care and telephone/online codes all land there
    deliberately: they are graded on the codes alone, which is exactly what a
    category we cannot model should do.
    """
    return _CODE_TO_CATEGORY.get(str(code or "").strip().upper(), OTHER)


def category_uses_mdm(category: str) -> bool:
    return category in MDM_CATEGORIES


def resolve_category(stored, em_code: str) -> str:
    """
    The key's own category if it states one, otherwise derived from its code.

    Keys written before categories existed have nothing stored, and every one
    of them was an office, ED or inpatient chart — the only kinds the form
    could express — so deriving gives them the category they already had.
    """
    c = str(stored or "").strip().lower()
    return c if c in EM_CATEGORIES else em_category(em_code)


def critical_care_units(cpt_rows: list, em_code: str) -> int:
    """
    How many add-on units of 99292 were claimed.

    The add-on can be a repeated line or one line carrying units; both mean the
    same thing on a claim, so both count the same here.
    """
    total = 0
    for row in cpt_rows:
        if str(row.get("code", "")).strip().upper() != CRITICAL_CARE_ADDON:
            continue
        total += norm_units_or_one(row.get("units"))
    if str(em_code or "").strip().upper() == CRITICAL_CARE_ADDON:
        total += 1
    return total


def _minutes(raw) -> Optional[int]:
    """A minute count, or None when nothing was stated. 0 is not a time."""
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None
    try:
        n = int(float(str(raw).strip()))
    except (TypeError, ValueError):
        return None
    return n if n > 0 else None


def norm_units_or_one(raw) -> int:
    from services.grading_engine import norm_units
    return norm_units(raw)


# ── Weight distribution ───────────────────────────────────────────────────────

def applicable_weights(cfg: dict, category: str, has_cpts: bool,
                       level_method: str = "MDM") -> dict:
    """
    The weights that apply to THIS chart, renormalised to total 100.

    Every chart has to be scored out of 100 or the numbers stop comparing. A
    preventive visit has no MDM component, so leaving the 30 reasoning points
    in the denominator caps it at 70% and a coder who did everything right
    fails; averaging it against an office chart then compares two different
    denominators and reports a difference that is purely arithmetic.

    Renormalising happens WITHIN each line, not across the whole chart. Line 1
    is coding accuracy and Line 2 is reasoning accuracy, and their 70/30 split
    is a statement about what the assessment values — rescaling across both
    would let a chart with no procedures quietly reweight reasoning upward.
    When Line 2 has nothing applicable at all, its points fold into Line 1,
    because the coding is then the whole of the work.

    Returns a dict of component -> weight. Components absent from the dict do
    not apply to this chart and are not scored.
    """
    line1 = {"em_level": cfg["em_level_weight"], "dx": cfg["dx_weight"]}
    # A chart with no procedures in the key has no CPT line to score.
    if has_cpts:
        line1["cpt"] = cfg["cpt_weight"]

    line2: dict = {}
    if category == CRITICAL_CARE:
        # No second line. The time is not separate reasoning to be scored on
        # its own — it is part of picking the code, so it rides along with the
        # E/M component and the chart comes out thirds: code+time, Dx, and
        # procedures where there are any.
        pass
    elif category_uses_mdm(category):
        if (level_method or "MDM").upper() == "TIME":
            line2["time"] = 1.0
        else:
            line2["copa"] = cfg["copa_weight"]
            line2["dr"] = cfg["dr_weight"]
            line2["risk"] = cfg["risk_weight"]
    # Preventive and Other get neither: nothing beyond the codes is assessed.

    l1_target = cfg["line1_weight"] + (0.0 if line2 else cfg["line2_weight"])
    l2_target = cfg["line2_weight"] if line2 else 0.0

    def _scale(parts, target):
        total = sum(parts.values())
        if total <= 0:
            return {}
        return {k: round(v * target / total, 4) for k, v in parts.items()}

    return {**_scale(line1, l1_target), **_scale(line2, l2_target)}


# ── Scoring engine ────────────────────────────────────────────────────────────

def _j(v):
    if isinstance(v, str):
        try:
            return json.loads(v)
        except Exception:
            return []
    return v or []


def _clean_codes(lst: list) -> list:
    return [str(c).strip().upper() for c in lst if c and str(c).strip().lower() not in ("", "none")]


# ── Time-based levelling (2021+ office/outpatient E/M) ───────────────────────
#
# A visit may be levelled by total time on the date of encounter OR by MDM —
# the coder chooses. ED codes (99281-99285) have no time option and are always
# MDM, so the Time path is never offered for them.
#
# Total time on date of encounter, per AMA CPT.
EM_TIME_BANDS = {
    # New patient
    "99202": (15, 29), "99203": (30, 44), "99204": (45, 59), "99205": (60, 74),
    # Established patient
    "99212": (10, 19), "99213": (20, 29), "99214": (30, 39), "99215": (40, 54),
}


def time_supports_code(total_time, em_code: str) -> bool:
    """Does the documented time fall in the band supporting this code?

    Graded on band rather than exact digits: coders read 'approximately 40
    minutes' off a chart, and failing them on transcription would test the
    wrong skill.
    """
    band = EM_TIME_BANDS.get((em_code or "").strip())
    if not band or total_time in (None, ""):
        return False
    try:
        mins = int(float(total_time))
    except (TypeError, ValueError):
        return False
    return band[0] <= mins <= band[1]


def code_supports_time(em_code: str) -> bool:
    """True when the code can be levelled by time at all (office/outpatient)."""
    return (em_code or "").strip() in EM_TIME_BANDS


def _canon_pointers(raw) -> list:
    """Shared cleaner — numeric, deduped, capped at four per line."""
    from services.grading_engine import canonical_pointers
    return canonical_pointers(raw)


def normalise_cpts(raw) -> list:
    """
    Normalise procedure CPTs to [{code, modifier, pointers, units}].

    Accepts both shapes so no migration is needed: legacy keys stored
    "code:modifier" strings inside the JSON column, newer ones store dicts
    carrying diagnosis pointers and units.

    `units` is only present when the source stated it. Absence means "this key
    predates units", which grading treats as "do not grade units" — different
    from an explicit 1.
    """
    out = []
    for item in _j(raw):
        if isinstance(item, dict):
            code = str(item.get("code") or "").strip().upper()
            if not code:
                continue
            row = {
                "code": code,
                "modifier": str(item.get("modifier") or "").strip().upper(),
                "pointers": _canon_pointers(item.get("pointers")),
            }
            if "units" in item and item.get("units") not in (None, ""):
                from services.grading_engine import norm_units
                row["units"] = norm_units(item.get("units"))
            out.append(row)
            continue
        s = str(item).strip().upper()
        if not s or s == "NONE":
            continue
        parts = s.split(":")
        ptrs = []
        if len(parts) > 2 and parts[2].strip():
            ptrs = _canon_pointers(parts[2].split(","))
        out.append({
            "code": parts[0].strip(),
            "modifier": parts[1].strip() if len(parts) > 1 else "",
            "pointers": ptrs,
        })
    return out


def _sanitise_level_method(method, em_code) -> str:
    """
    Normalise the levelling method, forcing MDM for codes that cannot be
    levelled by time. ED visit codes have no typical times in CPT, and the coder
    form hides the Time control for them — a TIME key would be unanswerable.
    """
    m = (method or "MDM").upper().strip()
    if m != "TIME":
        return "MDM"
    return "TIME" if code_supports_time(em_code) else "MDM"


def grade_em_chart(ak: dict, sub: dict, cfg: dict, overcoding_penalty: bool = True) -> dict:
    """
    Grade one E/M chart submission against its answer key.

    ak  — answer key dict (from em_answer_keys row)
    sub — coder submission dict
    cfg — em_scoring_configs row as dict
    Returns scoring breakdown dict.
    """
    l1w = cfg["line1_weight"]
    l2w = cfg["line2_weight"]
    em_w = cfg["em_level_weight"]
    cpt_w = cfg["cpt_weight"]
    dx_w = cfg["dx_weight"]
    copa_w = cfg["copa_weight"]
    dr_w = cfg["dr_weight"]
    risk_w = cfg["risk_weight"]

    # ── Derive submitted levels ───────────────────────────────────────────────
    derived_copa = derive_copa_level({k.replace("sub_", ""): v for k, v in sub.items() if k.startswith("sub_copa")})
    derived_dr = derive_dr_level({k.replace("sub_", ""): v for k, v in sub.items() if k.startswith("sub_dr")})
    derived_risk = derive_risk_level({k.replace("sub_", ""): v for k, v in sub.items() if k.startswith("sub_risk")})

    ak_copa = ak.get("copa_level", "")
    ak_dr = ak.get("dr_level", "")
    ak_risk = ak.get("risk_level", "")

    # ── What applies to this chart, and what it is worth ─────────────────────
    # The category decides whether MDM is assessed at all; the weights are then
    # renormalised so every chart is still scored out of 100.
    ak_cpts = _clean_codes(_j(ak.get("procedure_cpts", "[]")))
    category = resolve_category(ak.get("em_category"), ak.get("em_code"))
    ak_method_raw = (ak.get("level_method") or "MDM").upper().strip()
    weights = applicable_weights(cfg, category, bool(ak_cpts), ak_method_raw)
    em_w_adj = weights.get("em_level", 0.0)
    dx_w_adj = weights.get("dx", 0.0)
    cpt_w_adj = weights.get("cpt", 0.0)
    uses_mdm = "copa" in weights

    # ── Coding Accuracy ───────────────────────────────────────────────────────
    # E/M Level — complexity AND patient type must both match (when AK patient_type != NA)
    # Resolve code + modifier through the same helper CPT lines use, so the
    # comparison is identical: multi-modifier cells are order/separator
    # independent, and a modifier typed into the code cell ("99213-25" with the
    # modifier column left blank) is recovered rather than failing the match.
    from services.grading_engine import resolve_cpt_modifier

    ak_code_n, ak_mod_n = resolve_cpt_modifier(ak.get("em_code"), ak.get("em_modifier"))
    sub_code_n, sub_mod_n = resolve_cpt_modifier(sub.get("sub_em_code"), sub.get("sub_em_modifier"))

    sub_em_level = em_code_to_level(sub_code_n)
    ak_em_level = em_code_to_level(ak_code_n)
    ak_patient_type = (ak.get("patient_type") or "NA").upper().strip()
    sub_patient_type = (sub.get("sub_patient_type") or "NA").upper().strip()
    patient_type_ok = (ak_patient_type == "NA") or (sub_patient_type == ak_patient_type)
    # 99211 (nurse visit) carries no MDM level, so em_code_to_level returns None
    # and a level comparison can never succeed — an exactly-correct 99211 scored
    # zero. Fall back to code equality for any code outside the MDM table.
    level_match = bool(sub_em_level) and sub_em_level == ak_em_level
    code_match = bool(ak_code_n) and sub_code_n == ak_code_n
    # Modifier is graded exactly like a CPT modifier: the code+modifier PAIR has
    # to match, so a missing or wrong modifier 25 costs the E/M level component
    # the same way it costs a CPT line.
    modifier_ok = (ak_mod_n == sub_mod_n)
    em_level_score = em_w_adj if (
        (level_match or code_match) and patient_type_ok and modifier_ok) else 0.0
    patient_type_mismatch = (ak_patient_type != "NA") and (not patient_type_ok)

    # ── Critical care: the clock is part of picking the code ─────────────────
    # Total critical care time on the date of service is what selects 99291 and
    # how many units of 99292 follow, so it belongs to the E/M component rather
    # than being scored as separate reasoning. Coder against key, with no
    # threshold table applied in either direction.
    cc_minutes_ok = None
    cc_units_ok = None
    if category == CRITICAL_CARE:
        ak_cc = _minutes(ak.get("critical_care_minutes"))
        sub_cc = _minutes(sub.get("sub_critical_care_minutes"))
        if ak_cc is None:
            # The key never stated a time — nothing to grade against, and
            # withholding the marks would penalise the coder for a gap in it.
            cc_minutes_ok = None
        else:
            cc_minutes_ok = (sub_cc == ak_cc)

        # The add-on units say the same thing the clock says: 99292 is billed
        # once per further half hour. A coder can state the right total and
        # then claim the wrong number of units, or the reverse, and until now
        # only the clock was checked.
        #
        # Graded only when the key records its own CPT lines. An empty key
        # cannot distinguish "no add-on was due" from "nobody wrote it down",
        # and marking a coder down for a gap in the key is the same mistake
        # the minutes rule above already avoids.
        ak_cc_rows = normalise_cpts(ak.get("procedure_cpts", "[]"))
        if ak_cc_rows:
            ak_units = critical_care_units(ak_cc_rows, ak_code_n)
            sub_units = critical_care_units(
                normalise_cpts(sub.get("sub_procedure_cpts", "[]")), sub_code_n)
            cc_units_ok = (sub_units == ak_units)

        # ONE deduction, not two. Minutes and units are two statements of the
        # same quantity, so halving for each would cost a coder twice for a
        # single misread of the clock. Half the component, the same rule a
        # wrong Dx pointer or a wrong unit count gets: the service was
        # identified, the quantity behind it was not.
        if cc_minutes_ok is False or cc_units_ok is False:
            em_level_score = em_level_score / 2
    # Only flagged when everything else was right — so the feedback can say the
    # modifier alone was what cost the points.
    modifier_mismatch = (level_match or code_match) and patient_type_ok and not modifier_ok

    # ── Levelling method: MDM or Time ────────────────────────────────────────
    # Coding Accuracy is scored on the final code regardless of method — either
    # route can legitimately reach the right answer. The method mistake costs
    # Reasoning Accuracy instead, which is what "justify it correctly" means.
    ak_method = (ak.get("level_method") or "MDM").upper().strip()
    sub_method = (sub.get("sub_level_method") or "MDM").upper().strip()
    method_ok = (ak_method == sub_method)
    ak_is_time = (ak_method == "TIME")
    time_ok = False
    if ak_is_time and method_ok:
        time_ok = time_supports_code(sub.get("sub_total_time"), ak.get("em_code") or "")
    method_mismatch = not method_ok

    # ── Procedure CPTs, with diagnosis pointers on professional claims ───────
    # ED Profee and office E/M both bill on a CMS-1500, so each procedure line
    # points at the diagnoses justifying it. Pointers are checked exactly when
    # the answer key carries them, so a key without pointers grades as before.
    from services.grading_engine import resolve_pointers

    ak_cpt_rows = normalise_cpts(ak.get("procedure_cpts", "[]"))
    sub_cpt_rows = normalise_cpts(sub.get("sub_procedure_cpts", "[]"))
    ak_dx_ordered = _clean_codes(_j(ak.get("dx_codes", "[]")))
    sub_dx_ordered = _clean_codes(_j(sub.get("sub_dx_codes", "[]")))

    cpt_score = 0.0
    pointer_errors = []
    unit_errors = []
    if ak_cpt_rows and cpt_w_adj > 0:
        per_cpt = cpt_w_adj / len(ak_cpt_rows)
        used = [False] * len(sub_cpt_rows)
        matched = 0.0
        for a in ak_cpt_rows:
            for i, s in enumerate(sub_cpt_rows):
                if used[i] or a["code"] != s["code"] or a["modifier"] != s["modifier"]:
                    continue
                used[i] = True
                credit = 1.0
                if a["pointers"]:
                    # Pointers are positional, so compare the diagnosis CODES
                    # they resolve to — never the letters themselves.
                    if (resolve_pointers(a["pointers"], ak_dx_ordered)
                            != resolve_pointers(s["pointers"], sub_dx_ordered)):
                        # A linkage error is a lesser mistake than a wrong code,
                        # so it costs half the line — same rule as OP/Surgery.
                        credit = 0.5
                        pointer_errors.append({
                            "code": a["code"],
                            "ak": ",".join(a["pointers"]),
                            "sub": ",".join(s["pointers"]) or "(none)",
                        })

                # Graded only where the key states units, so older keys are
                # unaffected. Half the line, and it does not stack with a
                # pointer error — one line, one half-credit.
                if "units" in a:
                    from services.grading_engine import norm_units
                    ak_u, sub_u = norm_units(a.get("units")), norm_units(s.get("units"))
                    if ak_u != sub_u:
                        credit = min(credit, 0.5)
                        unit_errors.append({"code": a["code"], "ak": ak_u, "sub": sub_u})
                matched += credit
                break
        cpt_score = matched * per_cpt
        if overcoding_penalty:
            over = max(0, len(sub_cpt_rows) - len(ak_cpt_rows))
            cpt_score = max(0.0, cpt_score - over * per_cpt)

    # Dx (proportional, overcoding penalty)
    ak_dx = _clean_codes(_j(ak.get("dx_codes", "[]")))
    sub_dx = _clean_codes(_j(sub.get("sub_dx_codes", "[]")))
    dx_score = 0.0
    if ak_dx:
        per_dx = dx_w_adj / len(ak_dx)
        matched_dx = sum(1 for c in ak_dx if c in sub_dx)
        dx_score = matched_dx * per_dx
        if overcoding_penalty:
            over_dx = max(0, len(sub_dx) - len(ak_dx))
            dx_score = max(0.0, dx_score - over_dx * per_dx)
    # No AK dx → full dx score (chart has no expected dx)
    elif dx_w_adj > 0:
        dx_score = dx_w_adj

    coding_accuracy_total = em_level_score + cpt_score + dx_score

    # ── Reasoning Accuracy (element proportion) ───────────────────────────────
    copa_fields = [
        "copa_self_limited", "copa_stable_acute", "copa_stable_chronic",
        "copa_acute_uncomplicated", "copa_chronic_exacerbation",
        "copa_undiagnosed_new", "copa_acute_systemic",
        "copa_acute_complicated_injury", "copa_chronic_severe", "copa_threat_to_life",
    ]
    dr_fields = [
        "dr_prior_external_notes", "dr_review_test_results", "dr_order_tests",
        "dr_independent_historian", "dr_independent_interpretation", "dr_external_discussion",
    ]
    risk_fields = [
        "risk_low", "risk_prescription_drug_mgmt", "risk_minor_surgery_with_factors",
        "risk_elective_major_no_factors", "risk_hospitalization", "risk_sdoh",
        "risk_drug_intensive_monitoring", "risk_elective_major_with_factors",
        "risk_emergency_major_surgery", "risk_hospitalization_escalation",
        "risk_dnr_deescalate", "risk_parenteral_controlled",
    ]

    def _element_score(fields: list, ak_row: dict, sub_row: dict, weight: float) -> float:
        ak_vals = [int(ak_row.get(f, 0) or 0) for f in fields]
        sub_vals = [int(sub_row.get(f"sub_{f}", 0) or 0) for f in fields]
        total_ak = sum(min(v, 1) for v in ak_vals)  # count non-zero AK elements
        if total_ak == 0:
            return weight  # no elements expected → full score
        # Count only the elements the key actually expects. The previous version
        # counted every field where AK and submission agreed — including the many
        # both-zero fields — so `correct` ran far above `total_ak` and the score
        # blew past its weight (COPA alone returned 100 against a weight of 10).
        correct = sum(1 for a, s in zip(ak_vals, sub_vals) if a > 0 and a == s)
        return round((correct / total_ak) * weight, 2)

    if not category_uses_mdm(category):
        # Preventive and Other are levelled by things the MDM tables do not
        # describe — age and patient type, or a code set we do not model. Their
        # Line 2 weight has already been folded into Line 1 by
        # applicable_weights, so there is nothing left to award here.
        copa_element_score = dr_element_score = risk_element_score = 0.0
        reasoning_accuracy_total = 0.0
    elif ak_is_time:
        # Time-levelled chart: MDM elements are not the operative criteria, so
        # they are not scored. Reasoning credit turns on picking the Time route
        # and reading the time correctly.
        reasoning_w = weights.get("time", 0.0)
        if not method_ok:
            reasoning_earned = 0.0          # levelled by MDM when time was the basis
        elif time_ok:
            reasoning_earned = reasoning_w  # right route, time in band
        else:
            reasoning_earned = reasoning_w / 2   # right route, wrong time
        copa_element_score = dr_element_score = risk_element_score = 0.0
        reasoning_accuracy_total = round(reasoning_earned, 2)
    elif not method_ok:
        # MDM-levelled chart but the coder levelled by time — the MDM elements
        # they skipped are exactly what Reasoning Accuracy measures.
        copa_element_score = dr_element_score = risk_element_score = 0.0
        reasoning_accuracy_total = 0.0
    else:
        copa_element_score = _element_score(copa_fields, ak, sub, weights.get("copa", 0.0))
        dr_element_score = _element_score(dr_fields, ak, sub, weights.get("dr", 0.0))
        risk_element_score = _element_score(risk_fields, ak, sub, weights.get("risk", 0.0))
        reasoning_accuracy_total = copa_element_score + dr_element_score + risk_element_score

    total_score = round(coding_accuracy_total + reasoning_accuracy_total, 1)
    pass_threshold = cfg.get("pass_threshold", 80.0)
    pass_fail = "PASS" if total_score >= pass_threshold else "FAIL"

    return {
        "derived_copa_level": derived_copa,
        "derived_dr_level": derived_dr,
        "derived_risk_level": derived_risk,
        "ak_level_method": ak_method,
        "sub_level_method": sub_method,
        "method_mismatch": method_mismatch,
        "time_ok": time_ok,
        "em_category": category,
        "em_category_label": EM_CATEGORY_LABELS.get(category, "Other"),
        "uses_mdm": uses_mdm,
        "applied_weights": weights,
        "critical_care_minutes_ok": cc_minutes_ok,
        "critical_care_units_ok": cc_units_ok,
        "ak_critical_care_minutes": _minutes(ak.get("critical_care_minutes")),
        "sub_critical_care_minutes": _minutes(sub.get("sub_critical_care_minutes")),
        "pointer_errors": pointer_errors,
        "unit_errors": unit_errors,
        "modifier_mismatch": modifier_mismatch,
        "ak_em_modifier": ak_mod_n,
        "sub_em_modifier": sub_mod_n,
        "em_level_score": round(em_level_score, 2),
        "cpt_score": round(cpt_score, 2),
        "dx_score": round(dx_score, 2),
        "coding_accuracy_total": round(coding_accuracy_total, 2),
        "copa_element_score": copa_element_score,
        "dr_element_score": dr_element_score,
        "risk_element_score": risk_element_score,
        "reasoning_accuracy_total": round(reasoning_accuracy_total, 2),
        "total_score": total_score,
        "pass_fail": pass_fail,
        "patient_type_mismatch": patient_type_mismatch,
        "ak_patient_type": ak_patient_type,
        "sub_patient_type": sub_patient_type,
    }


# ── Pydantic models ───────────────────────────────────────────────────────────

class EMAnswerKeyPayload(BaseModel):
    chart_id: int
    copa_self_limited: int = 0
    copa_stable_acute: int = 0
    copa_stable_chronic: int = 0
    copa_acute_uncomplicated: int = 0
    copa_chronic_exacerbation: int = 0
    copa_undiagnosed_new: int = 0
    copa_acute_systemic: int = 0
    copa_acute_complicated_injury: int = 0
    copa_chronic_severe: int = 0
    copa_threat_to_life: int = 0
    copa_level_overridden: bool = False
    copa_level_override: Optional[str] = None
    dr_prior_external_notes: int = 0
    dr_review_test_results: int = 0
    dr_order_tests: int = 0
    dr_independent_historian: bool = False
    dr_independent_interpretation: bool = False
    dr_external_discussion: bool = False
    dr_level_overridden: bool = False
    dr_level_override: Optional[str] = None
    risk_low: bool = False
    risk_prescription_drug_mgmt: bool = False
    risk_minor_surgery_with_factors: bool = False
    risk_elective_major_no_factors: bool = False
    risk_hospitalization: bool = False
    risk_sdoh: bool = False
    risk_drug_intensive_monitoring: bool = False
    risk_elective_major_with_factors: bool = False
    risk_emergency_major_surgery: bool = False
    risk_hospitalization_escalation: bool = False
    risk_dnr_deescalate: bool = False
    risk_parenteral_controlled: bool = False
    risk_level_overridden: bool = False
    risk_level_override: Optional[str] = None
    em_code: str
    em_modifier: Optional[str] = None
    patient_type: Optional[str] = "NA"  # New / Established / NA
    level_method: Optional[str] = "MDM"  # MDM | TIME (office/outpatient only)
    total_time: Optional[int] = None     # total minutes on date of encounter
    # Blank means "derive from the E/M code", which is what every key written
    # before categories existed does.
    em_category: Optional[str] = None
    critical_care_minutes: Optional[int] = None
    dx_codes: list = []
    procedure_cpts: list = []
    entered_by: str


class EMScoringConfigPayload(BaseModel):
    line1_weight: float = 70.0
    line2_weight: float = 30.0
    em_level_weight: float = 23.33
    cpt_weight: float = 23.33
    dx_weight: float = 23.34
    copa_weight: float = 10.0
    dr_weight: float = 10.0
    risk_weight: float = 10.0
    pass_threshold: float = 80.0
    overcoding_penalty: bool = True
    passphrase: str
    updated_by: str


# ── Answer key endpoints ──────────────────────────────────────────────────────

@router.get("/em/answer-key/list")
def list_em_answer_keys(search: Optional[str] = None,
                        page: Optional[int] = Query(default=None, ge=1),
                        page_size: int = Query(default=50, ge=1, le=200),
                        db: Session = Depends(get_db)):
    where = ""
    params = {}
    if search and search.strip():
        where = """
        WHERE LOWER(c.chart_number) LIKE :needle
           OR LOWER(c.category) LIKE :needle
           OR LOWER(eak.em_code) LIKE :needle
           OR LOWER(eak.entered_by) LIKE :needle
        """
        params["needle"] = f"%{search.strip().lower()}%"
    total = db.execute(text(f"""
        SELECT COUNT(*)
        FROM em_answer_keys eak
        JOIN charts c ON c.id = eak.chart_id
        {where}
    """), params).scalar() or 0
    limit_clause = ""
    if page is not None:
        limit_clause = " LIMIT :limit OFFSET :offset"
        params["limit"] = page_size
        params["offset"] = (page - 1) * page_size
    rows = db.execute(text(f"""
        SELECT eak.chart_id, c.chart_number, c.category, c.specialty,
               eak.em_code, eak.copa_level, eak.dr_level, eak.risk_level,
               eak.entered_by, eak.entered_at,
               eak.em_category, eak.level_method, eak.patient_type,
               eak.em_modifier, eak.total_time, eak.critical_care_minutes,
               eak.dx_codes, eak.procedure_cpts,
               eak.copa_level_overridden, eak.dr_level_overridden, eak.risk_level_overridden
        FROM em_answer_keys eak
        JOIN charts c ON c.id = eak.chart_id
        {where}
        ORDER BY c.chart_number
        {limit_clause}
    """), params).mappings().fetchall()

    out = []
    for r in rows:
        d = dict(r)
        # Whether the three MDM levels mean anything on THIS chart. A preventive
        # visit or one levelled by time is not graded on medical decision making
        # at all, and the stored levels are the derivation's default rather than
        # a judgement anybody made. Printing "Minimal" beside such a chart says
        # a trainer decided something they never decided — and on a key screen
        # that is the kind of quiet wrongness that gets believed.
        category = resolve_category(d.get("em_category"), d.get("em_code"))
        method = (d.get("level_method") or "MDM").upper().strip()
        d["em_category"] = category
        d["uses_mdm"] = bool(category_uses_mdm(category) and method != "TIME")
        out.append(d)
    if page is not None:
        return {"total": total, "page": page, "page_size": page_size, "results": out}
    return out


# NOTE: every STATIC /em/answer-key/... path must be registered before the
# parameterised /{chart_id} route below — FastAPI matches in registration
# order, so /template was being parsed as a chart_id and 422'd.
@router.get("/em/answer-key/template")
def download_em_template():
    """Download the E/M answer key Excel template."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    from services.excel_style import NAVY

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EM Answer Key"  # "/" is illegal in an Excel sheet title

    hdr_font = Font(bold=True, color="FFFFFF")
    # Lower-cased, which is how this one escaped the search that
    # found the others. Same drift, same fix.
    hdr_fill = PatternFill("solid", fgColor=NAVY)
    section_fill = PatternFill("solid", fgColor="dbeafe")
    section_font = Font(bold=True, color=NAVY)

    def hdr(cell, val):
        ws[cell] = val
        ws[cell].font = hdr_font
        ws[cell].fill = hdr_fill
        ws[cell].alignment = Alignment(horizontal="center", wrap_text=True)

    def sec(cell, val):
        ws[cell] = val
        ws[cell].font = section_font
        ws[cell].fill = section_fill

    # Column headers come from EM_KEY_COLUMNS, the same table the parser reads
    # them back with — so the two cannot drift apart, and a column added in one
    # place is a column the other already understands.
    from openpyxl.utils import get_column_letter
    from services.excel_service import EM_KEY_COLUMNS

    for i, (_field, header) in enumerate(EM_KEY_COLUMNS, start=1):
        hdr(f"{get_column_letter(i)}1", header)

    # Sample row
    ws["A2"] = "EM001"
    ws["B2"] = 0
    ws["D2"] = 2
    ws["L2"] = ""
    ws["N2"] = 1
    ws["O2"] = 1
    ws["Q2"] = "Y"
    ws["U2"] = "Y"
    ws["AG2"] = "99214"
    ws["AH2"] = ""
    ws["AG2"] = "99214"
    ws["AI2"] = "Established"
    ws["BA2"] = "MDM"
    ws["BC2"] = "A"
    ws["BK2"] = ""
    ws["AJ2"] = "E11.9"
    ws["AK2"] = "I10"
    ws["AL2"] = "Z79.4"
    ws["AZ2"] = "Dr. Smith"

    ws.freeze_panes = "B2"  # freeze column A (chart number) and header row

    ws.row_dimensions[1].height = 60
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Stream it like every other template. Returning {filename, content:base64}
    # meant the frontend's window.open() rendered raw JSON in a tab instead of
    # downloading a workbook.
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=EM_AnswerKey_Template.xlsx"},
    )

@router.get("/em/answer-key/{chart_id}")
def get_em_answer_key(chart_id: int, db: Session = Depends(get_db)):
    row = db.execute(
        text("SELECT * FROM em_answer_keys WHERE chart_id = :c"), {"c": chart_id}
    ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="No E/M answer key for this chart")
    return dict(row)


@router.post("/em/answer-key")
def upsert_em_answer_key(payload: EMAnswerKeyPayload, db: Session = Depends(get_db)):
    d = payload.dict()
    chart_id = d.pop("chart_id")
    entered_by = d.pop("entered_by")

    # Derive levels (or use override)
    copa_level = d.pop("copa_level_override") or derive_copa_level(d)
    dr_level = d.pop("dr_level_override") or derive_dr_level(d)
    risk_level = d.pop("risk_level_override") or derive_risk_level(d)

    copa_overridden = d.pop("copa_level_overridden")
    dr_overridden = d.pop("dr_level_overridden")
    risk_overridden = d.pop("risk_level_overridden")

    dx_codes = json.dumps(d.pop("dx_codes"))
    procedure_cpts = json.dumps(d.pop("procedure_cpts"))
    # Stored explicitly rather than derived on every read: a trainer can key a
    # code we have not classified and still say which kind of encounter it is.
    _category = resolve_category(d.pop("em_category", None), payload.em_code)
    d.pop("critical_care_minutes", None)

    existing = db.execute(
        text("SELECT id FROM em_answer_keys WHERE chart_id = :c"), {"c": chart_id}
    ).first()

    if existing:
        db.execute(text("""
            UPDATE em_answer_keys SET
                copa_self_limited=:copa_self_limited,
                copa_stable_acute=:copa_stable_acute,
                copa_stable_chronic=:copa_stable_chronic,
                copa_acute_uncomplicated=:copa_acute_uncomplicated,
                copa_chronic_exacerbation=:copa_chronic_exacerbation,
                copa_undiagnosed_new=:copa_undiagnosed_new,
                copa_acute_systemic=:copa_acute_systemic,
                copa_acute_complicated_injury=:copa_acute_complicated_injury,
                copa_chronic_severe=:copa_chronic_severe,
                copa_threat_to_life=:copa_threat_to_life,
                copa_level=:copa_level, copa_level_overridden=:copa_overridden,
                dr_prior_external_notes=:dr_prior_external_notes,
                dr_review_test_results=:dr_review_test_results,
                dr_order_tests=:dr_order_tests,
                dr_independent_historian=:dr_independent_historian,
                dr_independent_interpretation=:dr_independent_interpretation,
                dr_external_discussion=:dr_external_discussion,
                dr_level=:dr_level, dr_level_overridden=:dr_overridden,
                risk_low=:risk_low,
                risk_prescription_drug_mgmt=:risk_prescription_drug_mgmt,
                risk_minor_surgery_with_factors=:risk_minor_surgery_with_factors,
                risk_elective_major_no_factors=:risk_elective_major_no_factors,
                risk_hospitalization=:risk_hospitalization,
                risk_sdoh=:risk_sdoh,
                risk_drug_intensive_monitoring=:risk_drug_intensive_monitoring,
                risk_elective_major_with_factors=:risk_elective_major_with_factors,
                risk_emergency_major_surgery=:risk_emergency_major_surgery,
                risk_hospitalization_escalation=:risk_hospitalization_escalation,
                risk_dnr_deescalate=:risk_dnr_deescalate,
                risk_parenteral_controlled=:risk_parenteral_controlled,
                risk_level=:risk_level, risk_level_overridden=:risk_overridden,
                em_code=:em_code, em_modifier=:em_modifier, patient_type=:patient_type,
                level_method=:level_method, total_time=:total_time,
                em_category=:em_category,
                critical_care_minutes=:critical_care_minutes,
                dx_codes=:dx_codes, procedure_cpts=:procedure_cpts,
                entered_by=:entered_by, entered_at=CURRENT_TIMESTAMP
            WHERE chart_id=:chart_id
        """), {**d, "chart_id": chart_id, "entered_by": entered_by,
               "copa_level": copa_level, "copa_overridden": copa_overridden,
               "dr_level": dr_level, "dr_overridden": dr_overridden,
               "risk_level": risk_level, "risk_overridden": risk_overridden,
               "patient_type": (payload.patient_type or "NA").upper(),
               "level_method": _sanitise_level_method(payload.level_method, payload.em_code),
               "total_time": payload.total_time,
               "em_category": _category,
               "critical_care_minutes": (payload.critical_care_minutes
                                         if _category == CRITICAL_CARE else None),
               "dx_codes": dx_codes, "procedure_cpts": procedure_cpts})
    else:
        db.execute(text("""
            INSERT INTO em_answer_keys (
                chart_id,
                copa_self_limited, copa_stable_acute, copa_stable_chronic,
                copa_acute_uncomplicated, copa_chronic_exacerbation, copa_undiagnosed_new,
                copa_acute_systemic, copa_acute_complicated_injury,
                copa_chronic_severe, copa_threat_to_life,
                copa_level, copa_level_overridden,
                dr_prior_external_notes, dr_review_test_results, dr_order_tests,
                dr_independent_historian, dr_independent_interpretation, dr_external_discussion,
                dr_level, dr_level_overridden,
                risk_low, risk_prescription_drug_mgmt, risk_minor_surgery_with_factors,
                risk_elective_major_no_factors, risk_hospitalization, risk_sdoh,
                risk_drug_intensive_monitoring, risk_elective_major_with_factors,
                risk_emergency_major_surgery, risk_hospitalization_escalation,
                risk_dnr_deescalate, risk_parenteral_controlled,
                risk_level, risk_level_overridden,
                em_code, em_modifier, patient_type, level_method, total_time,
                em_category, critical_care_minutes, dx_codes, procedure_cpts,
                entered_by
            ) VALUES (
                :chart_id,
                :copa_self_limited, :copa_stable_acute, :copa_stable_chronic,
                :copa_acute_uncomplicated, :copa_chronic_exacerbation, :copa_undiagnosed_new,
                :copa_acute_systemic, :copa_acute_complicated_injury,
                :copa_chronic_severe, :copa_threat_to_life,
                :copa_level, :copa_overridden,
                :dr_prior_external_notes, :dr_review_test_results, :dr_order_tests,
                :dr_independent_historian, :dr_independent_interpretation, :dr_external_discussion,
                :dr_level, :dr_overridden,
                :risk_low, :risk_prescription_drug_mgmt, :risk_minor_surgery_with_factors,
                :risk_elective_major_no_factors, :risk_hospitalization, :risk_sdoh,
                :risk_drug_intensive_monitoring, :risk_elective_major_with_factors,
                :risk_emergency_major_surgery, :risk_hospitalization_escalation,
                :risk_dnr_deescalate, :risk_parenteral_controlled,
                :risk_level, :risk_overridden,
                :em_code, :em_modifier, :patient_type, :level_method, :total_time,
                :em_category, :critical_care_minutes, :dx_codes, :procedure_cpts,
                :entered_by
            )
        """), {**d, "chart_id": chart_id, "entered_by": entered_by,
               "copa_level": copa_level, "copa_overridden": copa_overridden,
               "dr_level": dr_level, "dr_overridden": dr_overridden,
               "risk_level": risk_level, "risk_overridden": risk_overridden,
               "patient_type": (payload.patient_type or "NA").upper(),
               "level_method": _sanitise_level_method(payload.level_method, payload.em_code),
               "total_time": payload.total_time,
               "em_category": _category,
               "critical_care_minutes": (payload.critical_care_minutes
                                         if _category == CRITICAL_CARE else None),
               "dx_codes": dx_codes, "procedure_cpts": procedure_cpts})
    db.commit()
    return {"status": "ok", "copa_level": copa_level, "dr_level": dr_level,
            "risk_level": risk_level, "em_category": _category,
            "em_category_label": EM_CATEGORY_LABELS.get(_category, "Other"),
            "uses_mdm": category_uses_mdm(_category)}


@router.delete("/em/answer-key/{chart_id}")
def delete_em_answer_key(chart_id: int, passphrase: str, db: Session = Depends(get_db)):
    if passphrase != MASTER_PASSPHRASE:
        raise HTTPException(status_code=403, detail="Invalid passphrase")
    db.execute(text("DELETE FROM em_answer_keys WHERE chart_id = :c"), {"c": chart_id})
    db.commit()
    return {"status": "deleted"}


# ── Scoring config endpoints ──────────────────────────────────────────────────

@router.get("/em/scoring-config")
def get_em_scoring_config(db: Session = Depends(get_db)):
    row = db.execute(text("SELECT * FROM em_scoring_configs WHERE id = 1")).mappings().first()
    if not row:
        return {
            "line1_weight": 70.0, "line2_weight": 30.0,
            "em_level_weight": 23.33, "cpt_weight": 23.33, "dx_weight": 23.34,
            "copa_weight": 10.0, "dr_weight": 10.0, "risk_weight": 10.0,
            "pass_threshold": 80.0, "overcoding_penalty": True,
        }
    return dict(row)


@router.put("/em/scoring-config")
def update_em_scoring_config(payload: EMScoringConfigPayload, db: Session = Depends(get_db)):
    if payload.passphrase != MASTER_PASSPHRASE:
        raise HTTPException(status_code=403, detail="Invalid passphrase")
    d = payload.dict()
    if abs((payload.line1_weight + payload.line2_weight) - 100.0) > 0.1:
        raise HTTPException(
            status_code=400,
            detail="Coding Accuracy and Reasoning Accuracy weights must sum to 100",
        )
    if abs((payload.em_level_weight + payload.cpt_weight + payload.dx_weight)
           - payload.line1_weight) > 0.5:
        raise HTTPException(
            status_code=400,
            detail=(f"Coding Accuracy metric weights must sum to "
                    f"{payload.line1_weight:g}"),
        )
    if abs((payload.copa_weight + payload.dr_weight + payload.risk_weight)
           - payload.line2_weight) > 0.5:
        raise HTTPException(
            status_code=400,
            detail=(f"Reasoning Accuracy metric weights must sum to "
                    f"{payload.line2_weight:g}"),
        )
    d.pop("passphrase")
    d["updated_at"] = "CURRENT_TIMESTAMP"
    db.execute(text("""
        INSERT INTO em_scoring_configs
            (id, line1_weight, line2_weight, em_level_weight, cpt_weight, dx_weight,
             copa_weight, dr_weight, risk_weight, pass_threshold, overcoding_penalty,
             updated_by, updated_at)
        VALUES
            (1, :line1_weight, :line2_weight, :em_level_weight, :cpt_weight, :dx_weight,
             :copa_weight, :dr_weight, :risk_weight, :pass_threshold, :overcoding_penalty,
             :updated_by, CURRENT_TIMESTAMP)
        ON CONFLICT (id) DO UPDATE SET
            line1_weight=EXCLUDED.line1_weight,
            line2_weight=EXCLUDED.line2_weight,
            em_level_weight=EXCLUDED.em_level_weight,
            cpt_weight=EXCLUDED.cpt_weight,
            dx_weight=EXCLUDED.dx_weight,
            copa_weight=EXCLUDED.copa_weight,
            dr_weight=EXCLUDED.dr_weight,
            risk_weight=EXCLUDED.risk_weight,
            pass_threshold=EXCLUDED.pass_threshold,
            overcoding_penalty=EXCLUDED.overcoding_penalty,
            updated_by=EXCLUDED.updated_by,
            updated_at=EXCLUDED.updated_at
    """), {k: v for k, v in d.items() if k != "updated_at"})
    db.commit()
    return {"status": "saved"}


# ── Bulk Excel upload ─────────────────────────────────────────────────────────

@router.post("/em/answer-key/upload")
def upload_em_answer_keys(
    file: UploadFile = File(...),
    entered_by: str = Form(...),
    replace: bool = Form(False),
    passphrase: str = Form(""),
    db: Session = Depends(get_db),
):
    if replace and passphrase != MASTER_PASSPHRASE:
        raise HTTPException(status_code=403, detail="Invalid passphrase")

    try:
        rows = parse_em_answer_key_upload(file.file.read())
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Could not parse file: {e}")

    stored, replaced, skipped, not_found, wrong_specialty = [], [], [], [], []

    # The same two rules the IP/OP upload enforces, applied here rather than
    # only where they were written. This endpoint had neither.
    _seen: dict = {}
    for row in rows:
        _k = (row["chart_number"] or "").strip().upper()
        _seen[_k] = _seen.get(_k, 0) + 1
    _repeated = sorted(k for k, n in _seen.items() if n > 1)
    if _repeated:
        raise HTTPException(
            status_code=400,
            detail=(f"The file has more than one row for: {', '.join(_repeated[:10])}"
                    + (f" (and {len(_repeated) - 10} more)" if len(_repeated) > 10 else "")
                    + ". Remove the duplicates and upload again."),
        )

    for row in rows:
        chart_num = row["chart_number"]
        chart = _find_chart(db, chart_num)
        if not chart:
            not_found.append(chart_num)
            continue
        matched_chart_num = chart.chart_number

        if chart.specialty.value not in ("E/M", "ED Profee"):
            wrong_specialty.append(matched_chart_num)
            continue

        em_code = row.get("em_code", "").strip()
        if not em_code:
            skipped.append(matched_chart_num)
            continue

        entered_by_val = row.get("entered_by") or entered_by

        copa_override_raw = row.get("copa_level_override", "") or ""
        dr_override_raw = row.get("dr_level_override", "") or ""
        risk_override_raw = row.get("risk_level_override", "") or ""

        copa_level = copa_override_raw or derive_copa_level(row)
        dr_level = dr_override_raw or derive_dr_level(row)
        risk_level = risk_override_raw or derive_risk_level(row)

        dx_codes = json.dumps(row.get("dx_codes", []))
        procedure_cpts = json.dumps(row.get("procedure_cpts", []))
        _row_category = resolve_category(row.get("em_category"), em_code)

        existing = db.execute(
            text("SELECT id FROM em_answer_keys WHERE chart_id = :c"), {"c": chart.id}
        ).first()

        params = {
            "chart_id": chart.id,
            "copa_self_limited": row.get("copa_self_limited", 0),
            "copa_stable_acute": row.get("copa_stable_acute", 0),
            "copa_stable_chronic": row.get("copa_stable_chronic", 0),
            "copa_acute_uncomplicated": row.get("copa_acute_uncomplicated", 0),
            "copa_chronic_exacerbation": row.get("copa_chronic_exacerbation", 0),
            "copa_undiagnosed_new": row.get("copa_undiagnosed_new", 0),
            "copa_acute_systemic": row.get("copa_acute_systemic", 0),
            "copa_acute_complicated_injury": row.get("copa_acute_complicated_injury", 0),
            "copa_chronic_severe": row.get("copa_chronic_severe", 0),
            "copa_threat_to_life": row.get("copa_threat_to_life", 0),
            "copa_level": copa_level,
            "copa_level_overridden": bool(copa_override_raw),
            "dr_prior_external_notes": row.get("dr_prior_external_notes", 0),
            "dr_review_test_results": row.get("dr_review_test_results", 0),
            "dr_order_tests": row.get("dr_order_tests", 0),
            "dr_independent_historian": row.get("dr_independent_historian", False),
            "dr_independent_interpretation": row.get("dr_independent_interpretation", False),
            "dr_external_discussion": row.get("dr_external_discussion", False),
            "dr_level": dr_level,
            "dr_level_overridden": bool(dr_override_raw),
            "risk_low": row.get("risk_low", False),
            "risk_prescription_drug_mgmt": row.get("risk_prescription_drug_mgmt", False),
            "risk_minor_surgery_with_factors": row.get("risk_minor_surgery_with_factors", False),
            "risk_elective_major_no_factors": row.get("risk_elective_major_no_factors", False),
            "risk_hospitalization": row.get("risk_hospitalization", False),
            "risk_sdoh": row.get("risk_sdoh", False),
            "risk_drug_intensive_monitoring": row.get("risk_drug_intensive_monitoring", False),
            "risk_elective_major_with_factors": row.get("risk_elective_major_with_factors", False),
            "risk_emergency_major_surgery": row.get("risk_emergency_major_surgery", False),
            "risk_hospitalization_escalation": row.get("risk_hospitalization_escalation", False),
            "risk_dnr_deescalate": row.get("risk_dnr_deescalate", False),
            "risk_parenteral_controlled": row.get("risk_parenteral_controlled", False),
            "risk_level": risk_level,
            "risk_level_overridden": bool(risk_override_raw),
            "em_code": em_code,
            "em_modifier": row.get("em_modifier", "") or "",
            "patient_type": (row.get("patient_type") or "NA").upper().strip(),
            # ED codes (99281-99285) have no time-based option, so a TIME key
            # would be unanswerable — the coder form correctly hides Time there.
            "level_method": _sanitise_level_method(
                row.get("level_method"), row.get("em_code")),
            "total_time": row.get("total_time"),
            "em_category": _row_category,
            "critical_care_minutes": (row.get("critical_care_minutes")
                                      if _row_category == CRITICAL_CARE else None),
            "dx_codes": dx_codes,
            "procedure_cpts": procedure_cpts,
            "entered_by": entered_by_val,
        }

        if existing:
            if not replace:
                skipped.append(matched_chart_num)
                continue
            db.execute(text("""
                UPDATE em_answer_keys SET
                    copa_self_limited=:copa_self_limited, copa_stable_acute=:copa_stable_acute,
                    copa_stable_chronic=:copa_stable_chronic, copa_acute_uncomplicated=:copa_acute_uncomplicated,
                    copa_chronic_exacerbation=:copa_chronic_exacerbation, copa_undiagnosed_new=:copa_undiagnosed_new,
                    copa_acute_systemic=:copa_acute_systemic, copa_acute_complicated_injury=:copa_acute_complicated_injury,
                    copa_chronic_severe=:copa_chronic_severe, copa_threat_to_life=:copa_threat_to_life,
                    copa_level=:copa_level, copa_level_overridden=:copa_level_overridden,
                    dr_prior_external_notes=:dr_prior_external_notes, dr_review_test_results=:dr_review_test_results,
                    dr_order_tests=:dr_order_tests, dr_independent_historian=:dr_independent_historian,
                    dr_independent_interpretation=:dr_independent_interpretation, dr_external_discussion=:dr_external_discussion,
                    dr_level=:dr_level, dr_level_overridden=:dr_level_overridden,
                    risk_low=:risk_low, risk_prescription_drug_mgmt=:risk_prescription_drug_mgmt,
                    risk_minor_surgery_with_factors=:risk_minor_surgery_with_factors,
                    risk_elective_major_no_factors=:risk_elective_major_no_factors,
                    risk_hospitalization=:risk_hospitalization, risk_sdoh=:risk_sdoh,
                    risk_drug_intensive_monitoring=:risk_drug_intensive_monitoring,
                    risk_elective_major_with_factors=:risk_elective_major_with_factors,
                    risk_emergency_major_surgery=:risk_emergency_major_surgery,
                    risk_hospitalization_escalation=:risk_hospitalization_escalation,
                    risk_dnr_deescalate=:risk_dnr_deescalate, risk_parenteral_controlled=:risk_parenteral_controlled,
                    risk_level=:risk_level, risk_level_overridden=:risk_level_overridden,
                    em_code=:em_code, em_modifier=:em_modifier, patient_type=:patient_type,
                    level_method=:level_method, total_time=:total_time,
                    em_category=:em_category,
                    critical_care_minutes=:critical_care_minutes,
                    dx_codes=:dx_codes, procedure_cpts=:procedure_cpts,
                    entered_by=:entered_by, entered_at=CURRENT_TIMESTAMP
                WHERE chart_id=:chart_id
            """), params)
            replaced.append(matched_chart_num)
        else:
            db.execute(text("""
                INSERT INTO em_answer_keys (
                    chart_id,
                    copa_self_limited, copa_stable_acute, copa_stable_chronic,
                    copa_acute_uncomplicated, copa_chronic_exacerbation, copa_undiagnosed_new,
                    copa_acute_systemic, copa_acute_complicated_injury, copa_chronic_severe, copa_threat_to_life,
                    copa_level, copa_level_overridden,
                    dr_prior_external_notes, dr_review_test_results, dr_order_tests,
                    dr_independent_historian, dr_independent_interpretation, dr_external_discussion,
                    dr_level, dr_level_overridden,
                    risk_low, risk_prescription_drug_mgmt, risk_minor_surgery_with_factors,
                    risk_elective_major_no_factors, risk_hospitalization, risk_sdoh,
                    risk_drug_intensive_monitoring, risk_elective_major_with_factors,
                    risk_emergency_major_surgery, risk_hospitalization_escalation,
                    risk_dnr_deescalate, risk_parenteral_controlled,
                    risk_level, risk_level_overridden,
                    em_code, em_modifier, patient_type, level_method, total_time,
                    em_category, critical_care_minutes, dx_codes, procedure_cpts, entered_by
                ) VALUES (
                    :chart_id,
                    :copa_self_limited, :copa_stable_acute, :copa_stable_chronic,
                    :copa_acute_uncomplicated, :copa_chronic_exacerbation, :copa_undiagnosed_new,
                    :copa_acute_systemic, :copa_acute_complicated_injury, :copa_chronic_severe, :copa_threat_to_life,
                    :copa_level, :copa_level_overridden,
                    :dr_prior_external_notes, :dr_review_test_results, :dr_order_tests,
                    :dr_independent_historian, :dr_independent_interpretation, :dr_external_discussion,
                    :dr_level, :dr_level_overridden,
                    :risk_low, :risk_prescription_drug_mgmt, :risk_minor_surgery_with_factors,
                    :risk_elective_major_no_factors, :risk_hospitalization, :risk_sdoh,
                    :risk_drug_intensive_monitoring, :risk_elective_major_with_factors,
                    :risk_emergency_major_surgery, :risk_hospitalization_escalation,
                    :risk_dnr_deescalate, :risk_parenteral_controlled,
                    :risk_level, :risk_level_overridden,
                    :em_code, :em_modifier, :patient_type, :level_method, :total_time,
                    :em_category, :critical_care_minutes, :dx_codes, :procedure_cpts, :entered_by
                )
            """), params)
            stored.append(matched_chart_num)

    db.commit()
    return {
        "stored": stored,
        "replaced": replaced,
        "skipped_duplicates": skipped,
        "not_found": not_found,
        "wrong_specialty": wrong_specialty,
    }


# ── Excel template download ───────────────────────────────────────────────────
