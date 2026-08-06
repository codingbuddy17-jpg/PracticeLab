"""
Not every E/M encounter is levelled by MDM.

A preventive visit is levelled by the patient's age and whether they are new or
established. Critical care is levelled by total time. Neither has anything to
do with the COPA / Data Review / Risk tables — but the key demanded all three,
so a trainer keying a well-woman exam had to invent an MDM level and the coder
then had to guess the same invention to earn thirty points. Both were being
tested on a fiction.

The category comes from the E/M code, which is the one thing the key and the
coder both always state.
"""
import pytest

from routers.practicelab_pkg.em_grading import (
    CRITICAL_CARE, EMERGENCY, INPATIENT_OBSERVATION, OFFICE, OTHER, PREVENTIVE,
    applicable_weights, category_uses_mdm, critical_care_units, em_category,
    grade_em_chart, resolve_category,
)

CFG = {
    "line1_weight": 70.0, "line2_weight": 30.0,
    "em_level_weight": 23.33, "cpt_weight": 23.33, "dx_weight": 23.34,
    "copa_weight": 10.0, "dr_weight": 10.0, "risk_weight": 10.0,
    "pass_threshold": 80.0, "overcoding_penalty": True,
}


# ── classification ────────────────────────────────────────────────────────────

@pytest.mark.parametrize("code,expected", [
    ("99213", OFFICE), ("99205", OFFICE), ("99211", OFFICE),
    ("99223", INPATIENT_OBSERVATION), ("99232", INPATIENT_OBSERVATION),
    ("99236", INPATIENT_OBSERVATION), ("99238", INPATIENT_OBSERVATION),
    ("99284", EMERGENCY), ("99281", EMERGENCY),
    ("99396", PREVENTIVE), ("99385", PREVENTIVE), ("99401", PREVENTIVE),
    ("99291", CRITICAL_CARE), ("99292", CRITICAL_CARE), ("99471", CRITICAL_CARE),
])
def test_codes_land_in_their_category(code, expected):
    assert em_category(code) == expected


@pytest.mark.parametrize("code", ["99347", "99308", "99495", "99442", "", "banana"])
def test_anything_unmodelled_is_other_rather_than_a_guess(code):
    """
    Home visits, nursing facility, transitional care and telephone codes all
    land here on purpose: graded on the codes alone, which is what a category
    we do not model should do.
    """
    assert em_category(code) == OTHER


def test_only_the_mdm_driven_categories_use_mdm():
    assert category_uses_mdm(OFFICE)
    assert category_uses_mdm(INPATIENT_OBSERVATION)
    assert category_uses_mdm(EMERGENCY)
    assert not category_uses_mdm(PREVENTIVE)
    assert not category_uses_mdm(CRITICAL_CARE)
    assert not category_uses_mdm(OTHER)


def test_a_key_without_a_category_gets_one_from_its_code():
    """Every key written before categories existed keeps the category it had."""
    assert resolve_category(None, "99214") == OFFICE
    assert resolve_category("", "99284") == EMERGENCY


def test_a_stated_category_beats_the_code():
    """So a trainer can classify a code the app has never seen."""
    assert resolve_category("preventive", "99213") == PREVENTIVE


def test_a_nonsense_category_falls_back_to_the_code():
    assert resolve_category("whatever", "99213") == OFFICE


# ── weights ───────────────────────────────────────────────────────────────────

def _sum(w):
    return round(sum(w.values()), 2)


@pytest.mark.parametrize("category", [OFFICE, EMERGENCY, PREVENTIVE, CRITICAL_CARE, OTHER])
@pytest.mark.parametrize("has_cpts", [True, False])
def test_every_chart_is_scored_out_of_one_hundred(category, has_cpts):
    """
    The point of renormalising. Leave the reasoning weight in the denominator
    of a preventive chart and a coder who did everything right caps at 70%;
    average that against an office chart and the difference reported is purely
    arithmetic.
    """
    assert _sum(applicable_weights(CFG, category, has_cpts)) == 100.0


def test_an_office_chart_is_weighted_exactly_as_before():
    w = applicable_weights(CFG, OFFICE, has_cpts=True)
    assert w == {"em_level": 23.33, "cpt": 23.33, "dx": 23.34,
                 "copa": 10.0, "dr": 10.0, "risk": 10.0}


def test_a_chart_with_no_procedures_splits_the_cpt_weight_across_the_codes():
    """Line 1 renormalises within itself; the 70/30 split does not move."""
    w = applicable_weights(CFG, OFFICE, has_cpts=False)
    assert "cpt" not in w
    assert round(w["em_level"], 1) == 35.0
    assert round(w["dx"], 1) == 35.0
    assert w["copa"] == 10.0, "reasoning must not be rescaled by a Line 1 change"


def test_a_preventive_chart_has_no_mdm_component():
    w = applicable_weights(CFG, PREVENTIVE, has_cpts=True)
    assert "copa" not in w and "dr" not in w and "risk" not in w
    assert _sum(w) == 100.0


def test_a_preventive_chart_with_no_procedures_is_codes_alone():
    w = applicable_weights(CFG, PREVENTIVE, has_cpts=False)
    assert round(w["em_level"], 1) == 50.0
    assert round(w["dx"], 1) == 50.0


def test_critical_care_is_scored_in_thirds():
    """
    Code-and-time, diagnoses, procedures. The clock is not a fourth component:
    it is part of picking the code, so it rides with the E/M weight.
    """
    w = applicable_weights(CFG, CRITICAL_CARE, has_cpts=True)
    assert "copa" not in w and "critical_care_time" not in w
    assert round(w["em_level"], 1) == 33.3
    assert round(w["dx"], 1) == 33.3
    assert round(w["cpt"], 1) == 33.3
    assert _sum(w) == 100.0


def test_critical_care_without_procedures_is_scored_in_halves():
    w = applicable_weights(CFG, CRITICAL_CARE, has_cpts=False)
    assert "cpt" not in w
    assert round(w["em_level"], 1) == 50.0
    assert round(w["dx"], 1) == 50.0


def test_the_time_route_replaces_the_mdm_elements_not_the_line():
    w = applicable_weights(CFG, OFFICE, has_cpts=True, level_method="TIME")
    assert w["time"] == 30.0
    assert "copa" not in w
    assert _sum(w) == 100.0


# ── grading ───────────────────────────────────────────────────────────────────

def _key(**over):
    ak = {
        "em_code": "99214", "em_modifier": "", "patient_type": "NA",
        "dx_codes": '["E11.9"]', "procedure_cpts": "[]",
        "copa_level": "Moderate", "dr_level": "Moderate", "risk_level": "Moderate",
        "copa_stable_chronic": 2, "level_method": "MDM",
    }
    ak.update(over)
    return ak


def _sub(**over):
    s = {"sub_em_code": "99214", "sub_em_modifier": "", "sub_patient_type": "NA",
         "sub_dx_codes": ["E11.9"], "sub_procedure_cpts": [],
         "sub_level_method": "MDM"}
    s.update(over)
    return s


def test_a_preventive_chart_coded_correctly_scores_full_marks():
    """
    The bug this whole change exists for: the coder gets the code and the
    diagnosis right, and used to lose thirty points to MDM boxes that describe
    nothing about a well-patient exam.
    """
    ak = _key(em_code="99396", em_category="preventive")
    sub = _sub(sub_em_code="99396")
    assert grade_em_chart(ak, sub, CFG)["total_score"] == 100.0


def test_a_preventive_chart_is_not_scored_on_mdm_elements():
    """Ticking every MDM box, or none, changes nothing on a preventive chart."""
    ak = _key(em_code="99396", em_category="preventive")
    plain = grade_em_chart(ak, _sub(sub_em_code="99396"), CFG)
    busy = grade_em_chart(ak, _sub(sub_em_code="99396", sub_copa_threat_to_life=1,
                                   sub_risk_hospitalization=True), CFG)
    assert plain["total_score"] == busy["total_score"]
    assert busy["reasoning_accuracy_total"] == 0.0
    assert busy["uses_mdm"] is False


def test_an_office_chart_still_scores_on_mdm():
    got = grade_em_chart(_key(), _sub(sub_copa_stable_chronic=2), CFG)
    assert got["uses_mdm"] is True
    assert got["copa_element_score"] > 0


def test_a_wrong_code_still_costs_a_preventive_chart():
    """Removing MDM must not remove the grading."""
    ak = _key(em_code="99396", em_category="preventive")
    got = grade_em_chart(ak, _sub(sub_em_code="99395"), CFG)
    assert got["em_level_score"] == 0.0
    assert got["total_score"] < 100.0


def test_an_unmodelled_code_is_graded_on_the_codes_alone():
    """A home visit: right code, right diagnosis, nothing else asked of them."""
    ak = _key(em_code="99347")
    got = grade_em_chart(ak, _sub(sub_em_code="99347"), CFG)
    assert got["em_category"] == OTHER
    assert got["total_score"] == 100.0


# ── critical care ─────────────────────────────────────────────────────────────

def _cc_key(minutes=75, **over):
    return _key(em_code="99291", em_category="critical_care",
                critical_care_minutes=minutes, **over)


def test_critical_care_scores_on_the_time_the_coder_read():
    got = grade_em_chart(_cc_key(75), _sub(sub_em_code="99291",
                                           sub_critical_care_minutes=75), CFG)
    assert got["em_category"] == CRITICAL_CARE
    assert got["critical_care_minutes_ok"] is True
    assert got["total_score"] == 100.0


def test_misreading_the_clock_costs_half_the_code_component():
    """The service was identified; the quantity behind it was not — same rule
    a wrong Dx pointer and a wrong unit count already get."""
    right = grade_em_chart(_cc_key(75), _sub(sub_em_code="99291",
                                             sub_critical_care_minutes=75), CFG)
    wrong = grade_em_chart(_cc_key(75), _sub(sub_em_code="99291",
                                             sub_critical_care_minutes=40), CFG)
    assert wrong["critical_care_minutes_ok"] is False
    assert wrong["em_level_score"] == pytest.approx(right["em_level_score"] / 2, abs=0.02)


def test_not_documenting_the_time_at_all_costs_the_same():
    got = grade_em_chart(_cc_key(75), _sub(sub_em_code="99291"), CFG)
    assert got["critical_care_minutes_ok"] is False
    # Half of the 50 that a chart with no procedures carries.
    assert got["em_level_score"] == pytest.approx(25.0, abs=0.02)


def test_a_key_that_states_no_time_cannot_withhold_the_points():
    """The coder is not penalised for a gap in the key."""
    got = grade_em_chart(_cc_key(None), _sub(sub_em_code="99291"), CFG)
    assert got["critical_care_minutes_ok"] is None
    assert got["total_score"] == 100.0


def test_a_wrong_code_is_not_rescued_by_the_right_time():
    got = grade_em_chart(_cc_key(75), _sub(sub_em_code="99292",
                                           sub_critical_care_minutes=75), CFG)
    assert got["em_level_score"] == 0.0


def test_critical_care_ignores_the_mdm_boxes():
    ak = _cc_key(60)
    a = grade_em_chart(ak, _sub(sub_em_code="99291", sub_critical_care_minutes=60), CFG)
    b = grade_em_chart(ak, _sub(sub_em_code="99291", sub_critical_care_minutes=60,
                                sub_copa_threat_to_life=1), CFG)
    assert a["total_score"] == b["total_score"]


def test_the_add_on_is_counted_whether_repeated_or_carried_as_units():
    """Two 99292 lines and one line of two units are the same claim."""
    repeated = [{"code": "99292", "units": 1}, {"code": "99292", "units": 1}]
    carried = [{"code": "99292", "units": 2}]
    assert critical_care_units(repeated, "99291") == 2
    assert critical_care_units(carried, "99291") == 2


def test_an_encounter_with_no_add_on_counts_none():
    assert critical_care_units([{"code": "99291"}], "99291") == 0


# ── nothing moved for the charts that already worked ─────────────────────────

def test_an_ed_chart_grades_exactly_as_it_did():
    ak = _key(em_code="99284", copa_level="Moderate")
    got = grade_em_chart(ak, _sub(sub_em_code="99284", sub_copa_stable_chronic=2), CFG)
    assert got["em_category"] == EMERGENCY
    assert got["uses_mdm"] is True
    # No CPTs in the key, so Line 1's 70 points split across code and Dx.
    # Renormalising by ratio lands a hundredth off the old add-half-each
    # arithmetic; the split itself is unchanged.
    assert got["em_level_score"] == pytest.approx(35.0, abs=0.02)


# ── the E/M spreadsheet can state a category and a clock ──────────────────────

def test_the_em_template_carries_the_new_columns(client):
    import io
    from openpyxl import load_workbook

    r = client.get("/practicelab/em/answer-key/template")
    assert r.status_code == 200
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    headers = [c.value for c in ws[1]]
    assert any("Encounter Category" in str(h) for h in headers)
    assert any("Critical Care Total Minutes" in str(h) for h in headers)


def _em_key_workbook(**cells):
    """A one-row E/M key, written by column letter so indices stay honest."""
    import io
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Chart Number"
    ws["A2"] = "EM001"
    ws["AG2"] = "99291"
    for ref, val in cells.items():
        ws[f"{ref}2"] = val
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_the_category_and_clock_parse_out_of_the_spreadsheet():
    from services.excel_service import parse_em_answer_key_upload

    row = parse_em_answer_key_upload(
        _em_key_workbook(BK="Critical Care", BL=75))[0]
    assert row["em_category"] == "critical_care"
    assert row["critical_care_minutes"] == 75


def test_a_key_filled_before_categories_existed_states_neither():
    """Blank means "work it out from the code", not a category of its own."""
    from services.excel_service import parse_em_answer_key_upload

    row = parse_em_answer_key_upload(_em_key_workbook())[0]
    assert row["em_category"] == ""
    assert row["critical_care_minutes"] is None
    assert resolve_category(row["em_category"], "99291") == CRITICAL_CARE


def test_the_units_columns_did_not_move_when_the_category_columns_landed():
    """
    Every one of these columns was appended rather than inserted, and this is
    the assertion that keeps it that way: read the units back from BG.
    """
    from services.excel_service import parse_em_answer_key_upload

    row = parse_em_answer_key_upload(
        _em_key_workbook(AR="99292", BG=3, BK="Critical Care", BL=105))[0]
    assert row["procedure_cpts"][0] == {"code": "99292", "modifier": "", "pointers": [], "units": 3}
    assert row["critical_care_minutes"] == 105


def test_the_feedback_names_both_clocks(db):
    from routers.practicelab_pkg.practice_sessions import _em_feedback_items

    scoring = grade_em_chart(_cc_key(75), _sub(sub_em_code="99291",
                                               sub_critical_care_minutes=40), CFG)
    items = _em_feedback_items(scoring, _cc_key(75),
                               {"em_code": "99291"}, CFG)
    row = next(i for i in items if "Critical care time" in i["issue"])
    assert "75" in row["issue"] and "40" in row["issue"]
