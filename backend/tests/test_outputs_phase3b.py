"""
Phase 3B downstream output regression.

The setup deliberately submits real coder/auditor sessions first, then opens
the outputs that depend on those rows: review endpoints, analytics views, PDFs,
and workbook downloads.
"""

import io

from openpyxl import load_workbook

from models import Specialty
from tests.test_practice_submission_phase3a import (
    _add_answer_key,
    _audit_batch,
    _audit_chart,
    _audit_work,
    _coder_session,
    _em_entry,
    _em_key_payload,
    _entry,
    _truth,
)
from conftest import make_chart

XLSX_MIME = "spreadsheetml"


def _assert_pdf(resp):
    assert resp.status_code == 200, f"{resp.status_code}: {resp.text[:200]}"
    assert "application/pdf" in resp.headers.get("content-type", "")
    assert resp.content.startswith(b"%PDF")
    assert len(resp.content) > 500


def _assert_workbook(resp):
    assert resp.status_code == 200, f"{resp.status_code}: {resp.text[:200]}"
    assert XLSX_MIME in resp.headers.get("content-type", "")
    return load_workbook(io.BytesIO(resp.content))


def _submitted_coder(client, db, spec=Specialty.IP_DRG):
    chart = make_chart(db, specialty=spec.value, chart_number=f"P3B{spec.name[:8]}")
    _add_answer_key(db, chart, spec)
    session = _coder_session(client, spec, chart.id)
    entry = _entry(chart.id, spec)
    r = client.post(f"/practicelab/practice-sessions/{session['session_id']}/submit",
                    json={"entries": [entry]})
    assert r.status_code == 200, r.text
    return session, chart


def _submitted_em_coder(client, db, spec=Specialty.EM):
    chart = make_chart(db, specialty=spec.value, chart_number=f"P3B{spec.name}")
    db.commit()
    r = client.post("/practicelab/em/answer-key", json=_em_key_payload(chart.id, spec))
    assert r.status_code == 200, r.text
    session = _coder_session(client, spec, chart.id)
    submitted = client.post(
        f"/practicelab/practice-sessions/{session['session_id']}/submit",
        json={"entries": [_em_entry(chart.id, spec)]},
    )
    assert submitted.status_code == 200, submitted.text
    return session, chart


def _submitted_audit(client, db, clean=False):
    chart = _audit_chart(db, 10 if clean else 11)
    opened = _audit_batch(client, [chart.id], clean_share=100 if clean else 0)
    truth = _truth(client, opened["batch_id"])
    if clean:
        # Deterministic clean chart; the allocator keeps at least one
        # opportunity chart in very small batches.
        from models import AuditAssignment

        assignment = db.query(AuditAssignment).filter_by(
            batch_id=opened["batch_id"]).one()
        assignment.ground_truth = []
        db.commit()
        truth = {chart.id: []}
    r = client.post(f"/auditor/sessions/{opened['session_id']}/submit",
                    json=_audit_work(opened, truth, perfect=not clean))
    assert r.status_code == 200, r.text
    return opened, chart


def test_coder_standard_outputs_survive_real_submission(client, db):
    session, chart = _submitted_coder(client, db, Specialty.IP_DRG)

    review = client.get(f"/practicelab/practice-sessions/{session['session_id']}/review")
    assert review.status_code == 200, review.text
    assert review.json()["charts"][0]["total_score"] == 100

    _assert_pdf(client.get(
        f"/practicelab/practice-sessions/{session['session_id']}/coder-report.pdf"))
    _assert_pdf(client.get(
        f"/practicelab/batches/{session['batch_id']}/insights/report.pdf"))

    batch_wb = _assert_workbook(client.get(
        f"/practicelab/batches/{session['batch_id']}/results/export"))
    assert batch_wb.sheetnames

    perf_wb = _assert_workbook(client.get(
        "/practicelab/analytics/coder-performance.xlsx"))
    assert "Results" in perf_wb.sheetnames

    analytics_checks = [
        client.get("/practicelab/analytics/overview"),
        client.get("/practicelab/analytics/by-batch"),
        client.get("/practicelab/analytics/by-chart"),
        client.get("/practicelab/analytics/by-category"),
        client.get("/practicelab/analytics/coder-summary",
                   params={"coder_name": "Coder One"}),
        client.get(f"/practicelab/analytics/chart-detail/{chart.chart_number}"),
    ]
    for resp in analytics_checks:
        assert resp.status_code == 200, resp.text


def test_coder_em_outputs_survive_real_submission(client, db):
    session, _chart = _submitted_em_coder(client, db, Specialty.EM)

    review = client.get(f"/practicelab/practice-sessions/{session['session_id']}/review")
    assert review.status_code == 200, review.text
    assert review.json()["charts"][0]["total_score"] == 100

    _assert_pdf(client.get(
        f"/practicelab/practice-sessions/{session['session_id']}/coder-report.pdf"))

    mdm = client.get("/practicelab/analytics/em-mdm")
    assert mdm.status_code == 200, mdm.text
    assert mdm.json()["team"]["chart_count"] >= 1


def test_auditor_outputs_survive_real_submission(client, db):
    opened, _chart = _submitted_audit(client, db)

    review = client.get(f"/auditor/sessions/{opened['session_id']}/review")
    assert review.status_code == 200, review.text
    assert review.json()["summary"]["audit_score"] == 100.0

    overview = client.get("/auditor/analytics/overview")
    assert overview.status_code == 200, overview.text
    assert overview.json()["audit_score"] == 100.0

    analytics_checks = [
        client.get("/auditor/analytics/by-batch"),
        client.get("/auditor/analytics/by-auditor"),
        client.get("/auditor/analytics/by-specialty"),
        client.get("/auditor/analytics/chart-signals"),
        client.get("/auditor/analytics/detection"),
        client.get("/auditor/analytics/pattern", params={"kind": "omit_sdx"}),
    ]
    for resp in analytics_checks:
        assert resp.status_code == 200, resp.text

    _assert_workbook(client.get(f"/auditor/batches/{opened['batch_id']}/export"))
    _assert_workbook(client.get("/auditor/analytics/export"))
    _assert_pdf(client.get(f"/auditor/batches/{opened['batch_id']}/report.pdf"))
    _assert_pdf(client.get("/auditor/analytics/auditor-report.pdf",
                           params={"auditor": "Auditor One"}))
