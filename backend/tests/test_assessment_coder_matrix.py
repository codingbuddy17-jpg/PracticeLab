"""
The Coder Matrix identifies people the way the rest of the module does, and
exports the whole grid rather than the page on screen.
"""
import io
from datetime import datetime, timedelta, timezone

from openpyxl import load_workbook

from conftest import seed_question_pool
from models import (AssessmentResponse, AssessmentResult, AssessmentSession,
                    GeneratedAssessment, GeneratedAssessmentStudent)


def _payload(name, coders):
    return {
        "assessment_name": name,
        "coders": coders,
        "duration_minutes": 60,
        "total_questions": 3,
        "specialty_mix": [{"specialty": "IP-DRG", "pct": 1.0, "topic_filter": ""}],
        "difficulty_mode": "auto",
        "generated_by": "trainer",
        "save_config": False,
        "randomise": True,
    }


def _sit(client, token):
    started = client.post(f"/assessment/take/{token}/start").json()
    for i, q in enumerate(started["questions"]):
        client.post(f"/assessment/take/{token}/answer", json={
            "question_index": i, "question_id": q["question_id"], "selected_answer": "A",
        })
    client.post(f"/assessment/take/{token}/submit", json={"auto_submitted": False})


def _run(client, db, coders):
    seed_question_pool(db)
    gen = client.post("/assessment/generate", json=_payload("Paper", coders)).json()
    for s in gen["sessions"]:
        _sit(client, s["session_token"])


def test_same_named_coders_are_separate_rows(client, db):
    """Grouping on the typed name alone merged two people into one row."""
    _run(client, db, [{"coder_name": "Alice", "employee_id": "E001"},
                      {"coder_name": "Alice", "employee_id": "E002"}])
    d = client.get("/assessment/analytics/coder-matrix").json()
    assert len(d["coders"]) == 2
    assert sorted(c["employee_id"] for c in d["coders"]) == ["E001", "E002"]


def test_matrix_publishes_the_bar_for_below_threshold(client, db):
    _run(client, db, [{"coder_name": "Alice", "employee_id": "E001"}])
    d = client.get("/assessment/analytics/coder-matrix").json()
    assert d["default_pass_threshold"] == 90.0


def test_matrix_cells_carry_counts_and_gap_summary(client, db):
    a = GeneratedAssessment(assessment_name="Matrix detail", student_count=1, generated_by="trainer")
    db.add(a)
    db.commit()
    questions = [
        {"question_id": "M1", "question_text": "IP question", "specialty": "IP-DRG", "topic": "DRG"},
        {"question_id": "M2", "question_text": "E/M question", "specialty": "E/M", "topic": "MDM"},
    ]
    slot = GeneratedAssessmentStudent(assessment_id=a.id, student_label="Alice", questions_json=questions)
    db.add(slot)
    db.commit()
    s = AssessmentSession(
        session_token="MATRIX1",
        assessment_id=a.id,
        student_slot_id=slot.id,
        coder_name="Alice",
        employee_id="E001",
        duration_minutes=30,
        expires_at=datetime.now(timezone.utc) + timedelta(hours=8),
        status="submitted",
    )
    db.add(s)
    db.commit()
    db.add_all([
        AssessmentResult(session_id=s.id, total_questions=2, correct_count=1, score_pct=50),
        AssessmentResponse(session_id=s.id, question_index=0, question_id="M1",
                           selected_answer="A", is_correct=False),
        AssessmentResponse(session_id=s.id, question_index=1, question_id="M2",
                           selected_answer="A", is_correct=True),
    ])
    db.commit()

    row = client.get("/assessment/analytics/coder-matrix").json()["coders"][0]

    assert row["specialty_counts"]["IP-DRG"] == {"correct": 0, "total": 1}
    assert row["specialty_counts"]["E/M"] == {"correct": 1, "total": 1}
    assert row["gap_count"] == 1
    assert row["weakest_specialty"]["specialty"] == "IP-DRG"


def test_export_returns_a_readable_workbook(client, db):
    _run(client, db, [{"coder_name": "Alice", "employee_id": "E001"},
                      {"coder_name": "Bob", "employee_id": "E002"}])
    r = client.get("/assessment/analytics/coder-matrix.xlsx")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    text = "\n".join(
        " ".join(str(c) for c in row if c is not None)
        for row in wb.active.iter_rows(values_only=True)
    )
    assert "Alice" in text and "Bob" in text
    assert "E001" in text and "E002" in text


def test_export_honours_the_window(client, db):
    """The export takes the same filters as the screen."""
    seed_question_pool(db)
    for batch in ("Wave 1", "Wave 2"):
        p = _payload(f"{batch} paper", [{"coder_name": f"C-{batch}", "employee_id": batch}])
        p["batch_name"] = batch
        gen = client.post("/assessment/generate", json=p).json()
        _sit(client, gen["sessions"][0]["session_token"])

    r = client.get("/assessment/analytics/coder-matrix.xlsx", params={"batch_name": "Wave 1"})
    wb = load_workbook(io.BytesIO(r.content))
    text = "\n".join(
        " ".join(str(c) for c in row if c is not None)
        for row in wb.active.iter_rows(values_only=True)
    )
    assert "C-Wave 1" in text
    assert "C-Wave 2" not in text
