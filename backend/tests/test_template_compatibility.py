"""
Every answer key template, for every specialty, round-trips.

Adding units put a new column inside the CPT block. A key a trainer filled from
last month's download has one layout and a key filled today has another, and
both land on the same upload endpoint. This file walks all ten specialties and
pins three things per specialty:

  * the template's own columns are what the parser expects (download → fill →
    upload)
  * a file with the OLD column set still parses correctly (no units column)
  * the stored key exports back into a file that re-uploads to the same key

The generic OP layout is not one layout. Surgery/ED Profee/E/M carry Dx
pointers, ED Single Path carries two level columns before PDx, Ancillary has no
CPT block at all, and Edits/Denials now have no units column. A stride-counted
parser gets three of those wrong, which is the bug this file exists to catch.
"""
import io

import pytest
from openpyxl import Workbook, load_workbook

from models import Specialty
from routers.practicelab_pkg.shared import (
    _is_ip, _uses_pointers, _is_single_path, _is_dx_only, _uses_units,
)
from services.excel_service import (
    generate_answer_key_template, parse_answer_key_upload,
    generate_coder_sheet, parse_submission,
)

ALL = list(Specialty)


def _flags(spec):
    """Exactly what the download and upload endpoints derive, from one source."""
    return dict(
        with_pointers=_uses_pointers(spec),
        single_path=_is_single_path(spec),
        dx_only=_is_dx_only(spec),
    )


def _template(spec):
    f = _flags(spec)
    return generate_answer_key_template(
        "IP" if _is_ip(spec) else "OP", with_units=_uses_units(spec), **f)


def _headers(data):
    wb = load_workbook(io.BytesIO(data))
    return [str(c.value).split("\n")[0] if c.value else "" for c in wb.worksheets[0][1]]


def _parse(data, spec):
    return parse_answer_key_upload(
        data, "IP" if _is_ip(spec) else "OP", **_flags(spec))


def _fill(data, values: dict):
    """Write {header_name: value} into row 2 of a downloaded template."""
    wb = load_workbook(io.BytesIO(data))
    ws = wb.worksheets[0]
    at = {str(c.value).split("\n")[0]: c.column for c in ws[1] if c.value}
    for name, val in values.items():
        assert name in at, f"{name} is not a column in this template: {sorted(at)}"
        ws.cell(2, at[name], val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── the templates are internally consistent ───────────────────────────────────

@pytest.mark.parametrize("spec", ALL, ids=lambda s: s.value)
def test_every_template_generates_and_has_a_chart_number(spec):
    assert _headers(_template(spec))[0] == "Chart_Number"


@pytest.mark.parametrize("spec", ALL, ids=lambda s: s.value)
def test_the_units_column_appears_exactly_where_it_should(spec):
    headers = _headers(_template(spec))
    has = "CPT_1_Units" in headers
    # No CPT block at all → nothing to carry units, whatever the rule says.
    if _is_ip(spec) or _is_dx_only(spec):
        assert not has
    else:
        assert has is _uses_units(spec), f"{spec.value}: units column present={has}"


@pytest.mark.parametrize("spec", ALL, ids=lambda s: s.value)
def test_no_duplicate_columns(spec):
    """A repeated header makes the name-keyed parser read the wrong cell."""
    headers = [h for h in _headers(_template(spec)) if h]
    assert len(headers) == len(set(headers)), \
        f"{spec.value}: duplicates {[h for h in headers if headers.count(h) > 1]}"


@pytest.mark.parametrize("spec", ALL, ids=lambda s: s.value)
def test_single_path_level_columns_precede_the_diagnoses(spec):
    headers = _headers(_template(spec))
    if _is_single_path(spec):
        assert headers[1:4] == ["Facility_ED_Level", "Profee_ED_Level", "PDx_Code"]
    else:
        assert "Facility_ED_Level" not in headers


# ── download → fill → upload ──────────────────────────────────────────────────

@pytest.mark.parametrize("spec", [s for s in ALL if not _is_ip(s)], ids=lambda s: s.value)
def test_a_freshly_downloaded_template_round_trips(spec):
    values = {"Chart_Number": "X001", "PDx_Code": "E11.9", "SDx_1": "I10"}
    if _is_single_path(spec):
        values["Facility_ED_Level"] = "99283"
        values["Profee_ED_Level"] = "99284"
    if not _is_dx_only(spec):
        values["CPT_1"] = "11042"
        values["CPT_1_Modifier"] = "59"
        if _uses_units(spec):
            values["CPT_1_Units"] = 2
        if _uses_pointers(spec):
            values["CPT_1_DxPointers"] = "1,2"

    row = _parse(_fill(_template(spec), values), spec)[0]
    assert row["chart_number"] == "X001"
    assert row["pdx_code"] == "E11.9"
    assert [s["code"] for s in row["sdx"]] == ["I10"]

    if _is_single_path(spec):
        assert row["facility_level"] == "99283"
        assert row["profee_level"] == "99284"

    if _is_dx_only(spec):
        assert row["cpt"] == []
        return

    line = row["cpt"][0]
    assert line["code"] == "11042", f"{spec.value}: read the wrong cell as the code"
    assert line["modifier"] == "59", f"{spec.value}: read the wrong cell as the modifier"
    assert line.get("units") == (2 if _uses_units(spec) else None)
    if _uses_pointers(spec):
        assert line["pointers"] == ["1", "2"]


def test_the_inpatient_template_round_trips():
    spec = Specialty.IP_DRG
    data = _fill(_template(spec), {
        "Chart_Number": "IP001", "PDx_Code": "J18.9", "PDx_POA": "Y",
        "SDx_1": "I10", "SDx_1_POA": "Y", "SDx_1_CCMCC": "CC",
        "PCS_1": "0DTJ4ZZ",
    })
    row = _parse(data, spec)[0]
    assert row["pdx_code"] == "J18.9" and row["pdx_poa"] == "Y"
    assert row["sdx"][0] == {"code": "I10", "poa": "Y", "ccmcc": "CC"}
    assert [p["code"] for p in row["pcs"]] == ["0DTJ4ZZ"]


# ── keys filled before units existed ──────────────────────────────────────────

def _legacy_op_workbook(spec, with_pointers):
    """The OP layout as it stood before units: code, modifier[, pointers]."""
    wb = Workbook()
    ws = wb.active
    headers = ["Chart_Number"]
    if _is_single_path(spec):
        headers += ["Facility_ED_Level", "Profee_ED_Level"]
    headers.append("PDx_Code")
    headers += [f"SDx_{i}" for i in range(1, 21)]
    for i in range(1, 11):
        headers += [f"CPT_{i}", f"CPT_{i}_Modifier"]
        if with_pointers:
            headers.append(f"CPT_{i}_DxPointers")
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    at = {h: i + 1 for i, h in enumerate(headers)}
    ws.cell(2, at["Chart_Number"], "OLD1")
    ws.cell(2, at["PDx_Code"], "E11.9")
    ws.cell(2, at["SDx_1"], "I10")
    ws.cell(2, at["CPT_1"], "11042")
    ws.cell(2, at["CPT_1_Modifier"], "59")
    ws.cell(2, at["CPT_2"], "11043")
    if with_pointers:
        ws.cell(2, at["CPT_1_DxPointers"], "1,2")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.parametrize(
    "spec",
    [s for s in ALL if not _is_ip(s) and not _is_dx_only(s)],
    ids=lambda s: s.value,
)
def test_a_key_filled_from_the_pre_units_template_still_parses(spec):
    """
    The compatibility case that matters: units inserted a column in the middle
    of a repeating block, and half the filled keys in circulation predate it.
    """
    pointers = _uses_pointers(spec)
    row = _parse(_legacy_op_workbook(spec, pointers), spec)[0]
    assert row["chart_number"] == "OLD1"
    assert row["pdx_code"] == "E11.9"

    first, second = row["cpt"][0], row["cpt"][1]
    assert first["code"] == "11042"
    assert first["modifier"] == "59"
    assert "units" not in first, "a key that never mentioned units must not claim one"
    # The second line proves the block did not drift: a stride that mis-sized
    # one line puts every later line in the wrong column.
    assert second["code"] == "11043"
    assert second["modifier"] == ""
    if pointers:
        assert first["pointers"] == ["1", "2"]


# ── export → re-upload ────────────────────────────────────────────────────────

@pytest.mark.parametrize(
    "spec",
    [s for s in ALL if not _is_ip(s) and not _is_dx_only(s)],
    ids=lambda s: s.value,
)
def test_a_stored_key_exports_into_a_file_that_re_uploads_unchanged(spec, db):
    from models import Chart, AnswerKey, ChartStatus, Difficulty
    from services.excel_service import export_all_answer_keys

    cpt = [{"code": "11042", "modifier": "59", "pointers": ["1"] if _uses_pointers(spec) else []}]
    if _uses_units(spec):
        cpt[0]["units"] = 3

    chart = Chart(chart_number="RT001", specialty=spec, category="Test",
                  difficulty=Difficulty.INTERMEDIATE, status=ChartStatus.ACTIVE,
                  uploaded_by="t")
    db.add(chart)
    db.flush()
    ak = AnswerKey(chart_id=chart.id, specialty=spec, pdx_code="E11.9",
                   sdx=[{"code": "I10"}], cpt=cpt, entered_by="t")
    db.add(ak)
    db.flush()

    wb = load_workbook(io.BytesIO(export_all_answer_keys([(ak, chart)])))
    ws = wb.worksheets[0]
    buf = io.BytesIO()
    single = Workbook()
    single.remove(single.active)
    copy = single.create_sheet("k")
    for r in ws.iter_rows(values_only=True):
        copy.append(list(r))
    single.save(buf)

    row = _parse(buf.getvalue(), spec)[0]
    assert row["chart_number"] == "RT001"
    assert row["pdx_code"] == "E11.9"
    line = row["cpt"][0]
    assert line["code"] == "11042"
    assert line["modifier"] == "59"
    assert line.get("units") == (3 if _uses_units(spec) else None)
    if _uses_pointers(spec):
        assert line["pointers"] == ["1"]


# ── the coder's own workbook ──────────────────────────────────────────────────

@pytest.mark.parametrize("spec", ALL, ids=lambda s: s.value)
def test_the_coder_sheet_round_trips_for_every_specialty(spec):
    sheet = generate_coder_sheet("Alice", "Wave 1", [{
        "chart_number": "C001", "specialty": spec.value,
        "category": "Test", "difficulty": "Medium", "chart_url": "",
    }])
    wb = load_workbook(io.BytesIO(sheet))
    ws = wb["C001"]
    ip = _is_ip(spec)

    def _section(n):
        return next(r for r in range(1, ws.max_row + 1)
                    if f"SECTION {n}" in str(ws.cell(r, 1).value or ""))

    ws.cell(_section(1) + 2, 1, "E11.9" if not ip else "J18.9")
    if ip:
        ws.cell(_section(1) + 2, 2, "Y")
    ws.cell(_section(2) + 2, 2, "I10")
    proc = _section(3) + 2
    ws.cell(proc, 2, "0DTJ4ZZ" if ip else "11042")
    if not ip:
        ws.cell(proc, 3, "59")
        if _uses_units(spec):
            ws.cell(proc, 4, 4)

    buf = io.BytesIO()
    wb.save(buf)
    got = parse_submission(buf.getvalue())[0]
    assert got["chart_number"] == "C001"
    assert [s["code"] for s in got["sdx"]] == ["I10"]
    if ip:
        assert [p["code"] for p in got["pcs"]] == ["0DTJ4ZZ"]
        return
    line = got["cpt"][0]
    assert line["code"] == "11042" and line["modifier"] == "59"
    assert line.get("units") == (4 if _uses_units(spec) else None)


@pytest.mark.parametrize("spec", ALL, ids=lambda s: s.value)
def test_the_coder_sheet_offers_units_exactly_where_the_key_does(spec):
    sheet = generate_coder_sheet("Alice", "W", [{
        "chart_number": "C001", "specialty": spec.value,
        "category": "T", "difficulty": "Medium", "chart_url": "",
    }])
    ws = load_workbook(io.BytesIO(sheet))["C001"]
    text = " ".join(str(c.value or "") for row in ws.iter_rows() for c in row)
    expected = _uses_units(spec) and not _is_ip(spec)
    assert ("Units" in text) is expected, \
        f"{spec.value}: coder sheet and answer key disagree about units"
