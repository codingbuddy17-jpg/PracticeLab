"""
Units on a CPT line.

A coder can find the right procedure and still state the wrong quantity — two
units of a bilateral procedure, three of an add-on. That is a lesser mistake
than picking the wrong code: they read the chart correctly and described it
badly. It costs half the line, the same as a wrong diagnosis pointer, and for
the same reason.

Units are graded ONLY where the answer key states them, so every key written
before units existed grades exactly as it did.
"""
import io

import pytest

from services.grading_engine import norm_units, _match_cpt


def _line(code, modifier="", pointers=None, units=None):
    row = {"code": code, "modifier": modifier, "pointers": pointers or []}
    if units is not None:
        row["units"] = units
    return row


def _score(ak, cdr, **kw):
    """_match_cpt returns (matched, extra, feedback)."""
    return _match_cpt(ak, cdr, penalty=False, **kw)[0]


def _feedback(ak, cdr, **kw):
    return _match_cpt(ak, cdr, penalty=False, **kw)[2]


# ── norm_units ────────────────────────────────────────────────────────────────

def test_a_blank_cell_means_one_unit():
    """What a coder writing a single procedure leaves behind."""
    for blank in (None, "", "   "):
        assert norm_units(blank) == 1


def test_numbers_are_read_however_they_are_typed():
    assert norm_units(2) == 2
    assert norm_units("2") == 2
    assert norm_units(" 3 ") == 3
    assert norm_units("2.0") == 2


def test_nonsense_falls_back_to_one_rather_than_failing():
    """A stray character in a spreadsheet must not stop a chart being graded."""
    assert norm_units("two") == 1
    assert norm_units("-4") == 1
    assert norm_units(0) == 1


# ── grading ───────────────────────────────────────────────────────────────────

def test_matching_units_score_the_whole_line():
    assert _score([_line("11042", units=2)], [_line("11042", units=2)]) == 1.0


def test_a_units_mismatch_costs_half_the_line():
    """The procedure was right; the quantity was not."""
    assert _score([_line("11042", units=2)], [_line("11042", units=1)]) == 0.5


def test_a_key_without_units_grades_as_before(db):
    """
    Keys written before units existed must be untouched — absence means "do not
    grade units", which is different from an explicit 1.
    """
    assert _score([_line("11042")], [_line("11042", units=5)]) == 1.0


def test_an_explicit_one_is_still_graded():
    """Stating 1 is a claim about the line, not the same as saying nothing."""
    assert _score([_line("11042", units=1)], [_line("11042", units=3)]) == 0.5


def test_a_missing_coder_unit_reads_as_one():
    """The coder left it blank, which means one — and one is wrong here."""
    assert _score([_line("11042", units=2)], [_line("11042")]) == 0.5


def test_the_wrong_code_still_costs_the_whole_line():
    """Units must not soften a genuinely wrong procedure."""
    assert _score([_line("11042", units=2)], [_line("11043", units=2)]) == 0.0


def test_units_and_pointer_errors_do_not_stack(db):
    """
    One line, one half-credit. Zeroing it would price a line that was found and
    described badly the same as one that was missed entirely.
    """
    ak = [_line("11042", pointers=["1"], units=2)]
    cdr = [_line("11042", pointers=["2"], units=1)]
    got = _score(ak, cdr, ak_dx=["L97.909", "E11.9"],
                 cdr_dx=["L97.909", "E11.9"], check_pointers=True)
    assert got == 0.5


def test_the_feedback_names_both_numbers():
    """A coder cannot fix "units wrong" without being told which and what."""
    fb = _feedback([_line("11042", units=3)], [_line("11042", units=1)])
    rows = [f for f in fb if f.issue_type == "Wrong_Units"]
    assert len(rows) == 1
    assert "3" in rows[0].detail and "1" in rows[0].detail


# ── the E/M grader follows the same rule ──────────────────────────────────────

def test_the_em_normaliser_keeps_units_only_when_stated():
    from routers.practicelab_pkg.em_grading import normalise_cpts

    stated = normalise_cpts([{"code": "99213", "units": 2}])
    assert stated[0]["units"] == 2

    absent = normalise_cpts([{"code": "99213"}])
    assert "units" not in absent[0], "absence must stay absent, not become 1"

    legacy = normalise_cpts(["99213:25"])
    assert "units" not in legacy[0], "legacy string keys never stated units"


# ── the spreadsheets can actually state units ─────────────────────────────────

def _filled_key(headers, values):
    """A one-row answer key workbook with the given header row."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    for c, v in enumerate(values, 1):
        ws.cell(2, c, v)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_the_answer_key_template_offers_a_units_column():
    from openpyxl import load_workbook
    from services.excel_service import generate_answer_key_template

    wb = load_workbook(io.BytesIO(generate_answer_key_template("OP")))
    headers = [c.value for c in wb.worksheets[0][1]]
    assert "CPT_1_Units" in headers


def test_an_uploaded_key_carries_its_units_through():
    from services.excel_service import parse_answer_key_upload

    data = _filled_key(["Chart_Number", "PDx_Code", "CPT_1", "CPT_1_Modifier", "CPT_1_Units"],
                       ["OP001", "E11.9", "11042", "", 2])
    row = parse_answer_key_upload(data, "OP")[0]
    assert row["cpt"][0]["units"] == 2


def test_a_blank_units_cell_states_nothing():
    """Blank must not become an explicit 1 — that would grade units on every line."""
    from services.excel_service import parse_answer_key_upload

    data = _filled_key(["Chart_Number", "PDx_Code", "CPT_1", "CPT_1_Modifier", "CPT_1_Units"],
                       ["OP001", "E11.9", "11042", "", ""])
    assert "units" not in parse_answer_key_upload(data, "OP")[0]["cpt"][0]


def test_a_key_filled_from_the_old_template_still_parses():
    """
    Units added a column in the middle of the CPT block. A key filled from last
    month's download has no such column — read by stride, its modifier would be
    read as units and its pointers as a modifier.
    """
    from services.excel_service import parse_answer_key_upload

    data = _filled_key(["Chart_Number", "PDx_Code", "CPT_1", "CPT_1_Modifier", "CPT_1_DxPointers"],
                       ["OP001", "E11.9", "11042", "59", "1"])
    line = parse_answer_key_upload(data, "OP", with_pointers=True)[0]["cpt"][0]
    assert line["code"] == "11042"
    assert line["modifier"] == "59"
    assert line["pointers"] == ["1"]
    assert "units" not in line


def test_the_coder_sheet_round_trips_units():
    """What the coder writes has to survive back out of their own workbook."""
    from openpyxl import load_workbook
    from services.excel_service import generate_coder_sheet, parse_submission

    sheet = generate_coder_sheet("Alice", "Wave 1", [{
        "chart_number": "OP001", "specialty": "Surgery",
        "category": "Ortho", "difficulty": "Medium", "chart_url": "",
    }])
    wb = load_workbook(io.BytesIO(sheet))
    ws = wb["OP001"]
    proc = next(r for r in range(1, ws.max_row + 1)
                if "SECTION 3" in str(ws.cell(r, 1).value or ""))
    first = proc + 2
    ws.cell(first, 2, "11042")
    ws.cell(first, 3, "59")
    ws.cell(first, 4, 3)

    buf = io.BytesIO()
    wb.save(buf)
    line = parse_submission(buf.getvalue())[0]["cpt"][0]
    assert line["code"] == "11042" and line["modifier"] == "59"
    assert line["units"] == 3
