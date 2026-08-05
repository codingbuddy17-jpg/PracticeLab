"""
Uploading a question workbook.

The bank is read once before the row loop rather than queried per row. The old
shape called _next_qid on every blank-ID row, and each call fetched every
question_id for the specialty — millions of rows for one upload against a real
bank, and a request that can outlive the gateway. A proxy timeout carries no
JSON body, so that surfaces to the trainer as a bare "upload failed" with no
reason at all.
"""
import io

import openpyxl

from conftest import make_question

from models import AssessmentQuestion

MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _template(client, specialty="ICD10CM"):
    r = client.get("/assessment/questions/template", params={"specialty": specialty})
    assert r.status_code == 200
    return openpyxl.load_workbook(io.BytesIO(r.content))


def _upload(client, wb, specialty="ICD10CM", name="q.xlsx"):
    buf = io.BytesIO()
    wb.save(buf)
    return client.post(
        "/assessment/questions/upload",
        data={"specialty": specialty, "uploaded_by": "trainer"},
        files={"file": (name, buf.getvalue(), MIME)},
    )


def _row(text, correct="A", topic="General"):
    return ["", text, "a", "b", "c", "d", correct, "Medium", topic, "Conceptual", "Active", "Yes"]


def test_blank_ids_get_a_distinct_id_each(client, db):
    """Every new row needs its own id; a shared one would collide on commit."""
    wb = _template(client)
    ws = wb.active
    for i in range(25):
        ws.append(_row(f"Question {i}?"))
    r = _upload(client, wb)
    assert r.status_code == 200, r.text
    ids = r.json()["created_ids"]
    assert len(ids) == len(set(ids)) == 28   # 3 template samples + 25


def test_ids_continue_from_the_existing_bank(client, db):
    wb = _template(client)
    assert _upload(client, wb).status_code == 200
    wb2 = _template(client)
    ws = wb2.active
    for i in range(3):
        ws.append(_row(f"Second batch {i}?"))
    created = _upload(client, wb2).json()["created_ids"]
    # The three template samples repeat and are caught as duplicates; only the
    # genuinely new rows are stored, numbered after what already existed.
    assert created == ["ICDX-004", "ICDX-005", "ICDX-006"]


def test_a_duplicate_inside_one_file_is_caught(client, db):
    """
    Two identical questions in the same upload. The per-row query used to catch
    this only because SQLAlchemy autoflushed; reading the bank up front has to
    track texts added during this file explicitly.
    """
    wb = _template(client)
    ws = wb.active
    ws.append(_row("Exactly the same question?"))
    ws.append(_row("Exactly the same question?"))
    d = _upload(client, wb).json()
    assert d["created"] == 4          # 3 samples + one of the pair
    assert d["duplicates"] == 1
    assert "already exists" in d["duplicate_warnings"][0]


def test_reupload_updates_rather_than_duplicating(client, db):
    wb = _template(client)
    _upload(client, wb)
    wb2 = _template(client)
    ws = wb2.active
    ws.cell(row=2, column=1, value="ICDX-001")
    ws.cell(row=2, column=2, value="Edited question text?")
    d = _upload(client, wb2).json()
    assert "ICDX-001" in d["updated_ids"]
    q = db.query(AssessmentQuestion).filter(
        AssessmentQuestion.question_id == "ICDX-001").first()
    assert q.question_text == "Edited question text?"


def test_an_id_from_another_specialty_is_refused_not_stolen(client, db):
    """
    Caught by the prefix rule before the row is even looked up, so the message
    names the prefix rather than the owning specialty. Either way the row is
    refused rather than quietly rewriting another specialty's question.
    """
    wb = _template(client)
    _upload(client, wb)
    wb2 = _template(client, "Surgery")
    ws = wb2.active
    ws.cell(row=2, column=1, value="ICDX-001")
    d = _upload(client, wb2, "Surgery").json()
    assert d["created"] == 2   # the other two sample rows still import
    assert any("ICDX-001" in e for e in d["errors"]), d["errors"]


def test_a_non_xlsx_file_says_so_in_plain_words(client, db):
    r = client.post(
        "/assessment/questions/upload",
        data={"specialty": "ICD10CM", "uploaded_by": "trainer"},
        files={"file": ("q.csv", b"Question_Text\nHello?\n", "text/csv")},
    )
    assert r.status_code == 400
    detail = r.json()["detail"]
    assert "zip" not in detail.lower(), f"jargon leaked to the trainer: {detail}"
    assert ".xlsx" in detail


def test_the_bank_is_read_once_not_per_row(client, db):
    """A guard on the shape that made large uploads time out."""
    from sqlalchemy import event

    for _ in range(200):
        make_question(db, specialty="ICD10CM")
    db.commit()

    wb = _template(client)
    ws = wb.active
    for i in range(60):
        ws.append(_row(f"Perf question {i}?"))

    n = {"q": 0}
    bind = db.get_bind()

    def before(conn, cur, stmt, params, ctx, many):
        n["q"] += 1

    event.listen(bind, "before_cursor_execute", before)
    try:
        assert _upload(client, wb).status_code == 200
    finally:
        event.remove(bind, "before_cursor_execute", before)

    # Was ~4 queries per row plus a full scan inside each; now a small fixed
    # preamble and one insert per row.
    assert n["q"] < 2 * 63, f"query count grew back to per-row scanning: {n['q']}"
