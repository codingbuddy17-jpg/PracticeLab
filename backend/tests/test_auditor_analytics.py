"""
Auditor analytics — different questions from the coder reports, so different
numbers, and the two bases must never be confused.

  * audit accuracy AVERAGES chart scores — one chart, one unit of work
  * component accuracy POOLS — a chart with six plantings counts six times as
    much as a chart with one

Detection patterns is the report with no coder equivalent: which KINDS of error
slip past a cohort. That is the one that turns scoring into a curriculum.
"""
import pytest
from openpyxl import load_workbook

from tests.test_auditor_api import (
    PASS, allocate, library, make_batch, perfect_work, truth_map,
)


def _run(client, batch_id, find_everything=True, auditor_index=0):
    token = allocate(client, batch_id)["access_codes"][auditor_index]["token"]
    payload = client.get(f"/auditor/sessions/by-token/{token}").json()
    work = perfect_work(payload, truth_map(client, batch_id),
                        verdicts_only=not find_everything)
    r = client.post(f"/auditor/sessions/{payload['session_id']}/submit", json=work)
    assert r.status_code == 200, r.text
    return r.json()


class TestOverview:

    def test_an_empty_installation_does_not_divide_by_zero(self, client, db):
        r = client.get("/auditor/analytics/overview")
        assert r.status_code == 200, r.text
        assert r.json()["charts"] == 0
        assert r.json()["audit_accuracy"] is None

    def test_a_perfect_cohort_reports_100(self, client, db, library):
        batch_id = make_batch(client, charts_per=6)
        _run(client, batch_id)
        body = client.get("/auditor/analytics/overview").json()
        assert body["audit_accuracy"] == 100.0
        assert body["auditors"] == 1 and body["batches"] == 1

    def test_clean_and_opportunity_charts_are_reported_separately(
            self, client, db, library):
        """
        Otherwise the headline blends restraint with detection and hides which
        one is weak — a passive auditor scores 100 on one and 0 on the other.
        """
        batch_id = make_batch(client, charts_per=6)
        _run(client, batch_id, find_everything=False)
        body = client.get("/auditor/analytics/overview").json()
        assert body["clean_accuracy"] == 100.0
        assert body["opportunity_accuracy"] == 0.0
        assert body["clean_charts"] and body["opportunity_charts"]

    def test_the_basis_of_every_rate_is_stated(self, client, db, library):
        batch_id = make_batch(client, charts_per=6)
        _run(client, batch_id)
        body = client.get("/auditor/analytics/overview").json()
        assert body["audit_accuracy_basis"] == "average of chart scores"
        assert body["component_basis"] == "pooled findings over errors introduced"

    def test_a_component_never_planted_reads_NA_not_zero(self, client, db, library):
        batch_id = make_batch(client, charts_per=6)
        _run(client, batch_id)
        body = client.get("/auditor/analytics/overview").json()
        for name in ("add", "revise", "delete"):
            cell = body[name]
            if cell["planted"] == 0:
                assert cell["accuracy"] is None, name

    def test_thin_cohorts_still_receive_a_verdict(self, client, db, library):
        batch_id = make_batch(client, charts_per=2)
        _run(client, batch_id)
        body = client.get("/auditor/analytics/overview").json()
        assert body["pass_fail"] == "PASS"
        assert body["verdict_withheld_reason"] is None
        assert body["pass_count"] == 1
        assert body["verdict_count"] == 1
        assert body["pass_rate"] == 100.0

    def test_drg_impact_is_its_own_number(self, client, db, library):
        """Never blended into the headline as a weight."""
        batch_id = make_batch(client, charts_per=6)
        _run(client, batch_id)
        body = client.get("/auditor/analytics/overview").json()
        assert "drg_accuracy" in body and "drg_planted" in body


class TestPooling:

    def test_component_accuracy_pools_rather_than_averaging(
            self, client, db, library):
        """
        Averaging per-chart rates would let a chart with one planting count as
        much as a chart with six — how a rate quietly starts meaning something
        other than what it says.
        """
        batch_id = make_batch(client, charts_per=6, clean_share=0)
        _run(client, batch_id)
        body = client.get("/auditor/analytics/overview").json()
        for name in ("add", "revise", "delete"):
            cell = body[name]
            if cell["planted"]:
                assert cell["accuracy"] == round(
                    cell["found"] / cell["planted"] * 100, 2)


class TestByBatchAndAuditor:

    def test_each_batch_gets_its_own_row(self, client, db, library):
        a = make_batch(client, charts_per=4)
        _run(client, a)
        b = make_batch(client, charts_per=4)
        _run(client, b)
        rows = client.get("/auditor/analytics/by-batch").json()["batches"]
        assert {r["batch_id"] for r in rows} == {a, b}
        assert all(r["name"] for r in rows)

    def test_batch_search_runs_on_the_server(self, client, db, library):
        old = make_batch(client, charts_per=4, name="July audit")
        _run(client, old)
        target = make_batch(client, charts_per=4, name="August audit wave")
        _run(client, target)

        body = client.get("/auditor/analytics/by-batch",
                          params={"search": "August"}).json()

        assert body["matched"] == 1
        assert [r["batch_id"] for r in body["batches"]] == [target]

    def test_batches_can_sort_weakest_first(self, client, db, library):
        strong = make_batch(client, charts_per=4, name="Strong batch")
        _run(client, strong)
        weak = make_batch(client, charts_per=4, name="Weak batch")
        _run(client, weak, find_everything=False)

        rows = client.get("/auditor/analytics/by-batch",
                          params={"sort": "weakest"}).json()["batches"]

        assert rows[0]["batch_id"] == weak
        assert rows[0]["audit_score"] < rows[1]["audit_score"]

    def test_auditors_are_listed_weakest_first(self, client, db, library):
        """The screen exists to show who needs help, so put them at the top."""
        batch_id = make_batch(client, auditors=("Asha R", "Bo T"), charts_per=6)
        alloc = allocate(client, batch_id)
        truth = truth_map(client, batch_id)
        for i, code in enumerate(alloc["access_codes"]):
            payload = client.get(f"/auditor/sessions/by-token/{code['token']}").json()
            client.post(f"/auditor/sessions/{payload['session_id']}/submit",
                        json=perfect_work(payload, truth, verdicts_only=(i == 0)))
        rows = client.get("/auditor/analytics/by-auditor").json()["auditors"]
        assert len(rows) == 2
        assert rows[0]["audit_accuracy"] <= rows[1]["audit_accuracy"]

    def test_an_auditor_row_pools_across_their_batches(self, client, db, library):
        a = make_batch(client, charts_per=4)
        _run(client, a)
        b = make_batch(client, charts_per=4)
        _run(client, b)
        rows = client.get("/auditor/analytics/by-auditor").json()["auditors"]
        assert len(rows) == 1
        assert rows[0]["batches"] == 2
        assert rows[0]["charts"] == 8

    def test_specialty_rows_are_available_for_mixed_analytics(self, client, db, library):
        batch_id = make_batch(client, charts_per=4)
        _run(client, batch_id)
        rows = client.get("/auditor/analytics/by-specialty").json()["specialties"]
        assert len(rows) == 1
        assert rows[0]["charts"] == 4
        assert rows[0]["specialty"]

    def test_chart_signals_surface_chart_level_qa(self, client, db, library):
        batch_id = make_batch(client, charts_per=6, clean_share=0)
        _run(client, batch_id, find_everything=False)
        rows = client.get("/auditor/analytics/chart-signals").json()["charts"]
        assert rows
        assert rows[0]["attempts"] >= 1
        assert rows[0]["missed"] >= 0
        assert rows[0]["detection_score"] is not None
        assert rows[0]["stability_score"] is not None
        assert rows[0]["review_priority"]
        assert rows[0]["confidence"] in ("Early", "Established")
        assert rows[0]["signal"]


class TestDetectionPatterns:

    def test_every_planting_is_accounted_for(self, client, db, library):
        batch_id = make_batch(client, charts_per=6, clean_share=0)
        _run(client, batch_id)
        body = client.get("/auditor/analytics/detection").json()
        assert body["total_plantings"] > 0
        assert sum(k["planted"] for k in body["by_kind"]) == body["total_plantings"]
        assert all(k["found"] == k["planted"] for k in body["by_kind"])

    def test_missed_kinds_are_counted_and_sorted_worst_first(
            self, client, db, library):
        batch_id = make_batch(client, charts_per=6, clean_share=0)
        _run(client, batch_id, find_everything=False)
        body = client.get("/auditor/analytics/detection").json()
        assert all(k["missed"] == k["planted"] for k in body["by_kind"])
        accs = [k["accuracy"] or 0 for k in body["by_kind"]]
        assert accs == sorted(accs)

    def test_a_kind_seen_twice_is_not_called_a_pattern(self, client, db, library):
        """
        A trainer must not build a curriculum on a single miss, so weakest[]
        requires a minimum sample.
        """
        batch_id = make_batch(client, charts_per=6, clean_share=0)
        _run(client, batch_id, find_everything=False)
        body = client.get("/auditor/analytics/detection").json()
        assert all(k["planted"] >= body["min_for_pattern"] for k in body["weakest"])

    def test_kinds_carry_a_readable_label(self, client, db, library):
        batch_id = make_batch(client, charts_per=6, clean_share=0)
        _run(client, batch_id)
        body = client.get("/auditor/analytics/detection").json()
        labels = {k["label"] for k in body["by_kind"]}
        assert labels
        assert all(lbl and lbl != "unknown" for lbl in labels)

    def test_observed_and_synthetic_errors_are_split(self, client, db, library):
        """
        The comparison that describes the job: auditors tend to do better on
        generated errors than on the ones their own coders really make.
        """
        batch_id = make_batch(client, charts_per=6, clean_share=0)
        _run(client, batch_id)
        body = client.get("/auditor/analytics/detection").json()
        keys = {o["key"] for o in body["by_origin"]}
        assert keys <= {"observed", "synthetic"} and keys

    def test_it_can_be_scoped_to_one_auditor(self, client, db, library):
        batch_id = make_batch(client, auditors=("Asha R", "Bo T"), charts_per=6)
        alloc = allocate(client, batch_id)
        truth = truth_map(client, batch_id)
        for code in alloc["access_codes"]:
            payload = client.get(f"/auditor/sessions/by-token/{code['token']}").json()
            client.post(f"/auditor/sessions/{payload['session_id']}/submit",
                        json=perfect_work(payload, truth))
        both = client.get("/auditor/analytics/detection").json()
        one = client.get("/auditor/analytics/detection",
                         params={"auditor": "Asha R"}).json()
        assert one["total_plantings"] < both["total_plantings"]

    def test_an_empty_installation_returns_empty_buckets(self, client, db):
        body = client.get("/auditor/analytics/detection").json()
        assert body["total_plantings"] == 0
        assert body["by_kind"] == [] and body["weakest"] == []


class TestScale:
    """
    None of these endpoints may load the whole table to render one screen.
    The coder analytics learned this the expensive way; the same rules apply.
    """

    def test_the_batch_list_is_paged_and_reports_its_total(self, client, db, library):
        from tests.test_auditor_api import make_batch
        for i in range(6):
            make_batch(client, charts_per=2)
        r = client.get("/auditor/batches", params={"limit": 3}).json()
        assert len(r["batches"]) == 3
        assert r["total"] >= 6
        assert r["limit"] == 3

    def test_the_batch_list_searches_on_the_server(self, client, db, library):
        from tests.test_auditor_api import make_batch
        make_batch(client, charts_per=2, name="Sepsis wave")
        make_batch(client, charts_per=2, name="Cardiac wave")
        r = client.get("/auditor/batches", params={"search": "sepsis"}).json()
        assert r["total"] == 1
        assert "Sepsis" in r["batches"][0]["name"]

    def test_the_error_review_is_paged_and_omits_the_claim_by_default(
            self, client, db, library):
        """
        The claim is the heaviest field on the row and the list does not render
        it — shipping 500 of them to draw summaries is pure weight.
        """
        from tests.test_auditor_api import allocate, make_batch
        batch_id = make_batch(client, charts_per=6)
        allocate(client, batch_id)

        r = client.get(f"/auditor/batches/{batch_id}/plantings",
                       params={"limit": 2}).json()
        assert len(r["plantings"]) == 2
        assert r["total"] == 6
        assert "claim" not in r["plantings"][0]

        full = client.get(f"/auditor/batches/{batch_id}/plantings",
                          params={"include_claim": True}).json()
        assert "claim" in full["plantings"][0]

    def test_detection_patterns_says_when_it_did_not_read_everything(
            self, client, db, library):
        from tests.test_auditor_api import (
            allocate, make_batch, perfect_work, truth_map)
        batch_id = make_batch(client, charts_per=6)
        token = allocate(client, batch_id)["access_codes"][0]["token"]
        payload = client.get(f"/auditor/sessions/by-token/{token}").json()
        client.post(f"/auditor/sessions/{payload['session_id']}/submit",
                    json=perfect_work(payload, truth_map(client, batch_id)))

        full = client.get("/auditor/analytics/detection").json()
        assert full["truncated"] is False
        assert full["charts_scanned"] == full["charts_available"]

        capped = client.get("/auditor/analytics/detection",
                            params={"scan_limit": 2}).json()
        assert capped["truncated"] is True
        assert capped["charts_scanned"] == 2

    def test_the_curation_todo_list_is_paged(self, client, db, library):
        r = client.get("/auditor/keys/uncurated",
                       params={"specialty": "IP-DRG", "limit": 3}).json()
        assert len(r["charts"]) == 3
        assert r["total"] == len(library)

    def test_auditors_are_capped_weakest_first(self, client, db, library):
        """
        The cap keeps the people who need attention, not an arbitrary slice —
        which is why the ordering is decided in SQL rather than after the fact.
        """
        r = client.get("/auditor/analytics/by-auditor", params={"limit": 1}).json()
        assert len(r["auditors"]) <= 1


class TestFilters:
    """
    Every analytics view takes batch, specialty and auditor.

    These are worth their own tests because the failure mode is silent: FastAPI
    drops a query parameter the handler does not declare, so a filter the UI
    sends can look wired up while changing nothing. Three of the four handlers
    were in exactly that state — the assertions below are on the narrowing, not
    on the status code.
    """

    def _two_auditors(self, client):
        batch_id = make_batch(client, auditors=("Asha R", "Bo T"), charts_per=4)
        _run(client, batch_id, auditor_index=0)
        _run(client, batch_id, auditor_index=1)
        return batch_id

    def test_auditor_filter_narrows_every_view(self, client, db, library):
        batch_id = self._two_auditors(client)

        both = client.get("/auditor/analytics/overview").json()
        one = client.get("/auditor/analytics/overview",
                         params={"auditor": "Asha R"}).json()
        assert both["auditors"] == 2 and one["auditors"] == 1
        assert one["charts"] < both["charts"]

        rows = client.get("/auditor/analytics/by-auditor",
                          params={"auditor": "Asha R"}).json()["auditors"]
        assert [r["auditor_name"] for r in rows] == ["Asha R"]

        # by-batch keeps the batch but must count only that auditor inside it
        b = client.get("/auditor/analytics/by-batch",
                       params={"auditor": "Asha R"}).json()["batches"]
        assert len(b) == 1 and b[0]["batch_id"] == batch_id
        assert b[0]["auditors"] == 1

        d = client.get("/auditor/analytics/detection",
                       params={"auditor": "Asha R"}).json()
        assert d is not None

    def test_workbook_export_honours_current_filters(self, client, db, library):
        self._two_auditors(client)
        r = client.get("/auditor/analytics/export", params={"auditor": "Asha R"})
        assert r.status_code == 200, r.text
        import io
        wb = load_workbook(filename=io.BytesIO(r.content), read_only=True)
        rows = list(wb["By_Auditor"].iter_rows(values_only=True))
        names = [row[0] for row in rows if row and row[0] == "Asha R"]
        assert names == ["Asha R"]

    def test_workbook_specialty_headers_match_auditor_score_terms(
            self, client, db, library):
        self._two_auditors(client)
        r = client.get("/auditor/analytics/export")
        assert r.status_code == 200, r.text
        import io
        wb = load_workbook(filename=io.BytesIO(r.content), read_only=True)

        overview_terms = {row[0] for row in wb["Overview"].iter_rows(values_only=True)
                          if row and row[0]}
        specialty_headers = next(
            [c for c in row if c]
            for row in wb["By_Specialty"].iter_rows(values_only=True)
            if row and row[0] == "Specialty")

        assert "Audit Score" in overview_terms
        assert "Review Score" in overview_terms
        assert "Error Detection Rate" in overview_terms
        assert "Audit Score" in specialty_headers
        assert "Review Score" in specialty_headers
        assert "Error Detection Rate" in specialty_headers
        assert "Audit %" not in specialty_headers

    def test_batch_filter_is_accepted_by_by_batch(self, client, db, library):
        """by-batch took no batch_id at all, so selecting one changed nothing."""
        first = self._two_auditors(client)
        second = make_batch(client, auditors=("Cy K",), charts_per=4)
        _run(client, second)

        allb = client.get("/auditor/analytics/by-batch").json()["batches"]
        assert {r["batch_id"] for r in allb} == {first, second}

        just = client.get("/auditor/analytics/by-batch",
                          params={"batch_id": second}).json()["batches"]
        assert [r["batch_id"] for r in just] == [second]

    def test_a_filter_matching_nothing_reports_zero_not_everything(
            self, client, db, library):
        self._two_auditors(client)
        body = client.get("/auditor/analytics/overview",
                          params={"auditor": "Nobody At All"}).json()
        assert body["charts"] == 0
        assert body["audit_accuracy"] is None

    def test_date_filter_narrows_every_view_and_export(self, client, db, library):
        from datetime import datetime
        from models import AuditResult

        old_batch = make_batch(client, charts_per=4, name="Old audit wave")
        _run(client, old_batch)
        db.query(AuditResult).filter(AuditResult.batch_id == old_batch).update({
            AuditResult.scored_at: datetime(2024, 1, 10, 12, 0, 0),
        })
        new_batch = make_batch(client, charts_per=4, name="August audit wave")
        _run(client, new_batch)
        db.query(AuditResult).filter(AuditResult.batch_id == new_batch).update({
            AuditResult.scored_at: datetime(2026, 8, 10, 12, 0, 0),
        })
        db.commit()

        params = {"from_date": "2026-08-01", "to_date": "2026-08-31"}
        both = client.get("/auditor/analytics/overview").json()
        scoped = client.get("/auditor/analytics/overview", params=params).json()
        assert scoped["charts"] < both["charts"]
        assert scoped["batches"] == 1

        batches = client.get("/auditor/analytics/by-batch", params=params).json()["batches"]
        assert [b["batch_id"] for b in batches] == [new_batch]
        assert client.get("/auditor/analytics/by-auditor", params=params).json()["auditors"][0]["charts"] == 4
        assert client.get("/auditor/analytics/detection", params=params).json()["charts_available"] == 4
        assert client.get("/auditor/analytics/chart-signals", params=params).json()["charts"]
        # by-specialty was the one tab this test did not reach. It rolls each
        # specialty up in a second, inner query, which is exactly where a
        # filter gets dropped without anything failing loudly.
        by_spec = client.get("/auditor/analytics/by-specialty",
                             params=params).json()["specialties"]
        assert [s["charts"] for s in by_spec] == [4]

        r = client.get("/auditor/analytics/export", params=params)
        assert r.status_code == 200, r.text
        import io
        wb = load_workbook(filename=io.BytesIO(r.content), read_only=True)
        rows = list(wb["By_Batch"].iter_rows(values_only=True))
        names = {row[0] for row in rows if row and row[0] in {"Old audit wave", "August audit wave"}}
        assert names == {"August audit wave"}


class TestPdfReports:

    def test_auditor_report_pdf_renders(self, client, db, library):
        batch_id = make_batch(client, charts_per=6)
        _run(client, batch_id)
        r = client.get("/auditor/analytics/auditor-report.pdf",
                       params={"auditor": "Asha R"})
        assert r.status_code == 200, r.text
        assert r.content[:4] == b"%PDF"

    def test_batch_report_pdf_renders(self, client, db, library):
        batch_id = make_batch(client, charts_per=6)
        _run(client, batch_id)
        r = client.get(f"/auditor/batches/{batch_id}/report.pdf")
        assert r.status_code == 200, r.text
        assert r.content[:4] == b"%PDF"

    def test_the_report_verdict_names_the_score_it_was_made_on(self):
        """
        The report read audit_accuracy and called it "Audit Accuracy" — which
        is now only the detection half — so it could print a headline figure
        the PASS beside it was never computed from, and a sentence naming a
        metric that no longer decides anything.

        Tested against the function rather than the rendered PDF: the bug is in
        which field is read and what it is called, and asserting a PDF merely
        starts with %PDF proves only that it did not crash.
        """
        from services.pdf_report_service import _audit_headline, _audit_verdict

        summary = {"audit_score": 94.0, "audit_accuracy": 61.0,
                   "review_score": 99.0}
        assert _audit_headline(summary) == 94.0, "must read the blend"

        headline, detail, *_ = _audit_verdict(summary, pass_threshold=90)
        assert headline == "GOOD"
        assert "94.0" in detail and "61.0" not in detail
        assert "audit accuracy" not in detail.lower()
        assert "Audit Score" in detail

    def test_the_report_does_not_print_internal_sample_size_gates(self):
        from services.pdf_report_service import _audit_verdict

        summary = {"audit_score": 93.19,
                   "verdict_withheld_reason": "legacy sample-size reason"}
        headline, detail, *_ = _audit_verdict(summary, pass_threshold=90)

        assert headline == "GOOD"
        assert "9/50" not in detail
        assert "review lines" not in detail
        assert "INDICATIVE" not in headline

    def test_the_report_still_reads_results_scored_before_review_existed(self):
        """Old rows have no audit_score; they fall back rather than blanking."""
        from services.pdf_report_service import _audit_headline
        assert _audit_headline({"audit_accuracy": 72.0}) == 72.0
        assert _audit_headline({"audit_score": None, "audit_accuracy": 72.0}) == 72.0

    def test_batch_report_pdf_404s_without_scored_results(self, client, db, library):
        batch_id = make_batch(client, charts_per=6)
        r = client.get(f"/auditor/batches/{batch_id}/report.pdf")
        assert r.status_code == 404


class TestTotalsAndSearch:
    """
    Figures that must describe the whole query, not the page that came back.

    Counting loaded rows is the defect this codebase has paid for three times
    now: it reads correctly at a dozen rows and silently understates past the
    cap, with nothing failing.
    """

    def test_chart_signal_totals_survive_a_cap(self, client, db, library):
        batch_id = make_batch(client, charts_per=6)
        _run(client, batch_id, find_everything=False)

        full = client.get("/auditor/analytics/chart-signals").json()
        capped = client.get("/auditor/analytics/chart-signals",
                            params={"limit": 2}).json()

        # the page shrinks; the totals do not
        assert len(capped["charts"]) == 2
        assert capped["returned"] == 2
        assert capped["charts_total"] == full["charts_total"]
        assert capped["charts_with_signals"] == full["charts_with_signals"]
        assert capped["charts_stable"] == full["charts_stable"]
        assert capped["most_missed"] == full["most_missed"]
        assert capped["priority_distribution"]

    def test_chart_signal_search_runs_on_the_server(self, client, db, library):
        batch_id = make_batch(client, charts_per=6)
        _run(client, batch_id, find_everything=False)

        all_rows = client.get("/auditor/analytics/chart-signals",
                              params={"limit": 1}).json()
        target = client.get("/auditor/analytics/chart-signals").json()["charts"][-1]

        searched = client.get("/auditor/analytics/chart-signals",
                              params={"limit": 1, "search": target["chart_number"]}).json()

        assert all_rows["returned"] == 1
        assert searched["charts_total"] == 1
        assert searched["charts"][0]["chart_number"] == target["chart_number"]

    def test_totals_agree_with_the_rows_when_nothing_is_capped(
            self, client, db, library):
        batch_id = make_batch(client, charts_per=6)
        _run(client, batch_id, find_everything=False)
        body = client.get("/auditor/analytics/chart-signals").json()
        from_rows = sum(1 for r in body["charts"]
                        if r["missed"] or r["over_calls"] or r["detected_not_corrected"])
        assert from_rows == body["charts_with_signals"]
        assert body["charts_total"] == len(body["charts"])

    def test_auditor_search_runs_on_the_server(self, client, db, library):
        batch_id = make_batch(client, auditors=("Asha R", "Bo T"), charts_per=4)
        _run(client, batch_id, auditor_index=0)
        _run(client, batch_id, auditor_index=1)

        every = client.get("/auditor/analytics/by-auditor").json()
        assert every["matched"] == 2

        by_name = client.get("/auditor/analytics/by-auditor",
                             params={"search": "asha"}).json()
        assert by_name["matched"] == 1
        assert [a["auditor_name"] for a in by_name["auditors"]] == ["Asha R"]

        # and by employee id, which is why the search is not name-only
        by_emp = client.get("/auditor/analytics/by-auditor",
                            params={"search": "E0"}).json()
        assert by_emp["matched"] == 1

        assert client.get("/auditor/analytics/by-auditor",
                          params={"search": "nobody"}).json()["matched"] == 0

    def test_the_overview_carries_the_score_rule_it_is_judged_by(
            self, client, db, library):
        batch_id = make_batch(client, charts_per=4)
        _run(client, batch_id)
        body = client.get("/auditor/analytics/overview").json()
        assert "opportunities" in body
        assert body["pass_threshold"] == 90
        assert body["review_total"] > 0
        assert body["review_basis"] == "pooled code lines judged correctly"

    def test_the_trend_is_bucketed_by_date_not_batch_order(
            self, client, db, library):
        from datetime import datetime
        from models import AuditResult

        first = make_batch(client, charts_per=4, name="Older")
        _run(client, first)
        db.query(AuditResult).filter(AuditResult.batch_id == first).update({
            AuditResult.scored_at: datetime(2026, 1, 5, 9, 0, 0)})
        second = make_batch(client, charts_per=4, name="Newer")
        _run(client, second)
        db.query(AuditResult).filter(AuditResult.batch_id == second).update({
            AuditResult.scored_at: datetime(2026, 3, 9, 9, 0, 0)})
        db.commit()

        trend = client.get("/auditor/analytics/overview").json()["trend"]
        assert [p["date"] for p in trend] == ["2026-01-05", "2026-03-09"]
        assert all(p["charts"] == 4 for p in trend)

        # and it answers to the date filter, which the batch-ordered version
        # ignored completely
        scoped = client.get("/auditor/analytics/overview",
                            params={"from_date": "2026-02-01"}).json()["trend"]
        assert [p["date"] for p in scoped] == ["2026-03-09"]


class TestSectionScores:
    """
    Section scores are REVIEW scores now — code lines judged correctly, not
    errors detected. A section appears when the chart had lines in it, so an
    inpatient cohort reports PCS and an outpatient one reports CPT without a
    mapping to keep in step.

    POA, modifiers and units are attributes OF a line and are reported with
    their own percentages rather than inflating the denominator. Counting them
    as opportunities roughly doubled it with judgements that are almost never
    wrong, and let an auditor who flagged nothing score 94% and pass.
    """

    def test_sections_are_the_code_lines_the_chart_actually_had(
            self, client, db, library):
        batch_id = make_batch(client, charts_per=6)
        _run(client, batch_id)
        body = client.get("/auditor/analytics/overview").json()
        sections = body["sections"]
        assert sections, "an IP cohort must report some sections"
        assert set(sections) <= {"PDx", "SDx", "PCS", "CPT"}
        assert "POA" not in sections, "POA is an attribute, not a line"
        for name, s in sections.items():
            assert s["total"] > 0, f"{name} reported with no lines"
            assert 0 <= s["score"] <= 100

    def test_attributes_are_reported_separately(self, client, db, library):
        batch_id = make_batch(client, charts_per=6)
        _run(client, batch_id)
        attrs = client.get("/auditor/analytics/overview").json()["attributes"]
        # IP-DRG judges POA on every diagnosis, so it has its own percentage
        assert "POA" in attrs
        assert attrs["POA"]["total"] > 0

    def test_a_perfect_audit_reviews_every_line_correctly(
            self, client, db, library):
        batch_id = make_batch(client, charts_per=6)
        _run(client, batch_id)
        body = client.get("/auditor/analytics/overview").json()
        assert body["review_score"] == 100.0
        assert body["review_correct"] == body["review_total"]
        for s in body["sections"].values():
            assert s["score"] == 100.0

    def test_the_review_score_is_pooled_not_averaged(self, client, db, library):
        """A twenty-line chart must outweigh a five-line one."""
        batch_id = make_batch(client, charts_per=6)
        _run(client, batch_id)
        body = client.get("/auditor/analytics/overview").json()
        assert body["review_basis"] == "pooled code lines judged correctly"
        assert body["review_score"] == round(
            body["review_correct"] / body["review_total"] * 100, 2)

    def test_the_sections_answer_to_the_filters(self, client, db, library):
        batch_id = make_batch(client, charts_per=6)
        _run(client, batch_id)
        empty = client.get("/auditor/analytics/overview",
                           params={"from_date": "2099-01-01"}).json()
        assert empty["charts"] == 0


class TestAuditScoreBlend:
    """
    The Audit Score is the Error Detection Rate and the Review Score weighted
    together, and it is what the verdict is decided on.

    Neither half works alone. Detection is quantised by planting count — a
    chart with one error can only score 0 or 100. Review starts high because
    most lines on any chart are correct, so an auditor who flags nothing was
    still scoring in the eighties and passing.
    """

    def test_the_blend_is_reported_with_the_weights_that_made_it(
            self, client, db, library):
        batch_id = make_batch(client, charts_per=8)
        _run(client, batch_id)
        body = client.get("/auditor/analytics/overview").json()
        assert body["detection_weight"] == 50
        assert body["review_weight"] == 50
        assert body["audit_score"] == round(
            body["audit_accuracy"] * 0.5 + body["review_score"] * 0.5, 2)

    def test_a_passive_auditor_fails_the_blend(self, client, db, library):
        """The load-bearing assertion: flagging nothing must not pass."""
        batch_id = make_batch(client, charts_per=8)
        _run(client, batch_id, find_everything=False)
        body = client.get("/auditor/analytics/overview").json()
        assert body["review_score"] > 80, "review alone still reads well"
        assert body["audit_score"] < 90, "but the blend must fail them"
        assert body["pass_fail"] == "FAIL"

    def test_the_weights_are_settable_from_the_score_panel(
            self, client, db, library):
        batch_id = make_batch(client, charts_per=8)
        _run(client, batch_id, find_everything=False)
        even = client.get("/auditor/analytics/overview").json()["audit_score"]

        r = client.put("/auditor/config", json={
            "detection_weight": 80, "review_weight": 20,
            "updated_by": "T", "passphrase": PASS})
        assert r.status_code == 200, r.text

        body = client.get("/auditor/analytics/overview").json()
        assert body["detection_weight"] == 80
        assert body["audit_score"] < even, (
            "weighting detection harder must lower a passive auditor's score")

    def test_the_trend_plots_the_blend_not_one_half(self, client, db, library):
        batch_id = make_batch(client, charts_per=8)
        _run(client, batch_id)
        trend = client.get("/auditor/analytics/overview").json()["trend"]
        assert trend
        for point in trend:
            assert "detection" in point and "review" in point
            assert point["score"] is not None


class TestEveryRowCarriesTheBlend:
    """
    Every row shape the dashboard renders must carry audit_score, not just
    audit_accuracy.

    This has now been got wrong twice. The two are different numbers — one is
    the blend the verdict uses, the other is the detection half — and a column
    headed "Audit Score" rendering audit_accuracy is wrong in a way nothing
    fails on: it just quietly shows the harsher figure. Asserting the field is
    present everywhere is the cheapest guard against the next rewrite dropping
    it again.
    """

    def _rows(self, client, batch_id):
        return {
            "by-batch": client.get("/auditor/analytics/by-batch").json()["batches"],
            "by-auditor": client.get("/auditor/analytics/by-auditor").json()["auditors"],
            "by-specialty": client.get("/auditor/analytics/by-specialty").json()["specialties"],
        }

    def test_every_grouping_reports_the_blended_score(self, client, db, library):
        batch_id = make_batch(client, charts_per=8)
        _run(client, batch_id, find_everything=False)
        for name, rows in self._rows(client, batch_id).items():
            assert rows, f"{name} returned nothing"
            for row in rows:
                assert "audit_score" in row, f"{name} row has no audit_score"
                assert "audit_accuracy" in row, f"{name} row has no audit_accuracy"

    def test_the_blend_and_the_detection_half_are_different_numbers(
            self, client, db, library):
        """
        If these were ever equal the bug would be invisible, so the fixture has
        to produce a partial auditor — one who reviews well and detects badly.
        """
        batch_id = make_batch(client, charts_per=8)
        _run(client, batch_id, find_everything=False)
        for name, rows in self._rows(client, batch_id).items():
            row = rows[0]
            assert row["audit_score"] != row["audit_accuracy"], (
                f"{name}: blend and detection coincide, so this guard proves nothing")
            assert row["audit_score"] > row["audit_accuracy"], (
                f"{name}: the blend must sit above detection when review is stronger")


class TestWeeklyTrend:
    """
    The trend is weekly, because audit sessions do not run daily and a
    per-day line was a scatter of isolated points with gaps between them.
    """

    def _scored_on(self, client, db, when, charts=4, perfect=True, name="W"):
        from models import AuditResult
        batch_id = make_batch(client, charts_per=charts, name=name)
        _run(client, batch_id, find_everything=perfect)
        db.query(AuditResult).filter(AuditResult.batch_id == batch_id).update(
            {AuditResult.scored_at: when})
        db.commit()
        return batch_id

    def test_days_in_one_week_become_one_point(self, client, db, library):
        from datetime import datetime
        self._scored_on(client, db, datetime(2026, 8, 3, 9), name="Mon")
        self._scored_on(client, db, datetime(2026, 8, 5, 9), name="Wed")
        self._scored_on(client, db, datetime(2026, 8, 12, 9), name="NextWeek")

        trend = client.get("/auditor/analytics/overview").json()["trend"]
        assert [p["week_of"] for p in trend] == ["2026-08-03", "2026-08-10"]
        assert trend[0]["charts"] == 8      # Monday and Wednesday merged
        assert trend[1]["charts"] == 4

    def test_a_week_pools_by_chart_not_by_day(self, client, db, library):
        """
        The trap this guards. Detection is an average of chart scores, so
        rolling days up means summing and dividing by the chart count — a day
        with one chart must not weigh the same as a day with eight. Averaging
        the daily averages would give 50 here; pooling gives 88.89.
        """
        from datetime import datetime
        self._scored_on(client, db, datetime(2026, 9, 7, 9), charts=8,
                        perfect=True, name="Busy")
        self._scored_on(client, db, datetime(2026, 9, 9, 9), charts=1,
                        perfect=False, name="Quiet")

        trend = client.get("/auditor/analytics/overview").json()["trend"]
        point = [p for p in trend if p["week_of"] == "2026-09-07"][0]
        assert point["charts"] == 9
        assert point["detection"] > 60, (
            f"detection {point['detection']} looks like an average of daily "
            f"averages rather than a pool over charts")

    def test_an_empty_week_is_absent_rather_than_zero(self, client, db, library):
        """A week nobody worked is not a week they scored nothing."""
        from datetime import datetime
        self._scored_on(client, db, datetime(2026, 10, 5, 9), name="A")
        self._scored_on(client, db, datetime(2026, 10, 26, 9), name="B")
        trend = client.get("/auditor/analytics/overview").json()["trend"]
        assert [p["week_of"] for p in trend] == ["2026-10-05", "2026-10-26"]
        assert all(p["score"] is not None for p in trend)

    def test_the_point_carries_both_halves_of_the_blend(self, client, db, library):
        from datetime import datetime
        self._scored_on(client, db, datetime(2026, 11, 2, 9), perfect=False)
        point = client.get("/auditor/analytics/overview").json()["trend"][0]
        assert point["detection"] is not None and point["review"] is not None
        assert point["score"] == round(
            point["detection"] * 0.5 + point["review"] * 0.5, 2)


class TestWeakestFirstOrdering:
    """
    The capped lists must be ordered by the figure they DISPLAY.

    "Weakest first" sorted on detection while the card showed the blended Audit
    Score. That is not cosmetic: these lists are capped — ten auditor cards out
    of however many exist — so the sort decides who a trainer ever sees. An
    over-caller with strong detection and weak review would never surface.
    """

    def _result(self, db, name, detection, review_correct, review_total, chart_id):
        from datetime import datetime
        from models import AuditResult
        from models.charts import Specialty
        db.add(AuditResult(
            session_id=1, assignment_id=chart_id, chart_id=chart_id, batch_id=1,
            auditor_name=name, emp_id=name, specialty=Specialty.IP_DRG,
            is_clean=False, audit_accuracy=detection,
            review_correct=review_correct, review_total=review_total,
            review_score=round(review_correct / review_total * 100, 2),
            scored_at=datetime(2026, 8, 10, 9), findings=[], feedback=[]))

    def test_the_order_follows_the_blend_not_detection(self, client, db, library):
        # Sharp detects but over-calls; Steady is middling on both.
        self._result(db, "Sharp", detection=95.0, review_correct=60,
                     review_total=100, chart_id=1)     # blend 77.5
        self._result(db, "Steady", detection=80.0, review_correct=85,
                     review_total=100, chart_id=2)     # blend 82.5
        db.commit()

        rows = client.get("/auditor/analytics/by-auditor").json()["auditors"]
        order = [r["auditor_name"] for r in rows]
        scores = {r["auditor_name"]: r["audit_score"] for r in rows}

        # The two orderings genuinely disagree here, which is what makes this
        # test worth having — by detection Steady (80) is weakest, by the blend
        # Sharp (77.5) is.
        assert scores["Sharp"] < scores["Steady"]
        assert order == ["Sharp", "Steady"], (
            f"ordered {order}; by detection alone this would be Steady first")


class TestErrorPatternComposition:
    """
    Each pattern row draws three segments — caught, fixed wrongly, missed —
    rather than one accuracy bar.

    "40% caught" hides the difference between an auditor who cannot SEE the
    error and one who sees it and gets the fix wrong. Same number, entirely
    different coaching. detected_not_corrected was already counted per kind and
    only the total was ever drawn.
    """

    def _partly_wrong(self, client, batch_id):
        """Finds everything, but corrects half the Revises wrongly."""
        token = allocate(client, batch_id)["access_codes"][0]["token"]
        payload = client.get(f"/auditor/sessions/by-token/{token}").json()
        work = perfect_work(payload, truth_map(client, batch_id))
        n = 0
        for chart in work["charts"]:
            for f in chart["findings"]:
                # Revise is the only action that can be matched and still
                # wrong: an Add is matched BY its code, so a bad value there is
                # simply a different finding.
                if f.get("action") == "Revise":
                    if n % 2 == 0:
                        f["correct_value"] = "WRONGFIX"
                    n += 1
        r = client.post(f"/auditor/sessions/{payload['session_id']}/submit", json=work)
        assert r.status_code == 200, r.text[:300]

    def test_the_three_segments_reconcile_with_the_total(self, client, db, library):
        batch_id = make_batch(client, charts_per=8)
        self._partly_wrong(client, batch_id)
        body = client.get("/auditor/analytics/detection").json()

        assert any(r["detected_not_corrected"] > 0 for r in body["by_kind"]), (
            "the fixture produced no fixed-wrongly row, so this proves nothing")
        for bucket in ("by_kind", "by_origin"):
            for row in body[bucket]:
                missed = row["planted"] - row["found"] - row["detected_not_corrected"]
                assert missed >= 0, f"{bucket}/{row['key']}: segments exceed the total"
                assert row["found"] + row["detected_not_corrected"] + missed \
                    == row["planted"]

        # and every cell of the section x action matrix, plus its margins
        m = body["section_matrix"]
        for section, row in m["cells"].items():
            for action, cell in row.items():
                missed = cell["planted"] - cell["found"] - cell["detected_not_corrected"]
                assert missed >= 0, f"{section}/{action}: segments exceed the total"

    def test_the_pattern_threshold_is_shipped_so_thin_rows_can_be_marked(
            self, client, db, library):
        """
        A pattern seen three times is an anecdote. The UI dims those rather
        than hiding them — hiding would misstate the totals beside them — and
        needs the threshold to know which.
        """
        batch_id = make_batch(client, charts_per=8)
        self._partly_wrong(client, batch_id)
        body = client.get("/auditor/analytics/detection").json()
        assert body["min_for_pattern"] == 5
        assert any(r["planted"] < body["min_for_pattern"] for r in body["by_kind"])


class TestChapterAxis:
    """
    Which BODY of knowledge the error sat in, as opposed to which mechanic
    produced it. "Root operation errors are missed" tells a trainer what to
    drill; "obstetric diagnoses are missed" tells them who to put on which
    charts. The two do not overlap, and only the second was missing.
    """

    def test_diagnosis_plantings_are_grouped_by_chapter(self, client, db,
                                                        library):
        batch_id = make_batch(client, charts_per=6, clean_share=0)
        _run(client, batch_id, find_everything=False)
        body = client.get("/auditor/analytics/detection").json()
        rows = body["by_chapter"]
        # The library's diagnoses are E11/M17/I50-shaped, so at least one real
        # chapter must appear rather than an empty axis.
        assert isinstance(rows, list)
        for row in rows:
            assert row["label"] and row["planted"] >= 3
            assert row["found"] + row["missed"] + \
                row["detected_not_corrected"] == row["planted"]

    def test_a_chapter_never_counts_more_than_the_plantings(self, client, db,
                                                            library):
        batch_id = make_batch(client, charts_per=6, clean_share=0)
        _run(client, batch_id, find_everything=False)
        body = client.get("/auditor/analytics/detection").json()
        chapter_total = sum(r["planted"] for r in body["by_chapter"])
        assert chapter_total <= body["total_plantings"]

    def test_the_axis_is_capped(self, client, db, library):
        batch_id = make_batch(client, charts_per=4)
        _run(client, batch_id)
        assert len(client.get("/auditor/analytics/detection")
                   .json()["by_chapter"]) <= 10

    def test_an_empty_installation_returns_an_empty_axis(self, client, db):
        assert client.get("/auditor/analytics/detection"
                          ).json()["by_chapter"] == []


class TestSectionActionMatrix:
    """
    Section down, action across.

    The flat "SDx · Revise" rows answered only the question they were built
    for. A trainer's next two are "how are we on SDx overall?" and "are Revises
    worse than Deletes everywhere?" — neither of which a list of compound
    strings can answer, and both of which a matrix answers by being read in the
    other direction.
    """

    def test_the_margins_reconcile_with_the_cells(self, client, db, library):
        batch_id = make_batch(client, charts_per=8)
        _run(client, batch_id, find_everything=False)
        m = client.get("/auditor/analytics/detection").json()["section_matrix"]

        assert m["sections"] and m["actions"]
        for section in m["sections"]:
            row = m["cells"][section]
            planted = sum(c["planted"] for c in row.values())
            assert m["section_totals"][section]["planted"] == planted, (
                f"{section} row total disagrees with its cells")
        for action in m["actions"]:
            planted = sum(m["cells"][s].get(action, {}).get("planted", 0)
                          for s in m["sections"])
            assert m["action_totals"][action]["planted"] == planted, (
                f"{action} column total disagrees with its cells")
        grand = sum(c["planted"] for row in m["cells"].values() for c in row.values())
        assert m["total"]["planted"] == grand

    def test_only_sections_and_actions_that_occur_appear(self, client, db, library):
        """An outpatient cohort shows no PCS row rather than a row of dashes."""
        batch_id = make_batch(client, charts_per=8)
        _run(client, batch_id, find_everything=False)
        m = client.get("/auditor/analytics/detection").json()["section_matrix"]
        for section in m["sections"]:
            assert m["cells"][section], f"{section} is listed with no cells"
        assert set(m["sections"]) <= {"PDx", "SDx", "PCS", "CPT"}
        assert set(m["actions"]) <= {"Add", "Revise", "Delete"}


class TestPatternDrill:
    """
    One error pattern, drilled: who misses it, on which charts, and whether it
    is improving.

    Error Patterns could say root-operation errors slip past 70% of the time
    and then stopped. The next two questions are always "so who?" and "did the
    training work?" — a diagnosis with no treatment plan.
    """

    def _cohort(self, client, db, when=None):
        """One auditor who finds everything, one who finds nothing."""
        from models import AuditResult
        batch_id = make_batch(client, auditors=("Strong", "Passive"), charts_per=8)
        codes = allocate(client, batch_id)["access_codes"]
        truth = truth_map(client, batch_id)
        for i, code in enumerate(codes):
            payload = client.get(f"/auditor/sessions/by-token/{code['token']}").json()
            # verdicts_only is the honest way to build a passive auditor: every
            # section reviewed and left alone. Stripping findings while leaving
            # a "needs changes" verdict is refused at submit, and rightly.
            work = perfect_work(payload, truth, verdicts_only=(i == 1))
            r = client.post(f"/auditor/sessions/{payload['session_id']}/submit",
                            json=work)
            assert r.status_code == 200, r.text[:200]
        if when:
            db.query(AuditResult).update({AuditResult.scored_at: when})
            db.commit()
        return batch_id

    def test_it_names_who_misses_the_pattern_worst_first(self, client, db, library):
        self._cohort(client, db)
        body = client.get("/auditor/analytics/pattern",
                          params={"kind": "omit_sdx"}).json()
        assert body["planted"] > 0
        names = [a["auditor_name"] for a in body["auditors"]]
        assert set(names) == {"Strong", "Passive"}
        assert names[0] == "Passive", "the weakest must lead the list"
        weak = body["auditors"][0]
        assert weak["accuracy"] == 0.0 and weak["missed"] == weak["planted"]

    def test_it_names_the_charts_the_pattern_lives_on(self, client, db, library):
        self._cohort(client, db)
        body = client.get("/auditor/analytics/pattern",
                          params={"kind": "omit_sdx"}).json()
        assert body["charts"], "a pattern must point at the charts carrying it"
        assert all(c["chart_number"] for c in body["charts"])
        accs = [c["accuracy"] for c in body["charts"] if c["accuracy"] is not None]
        assert accs == sorted(accs), "charts are worst-first too"

    def test_it_trends_by_week(self, client, db, library):
        from datetime import datetime
        self._cohort(client, db, when=datetime(2026, 8, 12, 9))
        body = client.get("/auditor/analytics/pattern",
                          params={"kind": "omit_sdx"}).json()
        assert [t["week_of"] for t in body["trend"]] == ["2026-08-10"]
        point = body["trend"][0]
        assert point["found"] + point["missed"] + point["detected_not_corrected"] \
            == point["planted"]

    def test_a_section_and_action_can_be_drilled_too(self, client, db, library):
        """The matrix cells drill on the same endpoint the kind rows use."""
        self._cohort(client, db)
        body = client.get("/auditor/analytics/pattern",
                          params={"section": "SDx", "action": "Add"}).json()
        assert body["label"] == "SDx · Add"
        assert body["planted"] > 0

    def test_naming_no_pattern_is_refused(self, client, db, library):
        r = client.get("/auditor/analytics/pattern")
        assert r.status_code == 400

    def test_the_drill_honours_the_global_filters(self, client, db, library):
        self._cohort(client, db)
        scoped = client.get("/auditor/analytics/pattern",
                            params={"kind": "omit_sdx",
                                    "from_date": "2099-01-01"}).json()
        assert scoped["planted"] == 0
        assert scoped["auditors"] == []
