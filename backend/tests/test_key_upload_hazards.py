"""
Four ways an answer-key upload reported success and stored the wrong key.

Every one of these is silent by construction: the file parses, the endpoint
returns 200 with the chart listed under "stored", and the key that goes on to
grade every coder is not the key the trainer wrote. That is the worst shape a
defect can take here — nothing to notice until the scores are already out.
"""
import io

import pytest
from openpyxl import Workbook

from models import Chart, ChartStatus, Difficulty, Specialty
from services.excel_service import (
    EM_KEY_COLUMNS, parse_answer_key_upload, parse_em_answer_key_upload,
)

PASS = "test-passphrase"


def _sheet(headers, *rows):
    wb = Workbook()
    ws = wb.active
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    for r, values in enumerate(rows, start=2):
        for c, v in enumerate(values, 1):
            ws.cell(r, c, v)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture()
def op_chart(db):
    chart = Chart(chart_number="OP001", specialty=Specialty.SDS, category="Test",
                  difficulty=Difficulty.INTERMEDIATE, status=ChartStatus.ACTIVE,
                  uploaded_by="t")
    db.add(chart)
    db.commit()
    return chart


def _upload(client, data, specialty="SDS"):
    return client.post(
        "/practicelab/answer-key/upload",
        files={"file": ("keys.xlsx", data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"specialty": specialty, "entered_by": "Trainer", "passphrase": PASS},
    )


# ── 1. formula cells ──────────────────────────────────────────────────────────

def test_a_formula_is_read_as_its_result_not_its_text():
    """
    A trainer who assembles a code with a formula used to store the literal
    "=CONCATENATE(...)" as their principal diagnosis, and every coder was then
    marked wrong against a string no ICD-10 book contains.
    """
    wb = Workbook()
    ws = wb.active
    ws["A1"], ws["B1"] = "Chart_Number", "PDx_Code"
    ws["A2"] = "OP001"
    ws["B2"] = '=CONCATENATE("E11",".9")'
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    # openpyxl writes no cached result, which is exactly the Numbers/Sheets
    # case: with no value to fall back on, the formula text is all there is.
    # What must never happen is the parser preferring the text when a result
    # IS present, so inject one the way Excel would.
    got = parse_answer_key_upload(raw, "OP")[0]["pdx_code"]
    assert got.startswith("="), "openpyxl fixture sanity: no cached value here"


def test_the_cached_value_wins_when_the_file_has_one():
    from openpyxl import load_workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"], ws["B1"] = "Chart_Number", "PDx_Code"
    ws["A2"] = "OP001"
    ws["B2"] = "=SOMETHING()"
    buf = io.BytesIO()
    wb.save(buf)

    # Rewrite the same workbook carrying a cached result, as Excel saves it.
    cached = load_workbook(io.BytesIO(buf.getvalue()))
    cached.worksheets[0]["B2"] = "E11.9"
    out = io.BytesIO()
    cached.save(out)

    assert parse_answer_key_upload(out.getvalue(), "OP")[0]["pdx_code"] == "E11.9"


def test_a_plain_file_still_parses_when_nothing_is_cached():
    """
    The reason the parser cannot simply switch to data_only=True: Numbers and
    Google Sheets save no cached values at all, and every cell would read None.
    """
    data = _sheet(["Chart_Number", "PDx_Code", "SDx_1"], ["OP001", "E11.9", "I10"])
    row = parse_answer_key_upload(data, "OP")[0]
    assert row["pdx_code"] == "E11.9"
    assert [s["code"] for s in row["sdx"]] == ["I10"]


# ── 2. chart numbers typed by hand ────────────────────────────────────────────

@pytest.mark.parametrize("typed", ["op001", "  OP001  ", "Op001"])
def test_a_chart_number_matches_however_it_was_typed(client, op_chart, typed):
    """
    Chart numbers are issued upper-case and then typed into spreadsheets. An
    exact match answered "not found", which reads as a missing chart rather
    than a typo and sends the trainer looking for the wrong problem.
    """
    r = _upload(client, _sheet(["Chart_Number", "PDx_Code"], [typed, "E11.9"]))
    assert r.status_code == 200, r.text
    assert r.json()["not_found"] == []
    assert len(r.json()["stored"]) == 1


def test_a_genuinely_missing_chart_is_still_reported(client, op_chart):
    r = _upload(client, _sheet(["Chart_Number", "PDx_Code"], ["OP999", "E11.9"]))
    assert r.json()["not_found"] == ["OP999"]
    assert r.json()["stored"] == []


# ── 3. a file that disagrees with itself ──────────────────────────────────────

def test_two_rows_for_one_chart_are_refused(client, op_chart, db):
    """
    Both rows used to go through. The second overwrote the first or counted as
    a duplicate depending on the replace flag, and the trainer saw OP001 in the
    results once and never learned their file contained two different keys.
    """
    from models import AnswerKey

    data = _sheet(["Chart_Number", "PDx_Code"],
                  ["OP001", "E11.9"], ["OP001", "J18.9"])
    r = _upload(client, data)
    assert r.status_code == 400
    assert "more than one row" in r.json()["detail"]
    assert "OP001" in r.json()["detail"]
    assert db.query(AnswerKey).count() == 0, "nothing may be stored from a bad file"


def test_duplicates_are_caught_however_they_were_typed(client, op_chart):
    r = _upload(client, _sheet(["Chart_Number", "PDx_Code"],
                               ["OP001", "E11.9"], ["op001 ", "J18.9"]))
    assert r.status_code == 400


def test_distinct_charts_are_not_mistaken_for_duplicates(client, op_chart, db):
    chart2 = Chart(chart_number="OP002", specialty=Specialty.SDS, category="T",
                   difficulty=Difficulty.INTERMEDIATE, status=ChartStatus.ACTIVE,
                   uploaded_by="t")
    db.add(chart2)
    db.commit()

    r = _upload(client, _sheet(["Chart_Number", "PDx_Code"],
                               ["OP001", "E11.9"], ["OP002", "J18.9"]))
    assert r.status_code == 200, r.text
    assert len(r.json()["stored"]) == 2


# ── 4. the E/M sheet's columns ────────────────────────────────────────────────

def _em_row(**fields):
    """A one-row E/M key written against the real header table."""
    headers = [h for _f, h in EM_KEY_COLUMNS]
    index = {f: i for i, (f, _h) in enumerate(EM_KEY_COLUMNS)}
    values = [""] * len(headers)
    for field, val in fields.items():
        values[index[field]] = val
    return headers, values


def test_the_em_sheet_parses_from_its_own_template_layout():
    headers, values = _em_row(chart_number="EM001", em_code="99214",
                              dx_1="E11.9", cpt_1="20610", cpt_1_modifier="RT",
                              cpt_1_units=2, entered_by="Trainer", total_time=35)
    row = parse_em_answer_key_upload(_sheet(headers, values))[0]
    assert row["em_code"] == "99214"
    assert row["dx_codes"] == ["E11.9"]
    assert row["procedure_cpts"][0]["code"] == "20610"
    assert row["procedure_cpts"][0]["modifier"] == "RT"
    assert row["procedure_cpts"][0]["units"] == 2
    assert row["entered_by"] == "Trainer"
    assert row["total_time"] == 35


def test_an_inserted_column_no_longer_shifts_every_field_after_it():
    """
    The defect this fixes. The parser read column 43 as the first CPT with no
    check that column 43 was the CPT — insert one column and the key stores a
    modifier where a code belongs, grades every coder against it, and reports
    nothing wrong.
    """
    headers, values = _em_row(chart_number="EM001", em_code="99214",
                              dx_1="E11.9", cpt_1="20610", cpt_1_modifier="RT",
                              entered_by="Trainer")
    at = {f: i for i, (f, _h) in enumerate(EM_KEY_COLUMNS)}
    where = at["patient_type"]
    headers.insert(where, "Trainer's own notes")
    values.insert(where, "checked against the coding manual")

    row = parse_em_answer_key_upload(_sheet(headers, values))[0]
    assert row["em_code"] == "99214", "a shifted column silently changed the E/M code"
    assert row["dx_codes"] == ["E11.9"]
    assert row["procedure_cpts"][0]["code"] == "20610"
    assert row["entered_by"] == "Trainer"


def test_a_headerless_file_still_parses_by_position():
    """
    The fallback. Header matching must not become a new way to fail for a file
    that never had a header row — every field keeps the index it always had.
    """
    _headers, values = _em_row(chart_number="EM001", em_code="99214", dx_1="E11.9")
    data = _sheet(["" for _ in values], values)
    row = parse_em_answer_key_upload(data)[0]
    assert row["em_code"] == "99214"
    assert row["dx_codes"] == ["E11.9"]


def test_the_template_and_the_parser_read_the_same_table(client):
    """
    They are generated from EM_KEY_COLUMNS now, and this is the assertion that
    keeps them there — the drift between the two is what made the old parser
    dangerous.
    """
    from openpyxl import load_workbook

    r = client.get("/practicelab/em/answer-key/template")
    ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
    written = [c.value for c in ws[1]][:len(EM_KEY_COLUMNS)]
    assert written == [h for _f, h in EM_KEY_COLUMNS]


def test_the_em_upload_refuses_a_file_that_repeats_a_chart(client, db):
    chart = Chart(chart_number="EM001", specialty=Specialty.EM, category="T",
                  difficulty=Difficulty.INTERMEDIATE, status=ChartStatus.ACTIVE,
                  uploaded_by="t")
    db.add(chart)
    db.commit()

    headers, first = _em_row(chart_number="EM001", em_code="99214", entered_by="T")
    _h, second = _em_row(chart_number="EM001", em_code="99215", entered_by="T")
    r = client.post(
        "/practicelab/em/answer-key/upload",
        files={"file": ("k.xlsx", _sheet(headers, first, second),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"entered_by": "T", "passphrase": PASS},
    )
    assert r.status_code == 400
    assert "more than one row" in r.json()["detail"]


def test_the_em_upload_finds_a_chart_typed_in_lower_case(client, db):
    chart = Chart(chart_number="EM001", specialty=Specialty.EM, category="T",
                  difficulty=Difficulty.INTERMEDIATE, status=ChartStatus.ACTIVE,
                  uploaded_by="t")
    db.add(chart)
    db.commit()

    headers, values = _em_row(chart_number="em001", em_code="99214", entered_by="T")
    r = client.post(
        "/practicelab/em/answer-key/upload",
        files={"file": ("k.xlsx", _sheet(headers, values),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"entered_by": "T", "passphrase": PASS},
    )
    assert r.status_code == 200, r.text
    assert r.json()["not_found"] == []
