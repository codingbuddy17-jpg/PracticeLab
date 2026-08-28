"""
Every workbook this application produces looks like it came from one place.

Seven Excel writers grew independently and brought three header colours with
them: #4F46E5 on the assessment and auditor exports, #1E3A5F on two more, and
#1F3864 on the answer-key templates. The last two are near-identical navies
differing by a digit, which is drift rather than a decision — and one export,
the chart library report, had no formatting at all.

A grep, because that is how the drift arrives: nobody chooses a second navy on
purpose, they copy a line and adjust it.
"""
import re
from pathlib import Path

import pytest

BACKEND = Path(__file__).resolve().parents[1]

WRITERS = [
    "routers/assessment_pkg/export.py",
    "routers/assessment_pkg/questions.py",
    "routers/assessment_pkg/sessions.py",
    "routers/practicelab_pkg/em_grading.py",
    "routers/reports.py",
    "services/audit_export.py",
    "services/excel_service.py",
]

# Structure, from services/excel_style.
HOUSE = {"1F3864", "2E75B6", "FFFDE7", "FFFFFF"}
# Meaning. A passing row and a failing row should not look alike, so these are
# deliberately not the header palette and are not drift.
SEMANTIC = {"D1FAE5", "FEE2E2", "ECFDF5", "FFF1F2", "DBEAFE", "374151"}


def _fills(path):
    src = (BACKEND / path).read_text()
    return {m.group(1).upper()
            for m in re.finditer(r'fgColor\s*=\s*["\']([0-9A-Fa-f]{6})["\']', src)}


def test_the_writer_list_is_current():
    """Guards the guard — a stale list would check nothing."""
    found = {str(p.relative_to(BACKEND)) for p in BACKEND.rglob("*.py")
             if "test" not in str(p) and "__pycache__" not in str(p)
             and "PatternFill" in p.read_text()}
    missing = found - set(WRITERS) - {"services/excel_style.py"}
    assert not missing, "new Excel writer(s) not covered by this test: %s" % missing


@pytest.mark.parametrize("path", WRITERS)
def test_no_writer_invents_its_own_colour(path):
    stray = _fills(path) - HOUSE - SEMANTIC
    assert not stray, (
        "%s uses colour(s) %s that are neither the house palette nor a "
        "semantic cell fill — see services/excel_style" % (path, sorted(stray)))


def _excel_colour_lines(path):
    """
    Lines that colour a CELL. Excludes reportlab, which styles the PDF exports
    and is a separate palette question, and comments, where naming an old
    colour is describing it rather than using it.
    """
    out = []
    for line in (BACKEND / path).read_text().split("\n"):
        stripped = line.strip()
        if stripped.startswith("#") or "colors.HexColor" in line:
            continue
        if "fgColor" in line or re.search(r"Font\([^)]*color\s*=", line):
            out.append(line)
    return out


def test_the_two_near_identical_navies_are_gone():
    """
    The specific failure: #1E3A5F and #1F3864 side by side in one product.
    Named explicitly because the generic check above would pass the day someone
    reintroduces one of them as a 'house' constant somewhere else.
    """
    for path in WRITERS:
        for line in _excel_colour_lines(path):
            up = line.upper()
            assert "1E3A5F" not in up, "%s: %s" % (path, line.strip())
            assert "4F46E5" not in up, "%s: %s" % (path, line.strip())


def test_the_style_module_matches_the_word_documents():
    """
    A workbook and a handover document should not look like different products.
    The palette is lifted from docs/_spec_docx_build.js; if that changes, this
    says so rather than letting the two families drift apart quietly.
    """
    from services import excel_style
    docs = (BACKEND.parent / "docs" / "_spec_docx_build.js").read_text()
    m = re.search(r"const NAVY = '([0-9A-Fa-f]{6})', ACCENT = '([0-9A-Fa-f]{6})'", docs)
    assert m, "the document palette could not be read"
    assert excel_style.NAVY == m.group(1).upper()
    assert excel_style.ACCENT == m.group(2).upper()


def test_headers_are_frozen_and_filterable():
    """
    Colour is the least of it. A heading that scrolls away leaves a reader
    counting columns, and every export here is long enough for that.
    """
    from openpyxl import Workbook

    from services.excel_style import finish
    wb = Workbook()
    ws = wb.active
    ws.append(["Chart", "Coder", "Score"])
    ws.append(["IP001", "Asha", 89])
    finish(ws)
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:C2"
    assert ws["A1"].fill.fgColor.rgb.endswith("1F3864")


def test_a_long_value_cannot_stretch_a_column_off_the_screen():
    from openpyxl import Workbook

    from services.excel_style import finish
    wb = Workbook()
    ws = wb.active
    ws.append(["Note"])
    ws.append(["x" * 400])
    finish(ws)
    assert ws.column_dimensions["A"].width <= 52
