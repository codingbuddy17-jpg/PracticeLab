"""
Every file export, actually exercised.

Today produced three endpoints that had never once run — the E/M template was
shadowed by a parameterised route AND set an illegal sheet title, and the
coder report rendered raw JSON into a blank tab. None of that was visible from
reading the code. These tests open what comes back.
"""
import io
import json
import pytest
from openpyxl import load_workbook
from sqlalchemy import text

from models import Chart, Batch, Specialty, AnswerKey, GradingResult, PassFail, BatchCoder
from models.charts import ChartStatus, Difficulty
from models.practicelab import BatchStatus

PASSPHRASE = "test-passphrase"
XLSX_MIME = "spreadsheetml"


def _is_workbook(resp):
    """A real .xlsx, not a JSON error page or a base64 envelope."""
    assert resp.status_code == 200, f"{resp.status_code}: {resp.text[:200]}"
    assert XLSX_MIME in resp.headers.get("content-type", ""), resp.headers.get("content-type")
    return load_workbook(io.BytesIO(resp.content))     # raises on anything else


def _chart(db, number, specialty=Specialty.IP_DRG):
    c = Chart(chart_number=number, specialty=specialty, category="Sepsis",
              difficulty=Difficulty.BEGINNER, status=ChartStatus.ACTIVE,
              uploaded_by="t", view_count=0)
    db.add(c); db.commit()
    return c


@pytest.fixture()
def batch_with_results(db):
    b = Batch(name="Export Batch", specialty=Specialty.IP_DRG, status=BatchStatus.OPEN,
              created_by="t", charts_per_coder=1, is_direct_assignment=False,
              use_weighted=True, use_dpo=False, force_closed=False)
    db.add(b); db.commit()
    db.add(BatchCoder(batch_id=b.id, coder_name="Asha R", emp_id="E1"))
    c = _chart(db, "IP600")
    db.add(GradingResult(batch_id=b.id, coder_name="Asha R", emp_id="E1",
                         chart_id=c.id, specialty=Specialty.IP_DRG,
                         total_score=90, pass_fail=PassFail.PASS))
    db.commit()
    return b


class TestBlankTemplates:
    @pytest.mark.parametrize("specialty", [
        "IP-DRG", "SDS", "Surgery", "Ancillary", "ED Single Path", "ED Facility",
    ])
    def test_answer_key_template_downloads(self, client, specialty):
        wb = _is_workbook(client.get("/practicelab/answer-key/template",
                                     params={"specialty": specialty}))
        assert wb.worksheets[0].max_column > 1

    def test_template_filename_names_the_specialty(self, client):
        """Every non-IP template used to download as OP_AnswerKey_Template.xlsx."""
        r = client.get("/practicelab/answer-key/template", params={"specialty": "Surgery"})
        assert "Surgery" in r.headers.get("content-disposition", "")

    def test_em_template_downloads(self, client):
        """Was shadowed by /em/answer-key/{chart_id} and had an illegal sheet title."""
        wb = _is_workbook(client.get("/practicelab/em/answer-key/template"))
        headers = [c.value for c in wb.worksheets[0][1] if c.value]
        assert any("Dx Pointers" in str(h) for h in headers)

    def test_coder_list_template_downloads(self, client):
        wb = _is_workbook(client.get("/practicelab/coders/template"))
        assert [c.value for c in wb.worksheets[0][1]][:2] == ["Coder_Name", "Emp_ID"]

    def test_assessment_question_template_downloads(self, client):
        _is_workbook(client.get("/assessment/questions/template"))


class TestDataExports:
    def test_batch_results_export(self, client, batch_with_results):
        wb = _is_workbook(client.get(
            f"/practicelab/batches/{batch_with_results.id}/results/export"))
        assert wb.worksheets[0].max_row >= 2, "header plus at least one result row"

    def test_answer_key_export_requires_the_passphrase(self, client):
        r = client.get("/practicelab/answer-key/export", params={"passphrase": "wrong"})
        assert r.status_code == 403

    def test_answer_key_export_downloads(self, client, db):
        c = _chart(db, "IP601")
        db.add(AnswerKey(chart_id=c.id, specialty=Specialty.IP_DRG, pdx_code="J18.9",
                         pdx_poa="Y", sdx=[], pcs=[], cpt=[], entered_by="t"))
        db.commit()
        wb = _is_workbook(client.get("/practicelab/answer-key/export",
                                     params={"passphrase": PASSPHRASE}))
        assert "IP-DRG" in wb.sheetnames

    def test_export_keeps_specialty_specific_columns(self, client, db):
        """
        Export used to funnel every non-IP specialty into one generic OP sheet,
        so Surgery lost its pointer columns on the way out and a re-upload
        silently dropped them.
        """
        c = _chart(db, "SURG600", Specialty.SURGERY)
        db.add(AnswerKey(chart_id=c.id, specialty=Specialty.SURGERY, pdx_code="M17.11",
                         sdx=[], pcs=[],
                         cpt=[{"code": "27447", "modifier": "RT", "pointers": ["A"]}],
                         entered_by="t"))
        db.commit()
        wb = _is_workbook(client.get("/practicelab/answer-key/export",
                                     params={"passphrase": PASSPHRASE}))
        assert "Surgery" in wb.sheetnames
        headers = [str(x.value) for x in wb["Surgery"][1] if x.value]
        assert "CPT_1_DxPointers" in headers


class TestRoundTrip:
    """A template that cannot be filled in and read back is not an export."""

    def test_download_fill_reupload(self, client, db):
        _chart(db, "SDS600", Specialty.SDS)
        data = client.get("/practicelab/answer-key/template",
                          params={"specialty": "SDS"}).content
        wb = load_workbook(io.BytesIO(data))
        ws = wb.worksheets[0]
        ws.cell(2, 1, "SDS600"); ws.cell(2, 2, "R51")
        buf = io.BytesIO(); wb.save(buf)

        r = client.post("/practicelab/answer-key/upload",
                        files={"file": ("k.xlsx", buf.getvalue(), "application/vnd.ms-excel")},
                        data={"specialty": "SDS", "entered_by": "trainer"})
        assert r.status_code == 200, r.text
        assert "SDS600" in r.json()["stored"]


class TestAssessmentExports:
    """The Assessment module's exports, same treatment."""

    def test_export_all_requires_the_passphrase(self, client):
        assert client.get("/assessment/questions/export-all",
                          params={"passphrase": "wrong"}).status_code == 403

    def test_question_bank_export_all(self, client):
        r = client.get("/assessment/questions/export-all", params={"passphrase": PASSPHRASE})
        # Empty bank may legitimately 404; anything returned must be a workbook.
        if r.status_code == 200:
            _is_workbook(r)
        else:
            assert r.status_code in (404, 400), r.text

    def test_question_export_by_specialty(self, client):
        r = client.get("/assessment/questions/export",
                       params={"specialty": "IP-DRG", "passphrase": PASSPHRASE})
        if r.status_code == 200:
            _is_workbook(r)
        else:
            assert r.status_code in (404, 400), r.text

    def test_responses_export_for_a_missing_assessment_404s(self, client):
        """Must be a clean 404, not a 500 or a blank file."""
        r = client.get("/assessment/999999/export-responses.xlsx")
        assert r.status_code == 404

    def test_answer_key_export_for_a_missing_assessment_404s(self, client):
        r = client.get("/assessment/999999/export-answer-key")
        assert r.status_code == 404


class TestCoderPerformanceExport:
    """
    Long-format pivot source. The PDF answers "how is this coder doing"; this
    answers "let me slice it myself", which nothing else exposed.
    """

    def test_three_sheets(self, client, batch_with_results):
        wb = _is_workbook(client.get("/practicelab/analytics/coder-performance.xlsx"))
        assert wb.sheetnames == ["Results", "By Coder", "Errors"]

    def test_results_sheet_is_one_row_per_result(self, client, batch_with_results):
        ws = _is_workbook(client.get("/practicelab/analytics/coder-performance.xlsx"))["Results"]
        assert ws.max_row == 2, "header + one seeded result"
        headers = [c.value for c in ws[1]]
        for expected in ("Coder", "Emp ID", "Batch", "Assignment Type",
                         "Chart", "Specialty", "Category", "Score %", "Result"):
            assert expected in headers

    def test_pivot_dimensions_are_populated(self, client, batch_with_results):
        ws = _is_workbook(client.get("/practicelab/analytics/coder-performance.xlsx"))["Results"]
        row = {c.value: ws.cell(2, i + 1).value for i, c in enumerate(ws[1])}
        assert row["Coder"] == "Asha R"
        assert row["Emp ID"] == "E1"
        assert row["Score %"] == 90
        assert row["Result"] == "PASS"
        assert row["Category"] == "Sepsis", "category must be pivotable"

    def test_headers_freeze_and_filter(self, client, batch_with_results):
        """It is a pivot source — it has to be usable on a long sheet."""
        ws = _is_workbook(client.get("/practicelab/analytics/coder-performance.xlsx"))["Results"]
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref is not None

    def test_by_coder_sheet_aggregates(self, client, batch_with_results):
        ws = _is_workbook(client.get("/practicelab/analytics/coder-performance.xlsx"))["By Coder"]
        row = {c.value: ws.cell(2, i + 1).value for i, c in enumerate(ws[1])}
        assert row["Coder"] == "Asha R"
        assert row["Charts"] == 1
        assert row["Passed"] == 1
        assert row["Pass Rate %"] == 100.0

    def test_direct_assignments_are_included_but_labelled(self, client, db, batch_with_results):
        """
        A coder-level view includes direct work — but the trainer decides
        whether to count it, so it gets a column rather than a filter.
        """
        d = Batch(name="Direct One", specialty=Specialty.IP_DRG, status=BatchStatus.OPEN,
                  created_by="t", charts_per_coder=1, is_direct_assignment=True,
                  use_weighted=True, use_dpo=False, force_closed=False)
        db.add(d); db.commit()
        c = _chart(db, "IP602")
        db.add(GradingResult(batch_id=d.id, coder_name="Asha R", emp_id="E1",
                             chart_id=c.id, specialty=Specialty.IP_DRG,
                             total_score=40, pass_fail=PassFail.FAIL))
        db.commit()

        ws = _is_workbook(client.get("/practicelab/analytics/coder-performance.xlsx"))["Results"]
        headers = [c.value for c in ws[1]]
        col = headers.index("Assignment Type") + 1
        types = {ws.cell(r, col).value for r in range(2, ws.max_row + 1)}
        assert types == {"Batch", "Direct"}

    def test_filters_apply(self, client, batch_with_results):
        r = client.get("/practicelab/analytics/coder-performance.xlsx",
                       params={"specialty": "SDS"})     # nothing seeded for SDS
        ws = _is_workbook(r)["Results"]
        assert ws.cell(2, 1).value == "No results match the selected filters."

    def test_single_coder_filter(self, client, batch_with_results):
        r = client.get("/practicelab/analytics/coder-performance.xlsx",
                       params={"emp_id": "E1"})
        assert _is_workbook(r)["Results"].max_row == 2
        assert "Content-Disposition" in r.headers


class TestBatchListExport:
    """
    The batch list panel, exported as it is shown.

    Deliberately not the coder-performance export: that is one row per graded
    RESULT for pivoting, this is one row per BATCH for the list you are
    looking at. The distinction is the whole reason it exists.
    """

    def _mk(self, client, name, direct=False):
        r = client.post("/practicelab/batches", json={
            "name": name, "specialty": "IP-DRG", "categories": [], "difficulties": [],
            "charts_per_coder": 3, "coders": [{"name": "A", "emp_id": "E1"}],
            "created_by": "trainer", "is_direct_assignment": direct,
        })
        assert r.status_code == 200, r.text
        return r.json()["batch_id"]

    def test_one_row_per_batch_with_the_panel_columns(self, client):
        self._mk(client, "Wave 1")
        ws = _is_workbook(client.get("/practicelab/batches/export.xlsx")).worksheets[0]
        headers = [c.value for c in ws[1]]
        for expected in ("Name", "Type", "Specialty", "Status", "Coders",
                         "Charts / Coder", "Cycles", "Graded Results", "Created By"):
            assert expected in headers
        assert ws.max_row == 2, "header + one batch"
        row = {c.value: ws.cell(2, i + 1).value for i, c in enumerate(ws[1])}
        assert row["Name"] == "Wave 1"
        assert row["Coders"] == 1
        assert row["Charts / Coder"] == 3

    def test_the_route_is_not_shadowed_by_batch_id(self, client):
        """/batches/export.xlsx must not be parsed as /batches/{batch_id}."""
        r = client.get("/practicelab/batches/export.xlsx")
        assert r.status_code == 200, r.text
        assert XLSX_MIME in r.headers.get("content-type", "")

    def test_batches_and_direct_assignments_come_back_together(self, client):
        """The panel merges them, so the export of the panel must too."""
        self._mk(client, "Formal One")
        self._mk(client, "Direct One", direct=True)
        ws = _is_workbook(client.get("/practicelab/batches/export.xlsx")).worksheets[0]
        col = [c.value for c in ws[1]].index("Type") + 1
        assert {ws.cell(r, col).value for r in range(2, ws.max_row + 1)} == {"Batch", "Direct"}

    def test_search_filter_carries_over(self, client):
        self._mk(client, "Sepsis Wave")
        self._mk(client, "Other Work")
        ws = _is_workbook(client.get("/practicelab/batches/export.xlsx",
                                     params={"search": "sepsis"})).worksheets[0]
        assert ws.max_row == 2
        assert ws.cell(2, 1).value == "Sepsis Wave"

    def test_status_filter_carries_over(self, client):
        self._mk(client, "Open One")
        ws = _is_workbook(client.get("/practicelab/batches/export.xlsx",
                                     params={"status": "Closed"})).worksheets[0]
        assert ws.cell(2, 1).value == "No batches match the selected filters."

    def test_paging_does_not_carry_over(self, client):
        """
        A page is a screen limit, not a request. Exporting 25 of 340 because
        that is what fits on screen would be a trap.
        """
        for i in range(30):
            self._mk(client, f"Batch {i}")
        ws = _is_workbook(client.get("/practicelab/batches/export.xlsx")).worksheets[0]
        assert ws.max_row == 31, "all 30 rows, not one page"

    def test_headers_freeze_and_filter(self, client):
        self._mk(client, "Any")
        ws = _is_workbook(client.get("/practicelab/batches/export.xlsx")).worksheets[0]
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref is not None


class TestBatchAnalyticsExport:
    """
    The Analytics -> By Batch table. Third batch-shaped export, and the reason
    for a third is that they answer different questions: the batch LIST export
    is the roster (coders, cycles, status — how a batch was set up), this is
    performance (scores, pass rates — how it went).
    """

    def test_columns_are_the_performance_view_not_the_roster(self, client, batch_with_results):
        ws = _is_workbook(client.get("/practicelab/analytics/by-batch.xlsx")).worksheets[0]
        headers = [c.value for c in ws[1]]
        for expected in ("Batch", "Specialty", "Avg Grading Score %", "Chart Pass Rate %",
                         "Pass Mark %", "Graded Results", "Below Target"):
            assert expected in headers
        assert "Cycles" not in headers, "that is the roster export's column"

    def test_one_row_per_batch(self, client, batch_with_results):
        ws = _is_workbook(client.get("/practicelab/analytics/by-batch.xlsx")).worksheets[0]
        assert ws.max_row == 2
        row = {c.value: ws.cell(2, i + 1).value for i, c in enumerate(ws[1])}
        assert row["Batch"] == "Export Batch"
        assert row["Avg Grading Score %"] == 90
        assert row["Chart Pass Rate %"] == 100.0

    def test_pass_mark_travels_with_the_score(self, client, batch_with_results):
        """A score without its bar is not interpretable once it is in Excel."""
        ws = _is_workbook(client.get("/practicelab/analytics/by-batch.xlsx")).worksheets[0]
        row = {c.value: ws.cell(2, i + 1).value for i, c in enumerate(ws[1])}
        assert row["Pass Mark %"] == 80, "IP-DRG"

    def test_scope_carries_over(self, client, db, batch_with_results):
        d = Batch(name="Direct Row", specialty=Specialty.IP_DRG, status=BatchStatus.OPEN,
                  created_by="t", charts_per_coder=1, is_direct_assignment=True,
                  use_weighted=True, use_dpo=False, force_closed=False)
        db.add(d); db.commit()
        c = _chart(db, "IP610")
        db.add(GradingResult(batch_id=d.id, coder_name="B", emp_id="E2", chart_id=c.id,
                             specialty=Specialty.IP_DRG, total_score=50, pass_fail=PassFail.FAIL))
        db.commit()

        formal = _is_workbook(client.get("/practicelab/analytics/by-batch.xlsx")).worksheets[0]
        assert formal.max_row == 2, "default scope excludes direct"

        both = _is_workbook(client.get("/practicelab/analytics/by-batch.xlsx",
                                       params={"scope": "all"})).worksheets[0]
        assert both.max_row == 3
        col = [c.value for c in both[1]].index("Type") + 1
        assert {both.cell(r, col).value for r in range(2, 4)} == {"Batch", "Direct"}

    def test_empty_selection_says_so(self, client):
        ws = _is_workbook(client.get("/practicelab/analytics/by-batch.xlsx",
                                     params={"specialty": "SDS"})).worksheets[0]
        assert ws.cell(2, 1).value == "No batches match the selected filters."


class TestGridExports:
    """
    The screen caps columns and pages rows because that is what stays readable.
    These sheets answer the other question — everything at once — so the one
    thing they must not do is inherit the paging.
    """

    @pytest.fixture()
    def wide(self, db):
        from datetime import datetime
        from models.practicelab import BatchStatus
        b = Batch(name="Closed Wave", specialty=Specialty.IP_DRG, status=BatchStatus.CLOSED,
                  created_by="t", charts_per_coder=1, is_direct_assignment=False,
                  use_weighted=True, use_dpo=False, force_closed=False)
        db.add(b); db.commit()
        b.closed_at = datetime(2026, 6, 1); db.commit()
        for i in range(30):
            c = _chart(db, f"IPG{i:03d}")
            db.add(GradingResult(batch_id=b.id, coder_name=f"Coder {i:03d}", emp_id=f"G{i:03d}",
                                 chart_id=c.id, specialty=Specialty.IP_DRG,
                                 total_score=50 + i, pass_fail=PassFail.FAIL))
        db.commit()
        return b

    def test_matrix_grid_ignores_the_on_screen_paging(self, client, wide):
        """25 rows on screen; the sheet has all 30."""
        ws = _is_workbook(client.get("/practicelab/analytics/coder-matrix.xlsx",
                                     params={"group_by": "batch"})).worksheets[0]
        assert ws.max_row == 31, "header plus every coder"

    def test_matrix_grid_is_wide_not_long(self, client, wide):
        ws = _is_workbook(client.get("/practicelab/analytics/coder-matrix.xlsx",
                                     params={"group_by": "batch"})).worksheets[0]
        assert ws.cell(1, 1).value == "Coder"
        assert ws.cell(1, 2).value.startswith("Closed Wave")
        assert ws.cell(1, ws.max_column).value == "Overall"

    def test_a_missing_cell_is_blank_not_zero(self, client, db, wide):
        """A coder with no result in a batch did not score nothing, and a 0
        would drag any average taken down the column."""
        from datetime import datetime
        from models.practicelab import BatchStatus
        b2 = Batch(name="Second Wave", specialty=Specialty.IP_DRG, status=BatchStatus.CLOSED,
                   created_by="t", charts_per_coder=1, is_direct_assignment=False,
                   use_weighted=True, use_dpo=False, force_closed=False)
        db.add(b2); db.commit()
        b2.closed_at = datetime(2026, 6, 2); db.commit()
        c = _chart(db, "IPG900")
        db.add(GradingResult(batch_id=b2.id, coder_name="Coder 000", emp_id="G000",
                             chart_id=c.id, specialty=Specialty.IP_DRG,
                             total_score=90, pass_fail=PassFail.PASS))
        db.commit()
        ws = _is_workbook(client.get("/practicelab/analytics/coder-matrix.xlsx",
                                     params={"group_by": "batch"})).worksheets[0]
        # Headers carry the pass mark on a second line, so match the prefix.
        col = next(i for i, c in enumerate(ws[1], start=1)
                   if str(c.value).startswith("Second Wave"))
        blanks = [ws.cell(r, col).value for r in range(2, ws.max_row + 1)]
        assert None in blanks, "coders absent from that batch must be blank"

    def test_matrix_grid_filters_carry_over(self, client, wide):
        ws = _is_workbook(client.get("/practicelab/analytics/coder-matrix.xlsx",
                                     params={"group_by": "batch", "search": "Coder 001"})).worksheets[0]
        assert ws.max_row == 2

    def test_topic_heatmap_grid(self, client, wide):
        ws = _is_workbook(client.get("/practicelab/analytics/topic-heatmap.xlsx")).worksheets[0]
        assert ws.cell(1, 1).value == "Coder"
        assert ws.max_row == 31

    def test_chart_signals_is_a_list_not_a_grid(self, client, wide):
        """One row per chart — it is a list, so it exports as one."""
        ws = _is_workbook(client.get("/practicelab/analytics/chart-signals.xlsx")).worksheets[0]
        headers = [c.value for c in ws[1]]
        for expected in ("Chart", "Signal", "Pass Mark %", "Distinct Codes Missed"):
            assert expected in headers
        assert ws.max_row == 31, "all 30 charts, not the 24 the screen pages"

    def test_grid_exports_freeze_both_panes(self, client, wide):
        """Wide and long at once — row labels and headers must both survive."""
        ws = _is_workbook(client.get("/practicelab/analytics/coder-matrix.xlsx")).worksheets[0]
        assert ws.freeze_panes == "B2"

    def test_empty_selection_says_so(self, client):
        ws = _is_workbook(client.get("/practicelab/analytics/coder-matrix.xlsx",
                                     params={"search": "nobody"})).worksheets[0]
        assert ws.cell(2, 1).value == "No data matches the selected filters."


class TestErrorAnalysisExport:
    """
    Five sheets rather than one flat table: the tab answers several questions
    with different shapes, and a single sheet would need a column set that fits
    none of them.
    """

    URL = "/practicelab/analytics/error-analysis.xlsx"

    @pytest.fixture()
    def errors(self, db):
        from models import GradingFeedback
        b = Batch(name="Err Batch", specialty=Specialty.IP_DRG, status=BatchStatus.OPEN,
                  created_by="t", charts_per_coder=1, is_direct_assignment=False,
                  use_weighted=True, use_dpo=False, force_closed=False)
        db.add(b); db.commit()
        for i in range(6):
            c = _chart(db, f"IPXP{i}")
            r = GradingResult(batch_id=b.id, coder_name=f"Coder {i}", emp_id=f"E{i}",
                              chart_id=c.id, specialty=Specialty.IP_DRG,
                              total_score=40, pass_fail=PassFail.FAIL)
            db.add(r); db.commit()
            db.add(GradingFeedback(result_id=r.id, section="SDx", issue_type="Missed",
                                   ak_code="J18.9", coder_code=None, detail=""))
            db.commit()
        return b

    def test_the_five_sheets_are_present(self, client, errors):
        wb = _is_workbook(client.get(self.URL))
        for name in ("Insights", "By Issue Type", "By Section", "By Specialty", "Codes"):
            assert name in wb.sheetnames

    def test_insights_carry_the_written_findings(self, client, errors):
        ws = _is_workbook(client.get(self.URL))["Insights"]
        assert ws.cell(1, 1).value == "Finding"
        assert any("J18.9" in str(ws.cell(r, 1).value) or "SDx" in str(ws.cell(r, 1).value)
                   for r in range(2, 8))

    def test_codes_sheet_carries_the_pattern_verdict(self, client, errors):
        """A count without its spread cannot be acted on — the whole point."""
        ws = _is_workbook(client.get(self.URL))["Codes"]
        headers = [c.value for c in ws[1]]
        for expected in ("Code", "Times", "Coders", "Charts", "Pattern", "What it means"):
            assert expected in headers
        row = {c.value: ws.cell(2, i + 1).value for i, c in enumerate(ws[1])}
        assert row["Code"] == "J18.9"
        assert row["Coders"] == 6
        assert row["Pattern"] == "Team-wide"

    def test_the_specialty_filter_carries_into_the_export(self, client, errors):
        """An export taken while looking at SDS must describe SDS."""
        ws = _is_workbook(client.get(self.URL, params={"specialty": "SDS"}))["Codes"]
        assert ws.cell(2, 1).value == "No data for the selected filters."

    def test_the_filename_names_the_specialty(self, client, errors):
        r = client.get(self.URL, params={"specialty": "IP-DRG"})
        assert "IP-DRG" in r.headers.get("content-disposition", "")

    def test_paging_is_not_inherited(self, client, db, errors):
        """25 codes is a screenful; the reason to open Excel is the whole list."""
        from models import GradingFeedback
        b = db.query(Batch).first()
        for i in range(40):
            c = _chart(db, f"IPXQ{i}")
            r = GradingResult(batch_id=b.id, coder_name="Z", emp_id="EZ", chart_id=c.id,
                              specialty=Specialty.IP_DRG, total_score=40, pass_fail=PassFail.FAIL)
            db.add(r); db.commit()
            db.add(GradingFeedback(result_id=r.id, section="SDx", issue_type="Missed",
                                   ak_code=f"Q{i:03d}", coder_code=None, detail=""))
            db.commit()
        ws = _is_workbook(client.get(self.URL))["Codes"]
        assert ws.max_row > 26, "more than one page of codes"

    def test_an_issue_filter_narrows_the_export(self, client, errors):
        ws = _is_workbook(client.get(self.URL, params={"issue_type": "Over_coded"}))["Codes"]
        assert ws.cell(2, 1).value == "No data for the selected filters."

    def test_the_drilldown_behind_each_verdict_is_exported(self, client, errors):
        """
        The Codes sheet says "Team-wide". Without the names behind it that is a
        claim nobody can check, and on screen the drilldown opens one code at a
        time — in a sheet there is no reason to make someone click 200 times.
        """
        wb = _is_workbook(client.get(self.URL))
        assert "Code by Coder" in wb.sheetnames
        assert "Code by Chart" in wb.sheetnames

        ws = wb["Code by Coder"]
        assert [c.value for c in ws[1]] == ["Code", "Coder", "Emp ID", "Times", "Charts"]
        assert ws.max_row == 7, "six coders on J18.9, plus the header"
        assert {ws.cell(r, 1).value for r in range(2, 8)} == {"J18.9"}

    def test_code_by_chart_shows_where_a_code_concentrates(self, client, errors):
        ws = _is_workbook(client.get(self.URL))["Code by Chart"]
        headers = [c.value for c in ws[1]]
        for expected in ("Code", "Chart", "Specialty", "Topic", "Times", "Coders"):
            assert expected in headers
        assert ws.max_row == 7, "six charts carrying J18.9"

    def test_the_issue_mix_travels_with_each_code(self, client, errors):
        """It sat under the expanded row on screen and was absent from the sheet."""
        ws = _is_workbook(client.get(self.URL))["Codes"]
        headers = [c.value for c in ws[1]]
        assert "Issue Mix" in headers
        row = {c.value: ws.cell(2, i + 1).value for i, c in enumerate(ws[1])}
        assert row["Issue Mix"] == "Missed 6"

    def test_the_drilldown_respects_the_filters(self, client, errors):
        ws = _is_workbook(client.get(self.URL, params={"issue_type": "Over_coded"}))["Code by Coder"]
        assert ws.cell(2, 1).value == "No data for the selected filters."
