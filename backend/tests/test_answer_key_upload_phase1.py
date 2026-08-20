"""
Phase 1 answer-key upload regression.

These tests exercise the real upload endpoints, not just the parsers. A key
that parses but is invisible to PracticeLab or Auditor is still broken: coders
cannot be graded from it, and auditors cannot allocate from it.
"""
import io

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import text

from conftest import make_chart
from models import AnswerKey, AuditBatch, BatchStatus, Specialty
from routers.auditor_pkg.shared import audit_key_for
from routers.auditor_pkg.shared import chart_pool
from routers.practicelab_pkg.chart_grading import _grade_chart_for_sp
from routers.practicelab_pkg.em_grading import grade_em_chart
from routers.practicelab_pkg.shared import (
    _is_dx_only, _is_ip, _is_single_path, _uses_pointers, _uses_units,
)
from services.audit_mutation import Corpus, MUTATION_KINDS, MutationConfig, generate
from services.em_audit_key import EM_KEY_SPECIALTIES
from services.excel_service import (
    EM_KEY_COLUMNS, generate_answer_key_template,
)
from services.grading_engine import DEFAULT_EDSP_CFG, DEFAULT_IP_CFG, DEFAULT_OP_CFG

PASS = "test-passphrase"


STANDARD_KEY_SPECIALTIES = [
    Specialty.IP_DRG,
    Specialty.SDS,
    Specialty.ED_FACILITY,
    Specialty.SURGERY,
    Specialty.ED_SINGLE_PATH,
    Specialty.ANCILLARY,
]


def _headers(data: bytes) -> list[str]:
    ws = load_workbook(io.BytesIO(data)).worksheets[0]
    return [str(c.value).split("\n")[0] if c.value else "" for c in ws[1]]


def _template(spec: Specialty) -> bytes:
    return generate_answer_key_template(
        "IP" if _is_ip(spec) else "OP",
        with_pointers=_uses_pointers(spec),
        single_path=_is_single_path(spec),
        dx_only=_is_dx_only(spec),
        with_units=_uses_units(spec),
    )


def _fill(data: bytes, values: dict) -> bytes:
    wb = load_workbook(io.BytesIO(data))
    ws = wb.worksheets[0]
    at = {str(c.value).split("\n")[0]: c.column for c in ws[1] if c.value}
    for header, value in values.items():
        assert header in at, f"{header} not in template: {sorted(at)}"
        ws.cell(2, at[header], value)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _standard_key_file(spec: Specialty, chart_number: str) -> bytes:
    values = {"Chart_Number": chart_number, "PDx_Code": "E11.9", "SDx_1": "I10"}
    if _is_ip(spec):
        values.update({
            "PDx_POA": "Y",
            "SDx_1_POA": "Y",
            "SDx_1_CCMCC": "CC",
            "PCS_1": "0DTJ4ZZ",
        })
    else:
        if _is_single_path(spec):
            values["Facility_ED_Level"] = "99283"
            values["Profee_ED_Level"] = "99284"
        if not _is_dx_only(spec):
            values["CPT_1"] = "11042"
            values["CPT_1_Modifier"] = "59"
            if _uses_units(spec):
                values["CPT_1_Units"] = 2
            if _uses_pointers(spec):
                values["CPT_1_DxPointers"] = "1,2"
    return _fill(_template(spec), values)


def _upload_standard(client, spec: Specialty, data: bytes, replace=False,
                     passphrase=PASS):
    return client.post(
        "/practicelab/answer-key/upload",
        files={"file": ("keys.xlsx", data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "specialty": spec.value,
            "entered_by": "QA",
            "replace": str(replace).lower(),
            "passphrase": passphrase,
        },
    )


def _em_values(chart_number: str, code: str) -> tuple[list[str], list]:
    headers = [h for _field, h in EM_KEY_COLUMNS]
    at = {field: i for i, (field, _header) in enumerate(EM_KEY_COLUMNS)}
    values = [""] * len(headers)
    for field, value in {
        "chart_number": chart_number,
        "copa_stable_chronic": 2,
        "dr_review_test_results": 1,
        "dr_independent_interpretation": "Y",
        "risk_prescription_drug_mgmt": "Y",
        "em_code": code,
        "patient_type": "Established",
        "level_method": "MDM",
        "dx_1": "E11.9",
        "dx_2": "I10",
        "cpt_1": "20610",
        "cpt_1_modifier": "RT",
        "cpt_1_units": 1,
        "cpt_1_pointers": "A",
        "em_category": "office" if code.startswith("9921") else "emergency",
        "entered_by": "QA",
    }.items():
        values[at[field]] = value
    return headers, values


def _workbook(headers, *rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    for c, header in enumerate(headers, 1):
        ws.cell(1, c, header)
    for r, row in enumerate(rows, 2):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _em_key_file(chart_number: str, spec: Specialty) -> bytes:
    code = "99214" if spec == Specialty.EM else "99284"
    headers, values = _em_values(chart_number, code)
    return _workbook(headers, values)


def _upload_em(client, data: bytes, replace=False, passphrase=PASS):
    return client.post(
        "/practicelab/em/answer-key/upload",
        files={"file": ("em_keys.xlsx", data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "entered_by": "QA",
            "replace": str(replace).lower(),
            "passphrase": passphrase,
        },
    )


def _only_mutation(kind: str) -> MutationConfig:
    zeros = {field: 0 for _kind, field in MUTATION_KINDS}
    zeros[dict(MUTATION_KINDS)[kind]] = 100
    return MutationConfig(**zeros)


def _perfect_standard_submission(ak: AnswerKey, spec: Specialty) -> dict:
    out = {
        "pdx_code": ak.pdx_code or "",
        "pdx_poa": ak.pdx_poa or "",
        "sdx": ak.sdx or [],
        "pcs": ak.pcs or [],
        "cpt": ak.cpt or [],
    }
    if _is_single_path(spec):
        out["facility_level"] = ak.facility_level or ""
        out["profee_level"] = ak.profee_level or ""
    return out


def _em_scoring_config() -> dict:
    return {
        "line1_weight": 70,
        "line2_weight": 30,
        "em_level_weight": 23.33,
        "cpt_weight": 23.33,
        "dx_weight": 23.34,
        "copa_weight": 10.0,
        "dr_weight": 10.0,
        "risk_weight": 10.0,
        "pass_threshold": 80,
        "overcoding_penalty": True,
    }


@pytest.mark.parametrize("spec", STANDARD_KEY_SPECIALTIES, ids=lambda s: s.value)
def test_standard_specialty_upload_is_visible_to_coder_and_auditor(client, db, spec):
    chart = make_chart(db, specialty=spec.value, chart_number=f"QA{spec.name[:5]}")
    db.commit()

    r = _upload_standard(client, spec, _standard_key_file(spec, chart.chart_number))

    assert r.status_code == 200, r.text
    body = r.json()
    assert body["stored"] == [chart.chart_number]
    assert body["wrong_specialty"] == []
    assert body["not_found"] == []

    status = client.get("/practicelab/answer-key/status",
                        params={"specialty": spec.value}).json()
    assert status["with_answer_key"] == 1
    assert status["without_answer_key"] == 0

    audit_status = client.get("/auditor/keys/status",
                              params={"specialty": spec.value}).json()
    assert audit_status["auditable"] == 1
    assert audit_status["no_answer_key"] == 0

    batch = AuditBatch(name="QA", specialty=spec, charts_per_auditor=1,
                       created_by="QA", status=BatchStatus.OPEN)
    db.add(batch)
    db.commit()
    assert [c.id for c in chart_pool(db, batch)] == [chart.id]


@pytest.mark.parametrize("spec", STANDARD_KEY_SPECIALTIES, ids=lambda s: s.value)
def test_standard_uploaded_key_can_grade_a_perfect_coder_submission(client, db, spec):
    chart = make_chart(db, specialty=spec.value, chart_number=f"GR{spec.name[:5]}")
    db.commit()
    r = _upload_standard(client, spec, _standard_key_file(spec, chart.chart_number))
    assert r.status_code == 200, r.text

    ak = db.query(AnswerKey).filter(AnswerKey.chart_id == chart.id).one()
    result, feedback = _grade_chart_for_sp(
        chart, ak, _perfect_standard_submission(ak, spec),
        DEFAULT_IP_CFG, DEFAULT_OP_CFG, DEFAULT_EDSP_CFG,
    )

    assert result["weighted_score"] == 100
    assert result["pass_fail"] == "PASS"
    assert feedback == []


@pytest.mark.parametrize("spec", STANDARD_KEY_SPECIALTIES, ids=lambda s: s.value)
def test_standard_uploaded_key_can_drive_auditor_mutation(client, db, spec):
    chart = make_chart(db, specialty=spec.value, chart_number=f"MU{spec.name[:5]}")
    db.commit()
    r = _upload_standard(client, spec, _standard_key_file(spec, chart.chart_number))
    assert r.status_code == 200, r.text

    key = audit_key_for(db, chart)
    claim, ground_truth = generate(
        key, spec, seed=7, cfg=_only_mutation("spurious"), budget=1,
        corpus=Corpus(dx_codes=["E11.9", "I10", "Z99.89"]),
    )

    assert claim["pdx_code"]
    assert ground_truth
    assert ground_truth[0]["section"] == "SDx"
    assert ground_truth[0]["action"] == "Delete"


@pytest.mark.parametrize("spec", sorted(EM_KEY_SPECIALTIES, key=lambda s: s.value),
                         ids=lambda s: s.value)
def test_em_store_upload_is_visible_to_coder_and_auditor_pool(client, db, spec):
    chart = make_chart(db, specialty=spec.value, chart_number=f"QA{spec.name}")
    db.commit()

    r = _upload_em(client, _em_key_file(chart.chart_number, spec))

    assert r.status_code == 200, r.text
    assert r.json()["stored"] == [chart.chart_number]

    status = client.get("/practicelab/answer-key/status",
                        params={"specialty": spec.value}).json()
    assert status["key_store"] == "em"
    assert status["with_answer_key"] == 1
    assert status["without_answer_key"] == 0

    batch = AuditBatch(name="QA EM", specialty=spec, charts_per_auditor=1,
                       created_by="QA", status=BatchStatus.OPEN)
    db.add(batch)
    db.commit()
    assert [c.id for c in chart_pool(db, batch)] == [chart.id]


@pytest.mark.parametrize("spec", sorted(EM_KEY_SPECIALTIES, key=lambda s: s.value),
                         ids=lambda s: s.value)
def test_em_uploaded_key_can_grade_a_perfect_coder_submission(client, db, spec):
    chart = make_chart(db, specialty=spec.value, chart_number=f"GREM{spec.name}")
    db.commit()
    r = _upload_em(client, _em_key_file(chart.chart_number, spec))
    assert r.status_code == 200, r.text

    ak = db.execute(
        text("SELECT * FROM em_answer_keys WHERE chart_id = :c"),
        {"c": chart.id},
    ).mappings().one()
    sub = {
        "sub_em_code": ak["em_code"],
        "sub_em_modifier": ak["em_modifier"] or "",
        "sub_patient_type": ak["patient_type"],
        "sub_level_method": ak["level_method"],
        "sub_dx_codes": ak["dx_codes"],
        "sub_procedure_cpts": ak["procedure_cpts"],
        "sub_copa_stable_chronic": 2,
        "sub_dr_review_test_results": 1,
        "sub_dr_independent_interpretation": True,
        "sub_risk_prescription_drug_mgmt": True,
    }

    scoring = grade_em_chart(dict(ak), sub, _em_scoring_config())

    assert round(scoring["total_score"], 1) == 100.0
    assert scoring["em_level_score"] > 0
    assert scoring["dx_score"] > 0


@pytest.mark.parametrize("spec", sorted(EM_KEY_SPECIALTIES, key=lambda s: s.value),
                         ids=lambda s: s.value)
def test_em_uploaded_key_can_drive_mdm_auditor_mutation(client, db, spec):
    chart = make_chart(db, specialty=spec.value, chart_number=f"MUEM{spec.name}")
    db.commit()
    r = _upload_em(client, _em_key_file(chart.chart_number, spec))
    assert r.status_code == 200, r.text

    key = audit_key_for(db, chart)
    claim, ground_truth = generate(
        key, spec, seed=4, cfg=_only_mutation("mdm_shift"), budget=1,
    )

    assert set(claim["mdm"]) == {"copa", "dr", "risk"}
    assert ground_truth
    assert ground_truth[0]["section"] == "MDM"
    assert ground_truth[0]["action"] == "Revise"


def test_standard_upload_reports_wrong_specialty_without_storing(client, db):
    chart = make_chart(db, specialty="Surgery", chart_number="QAWRONG")
    db.commit()

    r = _upload_standard(client, Specialty.SDS,
                         _standard_key_file(Specialty.SDS, chart.chart_number))

    assert r.status_code == 200, r.text
    assert r.json()["stored"] == []
    assert r.json()["wrong_specialty"] == [chart.chart_number]
    assert db.query(AnswerKey).count() == 0


def test_standard_duplicate_key_is_skipped_then_replaced_with_passphrase(client, db):
    chart = make_chart(db, specialty="SDS", chart_number="QADUP")
    db.commit()
    first = _standard_key_file(Specialty.SDS, chart.chart_number)
    assert _upload_standard(client, Specialty.SDS, first).json()["stored"] == ["QADUP"]

    second = _fill(_template(Specialty.SDS), {
        "Chart_Number": chart.chart_number,
        "PDx_Code": "J18.9",
        "SDx_1": "I10",
        "CPT_1": "11042",
        "CPT_1_Modifier": "59",
        "CPT_1_Units": 1,
    })
    skipped = _upload_standard(client, Specialty.SDS, second)
    assert skipped.status_code == 200, skipped.text
    assert skipped.json()["skipped_duplicates"] == ["QADUP"]
    assert db.query(AnswerKey).filter_by(chart_id=chart.id).one().pdx_code == "E11.9"

    blocked = _upload_standard(client, Specialty.SDS, second, replace=True,
                               passphrase="wrong")
    assert blocked.status_code == 403

    replaced = _upload_standard(client, Specialty.SDS, second, replace=True)
    assert replaced.status_code == 200, replaced.text
    assert replaced.json()["replaced"] == ["QADUP"]
    assert db.query(AnswerKey).filter_by(chart_id=chart.id).one().pdx_code == "J18.9"


def test_standard_upload_fails_clearly_when_chart_number_header_is_missing(client, db):
    make_chart(db, specialty="SDS", chart_number="QAMISS")
    db.commit()
    data = _workbook(["PDx_Code", "SDx_1"], ["E11.9", "I10"])

    r = _upload_standard(client, Specialty.SDS, data)

    assert r.status_code in (400, 422)
    assert "Chart_Number" in r.json()["detail"]


def test_standard_upload_matches_chart_number_case_and_spacing(client, db):
    chart = make_chart(db, specialty="SDS", chart_number="QACASE")
    db.commit()

    r = _upload_standard(client, Specialty.SDS,
                         _standard_key_file(Specialty.SDS, "  qacase  "))

    assert r.status_code == 200, r.text
    assert r.json()["stored"] == [chart.chart_number]
    assert r.json()["not_found"] == []


def test_standard_upload_with_only_blank_rows_stores_nothing(client, db):
    make_chart(db, specialty="SDS", chart_number="QABLANK")
    db.commit()

    r = _upload_standard(client, Specialty.SDS,
                         _workbook(["Chart_Number", "PDx_Code"], ["", ""]))

    assert r.status_code == 200, r.text
    assert r.json()["stored"] == []
    assert r.json()["not_found"] == []
    assert db.query(AnswerKey).count() == 0


def test_standard_upload_missing_chart_number_reports_not_found_without_storing(client, db):
    make_chart(db, specialty="SDS", chart_number="QAFOUND")
    db.commit()

    r = _upload_standard(client, Specialty.SDS,
                         _workbook(["Chart_Number", "PDx_Code"], ["QANOTFOUND", "E11.9"]))

    assert r.status_code == 200, r.text
    assert r.json()["stored"] == []
    assert r.json()["not_found"] == ["QANOTFOUND"]
    assert db.query(AnswerKey).count() == 0


def test_ip_upload_without_poa_columns_fails_clearly(client, db):
    make_chart(db, specialty="IP-DRG", chart_number="QAIPPOA")
    db.commit()
    data = _workbook(["Chart_Number", "PDx_Code", "SDx_1", "PCS_1"],
                     ["QAIPPOA", "J18.9", "I10", "0DTJ4ZZ"])

    r = _upload_standard(client, Specialty.IP_DRG, data)

    assert r.status_code in (400, 422)
    assert "POA" in r.json()["detail"]


def test_ed_single_path_upload_without_level_columns_fails_clearly(client, db):
    make_chart(db, specialty="ED Single Path", chart_number="QAEDSPMISS")
    db.commit()
    data = _workbook(["Chart_Number", "PDx_Code", "SDx_1", "CPT_1"],
                     ["QAEDSPMISS", "R07.9", "I10", "93010"])

    r = _upload_standard(client, Specialty.ED_SINGLE_PATH, data)

    assert r.status_code in (400, 422)
    assert "Facility_ED_Level" in r.json()["detail"]
    assert "Profee_ED_Level" in r.json()["detail"]


def test_em_upload_reports_wrong_specialty_separately(client, db):
    chart = make_chart(db, specialty="Surgery", chart_number="QAEMWRONG")
    db.commit()

    r = _upload_em(client, _em_key_file(chart.chart_number, Specialty.EM))

    assert r.status_code == 200, r.text
    assert r.json()["stored"] == []
    assert r.json()["not_found"] == []
    assert r.json()["wrong_specialty"] == [chart.chart_number]


def test_em_upload_with_missing_em_code_skips_without_storing(client, db):
    chart = make_chart(db, specialty="E/M", chart_number="QAEMNOCODE")
    db.commit()
    headers, values = _em_values(chart.chart_number, "99214")
    values[[field for field, _header in EM_KEY_COLUMNS].index("em_code")] = ""

    r = _upload_em(client, _workbook(headers, values))

    assert r.status_code == 200, r.text
    assert r.json()["stored"] == []
    assert r.json()["skipped_duplicates"] == [chart.chart_number]
    assert db.execute(text(
        "SELECT COUNT(*) FROM em_answer_keys WHERE chart_id=:c"
    ), {"c": chart.id}).scalar() == 0


def test_em_upload_sanitises_weird_method_and_category_values(client, db):
    chart = make_chart(db, specialty="E/M", chart_number="QAEMWEIRD")
    db.commit()
    headers, values = _em_values(chart.chart_number, "99214")
    fields = [field for field, _header in EM_KEY_COLUMNS]
    values[fields.index("level_method")] = "nonsense"
    values[fields.index("em_category")] = "not a category"

    r = _upload_em(client, _workbook(headers, values))

    assert r.status_code == 200, r.text
    row = db.execute(text(
        "SELECT level_method, em_category FROM em_answer_keys WHERE chart_id=:c"
    ), {"c": chart.id}).fetchone()
    assert row[0] == "MDM"
    assert row[1] == "office"
