"""
The auditor API end to end: author, allocate, audit, score.

The engines are proved in isolation elsewhere. What this file exists for is the
wiring between them, and three rules that live only at this layer:

  * The auditor is never told a chart's type. Not while working, not in their
    results. Charts recycle, so a chart labelled "clean" is an answer key the
    next time they meet it.
  * A claim is frozen when the auditor opens it. A trainer may reroll before
    that and not after — otherwise the question changes after the answer.
  * Every section needs a verdict before submission, which turns a blank chart
    into an explicit claim rather than an absence.
"""
import pytest

from models import (
    AnswerKey, AuditAssignment, AuditSession, AuditSource, Chart, ChartStatus,
    Difficulty, Specialty,
)

PASS = "test-passphrase"


# ── fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture()
def library(db):
    """A pool of inpatient charts, each with a rich answer key."""
    charts = []
    for i in range(8):
        c = Chart(chart_number=f"AUD{i:03d}", specialty=Specialty.IP_DRG,
                  category="Cardiology", difficulty=Difficulty.INTERMEDIATE,
                  status=ChartStatus.ACTIVE, uploaded_by="t")
        db.add(c)
        db.flush()
        db.add(AnswerKey(
            chart_id=c.id, specialty=Specialty.IP_DRG,
            pdx_code="J18.9", pdx_poa="Y",
            sdx=[{"code": f"E11.{j}", "poa": "Y",
                  "ccmcc": "CC" if j % 2 == 0 else "-"} for j in range(8)],
            pcs=[{"code": f"0DTJ{j}ZZ"} for j in range(3)],
            cpt=[], entered_by="t"))
        charts.append(c)
    db.commit()
    return charts


def make_batch(client, auditors=("Asha R",), charts_per=4, **kw):
    body = {"name": "Audit wave", "specialty": "IP-DRG",
            "charts_per_auditor": charts_per, "created_by": "Trainer",
            "auditors": [{"name": n, "emp_id": f"E{i}"}
                         for i, n in enumerate(auditors)]}
    body.update(kw)
    r = client.post("/auditor/batches", json=body)
    assert r.status_code == 200, r.text
    return r.json()["batch_id"]


def allocate(client, batch_id, **kw):
    r = client.post(f"/auditor/batches/{batch_id}/run-allocation",
                    json={"run_by": "Trainer", **kw})
    assert r.status_code == 200, r.text
    return r.json()


def perfect_work(session_payload, ground_truth_by_chart, verdicts_only=False):
    """Build a submission that finds everything, from the trainer-side truth."""
    charts = []
    for c in session_payload["charts"]:
        sections = [s["key"] for s in session_payload["form"]["sections"]]
        gt = ground_truth_by_chart.get(c["chart_id"], [])
        findings = []
        if not verdicts_only:
            for p in gt:
                f = {"section": p["section"], "action": p["action"]}
                if p["action"] == "Add":
                    f["correct_value"] = p["correct_value"]
                    if (p.get("entry") or {}).get("poa"):
                        f["poa"] = p["entry"]["poa"]
                    if (p.get("entry") or {}).get("ccmcc"):
                        f["ccmcc"] = p["entry"]["ccmcc"]
                elif p["action"] == "Delete":
                    f["line"] = p["line"]
                    f["claim_value"] = p["claim_value"]
                else:
                    f["line"] = p["line"]
                    f["field"] = p.get("field", "code")
                    f["correct_value"] = p["correct_value"]
                findings.append(f)
        touched = {f["section"] for f in findings}
        charts.append({
            "chart_id": c["chart_id"],
            "section_verdicts": {s: ("needs_changes" if s in touched else "no_changes")
                                 for s in sections},
            "findings": findings,
        })
    return {"charts": charts}


def truth_map(client, batch_id):
    r = client.get(f"/auditor/batches/{batch_id}/plantings")
    assert r.status_code == 200, r.text
    return {p["chart_id"]: p["ground_truth"] for p in r.json()["plantings"]}


# ── the whole flow ───────────────────────────────────────────────────────────

class TestEndToEnd:

    def test_a_perfect_audit_scores_100(self, client, db, library):
        batch_id = make_batch(client)
        alloc = allocate(client, batch_id)
        token = alloc["access_codes"][0]["token"]

        opened = client.get(f"/auditor/sessions/by-token/{token}")
        assert opened.status_code == 200, opened.text
        payload = opened.json()

        work = perfect_work(payload, truth_map(client, batch_id))
        r = client.post(f"/auditor/sessions/{payload['session_id']}/submit", json=work)
        assert r.status_code == 200, r.text
        summary = r.json()["summary"]
        assert summary["audit_accuracy"] == 100.0
        assert summary["over_calls"] == 0

    def test_an_auditor_who_records_nothing_fails_the_opportunity_charts(
            self, client, db, library):
        batch_id = make_batch(client)
        alloc = allocate(client, batch_id)
        token = alloc["access_codes"][0]["token"]
        payload = client.get(f"/auditor/sessions/by-token/{token}").json()

        work = perfect_work(payload, truth_map(client, batch_id), verdicts_only=True)
        r = client.post(f"/auditor/sessions/{payload['session_id']}/submit", json=work)
        assert r.status_code == 200, r.text
        summary = r.json()["summary"]
        assert summary["clean_accuracy"] == 100.0
        assert summary["opportunity_accuracy"] == 0.0
        # Inaction is not invention — nothing was over-called.
        assert summary["over_calls"] == 0

    def test_the_clean_quota_is_honoured_and_rounds_down(self, client, db, library):
        batch_id = make_batch(client, charts_per=5)
        allocate(client, batch_id)
        plantings = client.get(f"/auditor/batches/{batch_id}/plantings").json()["plantings"]
        clean = [p for p in plantings if p["source"] == "Clean"]
        assert len(clean) == 2                      # 5 * 50% floors to 2
        assert len(plantings) - len(clean) == 3     # always an opportunity chart

    def test_two_auditors_on_one_chart_see_identical_errors(self, client, db, library):
        """
        The seed comes from (chart, cycle) and not the auditor, which is the
        only way their answers on a shared chart can be compared.
        """
        # clean_share=0 so every chart is generated for both auditors. Which
        # charts come up CLEAN varies per auditor by design, so without this the
        # two may share no generated chart at all and the test proves nothing.
        batch_id = make_batch(client, auditors=("Asha R", "Bo T"), charts_per=8,
                              clean_share=0)
        allocate(client, batch_id)
        plantings = client.get(f"/auditor/batches/{batch_id}/plantings",
                               params={"limit": 500}).json()["plantings"]

        by_chart = {}
        for p in plantings:
            if p["source"] != "Auto":
                continue
            by_chart.setdefault(p["chart_id"], []).append(p["ground_truth"])
        shared = [v for v in by_chart.values() if len(v) > 1]
        assert shared, "expected at least one chart drawn by both auditors"
        for versions in shared:
            assert all(v == versions[0] for v in versions)


# ── what the auditor is allowed to know ──────────────────────────────────────

class TestAuditorNeverLearnsChartType:

    def test_the_open_payload_never_names_the_chart_type(self, client, db, library):
        batch_id = make_batch(client)
        token = allocate(client, batch_id)["access_codes"][0]["token"]
        body = client.get(f"/auditor/sessions/by-token/{token}").text.lower()
        for word in ("clean", "opportunity", "planting", "ground_truth", "source"):
            assert word not in body, f"{word!r} leaked to the auditor"

    def test_results_returned_to_the_auditor_never_name_it_either(
            self, client, db, library):
        """
        Charts recycle through cycles. Telling someone a chart was clean hands
        them the answer if they meet it again.
        """
        batch_id = make_batch(client)
        token = allocate(client, batch_id)["access_codes"][0]["token"]
        payload = client.get(f"/auditor/sessions/by-token/{token}").json()
        work = perfect_work(payload, truth_map(client, batch_id))
        body = client.post(f"/auditor/sessions/{payload['session_id']}/submit",
                           json=work).text.lower()
        assert "is_clean" not in body
        assert '"chart_type"' not in body
        assert '"source"' not in body

    def test_a_clean_chart_renders_exactly_like_a_planted_one(self, client, db, library):
        """
        Structure comes from the specialty and nothing else. A section drawn
        differently because nothing was mutated is a tell.
        """
        batch_id = make_batch(client, charts_per=6)
        token = allocate(client, batch_id)["access_codes"][0]["token"]
        payload = client.get(f"/auditor/sessions/by-token/{token}").json()
        shapes = {tuple(sorted(c["claim"].keys())) for c in payload["charts"]}
        assert len(shapes) == 1

    def test_the_trainer_view_does_name_the_type(self, client, db, library):
        """It is trainer vocabulary — this is the one place it belongs."""
        batch_id = make_batch(client)
        alloc = allocate(client, batch_id)
        token = alloc["access_codes"][0]["token"]
        payload = client.get(f"/auditor/sessions/by-token/{token}").json()
        client.post(f"/auditor/sessions/{payload['session_id']}/submit",
                    json=perfect_work(payload, truth_map(client, batch_id)))

        review = client.get(f"/auditor/sessions/{payload['session_id']}/review")
        assert review.status_code == 200, review.text
        types = {c["chart_type"] for c in review.json()["charts"]}
        assert types <= {"clean", "opportunity"} and types


# ── section verdicts ─────────────────────────────────────────────────────────

class TestSectionVerdicts:

    def test_submission_is_refused_without_a_verdict_on_every_section(
            self, client, db, library):
        batch_id = make_batch(client)
        token = allocate(client, batch_id)["access_codes"][0]["token"]
        payload = client.get(f"/auditor/sessions/by-token/{token}").json()
        r = client.post(f"/auditor/sessions/{payload['session_id']}/submit",
                        json={"charts": [{"chart_id": c["chart_id"],
                                          "section_verdicts": {"PDx": "no_changes"}}
                                         for c in payload["charts"]]})
        assert r.status_code == 400
        detail = r.json()["detail"]
        assert "verdict" in detail["message"]
        assert set(detail["charts"][0]["missing_sections"]) == {"SDx", "PCS"}

    def test_the_sections_required_follow_the_specialty(self, client, db):
        r = client.get("/auditor/form-spec")
        assert r.status_code == 200, r.text
        by_specialty = {s["specialty"]: [x["key"] for x in s["sections"]]
                        for s in r.json()["specialties"]}
        assert by_specialty["IP-DRG"] == ["PDx", "SDx", "PCS"]
        assert by_specialty["Ancillary"] == ["PDx", "SDx"]
        assert "CPT" in by_specialty["Surgery"]

    def test_pdx_offers_only_revise(self, client, db):
        spec = client.get("/auditor/form-spec").json()["specialties"][0]
        pdx = next(s for s in spec["sections"] if s["key"] == "PDx")
        assert pdx["actions"] == ["Revise"]


# ── freezing ─────────────────────────────────────────────────────────────────

class TestFreezing:

    def test_a_trainer_can_reroll_before_the_auditor_opens_it(self, client, db, library):
        batch_id = make_batch(client)
        allocate(client, batch_id)
        planted = [p for p in client.get(
            f"/auditor/batches/{batch_id}/plantings").json()["plantings"]
            if p["source"] == "Auto"]
        target = planted[0]
        before = target["ground_truth"]

        r = client.post(f"/auditor/assignments/{target['assignment_id']}/regenerate",
                        json={"run_by": "Trainer"})
        assert r.status_code == 200, r.text
        after = next(p for p in client.get(
            f"/auditor/batches/{batch_id}/plantings").json()["plantings"]
            if p["assignment_id"] == target["assignment_id"])["ground_truth"]
        assert after != before

    def test_a_reroll_is_refused_once_the_auditor_has_opened_it(
            self, client, db, library):
        batch_id = make_batch(client)
        token = allocate(client, batch_id)["access_codes"][0]["token"]
        client.get(f"/auditor/sessions/by-token/{token}")      # opens every chart

        p = client.get(f"/auditor/batches/{batch_id}/plantings").json()["plantings"][0]
        r = client.post(f"/auditor/assignments/{p['assignment_id']}/regenerate",
                        json={"run_by": "Trainer"})
        assert r.status_code == 409
        assert "already opened" in r.json()["detail"]

    def test_editing_a_key_set_does_not_change_a_live_assignment(
            self, client, db, library):
        """
        The freeze in one sentence: improving the library mid-batch must never
        disturb work in flight.
        """
        chart = library[0]
        made = client.post(f"/auditor/keys/chart/{chart.id}", json={
            "name": "Curated", "authored_by": "T", "passphrase": PASS,
            "always_plant": True,
            "mutations": [{"section": "SDx", "action": "Add", "correct_value": "E11.2"}]})
        assert made.status_code == 200, made.text
        set_id = made.json()["id"]

        batch_id = make_batch(client, charts_per=8)
        allocate(client, batch_id)
        before = [p for p in client.get(
            f"/auditor/batches/{batch_id}/plantings").json()["plantings"]
            if p["chart_id"] == chart.id][0]["ground_truth"]

        client.put(f"/auditor/keys/{set_id}", json={
            "name": "Curated", "authored_by": "T", "passphrase": PASS,
            "mutations": [{"section": "SDx", "action": "Add", "correct_value": "E11.6"}]})

        after = [p for p in client.get(
            f"/auditor/batches/{batch_id}/plantings").json()["plantings"]
            if p["chart_id"] == chart.id][0]["ground_truth"]
        assert after == before


# ── the key library ──────────────────────────────────────────────────────────

class TestKeySets:

    def test_authoring_a_set_and_previewing_the_claim_it_builds(self, client, db, library):
        chart = library[0]
        r = client.post(f"/auditor/keys/chart/{chart.id}/preview", json={
            "mutations": [{"section": "SDx", "action": "Add", "correct_value": "E11.3"},
                          {"section": "SDx", "action": "Delete", "line": 0,
                           "claim_value": "I10"}]})
        assert r.status_code == 200, r.text
        body = r.json()
        assert "E11.3" not in [s["code"] for s in body["claim"]["sdx"]]
        assert body["claim"]["sdx"][0]["code"] == "I10"
        assert body["planting_count"] == 2

    def test_a_planting_naming_a_code_not_on_the_key_is_dropped_with_a_warning(
            self, client, db, library):
        r = client.post(f"/auditor/keys/chart/{library[0].id}/preview", json={
            "mutations": [{"section": "SDx", "action": "Add",
                           "correct_value": "NOTHERE"}]})
        assert r.json()["planting_count"] == 0
        assert "not on this chart's answer key" in r.json()["warning"]

    def test_pdx_cannot_be_added_or_deleted(self, client, db, library):
        """It can be wrong; it cannot be absent."""
        for action in ("Add", "Delete"):
            r = client.post(f"/auditor/keys/chart/{library[0].id}", json={
                "name": "x", "authored_by": "T", "passphrase": PASS,
                "mutations": [{"section": "PDx", "action": action,
                               "correct_value": "J18.9", "claim_value": "J18.9"}]})
            assert r.status_code == 400
            assert "only be revised" in r.json()["detail"]

    def test_authoring_needs_the_passphrase(self, client, db, library):
        r = client.post(f"/auditor/keys/chart/{library[0].id}", json={
            "name": "x", "authored_by": "T", "passphrase": "wrong", "mutations": []})
        assert r.status_code == 403

    def test_a_chart_with_no_answer_key_cannot_be_curated(self, client, db):
        c = Chart(chart_number="NOKEY1", specialty=Specialty.IP_DRG, category="x",
                  difficulty=Difficulty.BEGINNER, status=ChartStatus.ACTIVE,
                  uploaded_by="t")
        db.add(c)
        db.commit()
        r = client.post(f"/auditor/keys/chart/{c.id}", json={
            "name": "x", "authored_by": "T", "passphrase": PASS, "mutations": []})
        assert r.status_code == 400
        assert "no answer key" in r.json()["detail"]

    def test_a_zero_mutation_set_with_a_query_is_a_curated_clean_chart(
            self, client, db, library):
        r = client.post(f"/auditor/keys/chart/{library[0].id}", json={
            "name": "Query only", "authored_by": "T", "passphrase": PASS,
            "mutations": [], "query_expected": True,
            "query_rationale": "Sepsis without organ dysfunction"})
        assert r.status_code == 200, r.text
        assert r.json()["planting_count"] == 0
        assert r.json()["query_expected"] is True


# ── scope ────────────────────────────────────────────────────────────────────

class TestScope:

    @pytest.mark.parametrize("specialty", ["Edits", "Denials", "E/M"])
    def test_rubric_and_em_specialties_are_refused(self, client, db, specialty):
        """
        Edits & Denials have no coded key to plant errors in; E/M needs its own
        audit design. Both are refused with the reason, not silently accepted.
        """
        r = client.post("/auditor/batches", json={
            "name": "x", "specialty": specialty, "created_by": "T",
            "auditors": [{"name": "A", "emp_id": "E1"}]})
        assert r.status_code == 400
        assert "cannot be audited" in r.json()["detail"]

    def test_a_chart_with_no_answer_key_is_never_allocated(self, client, db, library):
        """
        There is no truth to plant errors in, so the auditor could only ever be
        measured on restraint.
        """
        db.add(Chart(chart_number="AUDNOKEY", specialty=Specialty.IP_DRG,
                     category="Cardiology", difficulty=Difficulty.INTERMEDIATE,
                     status=ChartStatus.ACTIVE, uploaded_by="t"))
        db.commit()
        batch_id = make_batch(client, charts_per=9)
        allocate(client, batch_id)
        numbers = {p["chart_number"] for p in client.get(
            f"/auditor/batches/{batch_id}/plantings").json()["plantings"]}
        assert "AUDNOKEY" not in numbers


# ── config ───────────────────────────────────────────────────────────────────

class TestConfig:

    def test_defaults_are_the_agreed_numbers(self, client, db):
        c = client.get("/auditor/config").json()
        assert (c["add_weight"], c["revise_weight"], c["delete_weight"]) == (40, 40, 20)
        assert c["over_call_revenue_pct"] == 20
        assert c["query_missed_pct"] == c["query_unnecessary_pct"] == 20
        assert c["pass_threshold"] == 90
        assert c["mix_total"] == 100

    def test_component_weights_must_total_100(self, client, db):
        r = client.put("/auditor/config", json={
            "add_weight": 50, "revise_weight": 40, "delete_weight": 20,
            "updated_by": "T", "passphrase": PASS})
        assert r.status_code == 400
        assert "must total 100" in r.json()["detail"]

    def test_the_mutation_mix_must_total_100(self, client, db):
        r = client.put("/auditor/config", json={
            "mix": {"mix_omit_sdx": 90}, "updated_by": "T", "passphrase": PASS})
        assert r.status_code == 400
        assert "total 100" in r.json()["detail"]

    def test_config_changes_need_the_passphrase(self, client, db):
        r = client.put("/auditor/config", json={
            "pass_threshold": 50, "updated_by": "T", "passphrase": "nope"})
        assert r.status_code == 403


# ── planting what coders really got wrong ────────────────────────────────────

class TestObservedPlantings:
    """
    The loop that makes this module compound: coders practise, their mistakes
    become audit content, and auditors train on the errors their own
    organisation actually makes.
    """

    def _record_coder_work(self, db, chart, submitted_sdx, coders=3):
        """Write practice results as if coders had sat this chart."""
        import json
        from sqlalchemy import text
        for i in range(coders):
            db.execute(text("""
                INSERT INTO practice_sessions
                  (batch_id, coder_name, specialty, token, chart_ids, status)
                VALUES (NULL, :n, 'IP-DRG', :t, :ci, 'submitted')"""),
                {"n": f"Coder{i}", "t": f"OBS{chart.id}{i}",
                 "ci": json.dumps([chart.id])})
            sid = db.execute(text(
                "SELECT id FROM practice_sessions WHERE token=:t"),
                {"t": f"OBS{chart.id}{i}"}).fetchone()[0]
            db.execute(text("""
                INSERT INTO practice_results
                  (session_id, chart_id, specialty, total_score, pass_fail,
                   pdx_submitted, sdx_submitted, pcs_submitted, cpt_submitted, feedback)
                VALUES (:s, :c, 'IP-DRG', 70, 'FAIL', 'J18.9', :sdx,
                        :pcs, '[]', '[]')"""),
                {"s": sid, "c": chart.id,
                 "sdx": json.dumps([{"code": c, "poa": "Y"} for c in submitted_sdx]),
                 "pcs": json.dumps([{"code": f"0DTJ{j}ZZ"} for j in range(3)])})
        db.commit()

    def test_a_code_coders_kept_missing_is_planted_for_the_auditor(
            self, client, db, library):
        chart = library[0]
        # Every coder missed E11.4 and E11.6.
        self._record_coder_work(
            db, chart, [f"E11.{j}" for j in range(8) if j not in (4, 6)])

        # clean_share=0 so every chart is planted. Otherwise whether THIS chart
        # comes up clean depends on batch_id, which shifts with test ordering.
        batch_id = make_batch(client, charts_per=8, clean_share=0)
        allocate(client, batch_id)
        planting = next(p for p in client.get(
            f"/auditor/batches/{batch_id}/plantings").json()["plantings"]
            if p["chart_id"] == chart.id)

        observed = [g for g in planting["ground_truth"] if g.get("origin") == "observed"]
        assert observed, planting["ground_truth"]
        assert {g["correct_value"] for g in observed} <= {"E11.4", "E11.6"}
        assert all(g["observed_coders"] == 3 for g in observed)
        assert all(g["observed_trend"] is True for g in observed)

    def test_a_chart_no_coder_has_sat_falls_back_to_synthetic(
            self, client, db, library):
        """Coverage grows with time; thin coverage must degrade, not break."""
        batch_id = make_batch(client, charts_per=8, clean_share=0)
        allocate(client, batch_id)
        plantings = client.get(
            f"/auditor/batches/{batch_id}/plantings").json()["plantings"]
        planted = [p for p in plantings if p["ground_truth"]]
        assert planted
        assert all(g.get("origin") != "observed"
                   for p in planted for g in p["ground_truth"])

    def test_an_auditor_can_still_score_100_on_an_observed_planting(
            self, client, db, library):
        """
        The shape has to survive the new source — an observed planting is
        matched and corrected exactly like a synthetic one.
        """
        self._record_coder_work(
            db, library[0], [f"E11.{j}" for j in range(8) if j != 5])
        batch_id = make_batch(client, charts_per=4)
        token = allocate(client, batch_id)["access_codes"][0]["token"]
        payload = client.get(f"/auditor/sessions/by-token/{token}").json()
        r = client.post(f"/auditor/sessions/{payload['session_id']}/submit",
                        json=perfect_work(payload, truth_map(client, batch_id)))
        assert r.status_code == 200, r.text
        assert r.json()["summary"]["audit_accuracy"] == 100.0

    def test_a_corrected_answer_key_retires_the_stale_observation(
            self, client, db, library):
        """
        Coders were marked wrong for omitting E11.7. If the key is later
        corrected to drop E11.7, they were right all along — and the auditor
        must not be asked to add a code the chart no longer supports.
        """
        from models import AnswerKey
        chart = library[0]
        self._record_coder_work(
            db, chart, [f"E11.{j}" for j in range(8) if j != 7])

        key = db.query(AnswerKey).filter(AnswerKey.chart_id == chart.id).first()
        key.sdx = [s for s in key.sdx if s["code"] != "E11.7"]
        db.commit()

        batch_id = make_batch(client, charts_per=8, clean_share=0)
        allocate(client, batch_id)
        planting = next(p for p in client.get(
            f"/auditor/batches/{batch_id}/plantings").json()["plantings"]
            if p["chart_id"] == chart.id)
        assert planting["ground_truth"], "expected this chart to be planted"
        assert all(g.get("correct_value") != "E11.7"
                   for g in planting["ground_truth"])


class TestBatchListing:
    """
    The list is what a trainer lands on, so it has to stay legible as batches
    accumulate: open work first, aged so nothing quietly scrolls away.
    """

    def test_the_list_reports_how_long_an_open_batch_has_been_sitting(
            self, client, db, library):
        from sqlalchemy import text
        batch_id = make_batch(client)
        db.execute(text("UPDATE audit_batches SET created_at=:d WHERE id=:b"),
                   {"d": "2026-07-01 09:00:00", "b": batch_id})
        db.commit()
        row = next(b for b in client.get("/auditor/batches").json()["batches"]
                   if b["id"] == batch_id)
        assert row["days_open"] is not None and row["days_open"] > 14

    def test_a_closed_batch_has_no_ageing_figure(self, client, db, library):
        """
        Days-open on a closed batch would be a number that keeps growing after
        the work stopped, which says nothing true about it.
        """
        batch_id = make_batch(client)
        token = allocate(client, batch_id)["access_codes"][0]["token"]
        payload = client.get(f"/auditor/sessions/by-token/{token}").json()
        client.post(f"/auditor/sessions/{payload['session_id']}/submit",
                    json=perfect_work(payload, truth_map(client, batch_id)))
        client.post(f"/auditor/batches/{batch_id}/close", json={"closed_by": "T"})

        row = next(b for b in client.get("/auditor/batches").json()["batches"]
                   if b["id"] == batch_id)
        assert row["status"] == "Closed"
        assert row["days_open"] is None
        assert row["closed_at"]

    def test_batches_can_be_filtered_by_status(self, client, db, library):
        make_batch(client)
        assert client.get("/auditor/batches",
                          params={"status": "Open"}).json()["batches"]
        assert client.get("/auditor/batches",
                          params={"status": "Closed"}).json()["batches"] == []


class TestExports:
    """
    Every export carries its denominator. A rate on two opportunities and one on
    twenty look identical in a spreadsheet cell, and once pasted into a deck
    nobody can tell which they are reading.
    """

    def _scored_batch(self, client, db, library):
        batch_id = make_batch(client, charts_per=5)
        token = allocate(client, batch_id)["access_codes"][0]["token"]
        payload = client.get(f"/auditor/sessions/by-token/{token}").json()
        client.post(f"/auditor/sessions/{payload['session_id']}/submit",
                    json=perfect_work(payload, truth_map(client, batch_id)))
        return batch_id

    def _sheets(self, content):
        import io
        from openpyxl import load_workbook
        return load_workbook(io.BytesIO(content))

    def test_batch_results_export_has_the_three_sheets(self, client, db, library):
        batch_id = self._scored_batch(client, db, library)
        r = client.get(f"/auditor/batches/{batch_id}/export")
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers["content-type"]
        assert "attachment" in r.headers["content-disposition"]
        wb = self._sheets(r.content)
        assert set(wb.sheetnames) == {"Summary", "Chart_Results", "Findings_Detail"}

    def test_the_export_names_chart_types_because_it_is_for_the_trainer(
            self, client, db, library):
        batch_id = self._scored_batch(client, db, library)
        wb = self._sheets(client.get(f"/auditor/batches/{batch_id}/export").content)
        col = [c.value for c in wb["Chart_Results"]["D"]]
        assert "Clean" in col or "Opportunity" in col

    def test_an_absent_component_exports_as_NA_not_zero(self, client, db, library):
        """Zero would say the auditor missed something that was never there."""
        batch_id = self._scored_batch(client, db, library)
        wb = self._sheets(client.get(f"/auditor/batches/{batch_id}/export").content)
        values = [c.value for row in wb["Chart_Results"].iter_rows() for c in row]
        assert "NA" in values

    def test_findings_detail_records_what_happened_to_each_error(
            self, client, db, library):
        batch_id = self._scored_batch(client, db, library)
        wb = self._sheets(client.get(f"/auditor/batches/{batch_id}/export").content)
        ws = wb["Findings_Detail"]
        assert ws.max_row > 3
        outcomes = {c.value for c in ws["I"]}
        assert "correct" in outcomes

    def test_analytics_export_covers_every_tab(self, client, db, library):
        self._scored_batch(client, db, library)
        r = client.get("/auditor/analytics/export")
        assert r.status_code == 200, r.text
        wb = self._sheets(r.content)
        assert set(wb.sheetnames) == {
            "Overview", "By_Specialty", "By_Batch", "By_Auditor",
            "Detection_Patterns", "Real_vs_Generated", "Chart_Signals"}

    def test_the_key_library_exports_one_row_per_error(self, client, db, library):
        client.post(f"/auditor/keys/chart/{library[0].id}", json={
            "name": "Curated", "authored_by": "T", "passphrase": PASS,
            "mutations": [
                {"section": "SDx", "action": "Add", "correct_value": "E11.2"},
                {"section": "SDx", "action": "Delete", "line": 0, "claim_value": "I10"}]})
        assert client.get("/auditor/keys/export").status_code == 403, \
            "the key export IS the answers — it must be passphrase-gated"
        r = client.get("/auditor/keys/export", params={"passphrase": PASS})
        assert r.status_code == 200, r.text
        ws = self._sheets(r.content)["Audit_Keys"]
        assert ws.max_row >= 5           # note + blank + header + two errors
        assert "Add" in {c.value for c in ws["D"]}

    def test_exports_work_on_an_empty_installation(self, client, db):
        """A trainer clicking Export before any work exists gets a file, not a 500."""
        assert client.get("/auditor/analytics/export").status_code == 200
        assert client.get("/auditor/keys/export",
                          params={"passphrase": PASS}).status_code == 200


class TestKeyCoverage:
    """
    The Audit Keys screen is a coverage report, like the coder Answer Keys
    screen it mirrors: how much of the library has been curated, and what is
    still running on generated errors.
    """

    def test_coverage_counts_curated_against_auditable(self, client, db, library):
        before = client.get("/auditor/keys/status",
                            params={"specialty": "IP-DRG"}).json()
        assert before["auditable"] == len(library)
        assert before["curated"] == 0
        assert before["uncurated"] == len(library)

        client.post(f"/auditor/keys/chart/{library[0].id}", json={
            "name": "Curated", "authored_by": "T", "passphrase": PASS,
            "mutations": [{"section": "SDx", "action": "Add", "correct_value": "E11.2"}]})

        after = client.get("/auditor/keys/status",
                           params={"specialty": "IP-DRG"}).json()
        assert after["curated"] == 1
        assert after["uncurated"] == len(library) - 1

    def test_charts_without_an_answer_key_are_counted_separately(self, client, db):
        """
        They are not a curation gap a trainer can close on this screen — they
        are a chart-library gap, and lumping them in makes the to-do list look
        bigger than the work actually available.
        """
        c = Chart(chart_number="NOKEY9", specialty=Specialty.IP_DRG, category="x",
                  difficulty=Difficulty.BEGINNER, status=ChartStatus.ACTIVE,
                  uploaded_by="t")
        db.add(c)
        db.commit()
        st = client.get("/auditor/keys/status", params={"specialty": "IP-DRG"}).json()
        assert st["no_answer_key"] == 1
        assert st["auditable"] == 0
        assert st["uncurated"] == 0

    def test_the_todo_list_drops_a_chart_once_it_is_curated(self, client, db, library):
        target = library[0]
        before = client.get("/auditor/keys/uncurated",
                            params={"specialty": "IP-DRG"}).json()["charts"]
        assert target.chart_number in {c["chart_number"] for c in before}

        client.post(f"/auditor/keys/chart/{target.id}", json={
            "name": "Curated", "authored_by": "T", "passphrase": PASS,
            "mutations": [{"section": "SDx", "action": "Add", "correct_value": "E11.2"}]})

        after = client.get("/auditor/keys/uncurated",
                           params={"specialty": "IP-DRG"}).json()["charts"]
        assert target.chart_number not in {c["chart_number"] for c in after}

    def test_the_todo_list_never_offers_a_chart_with_no_answer_key(self, client, db, library):
        db.add(Chart(chart_number="NOKEY8", specialty=Specialty.IP_DRG, category="x",
                     difficulty=Difficulty.BEGINNER, status=ChartStatus.ACTIVE,
                     uploaded_by="t"))
        db.commit()
        charts = client.get("/auditor/keys/uncurated",
                            params={"specialty": "IP-DRG"}).json()["charts"]
        assert "NOKEY8" not in {c["chart_number"] for c in charts}

    def test_the_key_list_can_be_scoped_to_a_specialty(self, client, db, library):
        client.post(f"/auditor/keys/chart/{library[0].id}", json={
            "name": "Curated", "authored_by": "T", "passphrase": PASS,
            "mutations": []})
        assert client.get("/auditor/keys",
                          params={"specialty": "IP-DRG"}).json()["sets"]
        assert client.get("/auditor/keys",
                          params={"specialty": "Surgery"}).json()["sets"] == []


class TestVersionMemory:
    """
    A chart can carry several versions of its errors. Two rules make that
    usable rather than confusing: an auditor never sees one chart twice in a
    cycle, and across cycles they get a version they have not had.
    """

    def _versions(self, client, chart, names):
        for n in names:
            r = client.post(f"/auditor/keys/chart/{chart.id}", json={
                "name": n, "authored_by": "T", "passphrase": PASS,
                "always_plant": True,
                "mutations": [{"section": "SDx", "action": "Add",
                               "correct_value": f"E11.{names.index(n)}"}]})
            assert r.status_code == 200, r.text

    def test_one_chart_is_never_assigned_twice_in_a_cycle(self, client, db, library):
        """
        The worry versions raise: two versions of one chart landing on the same
        auditor in one sitting. The draw takes each chart from the pool once, so
        a chart yields at most one assignment and therefore one version.
        """
        self._versions(client, library[0], ["V1", "V2", "V3"])
        batch_id = make_batch(client, charts_per=8)
        allocate(client, batch_id)

        rows = client.get(f"/auditor/batches/{batch_id}/plantings",
                          params={"limit": 500}).json()["plantings"]
        for auditor in {r["auditor_name"] for r in rows}:
            charts = [r["chart_number"] for r in rows if r["auditor_name"] == auditor]
            assert len(charts) == len(set(charts)), f"{auditor} got a chart twice: {charts}"

    def test_two_auditors_may_share_a_chart_and_get_the_same_version(
            self, client, db, library):
        """
        Deliberate — it is the only way their answers on that chart can be
        compared. Versions vary across CYCLES, not across people.
        """
        self._versions(client, library[0], ["V1", "V2"])
        batch_id = make_batch(client, auditors=("Asha R", "Bo T"), charts_per=8)
        allocate(client, batch_id)

        rows = [r for r in client.get(f"/auditor/batches/{batch_id}/plantings",
                                      params={"limit": 500}).json()["plantings"]
                if r["chart_number"] == library[0].chart_number]
        assert len(rows) == 2
        assert rows[0]["set_id"] == rows[1]["set_id"]

    def test_one_auditor_works_through_every_version_before_repeating(
            self, client, db):
        """
        The scenario that found the bug: one auditor, two charts a cycle, and a
        pool that recycles every third cycle against three versions. Keying the
        rotation on the CYCLE number pinned the chart to the same version each
        time it came round — the auditor saw version 2 twice before ever
        reaching version 3.

        Counting the chart's own encounters is immune to the gap between them.
        """
        from models import AnswerKey, Chart, ChartStatus, Difficulty, Specialty
        charts = []
        for i in range(6):
            c = Chart(chart_number=f"ROT{i}", specialty=Specialty.IP_DRG,
                      category="Cardiology", difficulty=Difficulty.INTERMEDIATE,
                      status=ChartStatus.ACTIVE, uploaded_by="t")
            db.add(c)
            db.flush()
            db.add(AnswerKey(
                chart_id=c.id, specialty=Specialty.IP_DRG, pdx_code="J18.9",
                pdx_poa="Y",
                sdx=[{"code": f"E11.{j}", "poa": "Y", "ccmcc": "-"} for j in range(6)],
                pcs=[{"code": "0DTJ0ZZ"}], cpt=[], entered_by="t"))
            charts.append(c)
        db.commit()

        target = charts[0]
        for n in ("V1", "V2", "V3"):
            client.post(f"/auditor/keys/chart/{target.id}", json={
                "name": n, "authored_by": "T", "passphrase": PASS,
                "always_plant": True,
                "mutations": [{"section": "SDx", "action": "Add",
                               "correct_value": "E11.1"}]})

        batch_id = make_batch(client, auditors=("Solo",), charts_per=2,
                              clean_share=0)
        for _ in range(9):
            allocate(client, batch_id)

        rows = client.get(f"/auditor/batches/{batch_id}/plantings",
                          params={"limit": 500}).json()["plantings"]
        seen = [r["set_id"] for r in rows
                if r["chart_number"] == target.chart_number]
        assert len(seen) >= 3, f"chart came round only {len(seen)} time(s)"
        # Every version reached before any repeats.
        assert len(set(seen[:3])) == 3, f"repeated before exhausting versions: {seen}"

    def test_a_later_cycle_gives_a_version_they_have_not_seen(
            self, client, db, library):
        self._versions(client, library[0], ["V1", "V2"])
        batch_id = make_batch(client, charts_per=8)

        seen = []
        for _ in range(2):
            allocate(client, batch_id)
            rows = client.get(f"/auditor/batches/{batch_id}/plantings",
                              params={"limit": 500}).json()["plantings"]
            hits = [r for r in rows if r["chart_number"] == library[0].chart_number]
            seen.append(hits[-1]["set_id"])
        assert seen[0] != seen[1], "the second cycle repeated the first version"

    def test_once_every_version_is_used_it_reuses_rather_than_skipping_the_chart(
            self, client, db, library):
        """
        Running dry is not a reason to drop a chart from the rotation — the
        same rule the chart draw itself follows.
        """
        self._versions(client, library[0], ["Only"])
        batch_id = make_batch(client, charts_per=8)
        for _ in range(3):
            allocate(client, batch_id)
        rows = client.get(f"/auditor/batches/{batch_id}/plantings",
                          params={"limit": 500}).json()["plantings"]
        hits = [r for r in rows if r["chart_number"] == library[0].chart_number]
        assert len(hits) == 3
        assert all(h["set_id"] == hits[0]["set_id"] for h in hits)


class TestVersionCap:

    def _add(self, client, chart, name):
        return client.post(f"/auditor/keys/chart/{chart.id}", json={
            "name": name, "authored_by": "T", "passphrase": PASS,
            "mutations": [{"section": "SDx", "action": "Add", "correct_value": "E11.1"}]})

    def test_a_chart_takes_at_most_three_versions(self, client, db, library):
        """
        Versions extend a chart across CYCLES, not within one — the draw takes
        each chart from the pool once however many it carries. Past three they
        are variants nobody reaches before the chart recycles.
        """
        for n in ("A", "B", "C"):
            assert self._add(client, library[0], n).status_code == 200
        r = self._add(client, library[0], "D")
        assert r.status_code == 400
        assert "already has 3 versions" in r.json()["detail"]
        assert "A, B, C" in r.json()["detail"]

    def test_two_versions_cannot_share_a_name(self, client, db, library):
        """They exist to be told apart; two called the same thing cannot be."""
        assert self._add(client, library[0], "Missed CC").status_code == 200
        r = self._add(client, library[0], "  missed cc  ")
        assert r.status_code == 400
        assert "already has a version called" in r.json()["detail"]

    def test_the_cap_is_per_chart_not_global(self, client, db, library):
        for n in ("A", "B", "C"):
            self._add(client, library[0], n)
        assert self._add(client, library[1], "A").status_code == 200

    def test_the_chart_payload_reports_the_cap(self, client, db, library):
        body = client.get(f"/auditor/keys/chart/{library[0].id}").json()
        assert body["max_versions"] == 3


class TestSubmissionIntegrity:
    """
    Three ways a submission could be malformed and still be accepted, each of
    which produced a score that meant something other than what it said.
    """

    def _open(self, client, charts_per=3):
        batch_id = make_batch(client, charts_per=charts_per)
        token = allocate(client, batch_id)["access_codes"][0]["token"]
        return batch_id, client.get(f"/auditor/sessions/by-token/{token}").json()

    def test_a_needs_changes_verdict_must_say_what_is_wrong(self, client, db, library):
        """
        "Needs changes" is a claim. With nothing behind it, it scored as
        though the auditor had said nothing while reading as though they had
        found something.
        """
        _b, p = self._open(client)
        secs = [s["key"] for s in p["form"]["sections"]]
        r = client.post(f"/auditor/sessions/{p['session_id']}/submit", json={
            "charts": [{"chart_id": c["chart_id"],
                        "section_verdicts": {s: "needs_changes" for s in secs},
                        "findings": []} for c in p["charts"]]})
        assert r.status_code == 400
        assert "must say what is wrong" in r.json()["detail"]["message"]

    def test_a_half_written_finding_is_refused(self, client, db, library):
        """
        A finding with no code cannot match anything, so it scored as an
        over-call — punishing the auditor for an unfinished row rather than
        for a wrong judgement.
        """
        _b, p = self._open(client)
        secs = [s["key"] for s in p["form"]["sections"]]
        r = client.post(f"/auditor/sessions/{p['session_id']}/submit", json={
            "charts": [{"chart_id": c["chart_id"],
                        "section_verdicts": {**{s: "no_changes" for s in secs},
                                             "SDx": "needs_changes"},
                        "findings": [{"section": "SDx", "action": "Add",
                                      "correct_value": "  "}]}
                       for c in p["charts"]]})
        assert r.status_code == 400
        assert "incomplete" in r.json()["detail"]["message"]

    def test_ip_missing_diagnosis_requires_poa(self, client, db, library):
        _b, p = self._open(client)
        secs = [s["key"] for s in p["form"]["sections"]]
        r = client.post(f"/auditor/sessions/{p['session_id']}/submit", json={
            "charts": [{"chart_id": c["chart_id"],
                        "section_verdicts": {**{s: "no_changes" for s in secs},
                                             "SDx": "needs_changes"},
                        "findings": [{"section": "SDx", "action": "Add",
                                      "correct_value": "E11.2"}]}
                       for c in p["charts"]]})
        assert r.status_code == 400
        assert "POA" in r.json()["detail"]["findings"][0]["problem"]

    def test_a_revise_without_a_corrected_value_is_refused(self, client, db, library):
        _b, p = self._open(client)
        secs = [s["key"] for s in p["form"]["sections"]]
        r = client.post(f"/auditor/sessions/{p['session_id']}/submit", json={
            "charts": [{"chart_id": c["chart_id"],
                        "section_verdicts": {**{s: "no_changes" for s in secs},
                                             "SDx": "needs_changes"},
                        "findings": [{"section": "SDx", "action": "Revise",
                                      "field": "code", "line": 0}]}
                       for c in p["charts"]]})
        assert r.status_code == 400

    def test_a_draft_cannot_be_saved_against_a_foreign_chart(self, client, db, library):
        """
        An access code could otherwise write draft rows for any chart in the
        library, and they would sit there unscored and unexplained.
        """
        _b, p = self._open(client)
        r = client.post(f"/auditor/sessions/{p['session_id']}/save-draft", json={
            "charts": [{"chart_id": 999999, "section_verdicts": {}, "findings": []}]})
        assert r.status_code == 400
        assert "not part of this session" in r.json()["detail"]

    def test_a_complete_submission_still_goes_through(self, client, db, library):
        batch_id, p = self._open(client)
        r = client.post(f"/auditor/sessions/{p['session_id']}/submit",
                        json=perfect_work(p, truth_map(client, batch_id)))
        assert r.status_code == 200, r.text


class TestResultVisibility:

    def test_results_are_withheld_by_default(self, client, db, library):
        """
        Off unless the trainer turns it on, matching the coder batches.

        The missed-findings list names the errors that were in the charts the
        auditor has just seen, and those charts get recycled. A trainer can
        switch it on deliberately; it is not something to get by accident.
        """
        batch_id = make_batch(client, charts_per=4)
        token = allocate(client, batch_id)["access_codes"][0]["token"]
        p = client.get(f"/auditor/sessions/by-token/{token}").json()
        assert p["show_results"] is False

        r = client.post(f"/auditor/sessions/{p['session_id']}/submit",
                        json=perfect_work(p, truth_map(client, batch_id)))
        assert r.status_code == 200, r.text
        assert r.json()["results"] == []

    def test_a_batch_can_show_results_when_the_trainer_asks(self, client, db, library):
        batch_id = make_batch(client, charts_per=4, show_results_to_auditor=True)
        token = allocate(client, batch_id)["access_codes"][0]["token"]
        p = client.get(f"/auditor/sessions/by-token/{token}").json()
        assert p["show_results"] is True

        r = client.post(f"/auditor/sessions/{p['session_id']}/submit",
                        json=perfect_work(p, truth_map(client, batch_id)))
        assert r.status_code == 200, r.text
        assert r.json()["results"] != []


class TestBatchCounts:

    def test_status_counts_cover_every_batch_not_just_the_page(self, client, db, library):
        """
        Counting loaded rows told a trainer there were no closed batches
        whenever the closed ones fell past the first page.
        """
        for i in range(4):
            make_batch(client, charts_per=2, name=f"Wave {i}")
        r = client.get("/auditor/batches", params={"limit": 1}).json()
        assert len(r["batches"]) == 1
        assert r["counts"]["open"] >= 4
        assert r["counts"]["all"] >= 4

    def test_filtering_by_status_happens_on_the_server(self, client, db, library):
        make_batch(client, charts_per=2)
        r = client.get("/auditor/batches", params={"status": "Closed"}).json()
        assert r["batches"] == []
        assert r["total"] == 0
        # The counts still describe everything, so the tab can show "Open 1".
        assert r["counts"]["open"] >= 1


class TestEmpIdIdentity:

    def test_same_name_auditors_are_separated_by_emp_id(self, client, db, library):
        r = client.post("/auditor/batches", json={
            "name": "Same name wave", "specialty": "IP-DRG",
            "charts_per_auditor": 2, "created_by": "Trainer",
            "auditors": [
                {"name": "Alex Lee", "emp_id": "E100"},
                {"name": "Alex Lee", "emp_id": "E200"},
            ],
        })
        assert r.status_code == 200, r.text
        batch_id = r.json()["batch_id"]

        alloc = allocate(client, batch_id)
        assert len(alloc["access_codes"]) == 2
        assert {c["emp_id"] for c in alloc["access_codes"]} == {"E100", "E200"}

        detail = client.get(f"/auditor/batches/{batch_id}").json()
        assert set(detail["assignments"].keys()) == {"Alex Lee (E100)", "Alex Lee (E200)"}

        sessions = db.query(AuditSession).filter(AuditSession.batch_id == batch_id).all()
        assert {s.emp_id for s in sessions} == {"E100", "E200"}
        for s in sessions:
            opened = client.get(f"/auditor/sessions/by-token/{s.token}")
            assert opened.status_code == 200, opened.text
            assert len(opened.json()["charts"]) == 2


class TestDeepReviewFixes:
    """
    Issues a deep review of allocation and mutation raised. Each was verified
    against the running code before being fixed.
    """

    def test_guided_quotas_control_the_authored_split(self, client, db, library):
        """
        Only the CLEAN quota was ever honoured. "Yours = 2" fell through to
        "authored if the chart happens to have a version", so Guided told a
        trainer they were choosing the mix while the pool chose it.
        """
        for c in library[:4]:
            client.post(f"/auditor/keys/chart/{c.id}", json={
                "name": "Curated", "authored_by": "T", "passphrase": PASS,
                "mutations": [{"section": "SDx", "action": "Add",
                               "correct_value": "E11.1"}]})

        batch_id = make_batch(client, charts_per=6, allocation_mode="guided",
                              quota_clean=1, quota_manual=2, quota_auto=3)
        allocate(client, batch_id)
        rows = client.get(f"/auditor/batches/{batch_id}/plantings",
                          params={"limit": 500}).json()["plantings"]
        counts = {}
        for r in rows:
            counts[r["source"]] = counts.get(r["source"], 0) + 1
        assert counts.get("Clean", 0) == 1
        assert counts.get("Manual", 0) == 2
        assert counts.get("Auto", 0) == 3

    def test_a_shortfall_of_authored_charts_is_reported_not_absorbed(
            self, client, db, library):
        """A trainer who asked for three and got one should be told why."""
        client.post(f"/auditor/keys/chart/{library[0].id}", json={
            "name": "Curated", "authored_by": "T", "passphrase": PASS,
            "mutations": [{"section": "SDx", "action": "Add", "correct_value": "E11.1"}]})
        batch_id = make_batch(client, charts_per=6, allocation_mode="guided",
                              quota_clean=0, quota_manual=4, quota_auto=2)
        res = allocate(client, batch_id)
        assert any("fell back to system-generated" in n for n in res["pool_notes"])

    def test_an_always_used_version_ignores_the_quota(self, client, db, library):
        """That is the entire meaning of the flag."""
        client.post(f"/auditor/keys/chart/{library[0].id}", json={
            "name": "Forced", "authored_by": "T", "passphrase": PASS,
            "always_plant": True,
            "mutations": [{"section": "SDx", "action": "Add", "correct_value": "E11.1"}]})
        batch_id = make_batch(client, charts_per=8, allocation_mode="guided",
                              quota_clean=4, quota_manual=0, quota_auto=4)
        allocate(client, batch_id)
        row = next(r for r in client.get(f"/auditor/batches/{batch_id}/plantings",
                                         params={"limit": 500}).json()["plantings"]
                   if r["chart_id"] == library[0].id)
        assert row["source"] == "Manual"

    def test_an_error_naming_a_code_not_on_the_key_is_refused_at_save(
            self, client, db, library):
        """
        Preview warned and dropped it; save accepted it. A stored version could
        therefore shed half its errors when next allocated, and for an audit
        key silently dropping the answer is the worst possible failure.
        """
        r = client.post(f"/auditor/keys/chart/{library[0].id}", json={
            "name": "Bad", "authored_by": "T", "passphrase": PASS,
            "mutations": [{"section": "SDx", "action": "Add",
                           "correct_value": "NOTONKEY"}]})
        assert r.status_code == 400
        assert "not on this chart's answer key" in r.json()["detail"]

    def test_a_revision_beyond_the_end_of_a_section_is_refused(
            self, client, db, library):
        r = client.post(f"/auditor/keys/chart/{library[0].id}", json={
            "name": "Bad", "authored_by": "T", "passphrase": PASS,
            "mutations": [{"section": "SDx", "action": "Revise", "field": "code",
                           "line": 99, "claim_value": "ZZZ"}]})
        assert r.status_code == 400
        assert "does not exist" in r.json()["detail"]


class TestHandPickedMode:
    """
    The trainer chooses WHICH charts; the quotas still decide the mix within
    them. Before the picker existed the backend accepted chart ids that no
    screen ever sent, so the mode behaved as filtered-automatic.
    """

    def test_only_the_picked_charts_are_allocated(self, client, db, library):
        wanted = [library[0].id, library[1].id]
        batch_id = make_batch(client, charts_per=2, allocation_mode="manual")
        r = client.post(f"/auditor/batches/{batch_id}/run-allocation",
                        json={"run_by": "T", "manual_chart_ids": wanted})
        assert r.status_code == 200, r.text
        rows = client.get(f"/auditor/batches/{batch_id}/plantings",
                          params={"limit": 500}).json()["plantings"]
        assert {p["chart_id"] for p in rows} == set(wanted)

    def test_the_quotas_apply_within_the_selection(self, client, db, library):
        """
        Hand-picked used to fall back to the clean-share default, so the quota
        boxes were silently ignored the moment someone switched modes.
        """
        for c in library[:3]:
            client.post(f"/auditor/keys/chart/{c.id}", json={
                "name": "Curated", "authored_by": "T", "passphrase": PASS,
                "mutations": [{"section": "SDx", "action": "Add",
                               "correct_value": "E11.1"}]})
        batch_id = make_batch(client, charts_per=4, allocation_mode="manual",
                              quota_clean=1, quota_manual=2, quota_auto=1)
        client.post(f"/auditor/batches/{batch_id}/run-allocation",
                    json={"run_by": "T",
                          "manual_chart_ids": [c.id for c in library[:4]]})
        rows = client.get(f"/auditor/batches/{batch_id}/plantings",
                          params={"limit": 500}).json()["plantings"]
        counts = {}
        for r in rows:
            counts[r["source"]] = counts.get(r["source"], 0) + 1
        assert counts.get("Clean", 0) == 1
        assert counts.get("Manual", 0) == 2
        assert counts.get("Auto", 0) == 1

    def test_picking_an_unusable_chart_says_why(self, client, db, library):
        """
        Allocating a shorter list than the trainer asked for, without comment,
        is how a batch quietly ends up smaller than intended.
        """
        c = Chart(chart_number="NOKEY7", specialty=Specialty.IP_DRG, category="x",
                  difficulty=Difficulty.BEGINNER, status=ChartStatus.ACTIVE,
                  uploaded_by="t")
        db.add(c)
        db.commit()
        batch_id = make_batch(client, charts_per=2, allocation_mode="manual")
        r = client.post(f"/auditor/batches/{batch_id}/run-allocation",
                        json={"run_by": "T",
                              "manual_chart_ids": [library[0].id, c.id]})
        assert r.status_code == 400
        assert "NOKEY7" in r.json()["detail"]
        assert "answer key" in r.json()["detail"]


class TestQuotaSanity:
    """
    A quota bigger than the allocation does not shrink politely — clean is
    filled first, so it eats everything after it.
    """

    def test_an_over_total_quota_is_refused(self, client, db, library):
        r = client.post("/auditor/batches", json={
            "name": "x", "specialty": "IP-DRG", "created_by": "T",
            "charts_per_auditor": 5, "allocation_mode": "guided",
            "quota_clean": 5, "quota_manual": 1, "quota_auto": 1,
            "auditors": [{"name": "A", "emp_id": "E1"}]})
        assert r.status_code == 400
        assert "add up to 7" in r.json()["detail"]

    def test_an_all_clean_quota_is_refused(self, client, db, library):
        """
        A session with nothing to find reports restraint and no detection at
        all. clean_quota already floors to prevent it; an explicit quota must
        not be a way around it.
        """
        r = client.post("/auditor/batches", json={
            "name": "x", "specialty": "IP-DRG", "created_by": "T",
            "charts_per_auditor": 4, "allocation_mode": "guided",
            "quota_clean": 4, "quota_manual": 0, "quota_auto": 0,
            "auditors": [{"name": "A", "emp_id": "E1"}]})
        assert r.status_code == 400
        assert "nothing to find" in r.json()["detail"]

    def test_a_quota_that_fits_is_accepted(self, client, db, library):
        assert client.post("/auditor/batches", json={
            "name": "x", "specialty": "IP-DRG", "created_by": "T",
            "charts_per_auditor": 5, "allocation_mode": "guided",
            "quota_clean": 2, "quota_manual": 1, "quota_auto": 2,
            "auditors": [{"name": "A", "emp_id": "E1"}]}).status_code == 200

    def test_the_allocator_never_makes_every_chart_clean(self, client, db, library):
        """Belt and braces — creation rejects it, this holds for any caller."""
        import random
        from services.audit_allocation import assign_intents, resolve_quotas

        class _C:
            def __init__(self, i): self.id = i
        drawn = [_C(i) for i in range(5)]
        q = resolve_quotas("guided", 5, 50, 99, 0, 0)
        intents, _ = assign_intents(drawn, q, {}, random.Random(1))
        assert intents.count("clean") < len(drawn)

    def test_an_authored_version_is_used_when_no_quota_asks_otherwise(
            self, client, db, library):
        """
        Automatic mode threw authored error sets away.

        It resolved to manual=0, which assign_intents read as "the trainer
        asked for none", so a chart with a version the trainer had written was
        handed to the generator instead. Only always-use versions survived,
        because those are forced ahead of any quota — which is why the bug was
        invisible to anyone who ticked that box.
        """
        import random
        from services.audit_allocation import assign_intents, resolve_quotas

        class _C:
            def __init__(self, i): self.id = i

        class _V:
            def __init__(self, i): self.id = i; self.always_plant = False

        drawn = [_C(i) for i in range(4)]
        authored = {0: [_V(10)], 1: [_V(11)]}

        for mode in ("auto", "guided", "manual"):
            q = resolve_quotas(mode, 4, 25)
            intents, notes = assign_intents(drawn, q, authored, random.Random(0))
            used = {i for i in (0, 1) if intents[i] == "manual"}
            assert used == {0, 1}, f"{mode} discarded an authored version: {intents}"
            assert not notes, f"{mode} reported a shortfall that was not asked for"

    def test_an_explicit_zero_still_passes_authored_versions_over(
            self, client, db, library):
        """The other half: asking for none must remain a real instruction."""
        import random
        from services.audit_allocation import assign_intents, resolve_quotas

        class _C:
            def __init__(self, i): self.id = i

        class _V:
            def __init__(self, i): self.id = i; self.always_plant = False

        drawn = [_C(i) for i in range(4)]
        q = resolve_quotas("guided", 4, 25, quota_clean=1,
                           quota_manual=0, quota_auto=3)
        intents, _ = assign_intents(drawn, q, {0: [_V(10)]}, random.Random(0))
        assert "manual" not in intents


class TestHandPickedAllocation:
    """
    Hand-picked mirrors the coder screen's manual allocation: the charts the
    trainer ticks go to every auditor, and the per-auditor number is a CAP
    rather than a target.

    The allocation call never sent charts_per_auditor, so the batch's fixed
    value always won — picking eight charts for a batch created at two handed
    each auditor a random two of them and silently dropped the rest.
    """

    def test_every_auditor_gets_all_the_picked_charts(self, client, db, library):
        picked = [c.id for c in library[:5]]
        batch_id = make_batch(client, auditors=("Asha R", "Bo T"), charts_per=2,
                              allocation_mode="manual")
        allocate(client, batch_id, manual_chart_ids=picked,
                 charts_per_auditor=len(picked))

        got = client.get(f"/auditor/batches/{batch_id}").json()["assignments"]
        assert set(got) == {"Asha R", "Bo T"}
        for name, rows in got.items():
            assert {r["chart_id"] for r in rows} == set(picked), name

    def test_the_cap_still_narrows_a_wider_selection(self, client, db, library):
        """A trainer who wants a subset drawn from a wider pick keeps that."""
        picked = [c.id for c in library[:6]]
        batch_id = make_batch(client, auditors=("Asha R",), charts_per=6,
                              allocation_mode="manual")
        allocate(client, batch_id, manual_chart_ids=picked, charts_per_auditor=3)

        rows = client.get(f"/auditor/batches/{batch_id}").json()["assignments"]["Asha R"]
        assert len(rows) == 3
        assert {r["chart_id"] for r in rows} <= set(picked)

    def test_without_an_override_the_batch_default_still_applies(
            self, client, db, library):
        picked = [c.id for c in library[:5]]
        batch_id = make_batch(client, auditors=("Asha R",), charts_per=2,
                              allocation_mode="manual")
        allocate(client, batch_id, manual_chart_ids=picked)
        rows = client.get(f"/auditor/batches/{batch_id}").json()["assignments"]["Asha R"]
        assert len(rows) == 2


class TestPerCycleChartCount:
    """
    The number of charts a cycle hands out is a per-allocation choice in every
    mode, matching the coder screen.

    It used to be fixed at batch creation: the allocation call never sent
    charts_per_auditor, so every cycle in a batch was the same size and the
    only way to run a bigger one was to create another batch.
    """

    @pytest.mark.parametrize("mode", ["auto", "guided", "manual"])
    def test_a_cycle_can_differ_from_the_batch(self, client, db, library, mode):
        batch_id = make_batch(client, auditors=("Asha R",), charts_per=2,
                              allocation_mode=mode)
        kw = {"manual_chart_ids": [c.id for c in library[:5]]} if mode == "manual" else {}
        allocate(client, batch_id, charts_per_auditor=5, **kw)

        rows = client.get(f"/auditor/batches/{batch_id}").json()["assignments"]["Asha R"]
        assert len(rows) == 5

    def test_two_cycles_in_one_batch_can_be_different_sizes(
            self, client, db, library):
        batch_id = make_batch(client, auditors=("Asha R",), charts_per=4,
                              allocation_mode="guided")
        allocate(client, batch_id, charts_per_auditor=2)
        allocate(client, batch_id, charts_per_auditor=3)

        got = client.get(f"/auditor/batches/{batch_id}").json()
        by_cycle = {}
        for r in got["assignments"]["Asha R"]:
            by_cycle.setdefault(r["cycle_id"], 0)
            by_cycle[r["cycle_id"]] += 1
        assert sorted(by_cycle.values()) == [2, 3]

    def test_omitting_it_still_falls_back_to_the_batch(self, client, db, library):
        batch_id = make_batch(client, auditors=("Asha R",), charts_per=3,
                              allocation_mode="guided")
        allocate(client, batch_id)
        rows = client.get(f"/auditor/batches/{batch_id}").json()["assignments"]["Asha R"]
        assert len(rows) == 3


class TestAuditableChartPicker:

    def test_only_charts_that_can_be_audited_are_offered(self, client, db, library):
        """
        The picker used the general chart search, so a trainer could select a
        chart with no answer key and find out only when allocation refused the
        entire run.
        """
        db.add(Chart(chart_number="NOKEY5", specialty=Specialty.IP_DRG,
                     category="Cardiology", difficulty=Difficulty.INTERMEDIATE,
                     status=ChartStatus.ACTIVE, uploaded_by="t"))
        db.commit()
        r = client.get("/auditor/charts", params={"specialty": "IP-DRG"})
        assert r.status_code == 200, r.text
        numbers = {c["chart_number"] for c in r.json()["charts"]}
        assert "NOKEY5" not in numbers
        assert library[0].chart_number in numbers

    def test_it_marks_charts_that_carry_authored_errors(self, client, db, library):
        client.post(f"/auditor/keys/chart/{library[0].id}", json={
            "name": "Curated", "authored_by": "T", "passphrase": PASS,
            "mutations": [{"section": "SDx", "action": "Add", "correct_value": "E11.1"}]})
        charts = client.get("/auditor/charts",
                            params={"specialty": "IP-DRG"}).json()["charts"]
        by_num = {c["chart_number"]: c for c in charts}
        assert by_num[library[0].chart_number]["has_audit_key"] is True
        assert by_num[library[1].chart_number]["has_audit_key"] is False

    def test_it_can_be_filtered(self, client, db, library):
        r = client.get("/auditor/charts",
                       params={"specialty": "IP-DRG", "q": library[0].chart_number})
        assert [c["chart_number"] for c in r.json()["charts"]] == [library[0].chart_number]
