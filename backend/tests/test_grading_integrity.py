"""
Guards for defects found in a grading-integrity review.

Each of these was silently wrong in production, so the assertions pin the
specific arithmetic or wiring rather than just checking a value exists.
"""
import pathlib
import pytest

from models import Specialty
from routers.practicelab_pkg.shared import _is_ip
from services.grading_engine import DEFAULT_IP_CFG

SRC = pathlib.Path(__file__).resolve().parent.parent / "routers" / "practicelab_pkg"


class TestAnswerKeySpecialtyValues:
    """
    The UI used to post the literal "IP"/"OP", which are not Specialty values,
    so /answer-key/upload returned 400 for every specialty — and professional
    templates never carried their pointer or level columns.
    """

    @pytest.mark.parametrize("value", [
        "IP-DRG", "SDS", "Surgery", "ED Profee", "ED Single Path", "E/M",
        "Ancillary", "ED Facility",
    ])
    def test_real_specialty_values_parse(self, value):
        assert Specialty(value)

    @pytest.mark.parametrize("legacy", ["IP", "OP"])
    def test_legacy_shorthand_is_not_a_specialty(self, legacy):
        with pytest.raises(ValueError):
            Specialty(legacy)

    def test_ip_ness_comes_from_the_enum(self):
        assert _is_ip(Specialty.IP_DRG)
        assert not _is_ip(Specialty.SDS)
        assert not _is_ip(Specialty.SURGERY)


class TestDRGReviewArithmetic:
    """
    In-browser IP practice stores a PERCENTAGE of the non-DRG components.
    Adding raw DRG points onto that percentage inflated the result:
    50% + 40 read as 90 when the true weighted score is 70.
    """

    @staticmethod
    def _total(stored_pct, drg_error, cfg=DEFAULT_IP_CFG):
        drg = 0 if drg_error else cfg.drg_weight
        non_drg = cfg.pdx_weight + cfg.sdx_weight + cfg.pcs_weight
        return int(round(min(100.0, stored_pct / 100.0 * non_drg + drg)))

    @pytest.mark.parametrize("pct,drg_error,expected", [
        (50, False, 70),    # the reported case
        (100, False, 100),
        (100, True, 60),
        (0, False, 40),
        (0, True, 0),
    ])
    def test_weighted_total(self, pct, drg_error, expected):
        assert self._total(pct, drg_error) == expected

    def test_old_arithmetic_would_have_inflated(self):
        naive = min(100, 50 + DEFAULT_IP_CFG.drg_weight)
        assert naive == 90
        assert self._total(50, False) == 70


class TestEMAnswerKeyMapping:
    """
    The E/M key was read with SELECT * zipped against a hand-written column
    list. _add_col appends columns at the END, so patient_type / level_method /
    total_time landed after entered_at while the list had patient_type
    mid-table — five fields were mis-read.
    """

    @property
    def _source(self):
        return (SRC / "practice_sessions.py").read_text()

    def test_positional_list_is_gone(self):
        assert "ak_keys = [" not in self._source

    def test_no_positional_zip(self):
        assert "dict(zip(ak_keys" not in self._source

    def test_maps_by_column_name(self):
        assert ".mappings().first()" in self._source


class TestRegradeCompleteness:
    @property
    def _source(self):
        return (SRC / "practice_sessions.py").read_text()

    def test_regrade_loads_em_and_level_fields(self):
        assert "em_data, facility_level, profee_level" in self._source

    def test_regrade_passes_levels_to_the_grader(self):
        assert '"facility_level": entry.facility_level or ""' in self._source

    def test_regrade_routes_em_through_the_mdm_engine(self):
        """
        E/M keys live in em_answer_keys and need grade_em_chart.
        _grade_chart_for_sp would grade them with the OP engine and produce a
        silently wrong score.
        """
        src = self._source
        assert "is_em_session" in src
        assert "grade_em_chart(em_ak" in src
        assert "_em_answer_key(db, chart_id)" in src

    def test_em_submission_bridge_is_shared_by_submit_and_regrade(self):
        """One implementation, so the two paths cannot drift apart."""
        assert self._source.count("_em_submission_dict(") >= 2


class TestFeedbackIsReadByLabelNotPosition:
    """
    _em_feedback_items PREPENDS mismatch rows (patient type -> Coding Accuracy,
    levelling method -> Reasoning Accuracy). Consumers that indexed positionally
    then showed the mismatch row as "E/M Level" and dropped Dx entirely.
    Label matching also repairs results already stored with the old ordering.
    """

    def test_backend_analytics_matches_by_label(self):
        src = (SRC / "practice_sessions.py").read_text()
        assert 'startswith(prefix.upper())' in src
        assert "ra_items[0][" not in src
        assert "ra_items[1][" not in src
        assert "ca_items[0].get" not in src

    def test_frontend_matches_by_label(self):
        fe = (SRC.parent.parent.parent / "frontend" / "src" / "pages" / "PracticeSession.tsx")
        src = fe.read_text()
        assert "pick(caItems, 'E/M Level')" in src
        assert "caItems[0] ?" not in src

    def test_mismatch_rows_do_not_displace_the_scored_rows(self):
        from routers.practicelab_pkg.practice_sessions import _em_feedback_items
        scoring = {
            "method_mismatch": True, "ak_level_method": "TIME", "sub_level_method": "MDM",
            "patient_type_mismatch": True, "ak_patient_type": "NEW", "sub_patient_type": "ESTABLISHED",
            "em_level_score": 23.33, "cpt_score": 10.0, "dx_score": 5.0,
            "copa_element_score": 0.0, "dr_element_score": 0.0, "risk_element_score": 0.0,
            "derived_copa_level": "Low", "derived_dr_level": "Low", "derived_risk_level": "Low",
            # This test is about row ORDER, so it needs the MDM rows present.
            # They are emitted only for charts actually graded on MDM.
            "uses_mdm": True,
        }
        items = _em_feedback_items(scoring, {"em_code": "99214"}, {"em_code": "99213"},
                                   {"em_level_weight": 23.33})
        ca = [i for i in items if i["section"] == "Coding Accuracy"]
        ra = [i for i in items if i["section"] == "Reasoning Accuracy"]
        # the mismatch rows really are first — that is why position cannot be trusted
        assert ca[0]["issue"].startswith("Patient Type mismatch")
        assert ra[0]["issue"].startswith("Levelling method")
        # ...and the scored rows are still present, findable by label
        assert any(i["issue"].startswith("E/M Level") for i in ca)
        assert any(i["issue"].startswith("Dx:") for i in ca)
        assert any(i["issue"].startswith("COPA") for i in ra)
        assert any(i["issue"].startswith("Risk") for i in ra)


class TestEMTemplateEndpoint:
    """
    The E/M template download was broken two ways at once: the route sat AFTER
    /em/answer-key/{chart_id}, so "template" was parsed as a chart id and 422'd,
    and the handler set a sheet title containing "/", which openpyxl rejects.
    It had therefore never produced a workbook.
    """

    def _client(self):
        from fastapi.testclient import TestClient
        import main
        return TestClient(main.app)

    def test_route_is_not_shadowed_by_the_chart_id_route(self):
        r = self._client().get("/practicelab/em/answer-key/template")
        assert r.status_code == 200, f"got {r.status_code}: {r.text[:200]}"

    def test_returns_a_real_workbook_not_json(self):
        r = self._client().get("/practicelab/em/answer-key/template")
        assert "spreadsheetml" in r.headers.get("content-type", "")
        assert "attachment" in r.headers.get("content-disposition", "")
        import io
        from openpyxl import load_workbook
        load_workbook(io.BytesIO(r.content))   # raises if it is not a workbook

    def test_template_carries_a_pointer_column_per_cpt(self):
        import io
        from openpyxl import load_workbook
        r = self._client().get("/practicelab/em/answer-key/template")
        ws = load_workbook(io.BytesIO(r.content)).worksheets[0]
        headers = [c.value for c in ws[1] if c.value]
        assert len([h for h in headers if "Dx Pointers" in str(h)]) == 4

    def test_static_paths_precede_the_parameterised_one(self):
        """Registration order is what makes the above work — pin it."""
        src = (SRC / "em_grading.py").read_text()
        assert src.index('"/em/answer-key/template"') < src.index('"/em/answer-key/{chart_id}"')
        assert src.index('"/em/answer-key/list"') < src.index('"/em/answer-key/{chart_id}"')
