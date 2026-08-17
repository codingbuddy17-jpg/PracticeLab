"""
Specialty wiring, ED Single Path grading, and the Ancillary Dx-only weights.

The wiring checks exist because specialty values are duplicated across several
files with no shared source of truth — adding one to the enum has silently left
dropdowns, prefixes and grading routes behind before.
"""
import io
import pytest
from dataclasses import replace as dc_replace
from openpyxl import load_workbook

from models import Specialty, SPECIALTY_PREFIX, PREFIX_FOR_SPECIALTY
from models.practicelab import IssueType
from routers.practicelab_pkg.shared import (
    _is_ip, _is_ed, _is_surgery, _uses_pointers, _uses_dpo,
    _is_dx_only, _is_single_path,
)
from services.chart_service import detect_specialty_from_prefix
from services.excel_service import generate_answer_key_template, parse_answer_key_upload
from services.grading_engine import (
    grade_ed_single_path, EDSinglePathAnswerKey, EDSinglePathSubmission,
    DEFAULT_EDSP_CFG, EDSinglePathCfg, cfg_from_db,
    grade_op, OPAnswerKey, OPSubmission, DEFAULT_OP_CFG,
)


class TestSpecialtyWiring:
    @pytest.mark.parametrize("spec", [Specialty.SURGERY, Specialty.ED_SINGLE_PATH])
    def test_new_specialties_have_prefixes(self, spec):
        assert spec in PREFIX_FOR_SPECIALTY

    def test_wrong_pointer_issue_type_exists(self):
        assert IssueType.WRONG_POINTER.value == "Wrong_Pointer"

    @pytest.mark.parametrize("filename,expected", [
        ("SURG001.pdf", Specialty.SURGERY),
        ("EDSP014.pdf", Specialty.ED_SINGLE_PATH),
        ("EDP007.pdf", Specialty.ED_PROFEE),
        ("ED002.pdf", Specialty.ED_FACILITY),
        ("SDS003.pdf", Specialty.SDS),
        ("IP010.pdf", Specialty.IP_DRG),
    ])
    def test_chart_numbering_prefix_detection(self, filename, expected):
        """EDSP / EDP / ED must not collide — matching is longest-prefix-first."""
        assert detect_specialty_from_prefix(filename) == expected

    @pytest.mark.parametrize("spec", [
        Specialty.SURGERY, Specialty.ED_PROFEE, Specialty.EM,
    ])
    def test_professional_specialties_use_pointers(self, spec):
        assert _uses_pointers(spec)

    @pytest.mark.parametrize("spec", [
        Specialty.SDS, Specialty.ED_FACILITY, Specialty.IP_DRG, Specialty.ANCILLARY,
        Specialty.ED_SINGLE_PATH,
    ])
    def test_facility_specialties_do_not(self, spec):
        assert not _uses_pointers(spec)

    def test_routing_flags(self):
        assert _is_surgery(Specialty.SURGERY)
        assert not _is_ed(Specialty.SURGERY), "Surgery must not hit the Edits/Denials rubric"
        assert not _is_ip(Specialty.SURGERY)
        assert _is_dx_only(Specialty.ANCILLARY)
        assert _is_single_path(Specialty.ED_SINGLE_PATH)

    @pytest.mark.parametrize("spec", [
        Specialty.IP_DRG, Specialty.ED_FACILITY, Specialty.SDS,
        Specialty.SURGERY, Specialty.ANCILLARY, Specialty.ED_SINGLE_PATH,
    ])
    def test_dpo_eligible_specialties(self, spec):
        assert _uses_dpo(spec)

    @pytest.mark.parametrize("spec", [
        Specialty.ED_PROFEE, Specialty.EM, Specialty.EDITS, Specialty.DENIALS,
    ])
    def test_non_dpo_specialties(self, spec):
        assert not _uses_dpo(spec)


class TestAnswerKeyTemplates:
    def test_professional_template_has_pointer_columns(self):
        ws = load_workbook(io.BytesIO(
            generate_answer_key_template("OP", with_pointers=True))).worksheets[0]
        headers = [c.value for c in ws[1] if c.value]
        assert "CPT_1_DxPointers" in headers
        assert "CPT_10_DxPointers" in headers

    def test_facility_template_has_none(self):
        ws = load_workbook(io.BytesIO(
            generate_answer_key_template("OP", with_pointers=False))).worksheets[0]
        headers = [c.value for c in ws[1] if c.value]
        assert not any("DxPointer" in str(h) for h in headers)

    def test_single_path_template_leads_with_both_levels(self):
        ws = load_workbook(io.BytesIO(generate_answer_key_template(
            "OP", with_pointers=False, single_path=True))).worksheets[0]
        headers = [c.value for c in ws[1] if c.value]
        assert headers[1] == "Facility_ED_Level"
        assert headers[2] == "Profee_ED_Level"
        assert headers[3] == "PDx_Code"
        assert not any("DxPointer" in str(h) for h in headers)

    def test_single_path_round_trip(self):
        data = generate_answer_key_template("OP", with_pointers=False, single_path=True)
        ws = load_workbook(io.BytesIO(data)).worksheets[0]
        ws.cell(2, 1, "EDSP001"); ws.cell(2, 2, "99283"); ws.cell(2, 3, "99284")
        ws.cell(2, 4, "S61.401A"); ws.cell(2, 5, "W26.0XXA")
        cpt_col = 4 + 20 + 1
        ws.cell(2, cpt_col, "12001")
        buf = io.BytesIO(); ws.parent.save(buf)

        row = parse_answer_key_upload(buf.getvalue(), "OP",
                                      with_pointers=False, single_path=True)[0]
        assert row["facility_level"] == "99283"
        assert row["profee_level"] == "99284"
        assert row["pdx_code"] == "S61.401A"
        assert row["cpt"] == [{"code": "12001", "modifier": ""}]


class TestEDSinglePath:
    AK = dict(pdx_code="S61.401A", sdx=[{"code": "W26.0XXA"}],
              cpt=[{"code": "12001", "modifier": "", "pointers": ["A"]}],
              facility_level="99283", profee_level="99284")

    def _grade(self, **overrides):
        sub = dict(pdx_code="S61.401A", sdx=[{"code": "W26.0XXA"}],
                   cpt=[{"code": "12001", "modifier": "", "pointers": ["A"]}],
                   facility_level="99283", profee_level="99284")
        sub.update(overrides)
        return grade_ed_single_path(EDSinglePathAnswerKey(**self.AK),
                                    EDSinglePathSubmission(**sub), DEFAULT_EDSP_CFG)

    def test_all_correct(self):
        res = self._grade()
        assert res.total_score == 100 and res.pass_fail == "PASS"

    def test_facility_right_profee_wrong(self):
        res = self._grade(profee_level="99282")
        assert res.total_score == 80
        assert res.facility_level_ok and not res.profee_level_ok

    def test_profee_right_facility_wrong(self):
        """The two levels are scored independently — that is the whole point."""
        res = self._grade(facility_level="99281")
        assert res.total_score == 80
        assert res.profee_level_ok and not res.facility_level_ok

    def test_pointers_are_not_enforced_for_single_path(self):
        """
        DELIBERATE, confirmed by the product owner: ED Single Path does not
        score diagnosis pointers, even though its profee half is a CMS-1500
        claim and standalone ED Profee does score them.

        The single-path form already asks for two E/M levels plus shared
        diagnoses; adding per-line pointers on top made the interface and the
        grading harder to follow than the skill was worth. ED Profee remains the
        place pointers are practised.

        So ED_SINGLE_PATH is absent from POINTER_SPECIALTIES and
        grade_ed_single_path passes check_pointers=False. This test exists to
        stop a later reviewer "restoring" it — the inconsistency with ED Profee
        is the point, not an oversight.
        """
        res = self._grade(cpt=[{"code": "12001", "modifier": "", "pointers": ["B"]}])
        assert res.total_score == 100
        assert not any(f.issue_type == "Wrong_Pointer" for f in res.feedback)

    def test_custom_config_changes_single_path_weights(self):
        cfg = EDSinglePathCfg(pdx_weight=10, sdx_weight=10,
                              facility_level_weight=30, profee_level_weight=30,
                              cpt_weight=20, pass_threshold=90)
        res = self._grade(facility_level="99281")
        custom = grade_ed_single_path(
            EDSinglePathAnswerKey(**self.AK),
            EDSinglePathSubmission(
                pdx_code="S61.401A", sdx=[{"code": "W26.0XXA"}],
                cpt=[{"code": "12001", "modifier": ""}],
                facility_level="99281", profee_level="99284"),
            cfg,
        )
        assert res.total_score == 80
        assert custom.total_score == 70

    def test_edsp_config_maps_from_db_row(self):
        class Row:
            specialty_type = "EDSP"
            pdx_weight = 10
            sdx_weight = 10
            facility_level_weight = 30
            profee_level_weight = 30
            cpt_weight = 20
            pass_threshold = 85
            overcoding_penalty = False

        cfg = cfg_from_db(Row())
        assert isinstance(cfg, EDSinglePathCfg)
        assert cfg.facility_level_weight == 30
        assert cfg.profee_level_weight == 30
        assert cfg.pass_threshold == 85
        assert not cfg.overcoding_penalty


class TestAncillaryDxOnly:
    """
    grade_op awards the FULL cpt_weight when the key has no CPTs and the coder
    entered none. Sensible in general, but for diagnosis-only work it handed
    every chart a 50-point floor, so nobody could fail on diagnoses alone.
    """
    AK = OPAnswerKey(pdx_code="R51", sdx=[{"code": "G43.909"}], cpt=[])
    DX_ONLY_CFG = dc_replace(DEFAULT_OP_CFG, pdx_weight=50, sdx_weight=50, cpt_weight=0)

    def test_op_config_gives_a_free_floor(self):
        wrong = OPSubmission(pdx_code="WRONG", sdx=[{"code": "WRONG"}], cpt=[])
        assert grade_op(self.AK, wrong, DEFAULT_OP_CFG).total_score == 50

    def test_dx_only_config_can_score_zero(self):
        wrong = OPSubmission(pdx_code="WRONG", sdx=[{"code": "WRONG"}], cpt=[])
        assert grade_op(self.AK, wrong, self.DX_ONLY_CFG).total_score == 0

    def test_dx_only_config_can_score_full(self):
        right = OPSubmission(pdx_code="R51", sdx=[{"code": "G43.909"}], cpt=[])
        assert grade_op(self.AK, right, self.DX_ONLY_CFG).total_score == 100


class TestEDSinglePathNoCPTs:
    """
    DELIBERATE, confirmed by the product owner: when the key lists no additional
    CPTs and the coder enters none, the CPT block scores full.

    Reviewed and kept because it reads as "correctly recognised there was nothing
    to code". It is the same rule that distorted Ancillary — but there CPT was
    50 of 100 and created a floor nobody could fail through, whereas here it is
    20 of 100 and the two independently-scored levels still discriminate.

    This test exists to stop the behaviour being "fixed" by a later reviewer.
    """

    def test_no_cpts_either_side_scores_the_block_full(self):
        ak = EDSinglePathAnswerKey(pdx_code="S61.401A", sdx=[], cpt=[],
                                   facility_level="99283", profee_level="99284")
        sub = EDSinglePathSubmission(pdx_code="S61.401A", sdx=[], cpt=[],
                                     facility_level="99283", profee_level="99284")
        res = grade_ed_single_path(ak, sub, DEFAULT_EDSP_CFG)
        assert res.cpt_score == DEFAULT_EDSP_CFG.cpt_weight
        assert res.total_score == 100

    def test_but_inventing_cpts_still_costs(self):
        """Credit is for recognising there was nothing — not for coding anyway."""
        ak = EDSinglePathAnswerKey(pdx_code="S61.401A", sdx=[], cpt=[],
                                   facility_level="99283", profee_level="99284")
        sub = EDSinglePathSubmission(pdx_code="S61.401A", sdx=[],
                                     cpt=[{"code": "12001", "modifier": ""}],
                                     facility_level="99283", profee_level="99284")
        res = grade_ed_single_path(ak, sub, DEFAULT_EDSP_CFG)
        assert res.cpt_score == 0


class TestAnswerKeyTemplatePerSpecialty:
    """Each specialty's blank template must match what its grading actually uses."""

    def test_dx_only_template_has_no_cpt_columns(self):
        ws = load_workbook(io.BytesIO(
            generate_answer_key_template("OP", dx_only=True))).worksheets[0]
        headers = [str(c.value) for c in ws[1] if c.value]
        assert not any(h.startswith("CPT_") for h in headers)
        assert any(h.startswith("SDx_") for h in headers)

    def test_normal_op_template_still_has_them(self):
        ws = load_workbook(io.BytesIO(
            generate_answer_key_template("OP"))).worksheets[0]
        headers = [str(c.value) for c in ws[1] if c.value]
        assert any(h.startswith("CPT_") for h in headers)

    def test_dx_only_parser_ignores_trailing_cpt_cells(self):
        data = generate_answer_key_template("OP", dx_only=True)
        ws = load_workbook(io.BytesIO(data)).worksheets[0]
        ws.cell(2, 1, "ANC001"); ws.cell(2, 2, "R51"); ws.cell(2, 3, "G43.909")
        buf = io.BytesIO(); ws.parent.save(buf)
        row = parse_answer_key_upload(buf.getvalue(), "OP", dx_only=True)[0]
        assert row["pdx_code"] == "R51"
        assert row["sdx"] == [{"code": "G43.909"}]
        assert row["cpt"] == []


class TestAnswerKeyExportRoundTrip:
    """
    Export used to lump every non-IP specialty into one generic OP sheet, so
    Surgery lost its pointer columns and ED Single Path lost both level columns
    — an export → edit → re-upload cycle silently dropped them.
    """

    class _Chart:
        def __init__(self, number, specialty):
            self.chart_number, self.specialty = number, specialty

    class _Key:
        def __init__(self, **kw):
            self.pdx_code = kw.get("pdx_code", "")
            self.pdx_poa = kw.get("pdx_poa", "")
            self.sdx = kw.get("sdx", [])
            self.pcs = kw.get("pcs", [])
            self.cpt = kw.get("cpt", [])
            self.facility_level = kw.get("facility_level")
            self.profee_level = kw.get("profee_level")

    def _export(self, pairs):
        from services.excel_service import export_all_answer_keys
        return load_workbook(io.BytesIO(export_all_answer_keys(pairs)))

    def test_one_sheet_per_specialty(self):
        wb = self._export([
            (self._Key(pdx_code="J18.9"), self._Chart("IP001", Specialty.IP_DRG)),
            (self._Key(pdx_code="M17.11"), self._Chart("SURG001", Specialty.SURGERY)),
            (self._Key(pdx_code="R51"), self._Chart("ANC001", Specialty.ANCILLARY)),
        ])
        assert set(wb.sheetnames) == {"IP-DRG", "Surgery", "Ancillary"}

    def test_surgery_sheet_keeps_pointer_columns_and_values(self):
        wb = self._export([(
            self._Key(pdx_code="M17.11", sdx=[{"code": "E11.9"}],
                      cpt=[{"code": "27447", "modifier": "RT", "pointers": ["A", "B"]}]),
            self._Chart("SURG001", Specialty.SURGERY))])
        ws = wb["Surgery"]
        headers = [str(c.value) for c in ws[1] if c.value]
        assert "CPT_1_DxPointers" in headers
        assert "A,B" in [str(c.value) for c in ws[2] if c.value]

    def test_single_path_sheet_keeps_both_levels(self):
        wb = self._export([(
            self._Key(pdx_code="S61.401A", facility_level="99283", profee_level="99284"),
            self._Chart("EDSP001", Specialty.ED_SINGLE_PATH))])
        ws = wb["ED Single Path"]
        headers = [str(c.value) for c in ws[1] if c.value]
        assert headers[1] == "Facility_ED_Level" and headers[2] == "Profee_ED_Level"
        assert not any("DxPointer" in h for h in headers)
        vals = [str(c.value) for c in ws[2] if c.value]
        assert "99283" in vals and "99284" in vals

    def test_ancillary_sheet_omits_cpt_columns(self):
        wb = self._export([(self._Key(pdx_code="R51"),
                            self._Chart("ANC001", Specialty.ANCILLARY))])
        headers = [str(c.value) for c in wb["Ancillary"][1] if c.value]
        assert not any(h.startswith("CPT_") for h in headers)

    def test_em_sheet_name_has_the_slash_replaced(self):
        """'E/M' is an illegal Excel sheet title."""
        wb = self._export([(self._Key(pdx_code="E11.9"),
                            self._Chart("EM001", Specialty.EM))])
        assert "E-M" in wb.sheetnames


def test_the_auditable_list_matches_the_frontend_copy():
    """
    The auditor's specialty list exists twice — AUDITABLE_SPECIALTIES on the
    backend and AUDITABLE in the frontend constants — and the specialty-sync
    checker does not cover it. A specialty added to one and not the other is
    either invisible in the UI or offered and then refused by the API.
    """
    import pathlib
    import re
    from routers.auditor_pkg.shared import AUDITABLE_SPECIALTIES

    src = (pathlib.Path(__file__).resolve().parents[2] / "frontend" / "src"
           / "pages" / "auditor" / "constants.ts").read_text()
    block = re.search(r"export const AUDITABLE = \[(.*?)\]", src, re.S).group(1)
    frontend = set(re.findall(r"'([^']+)'", block))
    backend = {s.value for s in AUDITABLE_SPECIALTIES}
    assert frontend == backend, (
        f"only on the backend: {sorted(backend - frontend)}; "
        f"only in the frontend: {sorted(frontend - backend)}")
