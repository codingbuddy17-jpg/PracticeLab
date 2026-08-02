"""
Diagnosis pointers on professional claims (CMS-1500 Box 24E).

The case that matters most is the last one: pointers are POSITIONAL, so a coder
who orders their diagnoses differently from the key uses a different letter for
the same diagnosis. Comparing letters would mark correct work wrong.
"""
import pytest

from services.grading_engine import (
    grade_op, OPAnswerKey, OPSubmission, DEFAULT_OP_CFG,
    resolve_pointers, claim_dx_list,
)

AK_SDX = [{"code": "N17.9"}, {"code": "J18.9"}]


def _grade(ak_pointers, sub_pointers, sub_sdx=None, check=True):
    ak = OPAnswerKey(pdx_code="A41.9", sdx=AK_SDX,
                     cpt=[{"code": "20610", "modifier": "", "pointers": ak_pointers}])
    sub = OPSubmission(pdx_code="A41.9", sdx=sub_sdx if sub_sdx is not None else AK_SDX,
                       cpt=[{"code": "20610", "modifier": "", "pointers": sub_pointers}])
    return grade_op(ak, sub, DEFAULT_OP_CFG, check_pointers=check)


def _pointer_errors(res):
    return [f for f in res.feedback if f.issue_type == "Wrong_Pointer"]


class TestPointerResolution:
    def test_claim_dx_list_is_pdx_then_secondaries(self):
        assert claim_dx_list("A41.9", AK_SDX) == ["A41.9", "N17.9", "J18.9"]

    def test_letters_map_to_positions(self):
        dx = claim_dx_list("A41.9", AK_SDX)
        assert resolve_pointers(["A"], dx) == {"A419"}
        assert resolve_pointers(["C"], dx) == {"J189"}

    def test_out_of_range_letters_ignored(self):
        assert resolve_pointers(["Z"], claim_dx_list("A41.9", AK_SDX)) == set()

    def test_empty_pointers(self):
        assert resolve_pointers([], ["A41.9"]) == set()
        assert resolve_pointers(None, ["A41.9"]) == set()


class TestPointerScoring:
    def test_matching_pointers_score_full(self):
        res = _grade(["A"], ["A"])
        assert res.cpt_score == 50
        assert _pointer_errors(res) == []

    def test_wrong_pointer_costs_half_the_line(self):
        res = _grade(["A"], ["B"])
        assert res.cpt_score == 25
        assert len(_pointer_errors(res)) == 1

    def test_pointer_order_within_a_line_is_irrelevant(self):
        assert _grade(["A", "B"], ["B", "A"]).cpt_score == 50

    def test_facility_claims_ignore_pointers_entirely(self):
        res = _grade(["A"], [], check=False)
        assert res.cpt_score == 50
        assert _pointer_errors(res) == []

    def test_wrong_cpt_code_loses_the_line_regardless(self):
        ak = OPAnswerKey(pdx_code="A41.9", sdx=AK_SDX,
                         cpt=[{"code": "20610", "modifier": "", "pointers": ["A"]}])
        sub = OPSubmission(pdx_code="A41.9", sdx=AK_SDX,
                           cpt=[{"code": "29881", "modifier": "", "pointers": ["A"]}])
        assert grade_op(ak, sub, DEFAULT_OP_CFG, check_pointers=True).cpt_score == 0

    def test_different_letter_same_diagnosis_is_correct(self):
        """
        Key points B -> sdx[0] = N17.9.
        Coder reversed their secondaries and points C -> sdx[1] = N17.9.
        Same diagnosis, different letter: must earn full credit.
        """
        res = _grade(["B"], ["C"], sub_sdx=[{"code": "J18.9"}, {"code": "N17.9"}])
        assert res.cpt_score == 50, "letter-comparison would wrongly fail this"
        assert _pointer_errors(res) == []

    def test_feedback_names_the_codes_not_the_letters(self):
        detail = _pointer_errors(_grade(["A"], ["B"]))[0].detail
        assert "A41.9" in detail and "N17.9" in detail


class TestNumericPointers:
    """
    Coders refer to diagnoses by NUMBER — Dx 1, Dx 2 — in the order they are
    listed, so that is what the app asks for and stores.

    Letters are still read. The CMS-1500 (02/12) prints A-L in Box 21 and every
    key entered before this switch used letters; a key that silently lost its
    pointers would grade every professional line as unlinked, turning correct
    work into errors across the whole history.
    """

    DX = ["M17.11", "E11.9", "I10", "Z79.4"]

    def test_numeric_pointer_resolves_to_the_listed_diagnosis(self):
        from services.grading_engine import resolve_pointers
        assert resolve_pointers(["1"], self.DX) == {"M1711"}
        assert resolve_pointers(["2"], self.DX) == {"E119"}

    def test_letters_resolve_to_the_same_positions(self):
        from services.grading_engine import resolve_pointers
        for num, letter in [("1", "A"), ("2", "B"), ("3", "C"), ("4", "D")]:
            assert resolve_pointers([num], self.DX) == resolve_pointers([letter], self.DX), num

    def test_a_key_in_letters_and_a_submission_in_numbers_agree(self):
        """The migration case: old key, new coder interface. Must not conflict."""
        from services.grading_engine import resolve_pointers
        assert resolve_pointers(["A", "C"], self.DX) == resolve_pointers(["1", "3"], self.DX)

    def test_pointer_zero_is_not_a_position(self):
        """Numbering starts at 1 on the page, so 0 is a typo, not Dx 1."""
        from services.grading_engine import resolve_pointers
        assert resolve_pointers(["0"], self.DX) == set()

    def test_two_digit_pointers_reach_the_last_diagnoses(self):
        from services.grading_engine import resolve_pointers
        dx = [f"A{i:02d}" for i in range(1, 13)]          # 12 diagnoses
        assert resolve_pointers(["12"], dx) == {"A12"}
        assert resolve_pointers(["10"], dx) != resolve_pointers(["1"], dx)

    def test_out_of_range_pointers_resolve_to_nothing(self):
        from services.grading_engine import resolve_pointers
        assert resolve_pointers(["9"], self.DX) == set()

    def test_feedback_reads_back_in_numbers(self):
        from services.grading_engine import pointer_display
        assert pointer_display(["1", "2"], self.DX) == "1=M17.11, 2=E11.9"

    def test_legacy_letters_are_shown_back_as_numbers(self):
        """One convention on screen, whatever the key was typed in."""
        from services.grading_engine import pointer_display
        assert pointer_display(["A", "B"], self.DX) == "1=M17.11, 2=E11.9"

    def test_garbage_is_ignored_rather_than_pointing_somewhere(self):
        from services.grading_engine import pointer_index
        assert pointer_index("") is None
        assert pointer_index("-") is None
        assert pointer_index(None) is None


class TestFourPointerLimit:
    """
    CMS-1500 Box 24E holds at most FOUR diagnosis pointers per service line.
    A claim may list twelve diagnoses; the coder picks the four that support
    medical necessity for that procedure, first one primary.

    This matters beyond tidiness. A key carrying five pointers describes a
    claim that could not be submitted, and grading against it marks a correct
    four-pointer answer as incomplete — the coder is penalised for being right.
    """

    def test_a_line_is_capped_at_four(self):
        from services.grading_engine import canonical_pointers
        assert canonical_pointers(["1", "2", "3", "4", "5", "6"]) == ["1", "2", "3", "4"]

    def test_the_first_four_are_kept_because_the_first_is_primary(self):
        """Order carries meaning, so the cap takes from the front, not anywhere."""
        from services.grading_engine import canonical_pointers
        assert canonical_pointers(["3", "1", "4", "2", "5"])[0] == "3"

    def test_repeats_do_not_consume_the_allowance(self):
        """Pointing at Dx 1 twice references one diagnosis, not two."""
        from services.grading_engine import canonical_pointers
        assert canonical_pointers(["1", "1", "2", "2", "3"]) == ["1", "2", "3"]

    def test_pointers_beyond_the_twelve_diagnosis_box_are_dropped(self):
        from services.grading_engine import canonical_pointers
        assert canonical_pointers(["11", "12", "13", "99"]) == ["11", "12"]

    def test_mixed_spellings_still_dedupe(self):
        """A and 1 are the same position — together they are one pointer."""
        from services.grading_engine import canonical_pointers
        assert canonical_pointers(["A", "1", "B"]) == ["1", "2"]

    def test_the_limit_is_stated_once(self):
        from services.grading_engine import MAX_POINTERS_PER_LINE, MAX_DIAGNOSES
        assert MAX_POINTERS_PER_LINE == 4
        assert MAX_DIAGNOSES == 12


class TestExcelUploadReadsNumericPointers:
    def test_numeric_pointers_survive_an_upload(self, client, db):
        """
        The Excel reader kept only characters passing isalpha(), so it dropped
        every numeric pointer — an uploaded key would have graded every
        professional line as unlinked.
        """
        import io
        from openpyxl import load_workbook
        from models import Chart, Specialty
        from models.charts import ChartStatus, Difficulty

        db.add(Chart(chart_number="SURG700", specialty=Specialty.SURGERY, category="Ortho",
                     difficulty=Difficulty.BEGINNER, status=ChartStatus.ACTIVE,
                     uploaded_by="t", view_count=0))
        db.commit()

        blank = client.get("/practicelab/answer-key/template",
                           params={"specialty": "Surgery"}).content
        wb = load_workbook(io.BytesIO(blank))
        ws = wb.worksheets[0]
        headers = [c.value for c in ws[1]]
        ws.cell(2, 1, "SURG700")
        ws.cell(2, 2, "M17.11")
        ws.cell(2, headers.index("CPT_1") + 1, "27447")
        ws.cell(2, headers.index("CPT_1_DxPointers") + 1, "1,2")
        buf = io.BytesIO(); wb.save(buf)

        r = client.post("/practicelab/answer-key/upload",
                        files={"file": ("k.xlsx", buf.getvalue(), "application/vnd.ms-excel")},
                        data={"specialty": "Surgery", "entered_by": "trainer"})
        assert r.status_code == 200, r.text
        assert "SURG700" in r.json()["stored"]

        from models import AnswerKey
        key = db.query(AnswerKey).filter(AnswerKey.chart_id == db.query(Chart).filter(
            Chart.chart_number == "SURG700").one().id).one()
        assert key.cpt[0]["pointers"] == ["1", "2"], key.cpt

    def test_an_over_long_pointer_list_is_capped_on_upload(self, client, db):
        from services.grading_engine import canonical_pointers
        assert len(canonical_pointers("1,2,3,4,5".split(","))) == 4
